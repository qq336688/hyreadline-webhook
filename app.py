"""
HyRead 維修記錄管理系統 - Flask + Supabase 後端
版本：v1.0（對應雛型 v7）
"""

import os
import re
import bcrypt
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, render_template, session
from supabase_client import create_client, SupabaseHTTP as Client
from dotenv import load_dotenv

load_dotenv()

# ── App 設定 ─────────────────────────────────────────────
app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-secret-change-me')
app.permanent_session_lifetime = timedelta(hours=24)

# ── Supabase 客戶端 ───────────────────────────────────────
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_SECRET_KEY', '')
sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
_ATT_BUCKET = 'repair-attachments'  # Storage bucket for attachments

# ============================================================
# 工具函式
# ============================================================
def now_str():
    return datetime.utcnow().isoformat()

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def check_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def next_serial(key: str, prefix: str, digits: int) -> str:
    """從 system_config 取得下一個流水號並更新"""
    res = sb.table('system_config').select('value').eq('key', key).single().execute()
    current = res.data[0]['value']       # e.g. "N00001"
    num     = int(current[len(prefix):]) # 取數字部分
    new_val = f"{prefix}{str(num + 1).zfill(digits)}"
    sb.table('system_config').update({'value': new_val}).eq('key', key).execute()
    return current

# ============================================================
# 權限裝飾器
# ============================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': '請先登入'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': '請先登入'}), 401
        if not session.get('is_admin'):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return decorated

def check_module(module_key):
    """檢查當前使用者是否有模組存取權"""
    perms = session.get('permissions', {})
    return perms.get(module_key, False)

# ============================================================
# 頁面路由
# ============================================================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'version': '1.0'})

# ============================================================
# 登入 / 登出 / 目前使用者
# ============================================================
@app.route('/api/login', methods=['POST'])
def login():
    data     = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'error': '請填寫帳號與密碼'}), 400

    # 先查使用者
    res = sb.table('users').select(
        'id, username, display_name, password_hash, is_active, group_id'
    ).eq('username', username).execute()

    if not res.data:
        return jsonify({'error': '帳號或密碼錯誤'}), 401

    user = res.data[0]

    if not user['is_active']:
        return jsonify({'error': '帳號已停用，請聯繫管理員'}), 403
    if not check_password(password, user['password_hash']):
        return jsonify({'error': '帳號或密碼錯誤'}), 401

    # 更新最後登入時間 + 遞增登入次數
    _cur_cnt = user.get('login_count') or 0
    sb.table('users').update({'last_login': now_str(), 'login_count': _cur_cnt + 1}).eq('id', user['id']).execute()

    # 另外查權限群組
    grp = {}
    is_admin = False
    perms = {}
    if user.get('group_id'):
        grp_res = sb.table('permission_groups').select('*').eq('id', user['group_id']).execute()
        if grp_res.data:
            grp = grp_res.data[0]
            perms = {k: v for k, v in grp.items() if k.startswith('mod_')}
            is_admin = (grp.get('name') == '系統管理員')

    session.permanent = True
    session['user_id']      = user['id']
    session['username']     = user['username']
    session['display_name'] = user['display_name'] or user['username']
    session['group_id']     = user['group_id']
    session['permissions']  = perms
    session['is_admin']     = is_admin

    return jsonify({
        'id':           user['id'],
        'username':     user['username'],
        'display_name': user['display_name'],
        'permissions':  perms,
        'is_admin':     is_admin,
    })

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/me')
@login_required
def me():
    return jsonify({
        'id':           session['user_id'],
        'username':     session['username'],
        'display_name': session['display_name'],
        'permissions':  session.get('permissions', {}),
        'is_admin':     session.get('is_admin', False),
    })

# ============================================================
# 維修記錄（主模組）
# ============================================================
# ── 欄位對應表（範本欄名 → Supabase 欄位）──────────────────────
REPAIR_IMPORT_COLS = {
    '舊維修編號':     'old_repair_no',
    '富動編號':       'product_number',       # 原「富動/配件編號」
    '阿偉編號':       'awei_number',
    '填表人':         'form_filler',
    '資料來源':       'data_source',
    '填單日期':       'fill_date',
    '維修類型':       'repair_type',
    '型號':           'model',
    'SN碼':           'serial_no',
    '客戶姓名':       'customer_name',
    '帳號':           'customer_account',
    '電話1':          'customer_phone1',
    '電話2':          'customer_phone2',
    '信箱':           'customer_email',
    '地址':           'customer_address',
    '展碁備註':       'ebook_note',
    '展碁通路':       'ebook_channel',
    '福利品':         'is_welfare',
    '發票號碼':       'invoice_no',
    '發票日期':       'invoice_date',
    '歷次維修編號':   'prev_repair_nos',
    '收件包裹':       'received_package',
    '收回日期':       'received_date',
    '原商品出貨日期': 'original_ship_date',
    '訂單資訊':       'order_info',
    '客戶問題備註':   'customer_issue',
    '換機換貨SN':     'exchange_sn',
    '檢測費單號':     'payment_no1',          # v2.9 改名（原付款單號1）
    '檢測費':         'payment_amount1',       # v2.9 改名（原付款金額1）
    '維修費單號':     'payment_no2',          # v2.9 改名（原付款單號2）
    '維修費':         'payment_amount2',       # v2.9 改名（原付款金額2）
    '付款單號備註':   'payment_note',
    '其他備註':       'other_notes',
    '保固與否':       'warranty',
    '故障大項':       'fault_category',
    '故障細項':       'fault_detail',
    '破屏/線條':      'screen_damage',
    '實測故障':       'actual_fault',
    '配件':           '__accessories__',      # 特殊處理：逗號分隔 → list
    '更換零件':       'replaced_parts',
    '更換零件記錄':   '__parts_checklist__',  # 特殊處理：逗號分隔名稱 → JSON
    '維修紀錄':       'repair_record',
    '實收總費用':     'inspection_fee',        # v2.9 改名（原檢測費）；自動加總 = 檢測費+維修費
    # '維修費(repair_fee)' 已廢棄，不再匯入
    # '付款總額(total_payment)' 已廢棄，不再匯入
    '維修員':         'technician',
    '維修完成日期含備註': 'repair_date',    # v2.12 改名（原「維修日期」），改為文字型允許填備註
    '換下壞品':       'bad_part_removed',
    '維修備註':       'repair_notes',
    '帳單系統':       'billing_system',
    '細項統計':       'detail_stats',
    '年度統計':       'annual_stats',
    '委外廠商':       'outsource_vendor',
    '委外請款月份':   'outsource_month',
    '委外金額':       'outsource_amount',
    '進度狀態':       'progress_status',
    '結案方式':       'close_method',
}

# 更換零件記錄：中文名稱 → parts_checklist 欄位名稱對照
PARTS_NAME_MAP = {
    '未更換零件': 'no_part_replace',   '換機': 'machine_exchange',
    '更換屏幕': 'replace_screen',      '更換主板': 'replace_mainboard',
    '更換電池': 'replace_battery',     '更換SD卡座': 'replace_sd_slot',
    '更換天線': 'replace_antenna',     '更換背殼': 'replace_back_cover',
    '更換副板': 'replace_sub_board',   '更換喇叭': 'replace_speaker',
    '更換主板小板排線接副板FPC': 'replace_fpc',
    '更換電源線': 'replace_power_cable','更換電源鍵': 'replace_power_button',
    '更換電源排線': 'replace_power_ribbon','更換螺絲': 'replace_screw',
    '更換SIM卡座': 'replace_sim_tray', '更換側邊FPC': 'replace_side_fpc',
}

@app.route('/api/repair/template')
@login_required
def repair_template():
    """下載維修記錄批次匯入範本"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = '維修記錄範本'
    required = {'填表人','資料來源','填單日期','維修類型','型號','SN碼','進度狀態'}
    headers = list(REPAIR_IMPORT_COLS.keys())
    yellow = PatternFill('solid', fgColor='FFFF00')
    green  = PatternFill('solid', fgColor='C6EFCE')
    bold   = Font(bold=True)
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i)
        cell.font = bold
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill = yellow if h in required else green
    # 說明列（順序對應 REPAIR_IMPORT_COLS）
    # 舊維修編號,富動編號,阿偉編號,填表人,資料來源,填單日期,維修類型,型號,SN碼,
    # 客戶姓名,帳號,電話1,電話2,信箱,地址,展碁備註,展碁通路,福利品,發票號碼,發票日期,
    # 歷次維修編號,收件包裹,收回日期,原商品出貨日期,訂單資訊,客戶問題備註,
    # 換機換貨SN,付款單號1,付款金額1,付款單號2,付款金額2,付款單號備註,其他備註,
    # 保固與否,故障大項,故障細項,破屏/線條,實測故障,配件,更換零件,更換零件記錄,維修紀錄,
    # 檢測費,維修費,維修員,維修日期,換下壞品,維修備註,帳單系統,付款總額,
    # 細項統計,年度統計,委外廠商,委外請款月份,委外金額,進度狀態,結案方式
    notes = [
        '原系統序號','','',
        '*必填','*必填','*必填 YYYY-MM-DD','*必填','*必填','*必填',
        '','','','','','',
        '','','是/否','','YYYY-MM-DD',
        '','','YYYY-MM-DD','文字（可填日期）','','',
        '','','數字','','數字','','',
        '保固內/保固外','','','有/無','',
        '逗號分隔多項','','逗號分隔中文名','',
        '數字','數字','','文字（可填日期+備註）','','',
        '','數字',
        '見下方選項','','','YYYY-MM','數字',
        '見下方選項','見下方選項'
    ]
    ws.append(notes)
    ws.cell(2,1).font = Font(italic=True, color='808080')
    # 範例
    ws.append([
        '1001','','',
        'Stacy','電話','2026-05-17','保固維修','ebook 7','SN123456',
        '王小明','hyread001','0912345678','','user@email.com','台北市中正區',
        '','博客來','否','','',
        '','原廠紙箱','2026-05-18','2025-01-01','','螢幕破損',
        '','','','','','','',
        '保固內','螢幕','破屏','有','螢幕破裂',
        '配件A,配件B','','更換屏幕','更換螢幕完成',
        '0','0','阿偉','2026-05-19 更換屏幕完成','','',
        '','0',
        '','','','','0',
        '已收貨，資料登錄中',''
    ])
    # 選項說明
    ws.append([])
    ws.append(['【進度狀態選項】'])
    for s in ['待收貨，客服建單中','已收貨，資料登錄中','維修，評估檢測中','已結案']:
        ws.append(['', s])
    ws.append(['【結案方式選項（進度狀態=已結案時必填）】'])
    for s in ['原機寄還','原機寄還(已親取)','換機(來回件)','換機(舊換新)',
              '換機(已親取)','換機','配件補寄','已入庫','放棄閱讀器(不寄回)','手動結案','其他']:
        ws.append(['', s])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='維修記錄匯入範本.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/repair/batch-import', methods=['POST'])
@login_required
def batch_import_repair():
    """批次匯入維修記錄（對應範本欄位）"""
    if 'file' not in request.files:
        return jsonify({'error': '請上傳檔案'}), 400
    file = request.files['file']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # 舊欄名相容對照（舊名 → 新名）v2.9 補付款欄位改名
        LEGACY_HEADER_MAP = {
            '保固':    '保固與否',
            '破屏線條': '破屏/線條',
            '付款單號1': '檢測費單號',   # v2.9
            '付款金額1': '檢測費',        # v2.9
            '付款單號2': '維修費單號',   # v2.9
            '付款金額2': '維修費',        # v2.9
            '檢測費（舊）': '實收總費用', # 避免誤用
            '維修日期': '維修完成日期含備註',  # v2.12 改名相容
        }
        # 找出哪些欄有對應
        col_idx = {}
        for i, h in enumerate(headers):
            h = LEGACY_HEADER_MAP.get(h, h)   # 舊名自動轉新名
            if h in REPAIR_IMPORT_COLS:
                col_idx[REPAIR_IMPORT_COLS[h]] = i
        if 'serial_no' not in col_idx and 'form_filler' not in col_idx:
            return jsonify({'error': '找不到必要欄位，請使用範本格式'}), 400

        # 取得已存在的舊維修編號（避免重複）
        existing_res = sb.table('repair_records').select('old_repair_no').execute()
        existing_old = {r['old_repair_no'] for r in (existing_res.data or []) if r.get('old_repair_no')}

        imported = skipped = 0
        errors = []
        pending = []  # 待 INSERT 的 (old_no, rec) 清單
        date_fields    = {'fill_date','invoice_date','received_date'}  # original_ship_date / repair_date 已改為文字型，不做日期轉換
        numeric_fields = {'payment_amount1','payment_amount2','inspection_fee','outsource_amount'}
        # v2.9: inspection_fee / outsource_amount 改為 NUMERIC；total_payment / repair_fee 已廢棄
        # is_welfare 已改為 DB TEXT（v2.8），直接存字串，不再轉 boolean
        # 跳過說明列的關鍵字
        SKIP_HINTS = {'*必填','必填','YYYY','說明','請勿','欄位'}

        for row in ws.iter_rows(min_row=2, values_only=True):  # 第1行標題，第2行起為資料或說明列
            if all((v is None or str(v).strip() == '') for v in row):
                continue
            # 跳過說明列（第一個值含說明關鍵字）
            first_val = str(row[0] or '').strip()
            if any(h in first_val for h in SKIP_HINTS):
                continue

            rec = {}
            for field, idx in col_idx.items():
                val = row[idx] if idx < len(row) else None
                s   = str(val).strip() if val is not None else ''
                if not s:
                    continue
                # ── 特殊欄位：配件（逗號分隔 → list）──────────────
                if field == '__accessories__':
                    rec['accessories'] = [v.strip() for v in s.split(',') if v.strip()]
                    continue
                # ── 特殊欄位：更換零件記錄（逗號分隔名稱 → JSON）──
                if field == '__parts_checklist__':
                    names = [v.strip() for v in s.split(',') if v.strip()]
                    rec['parts_checklist'] = {PARTS_NAME_MAP[n]: True for n in names if n in PARTS_NAME_MAP}
                    continue
                if field in date_fields:
                    parsed = _parse_date(val)
                    if parsed:
                        rec[field] = parsed
                elif field in numeric_fields:
                    try:
                        rec[field] = float(s.replace(',', ''))
                    except ValueError:
                        pass  # 非數字就略過
                else:
                    # datetime 物件（Excel 日期格式）只取日期部分，避免存入 "2023-05-15 00:00:00"
                    if isinstance(val, datetime):
                        rec[field] = val.strftime('%Y-%m-%d')
                    else:
                        rec[field] = s

            # 跳過重複
            old_no = rec.get('old_repair_no', '')
            if old_no and old_no in existing_old:
                skipped += 1
                continue

            rec['created_by']      = session['user_id']
            rec['updated_by']      = session['user_id']
            rec['created_by_name'] = session.get('display_name', session.get('username', ''))
            rec['updated_by_name'] = session.get('display_name', session.get('username', ''))
            rec.setdefault('progress_status', '待收貨，客服建單中')
            pending.append((old_no, rec))

        # ── 一次讀流水號，本地端連續編號，最後一次更新（避免 N×2 次 DB 呼叫超時）──
        sn_res = sb.table('system_config').select('value').eq('key', 'repair_next_no').execute()
        sn_str = (sn_res.data or [{}])[0].get('value', 'N00001')
        sn_num = int(sn_str[1:])   # 去掉 'N' 前綴取數字

        for old_no, rec in pending:
            rec['repair_no'] = f"N{str(sn_num).zfill(5)}"
            sn_num += 1

        # 更新流水號（一次）
        sb.table('system_config').update({'value': f"N{str(sn_num).zfill(5)}"}).eq('key', 'repair_next_no').execute()

        # ── 批次 INSERT（每批 80 筆）──────────────────────────────────
        # PGRST102 修復：同批次所有物件必須有相同的 keys，先收集全部欄位再補 None
        all_keys = set()
        for _, rec in pending:
            all_keys.update(rec.keys())

        def normalize_batch(buf):
            """補齊缺少的 key 為 None，確保同批次 keys 一致"""
            batch_keys = set()
            for r in buf:
                batch_keys.update(r.keys())
            for r in buf:
                for k in batch_keys:
                    r.setdefault(k, None)
            return buf

        BATCH_SIZE = 80
        batch_buf = []
        for old_no, rec in pending:
            batch_buf.append(rec)
            if len(batch_buf) >= BATCH_SIZE:
                try:
                    sb.table('repair_records').insert(normalize_batch(batch_buf)).execute()
                    imported += len(batch_buf)
                    for r in batch_buf:
                        ono = r.get('old_repair_no','')
                        if ono: existing_old.add(ono)
                except Exception as e:
                    errors.append(f"批次錯誤（{len(batch_buf)} 筆）：{str(e)[:80]}")
                batch_buf = []
        # 最後一批
        if batch_buf:
            try:
                sb.table('repair_records').insert(normalize_batch(batch_buf)).execute()
                imported += len(batch_buf)
            except Exception as e:
                errors.append(f"批次錯誤（{len(batch_buf)} 筆）：{str(e)[:80]}")

        return jsonify({'imported': imported, 'skipped': skipped, 'errors': errors[:10],
                        'col_count': len(col_idx), 'matched_cols': list(col_idx.keys())[:10]})
    except Exception as e:
        return jsonify({'error': f'處理失敗：{str(e)}'}), 500

@app.route('/api/repair/records')
@login_required
def list_repair_records():
    p        = request.args
    page     = max(1, int(p.get('page', 1)))
    per_page = min(200, max(10, int(p.get('per_page', 20))))
    offset   = (page - 1) * per_page

    q = sb.table('repair_records').select('*', count='exact')

    # 關鍵字搜尋（v2.9 擴充）
    search = p.get('q', '').strip()
    if search:
        # 附件備註跨表查詢
        try:
            att_res = sb.table('repair_attachments').select('record_id').eq('record_type','repair').ilike('note',f'%{search}%').execute()
            _att_ids = [str(r["record_id"]) for r in (att_res.data or [])]
        except Exception:
            _att_ids = []
        q = q.or_(
            f"repair_no.ilike.%{search}%,"
            f"old_repair_no.ilike.%{search}%,"
            f"awei_number.ilike.%{search}%,"
            f"product_number.ilike.%{search}%,"
            f"prev_repair_nos.ilike.%{search}%,"
            f"customer_name.ilike.%{search}%,"
            f"customer_account.ilike.%{search}%,"
            f"customer_phone1.ilike.%{search}%,"
            f"customer_phone2.ilike.%{search}%,"
            f"customer_email.ilike.%{search}%,"
            f"customer_address.ilike.%{search}%,"
            f"customer_issue.ilike.%{search}%,"
            f"order_info.ilike.%{search}%,"
            f"serial_no.ilike.%{search}%,"
            f"exchange_sn.ilike.%{search}%,"
            f"payment_no1.ilike.%{search}%,"
            f"payment_no2.ilike.%{search}%,"
            f"payment_note.ilike.%{search}%,"
            f"actual_fault.ilike.%{search}%,"
            f"repair_record.ilike.%{search}%,"
            f"replaced_parts.ilike.%{search}%,"
            f"bad_part_removed.ilike.%{search}%,"
            f"repair_notes.ilike.%{search}%,"
            f"other_notes.ilike.%{search}%"
            + (f",id.in.({chr(40)}{chr(44).join(_att_ids)}{chr(41)})" if _att_ids else "")
        )

    # ── 下拉篩選（eq 完全比對）──────────────────────────────
    for field in ['progress_status', 'repair_type', 'model', 'form_filler',
                  'data_source', 'warranty', 'fault_category', 'technician',
                  'outsource_vendor', 'close_method',
                  'is_welfare', 'ebook_note', 'ebook_channel',   # 福利品、展碁備註、展碁通路
                  'fault_detail', 'screen_damage']:               # 故障細項、破屏/線條
        v = p.get(field, '').strip()
        if v:
            q = q.eq(field, v)

    # ── 文字關鍵字（ilike 模糊比對）────────────────────────
    for field, col in [('invoice_no',     'invoice_no'),
                       ('actual_fault',   'actual_fault'),
                       ('replaced_parts', 'replaced_parts'),
                       ('repair_record',  'repair_record'),
                       ('other_notes',    'other_notes')]:
        v = p.get(field, '').strip()
        if v:
            q = q.ilike(col, f'%{v}%')

    # ── 日期區間 ────────────────────────────────────────────
    def _range(col, f_key, t_key, cast=None):
        fv = p.get(f_key, '').strip()
        tv = p.get(t_key, '').strip()
        if cast:
            try: fv = str(cast(fv)) if fv else ''
            except: fv = ''
            try: tv = str(cast(tv)) if tv else ''
            except: tv = ''
        return fv, tv

    # ── 日期區間 ────────────────────────────────────────────
    for col, fk, tk in [
        ('fill_date',     'fill_date_from',    'fill_date_to'),
        ('received_date', 'return_date_from',  'return_date_to'),
        ('repair_date',   'repair_date_from',  'repair_date_to'),
        ('invoice_date',  'inv_date_from',     'inv_date_to'),
    ]:
        fv, tv = _range(col, fk, tk)
        if fv and tv:   q = q.and_cond(f'{col}.gte.{fv},{col}.lte.{tv}')
        elif fv:        q = q.gte(col, fv)
        elif tv:        q = q.lte(col, tv)

    # ── 編號區間（字串比較）────────────────────────────────
    import re as _re
    def _norm_awei(s):
        """A1 → A00001, A15 → A00015, A1001 → A01001（5位補零確保字串比較正確）"""
        s = (s or '').strip().upper()
        m = _re.match(r'^([A-Za-z]+)(\d+)$', s)
        return m.group(1) + m.group(2).zfill(5) if m else s

    awei_from = _norm_awei(p.get('awei_from', ''))
    awei_to   = _norm_awei(p.get('awei_to',   ''))
    if awei_from and awei_to: q = q.and_cond(f'awei_number.gte.{awei_from},awei_number.lte.{awei_to}')
    elif awei_from:           q = q.gte('awei_number', awei_from)
    elif awei_to:             q = q.lte('awei_number', awei_to)

    repno_from    = p.get('repno_from', '').strip()
    repno_to      = p.get('repno_to',   '').strip()
    if repno_from and repno_to: q = q.and_cond(f'repair_no.gte.{repno_from},repair_no.lte.{repno_to}')
    elif repno_from:            q = q.gte('repair_no', repno_from)
    elif repno_to:              q = q.lte('repair_no', repno_to)

    old_from = p.get('old_repno_from', '').strip()
    old_to   = p.get('old_repno_to',   '').strip()
    if old_from and old_to: q = q.and_cond(f'old_repair_no.gte.{old_from},old_repair_no.lte.{old_to}')
    elif old_from:          q = q.gte('old_repair_no', old_from)
    elif old_to:            q = q.lte('old_repair_no', old_to)

    # ── 付款金額區間 ────────────────────────────────────────
    if p.get('amt_min') and p.get('amt_max'):
        try:
            q = q.and_cond(f'payment_amount1.gte.{float(p["amt_min"])},payment_amount1.lte.{float(p["amt_max"])}')
        except ValueError: pass
    else:
        if p.get('amt_min'):
            try: q = q.gte('payment_amount1', float(p['amt_min']))
            except ValueError: pass
        if p.get('amt_max'):
            try: q = q.lte('payment_amount1', float(p['amt_max']))
            except ValueError: pass

    # 排序與分頁
    sort_by  = p.get('sort', 'id')
    sort_asc = p.get('dir', 'desc').lower() == 'asc'
    q = q.order(sort_by, desc=not sort_asc, nullslast=True).range(offset, offset + per_page - 1)

    res = q.execute()
    total = res.count or 0

    return jsonify({
        'total':    total,
        'page':     page,
        'per_page': per_page,
        'pages':    max(1, (total + per_page - 1) // per_page),
        'records':  res.data or [],
    })

def _apply_repair_filters(q, p):
    """將 request.args 的篩選條件套用到 repair_records query builder 上並回傳。
    供 status-counts / export-all 等需要完整篩選但自行控制分頁的端點使用。"""
    search = p.get('q', '').strip()
    if search:
        q = q.or_(
            f"repair_no.ilike.%{search}%,"
            f"old_repair_no.ilike.%{search}%,"
            f"awei_number.ilike.%{search}%,"
            f"product_number.ilike.%{search}%,"
            f"prev_repair_nos.ilike.%{search}%,"
            f"customer_name.ilike.%{search}%,"
            f"customer_account.ilike.%{search}%,"
            f"customer_phone1.ilike.%{search}%,"
            f"customer_phone2.ilike.%{search}%,"
            f"customer_email.ilike.%{search}%,"
            f"customer_address.ilike.%{search}%,"
            f"customer_issue.ilike.%{search}%,"
            f"order_info.ilike.%{search}%,"
            f"serial_no.ilike.%{search}%,"
            f"exchange_sn.ilike.%{search}%,"
            f"payment_no1.ilike.%{search}%,"
            f"payment_no2.ilike.%{search}%,"
            f"payment_note.ilike.%{search}%,"
            f"actual_fault.ilike.%{search}%,"
            f"repair_record.ilike.%{search}%,"
            f"replaced_parts.ilike.%{search}%,"
            f"bad_part_removed.ilike.%{search}%,"
            f"repair_notes.ilike.%{search}%,"
            f"other_notes.ilike.%{search}%"
        )
    for field in ['progress_status', 'repair_type', 'model', 'form_filler',
                  'data_source', 'warranty', 'fault_category', 'technician',
                  'outsource_vendor', 'close_method',
                  'is_welfare', 'ebook_note', 'ebook_channel',
                  'fault_detail', 'screen_damage']:
        v = p.get(field, '').strip()
        if v:
            q = q.eq(field, v)
    for field, col in [('invoice_no', 'invoice_no'), ('actual_fault', 'actual_fault'),
                       ('replaced_parts', 'replaced_parts'), ('repair_record', 'repair_record'),
                       ('other_notes', 'other_notes')]:
        v = p.get(field, '').strip()
        if v:
            q = q.ilike(col, f'%{v}%')
    if p.get('fill_date_from'):   q = q.gte('fill_date',      p['fill_date_from'])
    if p.get('fill_date_to'):     q = q.lte('fill_date',      p['fill_date_to'])
    if p.get('return_date_from'): q = q.gte('received_date',  p['return_date_from'])
    if p.get('return_date_to'):   q = q.lte('received_date',  p['return_date_to'])
    if p.get('repair_date_from'): q = q.gte('repair_date',    p['repair_date_from'])
    if p.get('repair_date_to'):   q = q.lte('repair_date',    p['repair_date_to'])
    if p.get('inv_date_from'):    q = q.gte('invoice_date',   p['inv_date_from'])
    if p.get('inv_date_to'):      q = q.lte('invoice_date',   p['inv_date_to'])
    if p.get('awei_from'):  q = q.gte('awei_number', p['awei_from'].strip().upper())
    if p.get('awei_to'):    q = q.lte('awei_number', p['awei_to'].strip().upper())
    if p.get('repno_from'): q = q.gte('repair_no',   p['repno_from'])
    if p.get('repno_to'):   q = q.lte('repair_no',   p['repno_to'])
    if p.get('amt_min'):
        try: q = q.gte('payment_amount1', float(p['amt_min']))
        except ValueError: pass
    if p.get('amt_max'):
        try: q = q.lte('payment_amount1', float(p['amt_max']))
        except ValueError: pass
    return q


@app.route('/api/repair/status-counts')
@login_required
def repair_status_counts():
    """回傳符合篩選條件的所有記錄，依 progress_status 分組計數（供前端狀態列使用）。
    使用分頁迴圈繞過 Supabase server-side max_rows=1000 限制。"""
    p = request.args
    counts = {}
    offset = 0
    BATCH = 1000
    while True:
        batch = _apply_repair_filters(
            sb.table('repair_records').select('progress_status'), p
        ).order('id', desc=False).range(offset, offset + BATCH - 1).execute().data or []
        for r in batch:
            s = r.get('progress_status') or '（未設定）'
            counts[s] = counts.get(s, 0) + 1
        if len(batch) < BATCH:
            break
        offset += BATCH
        if offset >= 50000:   # 安全上限
            break
    return jsonify({'counts': counts})


@app.route('/api/repair/export')
@login_required
def export_repair_records():
    """匯出維修記錄為 Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file

    p = request.args
    q = sb.table('repair_records').select('*')
    search = p.get('q', '').strip()
    if search:
        q = q.or_(f"repair_no.ilike.%{search}%,customer_name.ilike.%{search}%,serial_no.ilike.%{search}%")
    for field in ['progress_status', 'repair_type', 'model', 'form_filler', 'warranty']:
        v = p.get(field, '').strip()
        if v:
            q = q.eq(field, v)
    res = q.order('id', desc=True).limit(5000).execute()
    records = res.data or []

    wb = Workbook()
    ws = wb.active
    ws.title = '維修記錄'
    headers = [
        '新維修編號','舊維修編號','富動編號','阿偉編號','填表人','資料來源','填單日期','維修類型','型號','SN碼',
        '客戶姓名','帳號','電話1','電話2','信箱','地址',
        '展碁備註','展碁通路','福利品','發票號碼','發票日期',
        '訂單資訊','收件包裹','收回日期','客戶問題備註',
        '換機換貨SN','檢測費單號','檢測費','維修費單號','維修費','付款單號備註','其他備註',
        '配件',
        '保固與否','故障大項','故障細項','破屏/線條','實測故障','更換零件','更換零件記錄','維修紀錄',
        '實收總費用','維修員','維修完成日期含備註','維修備註',
        '帳單系統','細項統計','年度統計','委外廠商','委外請款月份','委外金額',
        '進度狀態','結案方式'
    ]
    field_map = [
        'repair_no','old_repair_no','product_number','awei_number','form_filler','data_source','fill_date','repair_type','model','serial_no',
        'customer_name','customer_account','customer_phone1','customer_phone2','customer_email','customer_address',
        'ebook_note','ebook_channel','is_welfare','invoice_no','invoice_date',
        'order_info','received_package','received_date','customer_issue',
        'exchange_sn','payment_no1','payment_amount1','payment_no2','payment_amount2','payment_note','other_notes',
        'accessories',
        'warranty','fault_category','fault_detail','screen_damage','actual_fault','replaced_parts','parts_checklist','repair_record',
        'inspection_fee','technician','repair_date','repair_notes',
        'billing_system','detail_stats','annual_stats','outsource_vendor','outsource_month','outsource_amount',
        'progress_status','close_method'
    ]
    # 標題列
    header_fill = PatternFill('solid', fgColor='1A5276')
    header_font = Font(bold=True, color='FFFFFF')
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    # 更換零件記錄反查表（field key → 中文名稱）
    PARTS_KEY_TO_NAME = {v: k for k, v in PARTS_NAME_MAP.items()}
    # 資料列
    for r in records:
        row = []
        for f in field_map:
            v = r.get(f, '') or ''
            # 配件 list → 逗號字串
            if f == 'accessories' and isinstance(v, list):
                v = ','.join(v)
            # 更換零件記錄 dict → 逗號分隔中文名稱
            elif f == 'parts_checklist':
                if isinstance(v, dict):
                    v = ','.join(PARTS_KEY_TO_NAME.get(k, k) for k, val in v.items() if val)
                else:
                    v = ''
            row.append(str(v))
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='維修記錄.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/repair/export-all')
@admin_required
def export_all_repair_records():
    """【管理員專用】匯出全部維修記錄（不篩選），用於資料核對"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    from datetime import datetime

    # 全部記錄，依 id 升冪，分頁迴圈繞過 Supabase max_rows=1000
    records = []
    offset = 0
    BATCH = 1000
    while True:
        batch = sb.table('repair_records').select('*').order('id', desc=False).range(offset, offset + BATCH - 1).execute().data or []
        records.extend(batch)
        if len(batch) < BATCH:
            break
        offset += BATCH
        if offset >= 50000:
            break

    wb = Workbook()
    ws = wb.active
    ws.title = '維修記錄全DB'
    headers = [
        '系統ID','新維修編號','舊維修編號','富動編號','阿偉編號',
        '填表人','資料來源','填單日期','維修類型','型號','SN碼',
        '客戶姓名','帳號','電話1','電話2','信箱','地址',
        '展碁備註','展碁通路','福利品','發票號碼','發票日期',
        '訂單資訊','收件包裹','收回日期','客戶問題備註',
        '換機換貨SN','檢測費單號','檢測費','維修費單號','維修費','付款單號備註','其他備註',
        '配件',
        '保固與否','故障大項','故障細項','破屏/線條','實測故障','更換零件','更換零件記錄','維修紀錄',
        '實收總費用','維修員','維修完成日期含備註','維修備註',
        '帳單系統','細項統計','年度統計','委外廠商','委外請款月份','委外金額',
        '進度狀態','結案方式',
        '建立時間','更新時間'
    ]
    field_map = [
        'id','repair_no','old_repair_no','product_number','awei_number',
        'form_filler','data_source','fill_date','repair_type','model','serial_no',
        'customer_name','customer_account','customer_phone1','customer_phone2','customer_email','customer_address',
        'ebook_note','ebook_channel','is_welfare','invoice_no','invoice_date',
        'order_info','received_package','received_date','customer_issue',
        'exchange_sn','payment_no1','payment_amount1','payment_no2','payment_amount2','payment_note','other_notes',
        'accessories',
        'warranty','fault_category','fault_detail','screen_damage','actual_fault','replaced_parts','parts_checklist','repair_record',
        'inspection_fee','technician','repair_date','repair_notes',
        'billing_system','detail_stats','annual_stats','outsource_vendor','outsource_month','outsource_amount',
        'progress_status','close_method',
        'created_at','updated_at'
    ]
    header_fill = PatternFill('solid', fgColor='1A5276')
    header_font = Font(bold=True, color='FFFFFF')
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    PARTS_KEY_TO_NAME = {v: k for k, v in PARTS_NAME_MAP.items()}
    for r in records:
        row = []
        for f in field_map:
            v = r.get(f, '') or ''
            if f == 'accessories' and isinstance(v, list):
                v = ','.join(v)
            elif f == 'parts_checklist':
                if isinstance(v, dict):
                    v = ','.join(PARTS_KEY_TO_NAME.get(k, k) for k, val in v.items() if val)
                else:
                    v = ''
            row.append(str(v))
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'維修記錄_全DB_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True,
                     download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/repair/records/<int:rid>')
@login_required
def get_repair_record(rid):
    res = sb.table('repair_records').select('*').eq('id', rid).single().execute()
    if not res.data:
        return jsonify({'error': '找不到記錄'}), 404
    return jsonify(res.data[0])   # _Result.data 永遠是 list，取 [0] 回傳單筆 dict

# ── repair_records 允許欄位白名單（過濾前端送來的未知欄位，避免 Supabase 400）──
REPAIR_VALID_COLS = {
    # 識別 / 流水號
    'repair_no', 'old_repair_no',
    # 基本資訊
    'product_number', 'awei_number', 'form_filler', 'data_source', 'fill_date',
    'repair_type', 'model', 'serial_no', 'order_info',
    # 展碁 / 發票
    'ebook_note', 'ebook_channel', 'is_welfare', 'invoice_no', 'invoice_date',
    'original_ship_date', 'prev_repair_nos',
    # 客戶
    'customer_name', 'customer_account', 'customer_phone1', 'customer_phone2',
    'customer_email', 'customer_address',
    # 收件 / 客訴
    'received_package', 'received_date', 'customer_issue',
    # 故障
    'fault_category', 'fault_detail', 'screen_damage', 'actual_fault',
    'warranty',
    # 維修
    'replaced_parts', 'parts_checklist', 'repair_record', 'accessories',
    'inspection_fee', 'repair_fee', 'technician', 'repair_date',
    'bad_part_removed', 'repair_notes',
    # 付款 / 結案
    'exchange_sn', 'payment_no1', 'payment_amount1', 'payment_no2', 'payment_amount2',
    'payment_note', 'total_payment', 'billing_system',
    'return_time', 'close_method', 'close_notes', 'other_notes',
    # 統計
    'detail_stats', 'annual_stats', 'all_years',
    # 委外
    'outsource_vendor', 'outsource_month', 'outsource_amount',
    # 進度
    'progress_status',
    # 系統欄位
    'created_by', 'updated_by', 'created_by_name', 'updated_by_name',
}

_NUMERIC_COLS = {
    # HTML type="number" 的欄位，空字串轉 None，避免 PG 22P02
    'payment_amount1', 'payment_amount2',
    'inspection_fee',    # v2.9 改為 NUMERIC（原 TEXT）
    'outsource_amount',  # v2.9 改為 NUMERIC（原 TEXT）
}
_DATE_COLS = {
    'fill_date', 'invoice_date', 'received_date', 'repair_date',
    'return_time',
}
_BOOL_COLS: set = set()
# is_welfare 已在 v2.8 從 boolean 改為 TEXT（ALTER COLUMN），不再需要特殊處理

def _filter_repair_data(data: dict) -> dict:
    """僅保留白名單欄位，防止 Supabase PGRST204/400 未知欄位錯誤；
    數字/日期/布林欄位空字串轉 None，避免 22P02/22007 錯誤。"""
    result = {}
    for k, v in data.items():
        if k not in REPAIR_VALID_COLS:
            continue
        if v == '' and k in (_NUMERIC_COLS | _DATE_COLS | _BOOL_COLS):
            v = None
        result[k] = v
    return result

@app.route('/api/repair/records', methods=['POST'])
@login_required
def create_repair_record():
    data = request.json or {}
    try:
        # 自動產生新維修編號
        repair_no = next_serial('repair_next_no', 'N', 5)
        data['repair_no']        = repair_no
        data['created_by']       = session['user_id']
        data['updated_by']       = session['user_id']
        data['created_by_name']  = session.get('display_name', session.get('username', ''))
        data['updated_by_name']  = session.get('display_name', session.get('username', ''))
        data.pop('id', None)
        clean = _filter_repair_data(data)
        res = sb.table('repair_records').insert(clean).execute()
        if res.data:
            return jsonify({'id': res.data[0]['id'], 'repair_no': repair_no, 'ok': True}), 201
        return jsonify({'error': '建立失敗（無回傳資料）'}), 500
    except Exception as e:
        return jsonify({'error': f'建立失敗：{str(e)[:300]}'}), 500

@app.route('/api/repair/records/<int:rid>', methods=['PUT'])
@login_required
def update_repair_record(rid):
    data = request.json or {}
    try:
        data['updated_by']       = session['user_id']
        data['updated_by_name']  = session.get('display_name', session.get('username', ''))
        data.pop('id', None)
        data.pop('repair_no', None)     # 不允許修改編號
        data.pop('created_by', None)
        data.pop('created_by_name', None)
        data.pop('created_at', None)
        clean = _filter_repair_data(data)
        sb.table('repair_records').update(clean).eq('id', rid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': f'更新失敗：{str(e)[:300]}'}), 500

@app.route('/api/repair/records/<int:rid>', methods=['DELETE'])
@admin_required
def delete_repair_record(rid):
    sb.table('repair_records').delete().eq('id', rid).execute()
    return jsonify({'ok': True})

@app.route('/api/repair/check-invoice')
@login_required
def check_duplicate_invoice():
    """即時檢查發票號碼是否重複"""
    invoice = request.args.get('invoice', '').strip()
    rid     = request.args.get('exclude_id', None)
    if not invoice:
        return jsonify({'duplicates': []})
    q = sb.table('repair_records').select('id, repair_no, customer_name, model').eq('invoice_no', invoice)
    if rid:
        q = q.neq('id', int(rid))
    res = q.execute()
    return jsonify({'duplicates': res.data or []})

@app.route('/api/repair/check-sn')
@login_required
def check_duplicate_sn():
    """即時檢查 SN 碼是否重複"""
    sn  = request.args.get('sn', '').strip()
    rid = request.args.get('exclude_id', None)
    if not sn:
        return jsonify({'duplicates': []})
    q = sb.table('repair_records').select('id, repair_no, customer_name, model').eq('serial_no', sn)
    if rid:
        q = q.neq('id', int(rid))
    res = q.execute()
    return jsonify({'duplicates': res.data or []})

# ============================================================
# 儀表板統計
# ============================================================
@app.route('/api/dashboard/stats')
@login_required
def dashboard_stats():
    # 總數
    total_res = sb.table('repair_records').select('id', count='exact').execute()
    total = total_res.count or 0

    # 各進度狀態數量（RPC 可能不存在，容錯處理）
    try:
        status_res = sb.rpc('count_by_field', {'tbl': 'repair_records', 'col': 'progress_status'}).execute()
    except Exception:
        status_res = type('R', (), {'data': []})()

    # 年度分佈（取 fill_date 年份）
    try:
        year_res = sb.rpc('repair_by_year').execute()
    except Exception:
        year_res = type('R', (), {'data': []})()


    # 各機型年度維修（repair_type='主機維修'）+ 機器瑕疵統計
    # Supabase max-rows=1000，需分頁循環抓完整資料
    model_year_data = {}   # {(year, model): cnt}  -- 只統計主機維修
    screen_dmg_year = {}   # {(year, screen_damage): cnt}  -- 全維修類型
    try:
        _PAGE = 1000
        _offset = 0
        _all_rows = []
        while True:
            _batch = sb.table('repair_records').select('model,fill_date,screen_damage,repair_type') \
                       .range(_offset, _offset + _PAGE - 1).execute()
            _rows = _batch.data or []
            _all_rows.extend(_rows)
            if len(_rows) < _PAGE:
                break
            _offset += _PAGE

        # 取得目前有效的 screen_damage 代碼清單（不加 is_active 過濾，v2.2 後已硬刪除）
        try:
            _sd_codes_res = sb.table('code_options').select('value')                               .eq('field_key', 'screen_damage').execute()
            _valid_sd = {row['value'].strip() for row in (_sd_codes_res.data or []) if row.get('value')}
        except Exception:
            _valid_sd = set()

        for r in _all_rows:
            fd   = r.get('fill_date') or ''
            year = fd[:4] if len(fd) >= 4 else '未知'

            # 各機型年度維修：只抓 repair_type='主機維修'
            if (r.get('repair_type') or '').strip() == '主機維修':
                m_raw = r.get('model') or ''
                m = re.sub(r'^\d+\.', '', m_raw).strip() or '未填'
                model_year_data[(year, m)] = model_year_data.get((year, m), 0) + 1

            # 機器瑕疵統計：只統計目前有效代碼的值，空值或已廢棄值略過
            sd = (r.get('screen_damage') or '').strip()
            if sd and (not _valid_sd or sd in _valid_sd):
                screen_dmg_year[(year, sd)] = screen_dmg_year.get((year, sd), 0) + 1
    except Exception:
        pass

    # 月份趨勢（近 12 個月）
    try:
        month_res = sb.rpc('repair_by_month').execute()
    except Exception:
        month_res = type('R', (), {'data': []})()


    # 委外追蹤進行中
    try:
        tracking_res = sb.table('repair_tracking').select('id', count='exact').is_('repair_complete_cs', 'null').execute()
        tracking_count = tracking_res.count or 0
    except Exception:
        tracking_count = 0

    # 換貨待確認（return_received = 未收回）
    try:
        exchange_res = sb.table('exchange_orders').select('id', count='exact').eq('return_received', '未收回').execute()
        exchange_pending = exchange_res.count or 0
    except Exception:
        exchange_pending = 0

    return jsonify({
        'total':           total,
        'tracking_active': tracking_count,
        'exchange_pending': exchange_pending,
        'by_year':         year_res.data or [],
        'by_month':        month_res.data or [],
        'by_model_year':         [{'year': y, 'model': m, 'cnt': c} for (y, m), c in sorted(model_year_data.items())],
        'by_screen_damage_year': [{'year': y, 'screen_damage': sd, 'cnt': c} for (y, sd), c in sorted(screen_dmg_year.items())],
    })

# ============================================================
# 維修追蹤模組
# ============================================================
_TRK_FILTER_FIELDS = [
    'inspection_fee_received', 'given_to_zhuhan', 'sent_to_awei',
    'fault_parts', 'actual_fault', 'notified_cs_quote', 'payment_url_opened',
    'quoted_customer', 'repair_fee_received', 'notified_awei', 'repair_complete_cs',
]

@app.route('/api/tracking/records')
@login_required
def list_tracking():
    """進行中（交給客服日期為空）"""
    q = sb.table('repair_tracking').select('*').is_('repair_complete_cs', 'null')

    search = request.args.get('q', '').strip()
    if search:
        q = q.or_(
            f"repair_no.ilike.%{search}%,"
            f"model.ilike.%{search}%,"
            f"awei_no.ilike.%{search}%,"
            f"notes.ilike.%{search}%,"
            f"given_to_zhuhan.ilike.%{search}%,"
            f"sent_to_awei.ilike.%{search}%,"
            f"fault_parts.ilike.%{search}%,"
            f"actual_fault.ilike.%{search}%,"
            f"notified_cs_quote.ilike.%{search}%,"
            f"payment_url_opened.ilike.%{search}%,"
            f"quoted_customer.ilike.%{search}%,"
            f"repair_fee_received.ilike.%{search}%,"
            f"notified_awei.ilike.%{search}%,"
            f"repair_complete_cs.ilike.%{search}%"
        )

    # 11 個欄位「有/無」篩選：has_<field>=1 → NOT NULL；=0 → IS NULL
    for field in _TRK_FILTER_FIELDS:
        val = request.args.get(f'has_{field}', '').strip()
        if val == '1':
            q = q.not_.is_(field, 'null')
        elif val == '0':
            q = q.is_(field, 'null')

    q = q.order('created_at', desc=True)
    res = q.execute()
    return jsonify(res.data or [])

@app.route('/api/tracking/history')
@login_required
def list_tracking_history():
    """歷史追蹤（交給客服日期有值），唯讀"""
    q = sb.table('repair_tracking').select('*').not_.is_('repair_complete_cs', 'null')
    q = q.order('repair_complete_cs', desc=True)
    res = q.execute()
    return jsonify(res.data or [])

TRK_DATE_COLS = {
    'inspection_fee_received',   # 唯一保留日期型的欄位
}

def _clean_tracking(data: dict) -> dict:
    """空字串欄位處理：
    - 日期欄位（inspection_fee_received）空字串轉 None
    - repair_complete_cs 空字串轉 None（歷史追蹤依此欄是否為 NULL 判斷進行中/歷史）
    - awei_no 空字串轉 None（避免唯一性約束衝突）
    """
    for col in TRK_DATE_COLS:
        if col in data and (data[col] == '' or data[col] is None):
            data[col] = None
    # 交給客服欄位：空字串轉 None，確保歷史追蹤篩選正確
    if 'repair_complete_cs' in data and (data['repair_complete_cs'] == '' or data['repair_complete_cs'] is None):
        data['repair_complete_cs'] = None
    # 阿偉編號：空字串轉 None（避免多筆空字串觸發唯一性約束）
    if 'awei_no' in data and (data['awei_no'] == '' or data['awei_no'] is None):
        data['awei_no'] = None
    return data

@app.route('/api/tracking/next-awei')
@login_required
def next_awei_no():
    """產生下一個阿偉編號（由前端按鈕主動呼叫，不自動觸發）"""
    no = next_serial('awei_next_no', 'A', 4)
    return jsonify({'awei_no': no})

@app.route('/api/tracking/records', methods=['POST'])
@login_required
def create_tracking():
    data = request.json or {}
    # 阿偉編號：使用前端傳入值（人工填寫或點選「自動產號」），不自動產生
    awei_no = (data.get('awei_no') or '').strip()
    if not awei_no:
        data.pop('awei_no', None)  # 未填時不寫入（留空）
    data['created_by']       = session['user_id']
    data['created_by_name']  = session.get('display_name', session.get('username', ''))
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))

    # 自動帶入機型（依維修編號）
    repair_no = data.get('repair_no', '')
    if repair_no and not data.get('model'):
        rr = sb.table('repair_records').select('model').or_(
            f"repair_no.eq.{repair_no},old_repair_no.eq.{repair_no}"
        ).limit(1).execute()
        if rr.data:
            data['model'] = rr.data[0].get('model', '')

    data.pop('id', None)
    _clean_tracking(data)
    res = sb.table('repair_tracking').insert(data).execute()
    if res.data:
        return jsonify({'id': res.data[0]['id'], 'awei_no': awei_no, 'ok': True}), 201
    return jsonify({'error': '建立失敗'}), 500

@app.route('/api/tracking/records/<int:rid>', methods=['PUT'])
@login_required
def update_tracking(rid):
    data = request.json or {}
    data['updated_by_name'] = session.get('display_name', session.get('username', ''))
    data.pop('id', None)
    # 允許修改阿偉編號（人工填寫）
    data.pop('created_by', None)
    data.pop('created_by_name', None)
    data.pop('created_at', None)
    _clean_tracking(data)
    try:
        sb.table('repair_tracking').update(data).eq('id', rid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': f'更新失敗：{str(e)}'}), 500

@app.route('/api/tracking/template')
@login_required
def tracking_template():
    """下載維修追蹤匯入範本"""
    import io
    from openpyxl import Workbook
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = '維修追蹤範本'
    headers = ['維修編號','型號','阿偉編號','已收檢測費',
               '已給初檢','已寄委外','故障料件','實測故障',
               '已通知客服報價','已開付款網址','已報價客人','已收維修費',
               '已通知委外','交給客服','備註']
    ws.append(headers)
    ws.append(['N00001','ebook 7','A2001','2026-05-01',
               '2026-05-02','2026-05-03','螢幕','破屏',
               '','','','','','','測試資料'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='維修追蹤匯入範本.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/tracking/import', methods=['POST'])
@login_required
def import_tracking():
    """批次匯入維修追蹤 xlsx"""
    if 'file' not in request.files:
        return jsonify({'error': '請上傳檔案'}), 400
    file = request.files['file']
    try:
        rows, skipped, errors = _process_tracking_xlsx(file)
        return jsonify({'imported': rows, 'skipped': skipped, 'errors': errors})
    except Exception as e:
        return jsonify({'error': f'處理失敗：{str(e)}'}), 500

# ============================================================
# 客服換貨模組
# ============================================================
@app.route('/api/exchange/orders')
@login_required
def list_exchange():
    p        = request.args
    page     = max(1, int(p.get('page', 1)))
    per_page = min(200, max(10, int(p.get('per_page', 20))))
    offset   = (page - 1) * per_page

    q = sb.table('exchange_orders').select('*', count='exact')

    search = p.get('q', '').strip()
    if search:
        q = q.or_(
            f"exchange_no.ilike.%{search}%,"
            f"repair_no.ilike.%{search}%,"
            f"customer_contact.ilike.%{search}%,"
            f"order_info.ilike.%{search}%,"
            f"original_sn.ilike.%{search}%,"
            f"item.ilike.%{search}%,"
            f"cs_staff.ilike.%{search}%"
        )

    if p.get('return_remark'):
        q = q.eq('return_remark', p['return_remark'])
    if p.get('fill_date_from'):
        q = q.gte('fill_date', p['fill_date_from'])
    if p.get('fill_date_to'):
        q = q.lte('fill_date', p['fill_date_to'])
    if p.get('item'):
        q = q.eq('item', p['item'])
    if p.get('shipping_method'):
        q = q.eq('shipping_method', p['shipping_method'])
    if p.get('exchange_status'):
        q = q.eq('exchange_status', p['exchange_status'])
    # exchange_status_filter：歷史客服換貨模組固定傳 '倉管已確認'
    if p.get('exchange_status_filter'):
        q = q.eq('exchange_status', p['exchange_status_filter'])

    _ALLOWED_EXC_SORT = {'id', 'exchange_no', 'fill_date', 'cs_staff', 'shipping_method',
                         'shipping_staff', 'exchange_status', 'return_remark'}
    sort_field = p.get('sort', 'id')
    if sort_field not in _ALLOWED_EXC_SORT:
        sort_field = 'id'
    sort_desc = (p.get('dir', 'desc').lower() != 'asc')
    q = q.order(sort_field, desc=sort_desc).range(offset, offset + per_page - 1)
    res = q.execute()
    total = res.count or 0

    return jsonify({
        'total': total, 'page': page, 'per_page': per_page,
        'pages': max(1, (total + per_page - 1) // per_page),
        'records': res.data or [],
    })

@app.route('/api/exchange/status-counts')
@login_required
def exchange_status_counts():
    """回傳符合篩選條件的換貨記錄，依 exchange_status 分組計數。
    使用分頁迴圈繞過 Supabase max_rows=1000 限制。"""
    p = request.args
    counts = {}
    offset = 0
    BATCH = 1000
    while True:
        q = sb.table('exchange_orders').select('exchange_status')
        search = p.get('q', '').strip()
        if search:
            q = q.or_(
                f"exchange_no.ilike.%{search}%,"
                f"repair_no.ilike.%{search}%,"
                f"customer_contact.ilike.%{search}%,"
                f"order_info.ilike.%{search}%,"
                f"original_sn.ilike.%{search}%,"
                f"item.ilike.%{search}%,"
                f"cs_staff.ilike.%{search}%"
            )
        if p.get('return_remark'):
            q = q.eq('return_remark', p['return_remark'])
        if p.get('fill_date_from'):
            q = q.gte('fill_date', p['fill_date_from'])
        if p.get('fill_date_to'):
            q = q.lte('fill_date', p['fill_date_to'])
        if p.get('exchange_status_filter'):
            q = q.eq('exchange_status', p['exchange_status_filter'])
        batch = q.range(offset, offset + BATCH - 1).execute()
        rows = batch.data or []
        for r in rows:
            s = r.get('exchange_status') or '（未填）'
            counts[s] = counts.get(s, 0) + 1
        if len(rows) < BATCH:
            break
        offset += BATCH
    total = sum(counts.values())
    return jsonify({'counts': counts, 'total': total})


@app.route('/api/repair/lookup')
@login_required
def repair_lookup():
    """依維修編號查詢機型（供維修追蹤自動帶入）"""
    repair_no = request.args.get('repair_no', '').strip()
    if not repair_no:
        return jsonify({'error': '請輸入維修編號'}), 400
    res = sb.table('repair_records').select('repair_no, model, awei_number, actual_fault').or_(
        f"repair_no.eq.{repair_no},old_repair_no.eq.{repair_no}"
    ).limit(1).execute()
    if not res.data:
        return jsonify({'error': '找不到維修編號'}), 404
    r = res.data[0]
    return jsonify({
        'repair_no':    r.get('repair_no', '') or '',
        'model':        r.get('model', '') or '',
        'awei_number':  r.get('awei_number', '') or '',
        'actual_fault': r.get('actual_fault', '') or '',
    })

@app.route('/api/exchange/lookup')
@login_required
def exchange_lookup():
    """依維修編號（新/舊皆可）帶入客戶資料"""
    repair_no = request.args.get('repair_no', '').strip()
    if not repair_no:
        return jsonify({'error': '請輸入維修編號'}), 400
    res = sb.table('repair_records').select(
        'repair_no, serial_no, customer_name, customer_issue, '
        'customer_phone1, customer_address, is_welfare, other_notes, order_info, '
        'model, accessories'
    ).or_(
        f"repair_no.eq.{repair_no},"
        f"old_repair_no.eq.{repair_no}"
    ).limit(1).execute()
    if not res.data:
        return jsonify({'error': '找不到維修編號'}), 404
    r = res.data[0]
    # accessories 可能是 list 或 JSON 字串，轉為逗號分隔文字
    acc_raw = r.get('accessories', '') or ''
    if isinstance(acc_raw, list):
        acc_text = ', '.join(str(a) for a in acc_raw if a)
    else:
        acc_text = str(acc_raw)
    return jsonify({
        'repair_no':        r.get('repair_no', '') or '',
        'serial_no':        r.get('serial_no', '') or '',
        'customer_name':    r.get('customer_name', '') or '',
        'customer_issue':   r.get('customer_issue', '') or '',
        'customer_phone1':  r.get('customer_phone1', '') or '',
        'customer_address': r.get('customer_address', '') or '',
        'is_welfare':       bool(r.get('is_welfare', False)),
        'other_notes':      r.get('other_notes', '') or '',
        'order_info':       r.get('order_info', '') or '',
        'model':            r.get('model', '') or '',
        'accessories':      acc_text,
    })

@app.route('/api/exchange/orders', methods=['POST'])
@login_required
def create_exchange():
    data = request.json or {}
    exchange_no = next_serial('exchange_next_no', 'C', 4)
    data['exchange_no']      = exchange_no
    data['created_by']       = session['user_id']
    data['updated_by']       = session['user_id']
    data['created_by_name']  = session.get('display_name', session.get('username', ''))
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    data.pop('id', None)
    res = sb.table('exchange_orders').insert(data).execute()
    if res.data:
        return jsonify({'id': res.data[0]['id'], 'exchange_no': exchange_no, 'ok': True}), 201
    return jsonify({'error': '建立失敗'}), 500

@app.route('/api/exchange/orders/<int:oid>', methods=['PUT'])
@login_required
def update_exchange(oid):
    data = request.json or {}
    data['updated_by']       = session['user_id']
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    data.pop('id', None)
    data.pop('exchange_no', None)   # 不允許修改編號
    data.pop('created_by', None)
    data.pop('created_by_name', None)
    data.pop('created_at', None)
    sb.table('exchange_orders').update(data).eq('id', oid).execute()
    return jsonify({'ok': True})

@app.route('/api/exchange/orders/<int:oid>', methods=['DELETE'])
@admin_required
def delete_exchange(oid):
    sb.table('exchange_orders').delete().eq('id', oid).execute()
    return jsonify({'ok': True})

@app.route('/api/exchange/export-all')
@admin_required
def export_all_exchange():
    """【管理員專用】匯出全部客服換貨記錄（不篩選），用於資料核對"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    from datetime import datetime

    # 全量讀取，分頁迴圈繞過 Supabase max_rows=1000
    records = []
    offset = 0
    BATCH = 1000
    while True:
        batch = sb.table('exchange_orders').select('*').order('id', desc=False).range(offset, offset + BATCH - 1).execute().data or []
        records.extend(batch)
        if len(batch) < BATCH:
            break
        offset += BATCH
        if offset >= 50000:
            break

    wb = Workbook()
    ws = wb.active
    ws.title = '客服換貨全DB'
    headers = [
        '系統ID','換貨編號','填表日期','客服人員','維修編號','品項(客服)','訂單資訊(客服)',
        '訂購人-客戶問題描述','訂購人-電話-收件地址','福利品','原SN',
        '出貨方式','客服寄出備註','是否拆封','是否拆封備註','預計出貨日期','出貨人員',
        '換貨SN資訊','出貨收回','收回狀態','出貨備註','換貨狀態'
    ]
    field_map = [
        'id','exchange_no','fill_date','cs_staff','repair_no','item','order_info',
        'customer_desc','customer_contact','welfare_product','original_sn',
        'shipping_method','shipping_remark','unpack_video','unpack_remark',
        'expected_ship_date','shipping_staff',
        'system_process','return_received','return_remark','shipping_notes','exchange_status'
    ]
    header_fill = PatternFill('solid', fgColor='1F618D')
    header_font = Font(bold=True, color='FFFFFF')
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for r in records:
        ws.append([str(r.get(f, '') or '') for f in field_map])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True,
                     download_name=f'客服換貨全DB_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/exchange/template')
@login_required
def exchange_template():
    """下載客服換貨批次匯入範本"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = '客服換貨範本'
    headers = [
        '換貨編號(客服)', '填表日期(客服)', '客服人員', '維修編號(客服)',
        '品項(客服)', '訂單資訊(客服)', '訂購人-客戶問題描述', '訂購人-電話-收件地址',
        '福利品(客服)', '原SN', '出貨方式', '客服寄出備註',
        '是否拆封', '是否拆封備註', '預計出貨日期(客服)',
        '出貨人員(客服)', '換貨SN資訊', '出貨收回', '收回狀態', '出貨備註', '換貨狀態'
    ]
    required = {'換貨編號(客服)', '換貨狀態'}
    yellow = PatternFill('solid', fgColor='FFFF00')
    green  = PatternFill('solid', fgColor='C6EFCE')
    bold   = Font(bold=True)
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i)
        cell.font = bold
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill = yellow if h in required else green
    # 說明列
    notes = [
        '*必填，如 C1001', 'YYYY-MM-DD', '客服人員姓名', '對應維修編號',
        '產品品項', '訂單資訊', '客戶問題描述', '訂購人/電話/地址',
        '是/否', 'SN碼', '出貨方式選項', '出貨備註文字',
        '是否拆封選項', '拆封備註', '文字，如 2026-05-20',
        '出貨人員姓名', '系統處理狀態', '出貨收回文字', '收回狀態選項', '家羽備註', '*必填，換貨狀態選項'
    ]
    ws.append(notes)
    ws.cell(2, 1).font = Font(italic=True, color='808080')
    # 範例列
    ws.append([
        'C1001', '2026-05-19', 'Stacy', 'N00001',
        'ebook 7', 'ORD-12345', '螢幕破損', '王小明/0912345678/台北市',
        '否', 'SN123456', '宅配', '',
        '是', '', '2026-05-20',
        '家羽', '', '已收回', '', '', '倉管已確認'
    ])
    ws.append([])
    ws.append(['【欄位說明】'])
    ws.append(['換貨編號(客服)', '必填，C 開頭數字，如 C1001；已存在則跳過'])
    ws.append(['出貨收回', '自由文字，直接填寫收回狀況（如：已收回、未收回等）'])
    ws.append(['收回狀態', '請對照代碼管理中「出貨收回(換貨)」的選項值'])
    ws.append(['出貨方式', '請對照代碼管理中「出貨方式」的選項值'])
    ws.append(['是否拆封', '請對照代碼管理中「是否拆封」的選項值'])
    ws.append(['換貨狀態', '必填。請對照代碼管理中「換貨狀態(換貨)」的選項值，如：已遞件申請、商品已提供，待回收、倉管已確認'])
    # 欄寬
    col_widths = [16, 14, 10, 12, 10, 16, 20, 24, 6, 12, 10, 14, 8, 12, 14, 10, 14, 10, 14, 12, 12]
    for col_letter, w in zip('ABCDEFGHIJKLMNOPQRSTU', col_widths):
        ws.column_dimensions[col_letter].width = w
    ws.row_dimensions[1].height = 30
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='客服換貨匯入範本.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/exchange/import', methods=['POST'])
@login_required
def import_exchange():
    """批次匯入客服換貨 xlsx"""
    if 'file' not in request.files:
        return jsonify({'error': '請上傳檔案'}), 400
    file = request.files['file']
    try:
        rows, skipped, errors = _process_exchange_xlsx(file)
        return jsonify({'imported': rows, 'skipped': skipped, 'errors': errors})
    except Exception as e:
        return jsonify({'error': f'處理失敗：{str(e)}'}), 500

# ============================================================
# 代碼管理
# ============================================================
@app.route('/api/codes')
@login_required
def list_all_codes():
    # ⚠️ 重要：此處故意不加 is_active 過濾，請勿擅自加入
    # 原因1：Python True 傳給 PostgREST 會變成 eq.True（大寫），無法匹配布林欄位，會導致全部代碼消失
    # 原因2：代碼刪除為軟刪除（is_active=False），管理介面需要看到全部記錄（含已停用）才能正確顯示
    # 若未來確需過濾，正確寫法為 .eq('is_active', 'true')（小寫字串），並同步更新前端 loadCodes 重試邏輯
    res = sb.table('code_options').select('*').order('field_key').order('sort_order').execute()
    result = {}
    for row in (res.data or []):
        fk = row['field_key']
        if fk not in result:
            result[fk] = {'label': row['field_label'], 'options': []}
        result[fk]['options'].append({'id': row['id'], 'value': row['value'], 'sort_order': row['sort_order']})
    return jsonify(result)

@app.route('/api/codes/<field_key>')
@login_required
def list_codes_by_field(field_key):
    res = sb.table('code_options').select('*').eq('field_key', field_key).order('sort_order').execute()
    return jsonify([r['value'] for r in (res.data or [])])

@app.route('/api/codes/template')
@admin_required
def code_template():
    """下載代碼管理匯入範本（含現有全部資料，未建資料欄位也會列出）"""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file

    # 系統定義的所有欄位（與前端 CODE_FIELDS_MAP 保持一致）
    # 格式：(field_key, 欄位名稱, 使用模組)
    ALL_FIELDS = [
        ('form_filler',      '填表人(客服)',    '維修記錄'),
        ('data_source',      '資料來源(客服)',  '維修記錄'),
        ('repair_type',      '維修類型(客服)',  '維修記錄'),
        ('model',            '型號(客服)',      '維修記錄、維修追蹤'),
        ('accessories',      '配件(客服)',      '維修記錄'),
        ('zhanqi_notes',     '展碁備註',        '維修記錄'),
        ('zhanqi_channel',   '展碁通路',        '維修記錄'),
        ('welfare_product',  '福利品',          '維修記錄、客服換貨'),
        ('package_contents', '收件包裹內容',    '維修記錄'),
        ('progress_status',  '進度狀態(客服)',  '維修記錄'),
        ('close_method',     '結案方式',        '維修記錄'),
        ('warranty',         '保固與否',        '維修記錄'),
        ('fault_category',   '故障大項',        '維修記錄'),
        ('detail_category',  '故障細項',        '維修記錄'),
        ('repair_staff',     '維修員',          '維修記錄'),
        ('parts_checklist',  '更換零件記錄',    '維修記錄'),
        ('billing_system',   '帳單系統(Stacy)', '維修記錄'),
        ('detail_stats',     '細項統計(Stacy)', '維修記錄'),
        ('screen_damage',    '破屏/線條',       '維修記錄'),
        ('outsource_vendor', '委外廠商',        '維修記錄'),
        ('cs_staff',         '客服人員',        '客服換貨'),
        ('shipping_staff',   '出貨人員',        '客服換貨'),
        ('shipping_method',  '出貨方式(換貨)',  '客服換貨'),
        ('unpack_video',     '是否拆封(換貨)',  '客服換貨'),
        ('return_received',  '出貨收回(換貨)',  '客服換貨'),
        ('exchange_status',  '換貨狀態(換貨)',  '客服換貨'),
        ('exchange_item',    '換貨品項',         '客服換貨'),
        ('welfare_cancel',   '是否取消(福利品)', '福利品模組'),
        # ── 退換貨檢測模組 ──────────────────────────────────
        ('lib_removed',   '圖書館版移除',   '退換檢測'),
        ('open_closed',   '封閉/開放',      '退換檢測'),
        ('discontinued',  '是否停產',       '退換檢測'),
        ('welfare_grade', '福利等級',        '退換檢測'),
        ('welfare_sold',  '福利機售出狀態',  '退換檢測'),
        ('final_stock',   '放現貨/料件更換', '退換檢測'),
    ]

    # 從 Supabase 取得所有現有資料
    res = sb.table('code_options').select('*').order('field_key').order('sort_order').execute()
    # 依 field_key 分組
    existing = {}
    for r in (res.data or []):
        existing.setdefault(r['field_key'], []).append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = '代碼選項'

    # 表頭（第一欄為使用模組）
    headers = ['使用模組', '欄位代碼', '欄位名稱', '選項值', '排序']
    hfill  = PatternFill('solid', fgColor='1A5276')
    hfont  = Font(bold=True, color='FFFFFF')
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')

    empty_fill  = PatternFill('solid', fgColor='FFF3CD')   # 淡黃：尚無資料欄位
    empty_font  = Font(color='856404', italic=True)

    for fk, label, module in ALL_FIELDS:
        rows = existing.get(fk)
        if rows:
            for r in rows:
                ws.append([module, r['field_key'], r['field_label'], r['value'], r['sort_order']])
        else:
            # 欄位有定義但 Supabase 尚無資料 → 填入提示列（淡黃底）
            row_idx = ws.max_row + 1
            ws.append([module, fk, label, '（尚無選項，請填入後匯入）', ''])
            for cell in ws[row_idx]:
                cell.fill = empty_fill
                cell.font = empty_font

    ws.append([])
    ws.append(['', '【說明】欄位代碼與欄位名稱須完整填寫；排序數字越小越優先；淡黃列為尚未建立選項的欄位；使用模組欄位匯入時系統會自動忽略'])

    for col, w in zip('ABCDE', [20, 24, 24, 36, 8]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    today = datetime.now().strftime('%Y%m%d')
    return send_file(buf, as_attachment=True,
                     download_name=f'代碼管理_現有資料_{today}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/codes/import', methods=['POST'])
@admin_required
def import_codes():
    """批次匯入代碼選項（mode=upsert 合併 / replace 覆蓋）"""
    if 'file' not in request.files:
        return jsonify({'error': '請上傳檔案'}), 400
    file = request.files['file']
    mode = request.form.get('mode', 'upsert')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        def find_col(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n in h: return i
            return None
        ci_key   = find_col(['欄位代碼', 'field_key'])
        ci_label = find_col(['欄位名稱', 'field_label'])
        ci_val   = find_col(['選項值', 'value'])
        ci_sort  = find_col(['排序', 'sort_order'])
        if ci_key is None or ci_val is None:
            return jsonify({'error': '找不到「欄位代碼」或「選項值」欄位，請使用範本格式'}), 400
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all((v is None or str(v).strip() == '') for v in row): continue
            fk  = str(row[ci_key]  or '').strip()
            val = str(row[ci_val]  or '').strip()
            if not fk or not val: continue
            if val == '（尚無選項，請填入後匯入）': continue   # 跳過範本佔位列
            lbl  = str(row[ci_label] or fk).strip() if ci_label is not None else fk
            sort = int(row[ci_sort]) if ci_sort is not None and row[ci_sort] is not None and str(row[ci_sort]).strip().isdigit() else 0
            rows.append({'field_key': fk, 'field_label': lbl, 'value': val, 'sort_order': sort, 'is_active': True})
        if not rows:
            return jsonify({'error': '檔案中找不到有效資料'}), 400

        # ── 模式說明 ────────────────────────────────────────────────────────────
        # upsert（合併，預設）：安全。新選項新增、已存在的更新排序，不刪任何資料。
        #   → 日常新增/修改選項請永遠使用此模式。
        #
        # replace（覆蓋）：⚠️ 高風險，請謹慎使用。
        #   步驟1：先將檔案內出現的 field_key 之現有選項全部標記 is_active=False（停用）
        #   步驟2：再逐筆 upsert 檔案內的選項（is_active 復原為 True）
        #   風險：若步驟1完成但步驟2因網路中斷失敗，舊選項會停留在 is_active=False
        #         雖然 list_all_codes 目前不過濾 is_active（所以管理介面還看得到），
        #         但未來若過濾邏輯有變動，停用的選項會從 UI 消失。
        #   建議：replace 模式僅在需要「完整重建某欄位所有選項順序」時使用，
        #         且匯入前務必先用「下載範本」備份現有資料。
        # ────────────────────────────────────────────────────────────────────────
        if mode == 'replace':
            for key in set(r['field_key'] for r in rows):
                sb.table('code_options').update({'is_active': False}).eq('field_key', key).execute()
        inserted = updated = 0
        for row in rows:
            ex = sb.table('code_options').select('id').eq('field_key', row['field_key']).eq('value', row['value']).execute()
            if ex.data:
                sb.table('code_options').update({
                    'sort_order': row['sort_order'], 'field_label': row['field_label'], 'is_active': True
                }).eq('id', ex.data[0]['id']).execute()
                updated += 1
            else:
                sb.table('code_options').insert(row).execute()
                inserted += 1
        return jsonify({'inserted': inserted, 'updated': updated, 'total': len(rows)})
    except Exception as e:
        return jsonify({'error': f'處理失敗：{str(e)}'}), 500

@app.route('/api/codes', methods=['POST'])
@admin_required
def create_code():
    data = request.json or {}
    res  = sb.table('code_options').insert(data).execute()
    return jsonify({'ok': True, 'id': res.data[0]['id']}), 201

@app.route('/api/codes/<int:cid>', methods=['PUT'])
@admin_required
def update_code(cid):
    data = request.json or {}
    data.pop('id', None)
    data.pop('field_key', None)    # 不允許更改欄位類型
    data.pop('created_at', None)
    sb.table('code_options').update(data).eq('id', cid).execute()
    return jsonify({'ok': True})

@app.route('/api/codes/<int:cid>', methods=['DELETE'])
@admin_required
def delete_code(cid):
    try:
        # ── 取得代碼資訊（用 execute 不用 single，避免找不到時拋 406）──
        code_res = sb.table('code_options').select('field_key, value').eq('id', cid).execute()
        if not code_res.data:
            return jsonify({'error': '找不到此代碼'}), 404
        field_key = code_res.data[0]['field_key']
        value     = code_res.data[0]['value']
    except Exception as e:
        return jsonify({'error': f'查詢代碼失敗：{str(e)}'}), 500

    # ── field_key → [(資料表, DB欄位, 顯示名稱), ...] ──────────
    # 注意：部分 field_key 與 DB 欄位名稱不同（見 CLAUDE.md「代碼欄位混淆」）
    FIELD_USAGE_MAP = {
        'form_filler':     [('repair_records',  'form_filler',      '維修記錄')],
        'data_source':     [('repair_records',  'data_source',      '維修記錄')],
        'repair_type':     [('repair_records',  'repair_type',      '維修記錄')],
        'model':           [('repair_records',  'model',            '維修記錄'),
                            ('repair_tracking', 'model',            '維修追蹤')],
        'accessories':     [('repair_records',  'accessories',      '維修記錄')],
        'zhanqi_notes':    [('repair_records',  'ebook_note',       '維修記錄')],
        'zhanqi_channel':  [('repair_records',  'ebook_channel',    '維修記錄')],
        'welfare_product': [('repair_records',  'welfare_product',  '維修記錄'),
                            ('exchange_orders', 'welfare_product',  '客服換貨'),
                            ('welfare_products','welfare_type',      '福利品模組')],
        'package_contents':[('repair_records',  'received_package', '維修記錄')],
        'progress_status': [('repair_records',  'progress_status',  '維修記錄')],
        'close_method':    [('repair_records',  'close_method',     '維修記錄')],
        'warranty':        [('repair_records',  'warranty',         '維修記錄')],
        'fault_category':  [('repair_records',  'fault_category',   '維修記錄')],
        'detail_category': [('repair_records',  'fault_detail',     '維修記錄')],
        'repair_staff':    [('repair_records',  'repair_staff',     '維修記錄')],
        'billing_system':  [('repair_records',  'billing_system',   '維修記錄')],
        'detail_stats':    [('repair_records',  'detail_stats',     '維修記錄')],
        'screen_damage':   [('repair_records',  'screen_damage',    '維修記錄')],
        'outsource_vendor':[('repair_records',  'outsource_vendor', '維修記錄')],
        'cs_staff':        [('exchange_orders', 'cs_staff',         '客服換貨')],
        'shipping_staff':  [('exchange_orders', 'shipping_staff',   '客服換貨')],
        'shipping_method': [('exchange_orders', 'shipping_method',  '客服換貨')],
        'unpack_video':    [('exchange_orders', 'unpack_video',     '客服換貨')],
        'return_received': [('exchange_orders', 'return_received',  '客服換貨')],
        'exchange_status': [('exchange_orders', 'exchange_status',  '客服換貨')],
        'exchange_item':   [('exchange_orders', 'item',            '客服換貨')],
        'welfare_cancel':  [('welfare_products', 'is_cancelled',    '福利品模組')],
        # ── 退換貨檢測模組 ──────────────────────────────────
        'lib_removed':    [('inspection_records', 'lib_removed',   '退換檢測')],
        'open_closed':    [('inspection_records', 'open_closed',   '退換檢測')],
        'discontinued':   [('inspection_records', 'discontinued',  '退換檢測')],
        'welfare_grade':  [('inspection_records', 'welfare_grade', '退換檢測')],
        'welfare_sold':   [('inspection_records', 'welfare_sold',  '退換檢測')],
        'final_stock':    [('inspection_records', 'final_stock',   '退換檢測')],
    }

    # ── 查詢各資料表使用量 ───────────────────────────────────────
    in_use_parts = []
    for table, column, label in FIELD_USAGE_MAP.get(field_key, []):
        try:
            cnt_res = sb.table(table).select('id', count='exact').eq(column, value).execute()
            cnt = cnt_res.count or 0
            if cnt > 0:
                in_use_parts.append(f'「{label}」{cnt} 筆')
        except Exception:
            pass  # 查詢失敗時忽略，不阻擋刪除

    if in_use_parts:
        return jsonify({
            'error': f'此代碼尚有記錄 {", ".join(in_use_parts)} 在使用，請先移除該選項，才能刪除！'
        }), 409

    # ── 確認無使用，執行刪除 ─────────────────────────────────────
    sb.table('code_options').delete().eq('id', cid).execute()
    return jsonify({'ok': True})

# ============================================================
# 帳號管理
# ============================================================
@app.route('/api/accounts/users')
@admin_required
def list_users():
    res = sb.table('users').select('id, username, display_name, is_active, last_login, login_count, created_at, group_id, permission_groups(name)').order('id').execute()
    return jsonify(res.data or [])

@app.route('/api/accounts/users', methods=['POST'])
@admin_required
def create_user():
    data        = request.json or {}
    username    = data.get('username', '').strip()
    password    = data.get('password', '').strip()
    display_name = data.get('display_name', '').strip()
    group_id    = data.get('group_id')
    if not username or not password:
        return jsonify({'error': '帳號與密碼為必填'}), 400
    payload = {
        'username':      username,
        'display_name':  display_name,
        'password_hash': hash_password(password),
        'group_id':      group_id,
        'is_active':     True,
    }
    try:
        res = sb.table('users').insert(payload).execute()
        return jsonify({'ok': True, 'id': res.data[0]['id']}), 201
    except Exception as e:
        return jsonify({'error': f'帳號已存在或建立失敗：{str(e)}'}), 409

@app.route('/api/accounts/users/batch', methods=['POST'])
@admin_required
def batch_create_users():
    """批次新增帳號：前端已解析並去重，後端再做一次帳號唯一性檢查"""
    data  = request.json or {}
    users = data.get('users', [])
    if not users:
        return jsonify({'error': '沒有可匯入的資料'}), 400

    # 取現有帳號清單
    existing_res = sb.table('users').select('username').execute()
    existing = {r['username'] for r in (existing_res.data or [])}

    imported = 0
    skipped  = 0
    errors   = []

    for u in users:
        username = (u.get('username') or '').strip()
        password = (u.get('password') or '').strip()
        if not username or not password:
            skipped += 1; continue
        if username in existing:
            skipped += 1; continue
        import re as _re
        if not _re.match(r'^[A-Za-z0-9!@#$%^&*()\-_+=\[\]{}|;:,.<>?]{8,12}$', password):
            errors.append(f'帳號「{username}」密碼格式錯誤（需 8~12 位，可含英數及特殊符號），跳過')
            skipped += 1; continue
        payload = {
            'username':      username,
            'display_name':  (u.get('display_name') or '').strip(),
            'password_hash': hash_password(password),
            'group_id':      u.get('group_id') or None,
            'is_active':     u.get('is_active', True),
        }
        try:
            sb.table('users').insert(payload).execute()
            existing.add(username)
            imported += 1
        except Exception as e:
            errors.append(f'帳號「{username}」建立失敗：{str(e)[:60]}')
            skipped += 1

    return jsonify({'imported': imported, 'skipped': skipped, 'errors': errors})

@app.route('/api/accounts/users/<int:uid>', methods=['PUT'])
@admin_required
def update_user(uid):
    data = request.json or {}
    payload = {
        'display_name': data.get('display_name', ''),
        'group_id':     data.get('group_id'),
        'is_active':    data.get('is_active', True),
    }
    if data.get('password'):
        payload['password_hash'] = hash_password(data['password'])
    sb.table('users').update(payload).eq('id', uid).execute()
    return jsonify({'ok': True})

@app.route('/api/accounts/users/<int:uid>', methods=['DELETE'])
@admin_required
def delete_user(uid):
    if uid == session.get('user_id'):
        return jsonify({'error': '不能刪除自己的帳號'}), 400
    sb.table('users').delete().eq('id', uid).execute()
    return jsonify({'ok': True})

@app.route('/api/accounts/groups')
@login_required
def list_groups():
    res = sb.table('permission_groups').select('*').order('id').execute()
    return jsonify(res.data or [])

# permission_groups 允許寫入的欄位白名單
_GROUP_ALLOWED_COLS = {
    'name',
    'mod_dashboard', 'mod_repair_records', 'mod_repair_query', 'mod_exchange',
    'mod_exchange_view', 'mod_exchange_history',
    'mod_welfare', 'mod_welfare_view',
    'mod_tracking', 'mod_tracking_view', 'mod_history',
    'mod_refund',
    'mod_product_view', 'mod_product_edit', 'mod_product_inspect',
    'mod_accounts', 'mod_permissions', 'mod_code_mgmt',
    'tab_view_t1', 'tab_edit_t1', 'tab_view_t2', 'tab_edit_t2',
    'tab_view_t4', 'tab_edit_t4', 'tab_view_t5', 'tab_edit_t5',
    'tab_view_t6', 'tab_edit_t6', 'tab_view_t7', 'tab_edit_t7',
    'tab_view_exc_wh', 'tab_edit_exc_wh',
    'mod_inspection', 'mod_inspection_view',
    'mod_sn', 'mod_sn_view',
    'mod_sn_upload',
    'mod_staff_work',
}

@app.route('/api/accounts/groups', methods=['POST'])
@admin_required
def create_group():
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '群組名稱為必填'}), 400
    # 只保留白名單欄位，防止 PGRST204
    payload = {k: v for k, v in data.items() if k in _GROUP_ALLOWED_COLS}
    try:
        res = sb.table('permission_groups').insert(payload).execute()
        return jsonify({'ok': True, 'id': res.data[0]['id']}), 201
    except Exception as e:
        return jsonify({'error': f'建立失敗：{str(e)}'}), 409

@app.route('/api/accounts/groups/<int:gid>', methods=['PUT'])
@admin_required
def update_group(gid):
    data = request.json or {}
    # 只保留白名單欄位，防止 PGRST204
    data = {k: v for k, v in data.items() if k in _GROUP_ALLOWED_COLS}
    # 各模組唯讀/主模組由管理員自行決定，不強制連動
    try:
        sb.table('permission_groups').update(data).eq('id', gid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': f'更新失敗：{str(e)}'}), 500

@app.route('/api/accounts/groups/<int:gid>', methods=['DELETE'])
@admin_required
def delete_group(gid):
    if gid == 1:
        return jsonify({'error': '不能刪除系統管理員群組'}), 400
    sb.table('permission_groups').delete().eq('id', gid).execute()
    return jsonify({'ok': True})

# ============================================================
# xlsx 批次匯入（內部函式）
# ============================================================
def _parse_date(val):
    """支援多種日期格式"""
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    # Excel 序號
    try:
        n = float(s)
        if 40000 < n < 50000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=n)).strftime('%Y-%m-%d')
    except ValueError:
        pass
    return None

def _process_tracking_xlsx(file):
    import openpyxl, re
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # 欄位模糊對應（同前端 matchHeader：去除括號與空白後做子字串比對）
    # 格式：'關鍵字': 'db欄位'，一個 db 欄位可對應多個關鍵字
    COL_RULES = [
        (['維修編號'],                          'repair_no'),
        (['機型', '型號'],                       'model'),
        (['阿偉編號'],                           'awei_no'),
        (['已收檢測費', '收檢測費'],              'inspection_fee_received'),
        (['已給初檢', '給初檢', '已給竹涵', '給竹涵'], 'given_to_zhuhan'),
        (['已寄委外', '寄委外', '已寄阿偉', '寄阿偉'], 'sent_to_awei'),
        (['故障料件', '料件'],                   'fault_parts'),
        (['實測故障'],                           'actual_fault'),
        (['已通知客服報價', '通知客服報價'],       'notified_cs_quote'),
        (['已開付款網址', '付款網址'],            'payment_url_opened'),
        (['已報價客人', '報價客人'],              'quoted_customer'),
        (['已收維修費', '收維修費'],              'repair_fee_received'),
        (['已通知委外', '通知委外', '已通知阿偉', '通知阿偉'], 'notified_awei'),
        (['交給客服'],                           'repair_complete_cs'),
        (['備註', '聯絡進度'],                   'notes'),
    ]
    DATE_FIELDS = {'inspection_fee_received'}

    def _norm(s):
        """去除空白與括號內容，轉小寫，用於模糊比對"""
        return re.sub(r'\s|\(.*?\)|\（.*?\）', '', str(s)).lower()

    def _match(header, candidates):
        h = _norm(header)
        return any(h == _norm(c) or _norm(c) in h or h in _norm(c) for c in candidates)

    col_idx = {}
    for i, h in enumerate(headers):
        for candidates, field in COL_RULES:
            if field not in col_idx and _match(h, candidates):
                col_idx[field] = i
                break

    if 'repair_no' not in col_idx:
        return 0, 0, ['找不到「維修編號」欄位']

    # 取得已存在的維修編號
    existing = {r['repair_no'] for r in (sb.table('repair_tracking').select('repair_no').execute().data or [])}

    imported = skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        repair_no = str(row[col_idx['repair_no']] or '').strip()
        if not repair_no:
            continue
        if repair_no in existing:
            skipped += 1
            continue

        rec = {'repair_no': repair_no, 'created_by_name': '批次匯入', 'updated_by_name': '批次匯入'}
        for field, idx in col_idx.items():
            if field == 'repair_no':
                continue
            val = row[idx] if idx < len(row) else None
            if field in DATE_FIELDS:
                rec[field] = _parse_date(val)
            else:
                # datetime 物件只取日期部分，避免顯示 00:00:00
                from datetime import datetime as _dt, date as _date
                if isinstance(val, (_dt, _date)):
                    rec[field] = val.strftime('%Y-%m-%d')
                else:
                    rec[field] = str(val).strip() if val is not None and val != '' else None

        # 自動帶入機型（用 limit(1) 避免 .single() 在找不到時回 406 讓匯入整批炸掉）
        if not rec.get('model'):
            try:
                rr = sb.table('repair_records').select('model').eq('repair_no', repair_no).limit(1).execute()
                if rr.data:
                    rec['model'] = rr.data[0].get('model', '')
            except Exception:
                pass   # 找不到機型不影響匯入，留空即可

        _clean_tracking(rec)   # 確保 repair_complete_cs 空字串轉 None
        try:
            sb.table('repair_tracking').insert(rec).execute()
            existing.add(repair_no)
            imported += 1
        except Exception as e:
            errors.append(f"{repair_no}: {str(e)}")

    return imported, skipped, errors


def _process_exchange_xlsx(file):
    import openpyxl, re
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    DATE_FIELDS = {'fill_date'}   # expected_ship_date 已改為文字型，不做日期轉換
    VLOOKUP_FIELDS = {'order_info', 'customer_desc', 'customer_contact', 'welfare_product', 'original_sn'}

    def _norm(s):
        """去除空白與括號內容，轉小寫"""
        return re.sub(r'\s|\(.*?\)|\（.*?\）', '', str(s)).lower()

    # 優先序規則：含「家羽」的先處理，避免「出貨備註」誤判
    def _map_header(h_orig):
        h = h_orig.lower()
        h_n = _norm(h_orig)
        # ── 含「家羽」的欄位（舊範本相容）──────────────────
        if '家羽' in h or 'jiayu' in h:
            if '備註' in h or '系統' in h or 'note' in h:
                return 'shipping_notes' if '備註' in h else 'system_process'
        # ── 換貨編號 ─────────────────────────────────────
        if '換貨編號' in h: return 'exchange_no'
        # ── 填表日期 ─────────────────────────────────────
        if '填表日期' in h: return 'fill_date'
        # ── 客服 / 出貨人員 ───────────────────────────────
        if ('客服' in h or 'cs' in h) and '人員' in h: return 'cs_staff'
        if h_n in ('客服',): return 'cs_staff'
        if '出貨人員' in h: return 'shipping_staff'
        # ── 維修編號 ─────────────────────────────────────
        if '維修編號' in h: return 'repair_no'
        # ── 品項 ─────────────────────────────────────────
        if '品項' in h and '訂' not in h: return 'item'
        # ── 訂單 ─────────────────────────────────────────
        if '訂單資訊' in h: return 'order_info'
        if '問題描述' in h or '客戶問題' in h: return 'customer_desc'
        if '電話' in h and ('地址' in h or '收件' in h): return 'customer_contact'
        # ── 福利品 / SN ───────────────────────────────────
        if '福利品' in h: return 'welfare_product'
        if '原sn' in h or ('sn' in h and '原' in h): return 'original_sn'
        # ── 出貨方式 ─────────────────────────────────────
        if '出貨方式' in h and '備註' not in h: return 'shipping_method'
        # ── 客服寄出備註 / 出貨備註（非家羽）──────────────
        if '客服寄出備註' in h: return 'shipping_remark'
        if '出貨方式備註' in h: return 'shipping_remark'
        if '客服寄出' in h: return 'shipping_remark'
        # ── 是否拆封（備註要先判斷）──────────────────────
        if ('拆封' in h or '拆膜' in h) and '備註' in h: return 'unpack_remark'
        if '拆封' in h or '拆膜' in h: return 'unpack_video'
        # ── 預計出貨日期 ─────────────────────────────────
        if '預計出貨' in h: return 'expected_ship_date'
        # ── 出貨系統處理 / 換貨SN資訊 ───────────────────────
        if '換貨sn' in h or 'sn資訊' in h: return 'system_process'
        if ('系統處理' in h or '出貨系統' in h) and '家羽' not in h: return 'system_process'
        # ── 出貨備註（新欄位名，對應 shipping_notes）──────────
        if '出貨備註' in h and '家羽' not in h and '客服' not in h and '寄出' not in h: return 'shipping_notes'
        # ── 出貨收回（狀態/備註先判斷）───────────────────
        if '出貨收回備註' in h or '收回備註' in h or '收回狀態' in h: return 'return_remark'
        if '出貨收回' in h: return 'return_received'
        # ── 換貨狀態 ─────────────────────────────────
        if '換貨狀態' in h: return 'exchange_status'
        return None

    col_idx = {}
    for i, h in enumerate(headers):
        field = _map_header(h)
        if field and field not in col_idx:
            col_idx[field] = i
        # 精確補漏：針對無法模糊比對的特殊欄名
        FALLBACK = {
            '換貨編號': 'exchange_no', '填表日期': 'fill_date',
            '客服': 'cs_staff', '維修編號': 'repair_no', '品項': 'item',
        }
        if not field and h in FALLBACK and FALLBACK[h] not in col_idx:
            col_idx[FALLBACK[h]] = i

    if 'exchange_no' not in col_idx:
        return 0, 0, ['找不到「換貨編號(客服)」欄位']

    existing = {r['exchange_no'] for r in (sb.table('exchange_orders').select('exchange_no').execute().data or [])}

    imported = skipped = 0
    errors = []
    BATCH_SIZE = 80
    batch = []

    def flush():
        nonlocal imported
        if not batch:
            return
        sb.table('exchange_orders').insert(batch).execute()
        imported += len(batch)
        batch.clear()

    ALL_EXC_FIELDS = [
        'exchange_no','fill_date','cs_staff','repair_no','item','order_info',
        'customer_desc','customer_contact','welfare_product','original_sn',
        'shipping_method','shipping_remark','unpack_video','unpack_remark',
        'expected_ship_date','shipping_staff','system_process',
        'return_received','return_remark','shipping_notes','exchange_status',
    ]

    for row in ws.iter_rows(min_row=2, values_only=True):
        exc_idx = col_idx.get('exchange_no', -1)
        exchange_no = str(row[exc_idx] if exc_idx >= 0 and exc_idx < len(row) else '').strip()
        if not exchange_no:
            continue
        if exchange_no in existing:
            skipped += 1
            continue
        existing.add(exchange_no)  # 即時更新，避免同批次重複
        rec = {}
        for field, ci in col_idx.items():
            if ci >= len(row):
                continue
            val = row[ci]
            if val is None:
                rec[field] = None
                continue
            if field in ('fill_date', 'expected_ship_date'):
                rec[field] = _parse_date(val)
            else:
                from datetime import datetime as _dt, date as _date
                if isinstance(val, (_dt, _date)):
                    rec[field] = val.strftime('%Y-%m-%d')
                else:
                    rec[field] = str(val).strip() if str(val).strip() else None
        # 補齊所有欄位（避免 PGRST102 JSON keys 不一致）
        for f in ALL_EXC_FIELDS:
            if f not in rec:
                rec[f] = None
        batch.append(rec)
        if len(batch) >= BATCH_SIZE:
            try:
                flush()
            except Exception as ex:
                errors.append(str(ex))
                return imported, skipped, errors

    try:
        flush()
    except Exception as ex:
        errors.append(str(ex))

    return imported, skipped, errors


# ============================================================
# 退費追蹤模組（v2.18）
# ============================================================

_ALL_REFUND_FIELDS = [
    'refund_no','repair_no','old_repair_no','form_filler','order_date',
    'return_type','customer_name','customer_account','order_amount','order_no',
    'payment_method','invoice_no','refund_amount','contact_source',
    'cancel_date','cancel_reason',
    'discount_invoice','ntt1_cancel','transaction_cancel','cancel_points',
    'notify_customer','return_warehouse','claim_apply',
    'claim_warehouse_no','claim_reconciled',
]

_REFUND_NUMERIC_COLS = {'order_amount', 'refund_amount'}


def _filter_refund_data(d):
    out = {}
    for k, v in d.items():
        if k in _REFUND_NUMERIC_COLS:
            out[k] = None if (v == '' or v is None) else v
        else:
            out[k] = v
    return out


@app.route('/api/refund/orders')
@login_required
def list_refund_orders():
    q        = request.args.get('q', '').strip()
    date_f   = request.args.get('date_from', '')
    date_t   = request.args.get('date_to', '')
    ret_type = request.args.get('return_type', '')
    page     = max(1, int(request.args.get('page', 1)))
    per_page = min(200, max(10, int(request.args.get('per_page', 50))))
    sort_f   = request.args.get('sort', 'refund_no')
    sort_d   = request.args.get('dir', 'desc')

    ALLOWED_SORT = {
        'refund_no','repair_no','order_date','cancel_date',
        'form_filler','return_type','customer_name','invoice_no',
    }
    if sort_f not in ALLOWED_SORT:
        sort_f = 'refund_no'

    def _build_qb():
        qb = sb.table('refund_tracking').select('*')
        if q:
            qb = qb.or_(
                f'refund_no.ilike.%{q}%,repair_no.ilike.%{q}%,'
                f'old_repair_no.ilike.%{q}%,customer_name.ilike.%{q}%,'
                f'customer_account.ilike.%{q}%,invoice_no.ilike.%{q}%,'
                f'order_no.ilike.%{q}%,cancel_reason.ilike.%{q}%,'
                f'contact_source.ilike.%{q}%'
            )
        if date_f:
            qb = qb.gte('order_date', date_f)
        if date_t:
            qb = qb.lte('order_date', date_t)
        if ret_type:
            qb = qb.eq('return_type', ret_type)
        return qb

    count_res = _build_qb().execute()
    total = len(count_res.data or [])
    offset = (page - 1) * per_page
    res = _build_qb().order(sort_f, desc=(sort_d == 'desc')).range(offset, offset + per_page - 1).execute()

    return jsonify({
        'records': res.data or [],
        'total': total,
        'page': page,
        'pages': max(1, -(-total // per_page)),
    })


@app.route('/api/refund/orders/<int:rid>', methods=['GET'])
@login_required
def get_refund_order(rid):
    res = sb.table('refund_tracking').select('*').eq('id', rid).execute()
    if not res.data:
        return jsonify({'error': '找不到記錄'}), 404
    return jsonify(res.data[0])


@app.route('/api/refund/orders', methods=['POST'])
@login_required
def create_refund_order():
    data = request.json or {}
    payload = _filter_refund_data({k: v for k, v in data.items()
                                   if k in _ALL_REFUND_FIELDS and k != 'refund_no'})
    payload['refund_no'] = next_serial('refund_next_no', 'RF', 4)
    payload['created_at'] = now_str()
    payload['updated_at'] = now_str()
    payload['created_by']       = session['user_id']
    payload['updated_by']       = session['user_id']
    payload['created_by_name']  = session.get('display_name', session.get('username', ''))
    payload['updated_by_name']  = session.get('display_name', session.get('username', ''))
    res = sb.table('refund_tracking').insert(payload).execute()
    if not res.data:
        return jsonify({'error': '新增失敗'}), 500
    return jsonify(res.data[0])


@app.route('/api/refund/orders/<int:rid>', methods=['PUT'])
@login_required
def update_refund_order(rid):
    data = request.json or {}
    payload = _filter_refund_data({k: v for k, v in data.items()
                                   if k in _ALL_REFUND_FIELDS and k != 'refund_no'})
    payload['updated_at'] = now_str()
    payload['updated_by']       = session['user_id']
    payload['updated_by_name']  = session.get('display_name', session.get('username', ''))
    sb.table('refund_tracking').update(payload).eq('id', rid).execute()
    return jsonify({'ok': True})


@app.route('/api/refund/orders/<int:rid>', methods=['DELETE'])
@admin_required
def delete_refund_order(rid):
    sb.table('refund_tracking').delete().eq('id', rid).execute()
    return jsonify({'ok': True})


@app.route('/api/refund/lookup')
@login_required
def refund_lookup():
    repair_no = request.args.get('repair_no', '').strip()
    if not repair_no:
        return jsonify({'error': '請輸入維修編號'}), 400
    res = sb.table('repair_records').select(
        'repair_no,old_repair_no,customer_name,customer_account,'
        'customer_phone1,customer_email,model,serial_no,invoice_no,'
        'fill_date,data_source,order_info,form_filler'
    ).or_(f'repair_no.eq.{repair_no},old_repair_no.eq.{repair_no}').limit(1).execute()
    if not res.data:
        return jsonify({'error': '找不到維修編號'}), 404
    return jsonify(res.data[0])


@app.route('/api/refund/status-counts')
@login_required
def refund_status_counts():
    _PAGE = 1000
    _offset = 0
    all_rows = []
    while True:
        batch = sb.table('refund_tracking').select(
            'discount_invoice,ntt1_cancel,transaction_cancel,cancel_points,'
            'notify_customer,return_warehouse,claim_apply,claim_reconciled'
        ).range(_offset, _offset + _PAGE - 1).execute()
        rows = batch.data or []
        all_rows.extend(rows)
        if len(rows) < _PAGE:
            break
        _offset += _PAGE

    def _cnt(field):
        return sum(1 for r in all_rows if r.get(field) and str(r[field]).strip())

    return jsonify({
        'total': len(all_rows),
        'discount_invoice':   _cnt('discount_invoice'),
        'ntt1_cancel':        _cnt('ntt1_cancel'),
        'transaction_cancel': _cnt('transaction_cancel'),
        'cancel_points':      _cnt('cancel_points'),
        'notify_customer':    _cnt('notify_customer'),
        'return_warehouse':   _cnt('return_warehouse'),
        'claim_apply':        _cnt('claim_apply'),
        'claim_reconciled':   _cnt('claim_reconciled'),
    })


@app.route('/api/refund/template')
@login_required
def refund_template():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = '退費追蹤範本'
    headers = [
        '新維修編號','舊維修編號','填表人','下單日期','退貨類型',
        '讀者姓名','讀者帳號','訂單金額','訂單編號','付款方式',
        '發票號碼','刷退金額','來源','取消日期','取消原因',
        '折讓發票(帳務)','NTT1(景鴻)','取消交易(Rick)','取消點數(Rick)',
        '已通知客人','歸還倉庫','求償申請','求償入庫單號','求償成功對帳完',
    ]
    required = {'填表人','下單日期'}
    yellow = PatternFill('solid', fgColor='FFFF00')
    green  = PatternFill('solid', fgColor='C6EFCE')
    bold   = Font(bold=True)
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i)
        cell.font = bold
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill = yellow if h in required else green
    notes = [
        '從維修記錄帶入','從維修記錄帶入','*必填','YYYY-MM-DD','閱讀器/瑞米...',
        '自動帶入或手動','','數字','','信用卡/LINE Pay...',
        '','數字','電話/mail...','YYYY-MM-DD','',
        'v/x 或備註','v/x 或備註','v/x 或備註','v/x 或備註',
        'v/x 或備註','v/x 或備註','申請號碼或v','入庫單號','v 或備註',
    ]
    ws.append(notes)
    ws.cell(2, 1).font = Font(italic=True, color='808080')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='退費追蹤範本.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/refund/import', methods=['POST'])
@login_required
def import_refund():
    if 'file' not in request.files:
        return jsonify({'error': '請上傳 Excel 檔案'}), 400
    f = request.files['file']
    imported, skipped, errors = _process_refund_xlsx(f)
    return jsonify({'imported': imported, 'skipped': skipped, 'errors': errors})


def _process_refund_xlsx(f):
    import io
    from openpyxl import load_workbook
    from datetime import datetime as _dt, date as _date
    buf = io.BytesIO(f.read())
    wb = load_workbook(buf, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]

    _HEADER_MAP = {
        '新維修編號': 'repair_no', '舊維修編號': 'old_repair_no',
        '填表人': 'form_filler', '下單日期': 'order_date',
        '退貨類型': 'return_type', '讀者姓名': 'customer_name',
        '資料來源': 'customer_name', '讀者帳號': 'customer_account',
        '訂單金額': 'order_amount', '訂單編號': 'order_no',
        '付款方式': 'payment_method', '發票號碼': 'invoice_no',
        '刷退金額': 'refund_amount', '來源': 'contact_source',
        '取消日期': 'cancel_date', '取消原因': 'cancel_reason',
        '折讓發票': 'discount_invoice', '折讓發票(帳務)': 'discount_invoice',
        '刷退': 'discount_invoice',
        'NTT1': 'ntt1_cancel', 'NTT1(景鴻)': 'ntt1_cancel', 'ntt1取消': 'ntt1_cancel',
        '取消交易': 'transaction_cancel', '取消交易(Rick)': 'transaction_cancel',
        '交易取消': 'transaction_cancel',
        '取消點數': 'cancel_points', '取消點數(Rick)': 'cancel_points',
        '已通知客人': 'notify_customer', '通知讀者完成': 'notify_customer',
        '歸還倉庫': 'return_warehouse', '放回F': 'return_warehouse',
        '求償申請': 'claim_apply', '求償入庫單號': 'claim_warehouse_no',
        '求償成功對帳完': 'claim_reconciled', '求償成功': 'claim_reconciled',
    }

    col_idx = {}
    for i, h in enumerate(headers):
        key = _HEADER_MAP.get(h)
        if key and key not in col_idx:
            col_idx[key] = i

    existing = {r['invoice_no'] for r in
                (sb.table('refund_tracking').select('invoice_no').execute().data or [])
                if r.get('invoice_no')}

    cfg = sb.table('system_config').select('value').eq('key', 'refund_next_no').execute()
    current_no_str = (cfg.data or [{}])[0].get('value', 'RF2001')
    prefix = 'RF'
    digits = 4
    current_num = int(current_no_str[len(prefix):])

    imported = skipped = 0
    errors = []
    BATCH_SIZE = 80
    batch = []

    def flush():
        nonlocal imported
        if not batch:
            return
        sb.table('refund_tracking').insert(batch).execute()
        imported += len(batch)
        batch.clear()

    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(v is None for v in row):
            continue
        rec = {}
        for field, ci in col_idx.items():
            if ci >= len(row):
                continue
            val = row[ci]
            if val is None:
                rec[field] = None
                continue
            if field in ('order_date', 'cancel_date'):
                if isinstance(val, (_dt, _date)):
                    rec[field] = val.strftime('%Y-%m-%d')
                else:
                    s = str(val).strip()
                    rec[field] = s if s else None
            elif field in _REFUND_NUMERIC_COLS:
                try:
                    rec[field] = float(val) if val != '' else None
                except (ValueError, TypeError):
                    rec[field] = None
            else:
                if isinstance(val, (_dt, _date)):
                    rec[field] = val.strftime('%Y-%m-%d')
                else:
                    s = str(val).strip()
                    rec[field] = s if s else None

        inv = rec.get('invoice_no') or ''
        if inv and inv in existing:
            skipped += 1
            continue
        if inv:
            existing.add(inv)

        rec['refund_no'] = f"{prefix}{str(current_num).zfill(digits)}"
        current_num += 1
        rec['created_at'] = now_str()
        rec['updated_at'] = now_str()
        for fld in _ALL_REFUND_FIELDS:
            if fld not in rec:
                rec[fld] = None
        batch.append(rec)
        if len(batch) >= BATCH_SIZE:
            try:
                flush()
            except Exception as ex:
                errors.append(str(ex))
                return imported, skipped, errors

    try:
        flush()
    except Exception as ex:
        errors.append(str(ex))

    if imported > 0:
        new_no = f"{prefix}{str(current_num).zfill(digits)}"
        sb.table('system_config').update({'value': new_no}).eq('key', 'refund_next_no').execute()

    return imported, skipped, errors


@app.route('/api/refund/export')
@login_required
def export_refund():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file

    _PAGE = 1000
    _offset = 0
    all_rows = []
    while True:
        batch = sb.table('refund_tracking').select('*').order('refund_no').range(
            _offset, _offset + _PAGE - 1).execute()
        rows = batch.data or []
        all_rows.extend(rows)
        if len(rows) < _PAGE:
            break
        _offset += _PAGE

    wb = Workbook()
    ws = wb.active
    ws.title = '退費追蹤'
    headers = [
        '退費編號','新維修編號','舊維修編號','填表人','下單日期','退貨類型',
        '讀者姓名','讀者帳號','訂單金額','訂單編號','付款方式',
        '發票號碼','刷退金額','來源','取消日期','取消原因',
        '折讓發票(帳務)','NTT1(景鴻)','取消交易(Rick)','取消點數(Rick)',
        '已通知客人','歸還倉庫','求償申請','求償入庫單號','求償成功對帳完',
        '建立時間',
    ]
    hdr_fill = PatternFill('solid', fgColor='1a5276')
    hdr_font = Font(bold=True, color='FFFFFF')
    ws.append(headers)
    for i in range(1, len(headers)+1):
        c = ws.cell(1, i)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')

    field_map = [
        'refund_no','repair_no','old_repair_no','form_filler','order_date','return_type',
        'customer_name','customer_account','order_amount','order_no','payment_method',
        'invoice_no','refund_amount','contact_source','cancel_date','cancel_reason',
        'discount_invoice','ntt1_cancel','transaction_cancel','cancel_points',
        'notify_customer','return_warehouse','claim_apply','claim_warehouse_no',
        'claim_reconciled','created_at',
    ]
    for r in all_rows:
        ws.append([r.get(f) for f in field_map])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = __import__('datetime').date.today().strftime('%Y%m%d')
    return send_file(buf, download_name=f'退費追蹤_{today}.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



# ============================================================
# 產品規格 API  (v2.18)
# ============================================================

_PRODUCT_COLS = [
    'id','release_order','product_name','model_code','ncc_no',
    'android_version','screen_size','has_sdcard','light_type','color_options',
    'has_color_display','has_stylus','has_speaker','is_open',
    'resolution','storage','max_storage','bluetooth','wifi',
    'ram','cpu','weight','screen_spec','file_formats',
    'back_cover','power_spec','is_active','created_at','updated_at',
    'created_by','created_by_name','updated_by','updated_by_name'
]

_PRODUCT_BOOL_COLS = {'has_sdcard','has_color_display','has_stylus','has_speaker','is_open'}

def _filter_product_data(data: dict) -> dict:
    """過濾只保留合法欄位，布林欄位空字串轉 None"""
    allowed = set(_PRODUCT_COLS) - {'id','created_at','updated_at','created_by','created_by_name','updated_by','updated_by_name'}
    out = {k: v for k, v in data.items() if k in allowed}
    for col in _PRODUCT_BOOL_COLS:
        if col in out and out[col] == '':
            out[col] = None
    return out


@app.route('/api/products')
@login_required
def list_products():
    """取得產品規格列表，支援關鍵字搜尋與下拉篩選"""
    keyword = request.args.get('q', '').strip()
    screen_size = request.args.get('screen_size', '').strip()
    android_ver = request.args.get('android_version', '').strip()
    has_stylus  = request.args.get('has_stylus', '').strip()
    has_color   = request.args.get('has_color_display', '').strip()
    is_active   = request.args.get('is_active', 'true').strip()
    sort_col    = request.args.get('sort', 'release_order')
    sort_asc    = request.args.get('asc', 'true') == 'true'

    # 白名單防 SQL injection
    _SORT_ALLOW = {'release_order','product_name','model_code','screen_size',
                   'android_version','ram','storage','weight'}
    if sort_col not in _SORT_ALLOW:
        sort_col = 'release_order'

    q = sb.table('product_specs').select(','.join(_PRODUCT_COLS))

    if is_active != 'all':
        q = q.eq('is_active', is_active == 'true')

    if keyword:
        q = q.or_(f'product_name.ilike.%{keyword}%,model_code.ilike.%{keyword}%,'
                  f'ncc_no.ilike.%{keyword}%,color_options.ilike.%{keyword}%,'
                  f'screen_spec.ilike.%{keyword}%,cpu.ilike.%{keyword}%')

    if screen_size:
        q = q.eq('screen_size', screen_size)
    if android_ver:
        q = q.eq('android_version', android_ver)
    if has_stylus in ('true', 'false'):
        q = q.eq('has_stylus', has_stylus == 'true')
    if has_color in ('true', 'false'):
        q = q.eq('has_color_display', has_color == 'true')

    q = q.order(sort_col, desc=not sort_asc)

    res = q.execute()
    return jsonify(res.data or [])


@app.route('/api/products/template')
@login_required
def product_template():
    """下載批次匯入範本"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '產品規格範本'

    headers = [
        ('release_order','上市順序','整數，如 1'),
        ('product_name','產品名稱','必填，如 Gaze Mini +'),
        ('model_code','型號','必填唯一，如 K06NU'),
        ('ncc_no','NCC','審驗碼'),
        ('android_version','Android版本','如 11、14'),
        ('screen_size','尺寸(吋)','如 6、7.8、10.3'),
        ('has_sdcard','擴充記憶卡','TRUE/FALSE'),
        ('light_type','螢幕光源','冷暖 / 冷'),
        ('color_options','顏色','多色用/分隔'),
        ('has_color_display','彩色顯示','TRUE/FALSE'),
        ('has_stylus','手寫','TRUE/FALSE'),
        ('has_speaker','喇叭','TRUE/FALSE'),
        ('is_open','開放式系統','TRUE/FALSE'),
        ('resolution','解析度','如 300dpi'),
        ('storage','儲存容量','如 32G'),
        ('max_storage','可擴充最高','如 1TB'),
        ('bluetooth','藍芽','如 5.0'),
        ('wifi','Wi-Fi','如 2.4G+5G'),
        ('ram','RAM','如 3GB'),
        ('cpu','CPU','如 四核處理器 1.8 GHz'),
        ('weight','重量','含g，如 195g'),
        ('screen_spec','螢幕規格','長文字'),
        ('file_formats','支援格式','長文字'),
        ('back_cover','背殼設計','外觀說明'),
        ('power_spec','電源規格','預留空白'),
    ]

    hdr_fill = PatternFill('solid', fgColor='4472C4')
    lbl_fill = PatternFill('solid', fgColor='D9E1F2')
    req_fill = PatternFill('solid', fgColor='FFE699')
    hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    lbl_font = Font(name='Arial', size=9, color='444444')
    note_font = Font(name='Arial', size=9, color='666666', italic=True)

    for col_i, (db_key, label, note) in enumerate(headers, 1):
        c1 = ws.cell(1, col_i, db_key)
        c1.font = hdr_font
        c1.fill = hdr_fill if db_key not in ('product_name','model_code') else PatternFill('solid', fgColor='C00000')
        c1.alignment = Alignment(horizontal='center')

        c2 = ws.cell(2, col_i, label)
        c2.font = lbl_font
        c2.fill = lbl_fill
        c2.alignment = Alignment(horizontal='center')

        c3 = ws.cell(3, col_i, note)
        c3.font = note_font
        c3.fill = PatternFill('solid', fgColor='F2F2F2')

    ws.freeze_panes = 'A4'

    col_widths = [10,20,12,20,12,10,12,10,15,10,8,8,10,20,10,10,8,12,8,25,8,35,50,20,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    from flask import send_file
    import io
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='產品規格_匯入範本.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/products/import', methods=['POST'])
@login_required
def import_products():
    """批次匯入產品規格 Excel（admin 限定）"""
    if not session.get('is_admin'):
        return jsonify({'error': '無權限'}), 403

    import openpyxl

    f = request.files.get('file')
    if not f:
        return jsonify({'error': '請上傳 Excel 檔案'}), 400

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({'error': '檔案無資料'}), 400

        # 第一列為 DB 欄位名稱
        headers = [str(h).strip() if h else '' for h in rows[0]]

        _BOOL_FIELDS = {'has_sdcard','has_color_display','has_stylus','has_speaker','is_open'}
        _INT_FIELDS  = {'release_order'}
        _ALLOWED     = set(_PRODUCT_COLS) - {'id','created_at','updated_at'}

        # 取得現有 model_code
        existing_res = sb.table('product_specs').select('model_code').execute()
        existing = {r['model_code'] for r in (existing_res.data or [])}

        ok_cnt = skip_cnt = err_cnt = 0
        ts = now_str()

        # 跳過說明列（第2列若含中文標題）
        data_start = 1
        if len(rows) >= 2 and rows[1][0] is not None:
            sample = str(rows[1][0]).strip()
            if not sample.lstrip('-').isdigit() and not sample.replace('.','',1).isdigit():
                data_start = 2  # 第2列是說明列，從第3列開始

        records = []
        for row in rows[data_start:]:
            if all(v is None for v in row):
                continue
            rec = {}
            for col_i, h in enumerate(headers):
                if h not in _ALLOWED:
                    continue
                val = row[col_i] if col_i < len(row) else None
                if val is None or str(val).strip() == '':
                    val = None
                elif h in _BOOL_FIELDS:
                    val = str(val).upper() in ('TRUE', '1', 'V', 'YES', '是')
                elif h in _INT_FIELDS:
                    try:
                        val = int(val)
                    except Exception:
                        val = None
                else:
                    val = str(val).strip()
                rec[h] = val

            if not rec.get('product_name') or not rec.get('model_code'):
                skip_cnt += 1
                continue

            model = rec['model_code']
            if model in existing:
                skip_cnt += 1
                continue

            existing.add(model)
            rec['created_at'] = ts
            rec['updated_at'] = ts
            records.append(rec)

        # 批次 INSERT
        BATCH = 50
        for i in range(0, len(records), BATCH):
            batch = records[i:i+BATCH]
            try:
                sb.table('product_specs').insert(batch).execute()
                ok_cnt += len(batch)
            except Exception as e:
                err_cnt += len(batch)

        return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt})

    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/api/products/<string:model_code>', methods=['GET'])
@login_required
def get_product(model_code):
    """取得單一機型詳細規格（by model_code）"""
    res = sb.table('product_specs').select(','.join(_PRODUCT_COLS)) \
            .eq('model_code', model_code).execute()
    rows = res.data or []
    if not rows:
        return jsonify({'error': f'找不到型號 {model_code}'}), 404
    return jsonify(rows[0])


@app.route('/api/products', methods=['POST'])
@login_required
def create_product():
    """新增產品規格（admin 限定）"""
    if not session.get('is_admin'):
        return jsonify({'error': '無權限'}), 403
    data = _filter_product_data(request.get_json() or {})
    if not data.get('product_name') or not data.get('model_code'):
        return jsonify({'error': '產品名稱與型號為必填'}), 400
    # 檢查型號重複
    chk = sb.table('product_specs').select('id').eq('model_code', data['model_code']).execute()
    if chk.data:
        return jsonify({'error': f'型號 {data["model_code"]} 已存在'}), 409
    data['created_at'] = now_str()
    data['updated_at'] = now_str()
    data['created_by']       = session['user_id']
    data['updated_by']       = session['user_id']
    data['created_by_name']  = session.get('display_name', session.get('username', ''))
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    res = sb.table('product_specs').insert(data).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/products/<int:pid>', methods=['PUT'])
@login_required
def update_product(pid):
    """修改產品規格（admin 限定）"""
    if not session.get('is_admin'):
        return jsonify({'error': '無權限'}), 403
    data = _filter_product_data(request.get_json() or {})
    data['updated_at'] = now_str()
    data['updated_by']       = session['user_id']
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    data.pop('created_by', None)
    data.pop('created_by_name', None)
    data.pop('created_at', None)
    res = sb.table('product_specs').update(data).eq('id', pid).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/products/<int:pid>', methods=['DELETE'])
@login_required
def delete_product(pid):
    """刪除產品規格（admin 限定）"""
    if not session.get('is_admin'):
        return jsonify({'error': '無權限'}), 403
    sb.table('product_specs').delete().eq('id', pid).execute()
    return jsonify({'ok': True})




# ════════════════════════════════════════════════════════════════
# 福利品模組 API
# ════════════════════════════════════════════════════════════════

_WELFARE_COLS = {'sn', 'welfare_type', 'sold_date', 'order_no', 'is_cancelled', 'notes'}

def _clean_welfare(data: dict) -> dict:
    """過濾欄位、將空字串轉 None"""
    out = {}
    for k, v in data.items():
        if k not in _WELFARE_COLS:
            continue
        if isinstance(v, str) and v.strip() == '':
            out[k] = None
        else:
            out[k] = v
    return out


@app.route('/api/welfare', methods=['GET'])
@login_required
def list_welfare():
    p = request.args
    q = sb.table('welfare_products').select('*')

    kw = (p.get('q') or '').strip()
    if kw:
        like = f'%{kw}%'
        q = q.or_(
            f'sn.ilike.{like},'
            f'welfare_type.ilike.{like},'
            f'order_no.ilike.{like},'
            f'is_cancelled.ilike.{like},'
            f'notes.ilike.{like}'
        )

    if p.get('welfare_type'):
        q = q.eq('welfare_type', p['welfare_type'])
    if p.get('is_cancelled'):
        q = q.eq('is_cancelled', p['is_cancelled'])
    if p.get('sold_date_f'):
        q = q.gte('sold_date', p['sold_date_f'])
    if p.get('sold_date_t'):
        q = q.lte('sold_date', p['sold_date_t'])

    sort_col  = p.get('sort', 'id')
    sort_desc = p.get('desc', '1') == '1'
    allowed_sort = {'id', 'sn', 'welfare_type', 'sold_date', 'order_no', 'is_cancelled', 'created_at'}
    if sort_col not in allowed_sort:
        sort_col = 'id'
    q = q.order(sort_col, desc=sort_desc)

    try:
        page     = max(1, int(p.get('page', 1)))
        per_page = min(200, max(10, int(p.get('per_page', 50))))
    except ValueError:
        page, per_page = 1, 50

    q_cnt = q
    total_res = q_cnt.execute()
    total = len(total_res.data or [])

    offset = (page - 1) * per_page
    rows = (total_res.data or [])[offset: offset + per_page]

    return jsonify({'total': total, 'page': page, 'per_page': per_page, 'rows': rows})


@app.route('/api/welfare/check-sn-batch', methods=['POST'])
@login_required
def welfare_check_sn_batch():
    """批次查詢多個 SN 是否為有效福利品（未取消），回傳 {SN大寫: {...}}"""
    sns = (request.json or {}).get('sns', [])
    if not sns:
        return jsonify({})
    try:
        res = sb.table('welfare_products').select('sn,welfare_type,sold_date,order_no,is_cancelled') \
                .in_('sn', sns).execute()
        rows = res.data or []
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    result = {}
    for r in rows:
        sn_raw = (r.get('sn') or '').strip()
        sn_up  = sn_raw.upper()
        if sn_up and r.get('is_cancelled') != '是' and sn_up not in result:
            result[sn_up] = {
                'found':        True,
                'sn':           sn_raw,
                'welfare_type': r.get('welfare_type', ''),
                'sold_date':    r.get('sold_date', ''),
                'order_no':     r.get('order_no', ''),
            }
    return jsonify(result)


@app.route('/api/welfare/check-sn')
@login_required
def welfare_check_sn():
    """查詢指定 SN 是否為已取消的福利品"""
    sn = (request.args.get('sn') or '').strip()
    if not sn:
        return jsonify({'found': False})
    res = sb.table('welfare_products').select('sn,welfare_type,sold_date,order_no,is_cancelled') \
            .ilike('sn', sn).execute()
    rows = res.data or []
    # 找到 sn 完全相符（不分大小寫）且「未取消」的紀錄（is_cancelled 不是 '是' = 仍為有效福利品）
    active = [r for r in rows if r.get('sn','').strip().upper() == sn.upper()
                              and r.get('is_cancelled') != '是']
    if active:
        r = active[0]
        return jsonify({
            'found': True,
            'sn':           r.get('sn',''),
            'welfare_type': r.get('welfare_type',''),
            'sold_date':    r.get('sold_date',''),
            'order_no':     r.get('order_no',''),
        })
    return jsonify({'found': False})


@app.route('/api/welfare', methods=['POST'])
@login_required
def create_welfare():
    data = _clean_welfare(request.json or {})
    for req_f in ('sn', 'welfare_type', 'is_cancelled'):
        if not data.get(req_f):
            return jsonify({'error': f'必填欄位缺漏：{req_f}'}), 400
    data['created_by']       = session['user_id']
    data['updated_by']       = session['user_id']
    data['created_by_name']  = session.get('display_name', session.get('username', ''))
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    try:
        res = sb.table('welfare_products').insert(data).execute()
        return jsonify({'ok': True, 'id': res.data[0]['id']}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/welfare/<int:wid>', methods=['GET'])
@login_required
def get_welfare(wid):
    try:
        res = sb.table('welfare_products').select('*').eq('id', wid).execute()
        rows = res.data or []
        if not rows:
            return jsonify({'error': '找不到該筆記錄'}), 404
        return jsonify(rows[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/welfare/<int:wid>', methods=['PUT'])
@login_required
def update_welfare(wid):
    data = _clean_welfare(request.json or {})
    for req_f in ('sn', 'welfare_type', 'is_cancelled'):
        if req_f in data and not data.get(req_f):
            return jsonify({'error': f'必填欄位不可為空：{req_f}'}), 400
    data['updated_by']       = session['user_id']
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    data.pop('created_by', None)
    data.pop('created_by_name', None)
    data.pop('created_at', None)
    try:
        sb.table('welfare_products').update(data).eq('id', wid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/welfare/<int:wid>', methods=['DELETE'])
@login_required
def delete_welfare(wid):
    try:
        sb.table('welfare_products').delete().eq('id', wid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/welfare/export', methods=['GET'])
@login_required
def export_welfare():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    import io, datetime

    p = request.args
    q = sb.table('welfare_products').select('*')
    kw = (p.get('q') or '').strip()
    if kw:
        like = f'%{kw}%'
        q = q.or_(
            f'sn.ilike.{like},'
            f'welfare_type.ilike.{like},'
            f'order_no.ilike.{like},'
            f'is_cancelled.ilike.{like},'
            f'notes.ilike.{like}'
        )
    if p.get('welfare_type'):
        q = q.eq('welfare_type', p['welfare_type'])
    if p.get('is_cancelled'):
        q = q.eq('is_cancelled', p['is_cancelled'])
    if p.get('sold_date_f'):
        q = q.gte('sold_date', p['sold_date_f'])
    if p.get('sold_date_t'):
        q = q.lte('sold_date', p['sold_date_t'])
    q = q.order('id', desc=False)

    # 分頁取全量
    _PAGE = 1000
    _offset = 0
    rows = []
    while True:
        batch = q.range(_offset, _offset + _PAGE - 1).execute()
        rows.extend(batch.data or [])
        if len(batch.data or []) < _PAGE:
            break
        _offset += _PAGE

    wb = Workbook()
    ws = wb.active
    ws.title = '福利品'
    headers = ['SN', '福利品', '售出時間', '訂單編號', '是否取消', '備註']
    hfill = PatternFill('solid', fgColor='1A5276')
    hfont = Font(bold=True, color='FFFFFF')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal='center')

    field_map = [
        ('sn', None), ('welfare_type', None), ('sold_date', None),
        ('order_no', None), ('is_cancelled', None), ('notes', None),
    ]
    for ri, r in enumerate(rows, 2):
        for ci, (fk, _) in enumerate(field_map, 1):
            ws.cell(ri, ci, r.get(fk) or '')

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'福利品_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/api/welfare/template', methods=['GET'])
@login_required
def welfare_template():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = '福利品匯入範本'

    headers    = ['SN',      '福利品',    '售出時間',       '訂單編號', '是否取消', '備註']
    required   = { 'SN', '福利品', '是否取消' }
    notes_row  = ['裝置SN碼', '官網福利品 / 經銷福利品', 'YYYY-MM-DD（可空白）', '（可空白）', '否 / 是', '（可空白）']

    hfill  = PatternFill('solid', fgColor='1A5276')
    rfill  = PatternFill('solid', fgColor='FFF2CC')   # 必填黃底
    hfont  = Font(bold=True, color='FFFFFF')
    rfont  = Font(bold=True, color='7B3F00')
    nfill  = PatternFill('solid', fgColor='F2F2F2')
    nfont  = Font(italic=True, color='888888')

    widths = [24, 16, 14, 16, 12, 30]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        if h in required:
            c.fill = rfill; c.font = rfont
        else:
            c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci - 1]

    for ci, n in enumerate(notes_row, 1):
        c = ws.cell(2, ci, n)
        c.fill = nfill; c.font = nfont
        c.alignment = Alignment(horizontal='center')

    # 範例資料
    ws.append(['TWHR10P2019120595', '官網福利品', '2024-09-07', '', '否', ''])
    ws.append(['TWHM08C2025011313', '經銷福利品', '2025-12-17', 'ORD12345', '是', '拍賣會用福利機'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='福利品匯入範本.xlsx')


@app.route('/api/welfare/batch-import', methods=['POST'])
@login_required
def batch_import_welfare():
    from openpyxl import load_workbook
    import io, datetime as _dt_mod
    from datetime import date as _date_cls, datetime as _dt_cls

    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400

    f = request.files['file']
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{str(e)}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'error': '檔案無資料列'}), 400

    # 解析表頭
    raw_headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    def _map_wh(h):
        h = h.strip()
        if 'SN' in h.upper() or 'sn' in h:         return 'sn'
        if '福利品' in h and '是否' not in h:       return 'welfare_type'
        if '售出' in h:                              return 'sold_date'
        if '訂單' in h:                              return 'order_no'
        if '是否取消' in h or '取消' in h:           return 'is_cancelled'
        if '備註' in h:                              return 'notes'
        return None

    col_map = {}
    for ci, h in enumerate(raw_headers):
        fk = _map_wh(h)
        if fk and fk not in col_map:
            col_map[fk] = ci

    def _parse_date_wf(val):
        if val is None:
            return None
        if isinstance(val, (_dt_cls, _date_cls)):
            return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y'):
            try:
                return _dt_cls.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
        return s or None

    # 取現有 SN 集合（避免重複匯入）
    existing_sns = set()
    _off = 0
    while True:
        res = sb.table('welfare_products').select('sn').range(_off, _off + 999).execute()
        for r in (res.data or []):
            if r.get('sn'):
                existing_sns.add(r['sn'])
        if len(res.data or []) < 1000:
            break
        _off += 1000

    ok_cnt = skip_cnt = err_cnt = 0
    batch = []

    for ri, row in enumerate(rows[1:], 2):
        # 跳過說明列
        first_val = str(row[0]).strip() if row[0] is not None else ''
        if first_val in ('', '裝置SN碼', '（說明）'):
            skip_cnt += 1
            continue

        def _get(fk):
            ci = col_map.get(fk)
            if ci is None or ci >= len(row):
                return None
            v = row[ci]
            if v is None:
                return None
            return str(v).strip() if not isinstance(v, (_dt_cls, _date_cls)) else v

        sn_val   = _get('sn')
        wt_val   = _get('welfare_type')
        ic_val   = _get('is_cancelled') or '否'
        sd_val   = _parse_date_wf(_get('sold_date'))
        ord_val  = _get('order_no')
        note_val = _get('notes')

        # 必填驗證
        if not sn_val or not wt_val:
            skip_cnt += 1
            continue

        # 重複 SN 跳過
        if sn_val in existing_sns:
            skip_cnt += 1
            continue
        existing_sns.add(sn_val)

        rec = {
            'sn':          sn_val,
            'welfare_type': wt_val,
            'sold_date':   sd_val,
            'order_no':    ord_val or None,
            'is_cancelled': ic_val,
            'notes':        note_val or None,
        }
        batch.append(rec)

        if len(batch) >= 80:
            try:
                sb.table('welfare_products').insert(batch).execute()
                ok_cnt += len(batch)
            except Exception:
                err_cnt += len(batch)
            batch = []

    if batch:
        try:
            sb.table('welfare_products').insert(batch).execute()
            ok_cnt += len(batch)
        except Exception:
            err_cnt += len(batch)

    return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt})


@app.route('/api/welfare/batch-delete', methods=['POST'])
@login_required
def batch_delete_welfare():
    ids = (request.json or {}).get('ids', [])
    if not ids:
        return jsonify({'error': 'no ids'}), 400
    try:
        sb.table('welfare_products').delete().in_('id', ids).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
# 退換貨檢測模組 (inspection_records)
# ============================================================

# 退換檢測所有 DB 欄位（不含 id / created_at / updated_at）
_INSP_FIELDS = [
    'insp_no', 'old_insp_no', 'repair_no', 'model', 'serial_no', 'sn_manual', 'sn_mismatch',
    'order_no', 'customer_issue',
    'exchange_order_no', 'cancel_date', 'cancel_staff',
    'lib_removed', 'lib_process_date', 'lib_process_staff',
    'open_closed', 'sn_consistent', 'emmc_version', 'ram_lcd',
    'screen_appearance', 'sd_touch', 'power_earphone', 'screen_wifi',
    'touch_bluetooth', 'battery_mic', 'led_battery', 'wifi_sd',
    'bluetooth_keyboard', 'gsensor_brightness', 'backlight_storage',
    'temperature_key', 'usb_other', 'earphone_gravity', 'waveform_note',
    'photo_test', 'wizard_pen', 'vcom', 'appearance_scratch',
    'magnet_position', 'usb_cable_pc', 'wifi_pair', 'bt_pair',
    'sim_card', 'charge_led', 'bookstore_trial', 'library_download',
    'case_sleep', 'case_cable', 'factory_reset', 'language_screen',
    'clean_reader', 'inspection_note', 'inspection_date', 'inspection_staff',
    'disassemble', 'disassemble_note', 'disassemble_date', 'disassemble_staff',
    'box_sn_consistent', 'box_damaged', 'sleeve_one_sc', 'screen_film_mini',
    'jelly_case', 'pvc_bag', 'warranty_card', 'usb_cable_pen',
    'stylus_replacement', 'seal_film', 'refurb_note', 'refurb_date', 'refurb_staff',
    'final_disassemble', 'final_stock', 'final_order_no', 'final_date', 'final_staff',
    'welfare_grade', 'welfare_grade_note', 'welfare_notify_date', 'welfare_sold', 'discontinued',
]
_INSP_DATE_FIELDS = {
    'cancel_date', 'lib_process_date', 'inspection_date',
    'disassemble_date', 'refurb_date', 'final_date', 'welfare_notify_date',
}

def _filter_insp_data(data):
    """過濾不在白名單或值為空字串的日期欄位，避免 PG 型別錯誤"""
    out = {k: v for k, v in data.items() if k in _INSP_FIELDS}
    for f in _INSP_DATE_FIELDS:
        if f in out and out[f] == '':
            out[f] = None
    return out


def _insp_next_no():
    """取得下一個退換檢測流水號並更新 system_config"""
    res = sb.table('system_config').select('value').eq('key', 'insp_next_no').execute()
    cur = (res.data or [{}])[0].get('value', 'D3001')
    prefix, num = cur[0], int(cur[1:])
    nxt = f'{prefix}{num + 1}'
    sb.table('system_config').update({'value': nxt}).eq('key', 'insp_next_no').execute()
    return cur


@app.route('/api/inspection')
@login_required
def list_inspection():
    """退換檢測列表（分頁 + 關鍵字 + 篩選 + 排序）"""
    page     = int(request.args.get('page', 1))
    per_page = min(int(request.args.get('per_page', 50)), 200)
    kw       = request.args.get('q', '').strip()
    sort_col = request.args.get('sort', 'created_at')
    sort_asc = request.args.get('order', 'desc') == 'asc'
    # 篩選
    f_repair_no        = request.args.get('repair_no', '').strip()
    f_model            = request.args.get('model', '').strip()
    f_serial_no        = request.args.get('serial_no', '').strip()
    f_inspection_staff = request.args.get('inspection_staff', '').strip()
    f_refurb_staff     = request.args.get('refurb_staff', '').strip()
    f_insp_date_from   = request.args.get('insp_date_from', '').strip()
    f_insp_date_to     = request.args.get('insp_date_to', '').strip()
    f_created_from     = request.args.get('created_from', '').strip()
    f_created_to       = request.args.get('created_to', '').strip()
    f_welfare_grade    = request.args.get('welfare_grade', '').strip()
    f_final_stock      = request.args.get('final_stock', '').strip()
    f_welfare_sold     = request.args.get('welfare_sold', '').strip()
    f_discontinued     = request.args.get('discontinued', '').strip()
    f_lib_removed      = request.args.get('lib_removed', '').strip()

    try:
        q = sb.table('inspection_records').select('*', count='exact')

        if kw:
            q = q.or_(
                f'insp_no.ilike.%{kw}%,repair_no.ilike.%{kw}%,'
                f'serial_no.ilike.%{kw}%,model.ilike.%{kw}%,'
                f'order_no.ilike.%{kw}%,customer_issue.ilike.%{kw}%,'
                f'exchange_order_no.ilike.%{kw}%,inspection_note.ilike.%{kw}%,'
                f'refurb_note.ilike.%{kw}%,final_order_no.ilike.%{kw}%'
            )
        if f_repair_no:        q = q.ilike('repair_no', f'%{f_repair_no}%')
        if f_model:            q = q.ilike('model',     f'%{f_model}%')
        if f_serial_no:        q = q.ilike('serial_no', f'%{f_serial_no}%')
        if f_inspection_staff: q = q.eq('inspection_staff', f_inspection_staff)
        if f_refurb_staff:     q = q.eq('refurb_staff',     f_refurb_staff)
        if f_insp_date_from:   q = q.gte('inspection_date', f_insp_date_from)
        if f_insp_date_to:     q = q.lte('inspection_date', f_insp_date_to)
        if f_created_from:     q = q.gte('created_at', f_created_from)
        if f_created_to:       q = q.lte('created_at', f_created_to + 'T23:59:59')
        if f_welfare_grade:    q = q.eq('welfare_grade', f_welfare_grade)
        if f_final_stock:      q = q.eq('final_stock', f_final_stock)
        if f_welfare_sold:     q = q.eq('welfare_sold', f_welfare_sold)
        if f_discontinued:     q = q.eq('discontinued', f_discontinued)
        if f_lib_removed:      q = q.eq('lib_removed', f_lib_removed)

        # 排序
        allowed_sorts = set(_INSP_FIELDS) | {'created_at', 'updated_at'}
        if sort_col not in allowed_sorts:
            sort_col = 'created_at'
        q = q.order(sort_col, desc=not sort_asc)

        offset = (page - 1) * per_page
        q = q.range(offset, offset + per_page - 1)
        res = q.execute()
        total = res.count or 0
        return jsonify({
            'records': res.data or [],
            'total': total,
            'page': page,
            'per_page': per_page,
            'pages': max(1, -(-total // per_page)),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/inspection', methods=['POST'])
@login_required
def create_inspection():
    """新增退換檢測記錄"""
    data = _filter_insp_data(request.json or {})
    if not data.get('insp_no'):
        data['insp_no'] = _insp_next_no()
    data['created_by']       = session['user_id']
    data['updated_by']       = session['user_id']
    data['created_by_name']  = session.get('display_name', session.get('username', ''))
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    try:
        res = sb.table('inspection_records').insert(data).execute()
        return jsonify({'ok': True, 'record': res.data[0]}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/inspection/<record_id>', methods=['PUT'])
@login_required
def update_inspection(record_id):
    """更新退換檢測記錄"""
    data = _filter_insp_data(request.json or {})
    data.pop('insp_no', None)   # 不允許修改流水號
    data['updated_by']       = session['user_id']
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    data.pop('created_by', None)
    data.pop('created_by_name', None)
    data.pop('created_at', None)
    try:
        sb.table('inspection_records').update(data).eq('id', record_id).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/inspection/<record_id>', methods=['DELETE'])
@admin_required
def delete_inspection(record_id):
    """刪除退換檢測記錄（admin 限定）"""
    sb.table('inspection_records').delete().eq('id', record_id).execute()
    return jsonify({'ok': True})


@app.route('/api/inspection/export')
@login_required
def export_inspection():
    """匯出退換檢測 Excel（目前頁篩選條件）"""
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    # 重用 list_inspection 取全量資料（不分頁）
    kw       = request.args.get('q', '').strip()
    f_repair_no        = request.args.get('repair_no', '').strip()
    f_model            = request.args.get('model', '').strip()
    f_serial_no        = request.args.get('serial_no', '').strip()
    f_inspection_staff = request.args.get('inspection_staff', '').strip()
    f_refurb_staff     = request.args.get('refurb_staff', '').strip()
    f_insp_date_from   = request.args.get('insp_date_from', '').strip()
    f_insp_date_to     = request.args.get('insp_date_to', '').strip()
    f_welfare_grade    = request.args.get('welfare_grade', '').strip()
    f_discontinued     = request.args.get('discontinued', '').strip()

    _PAGE = 1000
    _offset = 0
    all_rows = []
    while True:
        q = sb.table('inspection_records').select('*')
        if kw:
            q = q.or_(
                f'insp_no.ilike.%{kw}%,repair_no.ilike.%{kw}%,'
                f'serial_no.ilike.%{kw}%,model.ilike.%{kw}%,'
                f'order_no.ilike.%{kw}%,customer_issue.ilike.%{kw}%'
            )
        if f_repair_no:        q = q.ilike('repair_no', f'%{f_repair_no}%')
        if f_model:            q = q.ilike('model',     f'%{f_model}%')
        if f_serial_no:        q = q.ilike('serial_no', f'%{f_serial_no}%')
        if f_inspection_staff: q = q.eq('inspection_staff', f_inspection_staff)
        if f_refurb_staff:     q = q.eq('refurb_staff',     f_refurb_staff)
        if f_insp_date_from:   q = q.gte('inspection_date', f_insp_date_from)
        if f_insp_date_to:     q = q.lte('inspection_date', f_insp_date_to)
        if f_welfare_grade:    q = q.ilike('welfare_grade', f'%{f_welfare_grade}%')
        if f_discontinued:     q = q.eq('discontinued', f_discontinued)
        batch = q.order('created_at', desc=True).range(_offset, _offset + _PAGE - 1).execute()
        rows = batch.data or []
        all_rows.extend(rows)
        if len(rows) < _PAGE:
            break
        _offset += _PAGE

    headers_zh = [
        '退換檢測編號', '舊檢測碼', '維修編號', '型號', 'SN', 'SN(人工掃)', '原訂單編號', '客戶問題備註',
        '換貨訂單編號', '取消日期', '取消人員',
        '圖書館版移除', '圖書館版處理日', '圖書館版處理人員',
        '封閉/開放', '外盒機身SN保卡SN一致', 'emmc/Version', 'ram/LCD',
        '螢幕外觀', 'sd card/Touch', 'power key/Earphone', 'screen/Wifi',
        'touch/Bluetooth', 'battery/MIC', 'led/Battery', 'wifi/SD',
        'Bluetooth/Keyboard', 'gsensor/Brightness', 'backlight/Storage',
        'temperature/按鍵', 'usbtypec/其他', 'earphone/重力感應', 'waveform/筆記',
        '拍照測試', '巫師筆更換(10.3)', 'vcom值', '機器外觀是否明顯刮損',
        '磁吸位置', '傳輸線接電腦', 'Wifi配對', '藍芽配對', '插sim卡',
        '充電紅>綠燈', '書店電子書試閱', '圖書館借書下載',
        '保護殼/磁吸休眠功能', '保護殼/傳輸線插入拔出',
        '恢復原廠設定', '回到初始語言畫面關機', '清潔閱讀器',
        '檢測備註', '檢測日期', '檢測人員',
        '是否拆機', '拆機(福利機)備註', '拆機作業日期', '拆機人員',
        '外盒與保卡SN一致', '外盒凹折嚴重', '封套(ONE SC)', '螢幕透明膜(MINI)', '果凍套(X/NOTE)',
        'PVC塑膠袋', '保固卡', '傳輸線/筆芯組', '觸控筆更換並測試', '封膜',
        '整新備註', '整新日期', '整新人員',
        '看是否拆機', '放現貨/料件更換', '最終訂單/進貨編號', '最終處理日期', '最終處理人員',
        '福利等級', '福利等級備註', '通知福利上架日', '是否售出福利機', '是否停產',
        '建立時間',
    ]
    db_keys = _INSP_FIELDS + ['created_at']

    wb = Workbook(); ws = wb.active; ws.title = '退換貨檢測'
    hfill = PatternFill('solid', fgColor='1A237E')
    hfont = Font(bold=True, color='FFFFFF')
    ws.append(headers_zh)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
    _INSP_DATE_COLS = {'inspection_date','disassemble_date','refurb_date','final_date','welfare_notify_date'}
    for r in all_rows:
        row_vals = []
        for k in db_keys:
            v = r.get(k, '') or ''
            if k in _INSP_DATE_COLS and v and len(str(v)) > 10:
                v = str(v)[:10]
            row_vals.append(v)
        ws.append(row_vals)
    ws.freeze_panes = 'A2'
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from datetime import datetime
    fname = '退換貨檢測_{}.xlsx'.format(datetime.now().strftime('%Y%m%d%H%M'))
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/api/inspection/export-all')
@admin_required
def export_all_inspection():
    """【管理員專用】匯出全部退換貨檢測記錄（不篩選），含系統ID"""
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from datetime import datetime

    records = []
    offset = 0
    BATCH = 1000
    while True:
        batch = sb.table('inspection_records').select('*').order('id', desc=False).range(offset, offset + BATCH - 1).execute()
        rows = batch.data or []
        records.extend(rows)
        if len(rows) < BATCH:
            break
        offset += BATCH
        if offset >= 50000:
            break

    headers_zh = [
        '系統ID', '退換檢測編號', '舊檢測碼', '維修編號', '型號', 'SN', 'SN(人工掃)', '原訂單編號', '客戶問題備註',
        '換貨訂單編號', '取消日期', '取消人員',
        '圖書館版移除', '圖書館版處理日', '圖書館版處理人員',
        '封閉/開放', '外盒機身SN保卡SN一致', 'emmc/Version', 'ram/LCD',
        '螢幕外觀', 'sd card/Touch', 'power key/Earphone', 'screen/Wifi',
        'touch/Bluetooth', 'battery/MIC', 'led/Battery', 'wifi/SD',
        'Bluetooth/Keyboard', 'gsensor/Brightness', 'backlight/Storage',
        'temperature/按鍵', 'usbtypec/其他', 'earphone/重力感應', 'waveform/筆記',
        '拍照測試', '巫師筆更換(10.3)', 'vcom值', '機器外觀是否明顯刮損',
        '磁吸位置', '傳輸線接電腦', 'Wifi配對', '藍芽配對', '插sim卡',
        '充電紅>綠燈', '書店電子書試閱', '圖書館借書下載',
        '保護殼/磁吸休眠功能', '保護殼/傳輸線插入拔出',
        '恢復原廠設定', '回到初始語言畫面關機', '清潔閱讀器',
        '檢測備註', '檢測日期', '檢測人員',
        '是否拆機', '拆機(福利機)備註', '拆機作業日期', '拆機人員',
        '外盒與保卡SN一致', '外盒凹折嚴重', '封套(ONE SC)', '螢幕透明膜(MINI)', '果凍套(X/NOTE)',
        'PVC塑膠袋', '保固卡', '傳輸線/筆芯組', '觸控筆更換並測試', '封膜',
        '整新備註', '整新日期', '整新人員',
        '看是否拆機', '放現貨/料件更換', '最終訂單/進貨編號', '最終處理日期', '最終處理人員',
        '福利等級', '福利等級備註', '通知福利上架日', '是否售出福利機', '是否停產',
        '建立時間',
    ]
    db_keys = ['id'] + _INSP_FIELDS + ['created_at']

    wb = Workbook(); ws = wb.active; ws.title = '退換貨檢測全DB'
    hfill = PatternFill('solid', fgColor='1A237E')
    hfont = Font(bold=True, color='FFFFFF')
    ws.append(headers_zh)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
    _INSP_DATE_COLS = {'inspection_date','disassemble_date','refurb_date','final_date','welfare_notify_date'}
    for r in records:
        row_vals = []
        for k in db_keys:
            v = r.get(k, '') or ''
            if k in _INSP_DATE_COLS and v and len(str(v)) > 10:
                v = str(v)[:10]
            row_vals.append(v)
        ws.append(row_vals)
    ws.freeze_panes = 'A2'
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = '退換貨檢測全DB_{}.xlsx'.format(datetime.now().strftime('%Y%m%d%H%M'))
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/inspection/template')
@login_required
def insp_template():
    """下載退換貨檢測匯入範本"""
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    headers = [
        '舊檢測碼','維修編號','型號','SN','SN(人工掃)','原訂單編號','客戶問題備註',
        '換貨訂單編號','取消日期','取消人員',
        '圖書館版移除','圖書館版處理日','圖書館版處理人員',
        '封閉/開放','外盒機身SN保卡SN一致',
        'emmc/Version','ram/LCD','螢幕外觀','sd card/Touch',
        'power key/Earphone','screen/Wi-Fi','touch/Bluetooth','battery/MIC',
        'led/Battery','Wi-Fi/SD','Bluetooth/Keyboard','gsensor/Brightness',
        'backlight/Storage','temperature/按鍵','usbtypec/其他',
        'earphone/重力感應','waveform/筆記','拍照測試','巫師筆更換(10.3)','vcom值',
        '機器外觀是否明顯刮損','磁吸位置','傳輸線接電腦',
        'Wifi配對','藍芽配對','插sim卡','充電紅>綠燈',
        '書店電子書試閱','圖書館借書下載','保護殼/磁吸休眠',
        '保護殼/傳輸線插入拔出','恢復原廠設定','回到初始語言畫面關機','清潔閱讀器',
        '檢測備註','檢測日期','檢測人員',
        '是否拆機','拆機(福利機)備註','拆機作業日期','拆機人員',
        '外盒與保卡SN一致','外盒凹折嚴重','封套(ONE SC)','螢幕透明膜(MINI)',
        '果凍套(X/NOTE)','PVC塑膠袋','保固卡','傳輸線/筆芯組','觸控筆更換並測試','封膜',
        '整新備註','整新日期','整新人員',
        '看是否拆機','放現貨/料件更換','最終訂單/進貨編號','最終處理日期','最終處理人員',
        '福利等級','福利等級備註','通知福利上架日','是否售出福利機','是否停產',
    ]
    HEADER_MAP = {
        '舊檢測碼':'old_insp_no',
        '維修編號':'repair_no','型號':'model','SN':'serial_no','SN(人工掃)':'sn_manual',
        '原訂單編號':'order_no','客戶問題備註':'customer_issue',
        '換貨訂單編號':'exchange_order_no','取消日期':'cancel_date','取消人員':'cancel_staff',
        '圖書館版移除':'lib_removed','圖書館版處理日':'lib_process_date','圖書館版處理人員':'lib_process_staff',
        '封閉/開放':'open_closed','外盒機身SN保卡SN一致':'sn_consistent',
    }
    wb = Workbook(); ws = wb.active; ws.title = '退換貨檢測範本'
    hfill = PatternFill('solid', fgColor='1A237E')
    hfont = Font(bold=True, color='FFFFFF')
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 14
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='退換貨檢測_匯入範本.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/inspection/batch-import', methods=['POST'])
@login_required
def insp_batch_import():
    """批次匯入退換貨檢測"""
    import openpyxl
    from io import BytesIO
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '未上傳檔案'}), 400
    wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
    ws = wb.active
    raw_headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    HMAP = {
        '舊檢測碼':'old_insp_no',
        '維修編號':'repair_no','型號':'model','SN':'serial_no','SN(人工掃)':'sn_manual',
        '原訂單編號':'order_no','客戶問題備註':'customer_issue',
        '換貨訂單編號':'exchange_order_no','取消日期':'cancel_date','取消人員':'cancel_staff',
        '圖書館版移除':'lib_removed','圖書館版處理日':'lib_process_date','圖書館版處理人員':'lib_process_staff',
        '封閉/開放':'open_closed','外盒機身SN保卡SN一致':'sn_consistent',
        'emmc/Version':'emmc_version','ram/LCD':'ram_lcd','螢幕外觀':'screen_appearance',
        'sd card/Touch':'sd_touch','power key/Earphone':'power_earphone',
        'screen/Wi-Fi':'screen_wifi','screen/Wifi':'screen_wifi',
        'touch/Bluetooth':'touch_bluetooth','battery/MIC':'battery_mic',
        'led/Battery':'led_battery','Wi-Fi/SD':'wifi_sd','wifi/SD':'wifi_sd',
        'Bluetooth/Keyboard':'bluetooth_keyboard','gsensor/Brightness':'gsensor_brightness',
        'backlight/Storage':'backlight_storage','temperature/按鍵':'temperature_key',
        'usbtypec/其他':'usb_other','earphone/重力感應':'earphone_gravity',
        'waveform/筆記':'waveform_note','拍照測試':'photo_test',
        '巫師筆更換(10.3)':'wizard_pen','vcom值':'vcom',
        '機器外觀是否明顯刮損':'appearance_scratch','磁吸位置':'magnet_position',
        '傳輸線接電腦':'usb_cable_pc','Wifi配對':'wifi_pair','藍芽配對':'bt_pair',
        '插sim卡':'sim_card','充電紅>綠燈':'charge_led',
        '書店電子書試閱':'bookstore_trial','圖書館借書下載':'library_download',
        '保護殼/磁吸休眠':'case_sleep','保護殼/傳輸線插入拔出':'case_cable',
        '恢復原廠設定':'factory_reset','回到初始語言畫面關機':'language_screen',
        '清潔閱讀器':'clean_reader','檢測備註':'inspection_note',
        '檢測日期':'inspection_date','檢測人員':'inspection_staff',
        '是否拆機':'disassemble','拆機(福利機)備註':'disassemble_note',
        '拆機作業日期':'disassemble_date','拆機人員':'disassemble_staff',
        '外盒與保卡SN一致':'box_sn_consistent','外盒凹折嚴重':'box_damaged',
        '封套(ONE SC)':'sleeve_one_sc','螢幕透明膜(MINI)':'screen_film_mini',
        '果凍套(X/NOTE)':'jelly_case','PVC塑膠袋':'pvc_bag','保固卡':'warranty_card',
        '傳輸線/筆芯組':'usb_cable_pen','觸控筆更換並測試':'stylus_replacement','封膜':'seal_film',
        '整新備註':'refurb_note','整新日期':'refurb_date','整新人員':'refurb_staff',
        '看是否拆機':'final_disassemble','放現貨/料件更換':'final_stock',
        '最終訂單/進貨編號':'final_order_no','最終處理日期':'final_date','最終處理人員':'final_staff',
        '福利等級':'welfare_grade','福利等級備註':'welfare_grade_note',
        '通知福利上架日':'welfare_notify_date','是否售出福利機':'welfare_sold','是否停產':'discontinued',
    }
    col_map = {i: HMAP[h] for i, h in enumerate(raw_headers) if h in HMAP}
    # 取下一個流水號
    res = sb.table('system_config').select('value').eq('key','insp_next_no').execute()
    cur_no = (res.data or [{}])[0].get('value','D3001')
    prefix = 'D'; num = int(cur_no[1:])
    rows_to_insert = []; skip = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row): continue
        rec = {v: None for v in _INSP_FIELDS if v != 'insp_no'}
        for ci, db_field in col_map.items():
            if ci < len(row):
                val = row[ci]
                if val is not None:
                    rec[db_field] = str(val).strip() if not isinstance(val, str) else val.strip()
        rec['insp_no'] = f'{prefix}{num:04d}'; num += 1
        rows_to_insert.append(rec)
    if not rows_to_insert:
        return jsonify({'success':0,'skip':0,'error':0,'msg':'沒有可匯入的資料'})
    # 若有維修編號，自動帶入機型/SN（同時比對 repair_no 和 old_repair_no）
    repair_nos = list({r['repair_no'] for r in rows_to_insert if r.get('repair_no')})
    repair_map = {}  # key 可能是 repair_no 或 old_repair_no
    if repair_nos:
        for _field in ('repair_no', 'old_repair_no'):
            _off = 0
            while True:
                _batch = sb.table('repair_records').select('repair_no,old_repair_no,model,serial_no') \
                           .in_(_field, repair_nos).range(_off, _off+999).execute()
                for rr in (_batch.data or []):
                    # 以 repair_no 和 old_repair_no 兩個 key 都存入 map
                    if rr.get('repair_no'):
                        repair_map[rr['repair_no']] = rr
                    if rr.get('old_repair_no'):
                        repair_map[rr['old_repair_no']] = rr
                if len(_batch.data or []) < 1000: break
                _off += 1000
    for rec in rows_to_insert:
        rno = rec.get('repair_no')
        if rno and rno in repair_map:
            if not rec.get('model'):
                rec['model'] = repair_map[rno].get('model') or rec.get('model')
            if not rec.get('serial_no'):
                rec['serial_no'] = repair_map[rno].get('serial_no') or rec.get('serial_no')
    # 批次 INSERT 每批 50 筆
    BATCH = 50; ok = 0; err = 0
    for i in range(0, len(rows_to_insert), BATCH):
        batch = rows_to_insert[i:i+BATCH]
        try:
            sb.table('inspection_records').insert(batch).execute()
            ok += len(batch)
        except Exception as e:
            err += len(batch)
    # 更新流水號
    sb.table('system_config').update({'value': f'{prefix}{num:04d}'}).eq('key','insp_next_no').execute()
    return jsonify({'success':ok,'skip':skip,'error':err})

@app.route('/api/inspection/next-no')
@login_required
def insp_peek_next_no():
    res = sb.table('system_config').select('value').eq('key', 'insp_next_no').execute()
    cur = (res.data or [{}])[0].get('value', 'D3001')
    return jsonify({'next_no': cur})


@app.route('/api/sn/template', methods=['GET'])
@login_required
def sn_template():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = '新編SN匯入範本'

    headers  = ['建檔日期',    '填表人',   '維修編號(新/舊)', '富動/阿偉編號', '機型',              'SN',       '實測故障(因)']
    required = {'建檔日期', '填表人', 'SN'}
    notes    = ['YYYY-MM-DD（必填）', '填表人姓名（必填）', '對應維修記錄編號', '如：A123（可空白）',
                '自動帶入或手動填寫', '新編SN碼（必填）', '可空白，建議填寫']
    widths   = [14, 12, 18, 14, 30, 24, 40]

    hfill = PatternFill('solid', fgColor='1A5276')
    rfill = PatternFill('solid', fgColor='FFF2CC')
    hfont = Font(bold=True, color='FFFFFF')
    rfont = Font(bold=True, color='7B3F00')
    nfill = PatternFill('solid', fgColor='F2F2F2')
    nfont = Font(italic=True, color='888888')

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        if h in required:
            c.fill = rfill; c.font = rfont
        else:
            c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci - 1]

    for ci, n in enumerate(notes, 1):
        c = ws.cell(2, ci, n)
        c.fill = nfill; c.font = nfont
        c.alignment = Alignment(horizontal='center')

    # 範例資料
    ws.append(['2026-06-05', '許培菁', '7084', 'A1052', '07.Gaze Note Plus 7.8吋(黑)', 'TWHK08P2026130001', '無法開機'])
    ws.append(['2026-05-19', '鄭竹涵', '5455', 'A820',  '11.Gaze Mini 6吋(白)', 'TWHK06N2025130002', '無法開機'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='新編SN匯入範本.xlsx')


@app.route('/api/sn/batch-import', methods=['POST'])
@login_required
def batch_import_sn():
    from openpyxl import load_workbook
    import io
    from datetime import date as _date_cls, datetime as _dt_cls

    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400

    f = request.files['file']
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{str(e)}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'error': '檔案無資料列'}), 400

    raw_headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    def _map_h(h):
        h = h.strip()
        if '建檔' in h or '日期' in h:        return 'record_date'
        if '填表' in h:                        return 'form_filler'
        if '維修編號' in h or '維修' in h:    return 'repair_no'
        if '阿偉' in h or '富動' in h:        return 'awei_number'
        if '機型' in h:                        return 'model'
        if h.upper() == 'SN' or 'SN' in h.upper() and '阿偉' not in h: return 'sn'
        if '故障' in h or '實測' in h:        return 'actual_fault'
        return None

    col_map = {}
    for ci, h in enumerate(raw_headers):
        fk = _map_h(h)
        if fk and fk not in col_map:
            col_map[fk] = ci

    def _parse_date(val):
        if val is None:
            return None
        if isinstance(val, (_dt_cls, _date_cls)):
            return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y'):
            try:
                return _dt_cls.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
        return s or None

    def _get(row, fk):
        ci = col_map.get(fk)
        if ci is None or ci >= len(row):
            return None
        v = row[ci]
        if v is None:
            return None
        if isinstance(v, (_dt_cls, _date_cls)):
            return v
        s = str(v).strip()
        return s if s else None

    SKIP_NOTES = {'YYYY-MM-DD（必填）', '填表人姓名（必填）', '（說明）', ''}

    # ── 預先收集所有維修編號，批次查 repair_records 建立對照表 ──
    repair_nos = set()
    for row in rows[1:]:
        v = _get(row, 'repair_no')
        if v:
            repair_nos.add(str(v).strip())

    repair_map = {}  # repair_no / old_repair_no → {model, awei_number, actual_fault}
    if repair_nos:
        rno_list = list(repair_nos)
        _chunk = 50
        for i in range(0, len(rno_list), _chunk):
            chunk = rno_list[i:i+_chunk]
            try:
                res = sb.table('repair_records')                     .select('repair_no,old_repair_no,model,awei_number,actual_fault')                     .in_('repair_no', chunk).execute()
                for r in (res.data or []):
                    key = str(r.get('repair_no') or '').strip()
                    if key:
                        repair_map[key] = r
                # 也用 old_repair_no 查
                res2 = sb.table('repair_records')                     .select('repair_no,old_repair_no,model,awei_number,actual_fault')                     .in_('old_repair_no', chunk).execute()
                for r in (res2.data or []):
                    old_key = str(r.get('old_repair_no') or '').strip()
                    if old_key and old_key not in repair_map:
                        repair_map[old_key] = r
            except Exception:
                pass

    ok_cnt = skip_cnt = err_cnt = 0
    batch = []

    for row in rows[1:]:
        first = str(row[0]).strip() if row[0] is not None else ''
        if first in SKIP_NOTES:
            skip_cnt += 1
            continue

        sn_val   = _get(row, 'sn')
        rd_val   = _parse_date(_get(row, 'record_date'))
        ff_val   = _get(row, 'form_filler')

        if not sn_val or not rd_val or not ff_val:
            skip_cnt += 1
            continue

        rno_val = _get(row, 'repair_no')
        rno_key = str(rno_val).strip() if rno_val else ''
        matched = repair_map.get(rno_key, {})

        # Excel 有值優先，空白才從維修記錄帶入
        awei_val  = _get(row, 'awei_number')  or matched.get('awei_number') or None
        model_val = _get(row, 'model')         or matched.get('model')        or None
        fault_val = _get(row, 'actual_fault')  or matched.get('actual_fault') or None

        rec = {
            'record_date':  rd_val,
            'form_filler':  ff_val,
            'repair_no':    rno_val,
            'awei_number':  awei_val,
            'model':        model_val,
            'sn':           sn_val,
            'actual_fault': fault_val,
        }
        # 空字串轉 None
        rec = {k: (None if isinstance(v, str) and v == '' else v) for k, v in rec.items()}
        batch.append(rec)

        if len(batch) >= 80:
            try:
                sb.table('sn_records').insert(batch).execute()
                ok_cnt += len(batch)
            except Exception:
                err_cnt += len(batch)
            batch = []

    if batch:
        try:
            sb.table('sn_records').insert(batch).execute()
            ok_cnt += len(batch)
        except Exception:
            err_cnt += len(batch)

    return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt})



# ─────────────────── 資料上稿 API ───────────────────

# ── 機型對照表 ──────────────────────────────────────
@app.route('/api/sn-model-refs', methods=['GET'])
@login_required
def list_sn_model_refs():
    q = sb.table('sn_model_refs').select('*').order('seq', desc=False)
    kw = (request.args.get('q') or '').strip()
    if kw:
        like = f'%{kw}%'
        q = q.or_(f'model_name.ilike.{like},prefix.ilike.{like},type_letter.ilike.{like}')
    res = q.execute()
    return jsonify({'rows': res.data or []})

@app.route('/api/sn-model-refs', methods=['POST'])
@login_required
def create_sn_model_ref():
    data = {k: v for k, v in (request.json or {}).items()
            if k in {'seq','model_name','prefix','size','type_letter'}}
    try:
        res = sb.table('sn_model_refs').insert(data).execute()
        return jsonify({'ok': True, 'id': res.data[0]['id']}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sn-model-refs/<int:rid>', methods=['PUT'])
@login_required
def update_sn_model_ref(rid):
    data = {k: v for k, v in (request.json or {}).items()
            if k in {'seq','model_name','prefix','size','type_letter'}}
    try:
        sb.table('sn_model_refs').update(data).eq('id', rid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sn-model-refs/<int:rid>', methods=['DELETE'])
@admin_required
def delete_sn_model_ref(rid):
    try:
        sb.table('sn_model_refs').delete().eq('id', rid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sn-model-refs/template', methods=['GET'])
@login_required
def sn_model_refs_template():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    import io
    wb = Workbook(); ws = wb.active; ws.title = '機型對照表'
    headers  = ['#', '機型', '字母(TW=Taiwan、H=HyRead、R/K/M=CPU型號)', '機型尺寸', '字母(N非手寫/P可手寫/C彩機/CC進階彩機)']
    notes    = ['序號', '如：01.Gaze 7.8吋(白)', '如：TWHR / TWHK / TWHM', '如：06 / 08 / 10', '如：N / P / C / CC / NU']
    widths   = [6, 36, 38, 12, 38]
    hfill = PatternFill('solid', fgColor='1A5276'); hfont = Font(bold=True, color='FFFFFF')
    nfill = PatternFill('solid', fgColor='F2F2F2'); nfont = Font(italic=True, color='888888')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci-1]
    for ci, n in enumerate(notes, 1):
        c = ws.cell(2, ci, n); c.fill = nfill; c.font = nfont
        c.alignment = Alignment(horizontal='center')
    # 現有資料
    res = sb.table('sn_model_refs').select('*').order('seq', desc=False).execute()
    for r in (res.data or []):
        ws.append([r.get('seq',''), r.get('model_name',''), r.get('prefix',''), r.get('size',''), r.get('type_letter','')])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='機型對照表範本.xlsx')

@app.route('/api/sn-model-refs/import', methods=['POST'])
@login_required
def import_sn_model_refs():
    from openpyxl import load_workbook
    import io
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    try:
        wb = load_workbook(io.BytesIO(request.files['file'].read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析：{str(e)}'}), 400
    rows = list(ws.iter_rows(values_only=True))
    SKIP = {'#', '序號', '', None}
    ok_cnt = skip_cnt = err_cnt = 0
    batch = []
    for row in rows[1:]:
        first = str(row[0]).strip() if row[0] is not None else ''
        if first in SKIP or (len(first) > 0 and not first[0].isdigit()):
            skip_cnt += 1; continue
        def _v(i): return str(row[i]).strip() if i < len(row) and row[i] is not None else None
        model_val = _v(1); prefix_val = _v(2); size_val = _v(3); type_val = _v(4)
        if not model_val:
            skip_cnt += 1; continue
        try: seq_val = int(float(first))
        except: seq_val = None
        batch.append({'seq': seq_val, 'model_name': model_val, 'prefix': prefix_val or None,
                      'size': size_val or None, 'type_letter': type_val or None})
        if len(batch) >= 80:
            try: sb.table('sn_model_refs').insert(batch).execute(); ok_cnt += len(batch)
            except: err_cnt += len(batch)
            batch = []
    if batch:
        try: sb.table('sn_model_refs').insert(batch).execute(); ok_cnt += len(batch)
        except: err_cnt += len(batch)
    return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt})


# ── SN碼對照 ──────────────────────────────────────
@app.route('/api/sn-code-refs', methods=['GET'])
@login_required
def list_sn_code_refs():
    q = sb.table('sn_code_refs').select('*').order('id', desc=False)
    kw = (request.args.get('q') or '').strip()
    if kw:
        like = f'%{kw}%'
        q = q.or_(f'sn_code.ilike.{like},prod_year.ilike.{like},serial_no.ilike.{like}')
    res = q.execute()
    return jsonify({'rows': res.data or []})

@app.route('/api/sn-code-refs', methods=['POST'])
@login_required
def create_sn_code_ref():
    data = {k: v for k, v in (request.json or {}).items()
            if k in {'prod_year','prod_month','serial_no','sn_code'}}
    try:
        res = sb.table('sn_code_refs').insert(data).execute()
        return jsonify({'ok': True, 'id': res.data[0]['id']}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sn-code-refs/<int:rid>', methods=['DELETE'])
@admin_required
def delete_sn_code_ref(rid):
    try:
        sb.table('sn_code_refs').delete().eq('id', rid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sn-code-refs/template', methods=['GET'])
@login_required
def sn_code_refs_template():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    import io
    wb = Workbook(); ws = wb.active; ws.title = 'SN碼對照'
    headers = ['生產年份/重編年份', '生產月份/重編13', '流水號', 'SN碼']
    notes   = ['如：2026', '1-12=生產月份；13=重編', '如：0001', '自動組合或手動填入']
    widths  = [20, 20, 14, 28]
    hfill = PatternFill('solid', fgColor='1A5276'); hfont = Font(bold=True, color='FFFFFF')
    nfill = PatternFill('solid', fgColor='F2F2F2'); nfont = Font(italic=True, color='888888')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci-1]
    for ci, n in enumerate(notes, 1):
        c = ws.cell(2, ci, n); c.fill = nfill; c.font = nfont
        c.alignment = Alignment(horizontal='center')
    res = sb.table('sn_code_refs').select('*').order('id', desc=False).execute()
    for r in (res.data or []):
        ws.append([r.get('prod_year',''), r.get('prod_month',''), r.get('serial_no',''), r.get('sn_code','')])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='SN碼對照範本.xlsx')

@app.route('/api/sn-code-refs/import', methods=['POST'])
@login_required
def import_sn_code_refs():
    from openpyxl import load_workbook
    import io
    from datetime import date as _dc, datetime as _dtc
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    try:
        wb = load_workbook(io.BytesIO(request.files['file'].read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析：{str(e)}'}), 400
    rows = list(ws.iter_rows(values_only=True))
    SKIP_FIRST = {'生產年份/重編年份', 'yyyy', '', None}
    ok_cnt = skip_cnt = err_cnt = 0
    batch = []
    for row in rows[1:]:
        first = str(row[0]).strip() if row[0] is not None else ''
        if first in SKIP_FIRST:
            skip_cnt += 1; continue
        def _v(i):
            if i >= len(row) or row[i] is None: return None
            v = row[i]
            if isinstance(v, (_dc, _dtc)): return str(v.year)
            return str(v).strip() or None
        yr = _v(0); mo = _v(1); sn_serial = _v(2); sn_code = _v(3)
        if not yr:
            skip_cnt += 1; continue
        batch.append({'prod_year': yr, 'prod_month': mo, 'serial_no': sn_serial, 'sn_code': sn_code})
        if len(batch) >= 80:
            try: sb.table('sn_code_refs').insert(batch).execute(); ok_cnt += len(batch)
            except: err_cnt += len(batch)
            batch = []
    if batch:
        try: sb.table('sn_code_refs').insert(batch).execute(); ok_cnt += len(batch)
        except: err_cnt += len(batch)
    return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt})


# ─────────────────── 新編SN API ───────────────────
_SN_COLS = {'record_date', 'form_filler', 'repair_no', 'awei_number', 'model', 'sn', 'actual_fault', 'updated_at'}

def _clean_sn(data: dict) -> dict:
    out = {}
    for k, v in data.items():
        if k not in _SN_COLS:
            continue
        if isinstance(v, str) and v.strip() == '':
            out[k] = None
        else:
            out[k] = v
    return out

@app.route('/api/sn', methods=['GET'])
@login_required
def list_sn():
    p = request.args
    q = sb.table('sn_records').select('*')
    kw = (p.get('q') or '').strip()
    if kw:
        like = f'%{kw}%'
        q = q.or_(
            f'repair_no.ilike.{like},'
            f'awei_number.ilike.{like},'
            f'model.ilike.{like},'
            f'sn.ilike.{like},'
            f'actual_fault.ilike.{like},'
            f'form_filler.ilike.{like}'
        )
    if p.get('form_filler'):
        q = q.eq('form_filler', p['form_filler'])
    if p.get('date_f'):
        q = q.gte('record_date', p['date_f'])
    if p.get('date_t'):
        q = q.lte('record_date', p['date_t'])

    sort_col  = p.get('sort', 'id')
    sort_desc = p.get('desc', '1') == '1'
    allowed_sort = {'id', 'record_date', 'form_filler', 'repair_no', 'awei_number', 'model', 'sn', 'created_at'}
    if sort_col not in allowed_sort:
        sort_col = 'id'
    q = q.order(sort_col, desc=sort_desc)

    try:
        page     = max(1, int(p.get('page', 1)))
        per_page = min(200, max(10, int(p.get('per_page', 50))))
    except ValueError:
        page, per_page = 1, 50

    total_res = q.execute()
    total = len(total_res.data or [])
    offset = (page - 1) * per_page
    rows = (total_res.data or [])[offset: offset + per_page]
    return jsonify({'total': total, 'page': page, 'per_page': per_page, 'rows': rows})


@app.route('/api/sn', methods=['POST'])
@login_required
def create_sn():
    data = _clean_sn(request.json or {})
    data['created_by']       = session['user_id']
    data['updated_by']       = session['user_id']
    data['created_by_name']  = session.get('display_name', session.get('username', ''))
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    try:
        res = sb.table('sn_records').insert(data).execute()
        return jsonify({'ok': True, 'id': res.data[0]['id']}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sn/<int:sid>', methods=['PUT'])
@login_required
def update_sn(sid):
    data = _clean_sn(request.json or {})
    data['updated_by']       = session['user_id']
    data['updated_by_name']  = session.get('display_name', session.get('username', ''))
    data.pop('created_by', None)
    data.pop('created_by_name', None)
    try:
        sb.table('sn_records').update(data).eq('id', sid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sn/<int:sid>', methods=['DELETE'])
@admin_required
def delete_sn(sid):
    try:
        sb.table('sn_records').delete().eq('id', sid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sn/export', methods=['GET'])
@login_required
def export_sn():
    from flask import send_file
    import io, openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    p = request.args
    q = sb.table('sn_records').select('*')
    kw = (p.get('q') or '').strip()
    if kw:
        like = f'%{kw}%'
        q = q.or_(
            f'repair_no.ilike.{like},'
            f'awei_number.ilike.{like},'
            f'model.ilike.{like},'
            f'sn.ilike.{like},'
            f'actual_fault.ilike.{like},'
            f'form_filler.ilike.{like}'
        )
    if p.get('form_filler'):
        q = q.eq('form_filler', p['form_filler'])
    if p.get('date_f'):
        q = q.gte('record_date', p['date_f'])
    if p.get('date_t'):
        q = q.lte('record_date', p['date_t'])

    q = q.order('id', desc=False)

    _PAGE = 1000
    _offset = 0
    all_rows = []
    while True:
        batch = q.range(_offset, _offset + _PAGE - 1).execute()
        rows = batch.data or []
        all_rows.extend(rows)
        if len(rows) < _PAGE:
            break
        _offset += _PAGE

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '新編SN'
    headers = ['建檔日期', '填表人', '維修編號(新/舊)', '富動/阿偉編號', '機型', 'SN', '實測故障(因)', '異動日期']
    col_keys = ['record_date', 'form_filler', 'repair_no', 'awei_number', 'model', 'sn', 'actual_fault', 'updated_at']
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(all_rows, 2):
        for ci, key in enumerate(col_keys, 1):
            ws.cell(ri, ci, row.get(key, '') or '')
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = quote('新編SN記錄.xlsx')
    return send_file(buf, download_name='新編SN記錄.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# -- Attachment API --
@app.route('/api/attachments/<record_type>/<int:record_id>', methods=['GET'])
@login_required
def list_attachments(record_type, record_id):
    if record_type not in ('repair', 'exchange'):
        return jsonify({'error': 'invalid record_type'}), 400
    res = sb.table('repair_attachments').select('*') \
            .eq('record_type', record_type).eq('record_id', record_id) \
            .order('created_at').execute()
    rows = res.data or []
    for r in rows:
        r['url'] = sb.storage(_ATT_BUCKET).public_url(r['storage_path'])
    return jsonify(rows)

@app.route('/api/attachments/upload', methods=['POST'])
@login_required
def upload_attachments():
    record_type = request.form.get('record_type', '').strip()
    record_id   = request.form.get('record_id', '').strip()
    if record_type not in ('repair', 'exchange') or not record_id:
        return jsonify({'error': 'missing params'}), 400
    try:
        record_id = int(record_id)
    except ValueError:
        return jsonify({'error': 'record_id must be int'}), 400
    files = request.files.getlist('files')
    notes = request.form.getlist('notes')
    if not files:
        return jsonify({'error': 'no files'}), 400
    saved = []; errors = []
    import uuid, mimetypes
    for idx, f in enumerate(files):
        if not f or not f.filename: continue
        note = notes[idx] if idx < len(notes) else ''
        data = f.read()
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'bin'
        content_type = f.content_type or mimetypes.guess_type(f.filename)[0] or 'application/octet-stream'
        storage_path = f"{record_type}/{record_id}/{uuid.uuid4().hex}.{ext}"
        try:
            sb.storage(_ATT_BUCKET).upload(storage_path, data, content_type)
        except Exception as e:
            errors.append(f'{f.filename}: {e}'); continue
        row = sb.table('repair_attachments').insert({
            'record_type': record_type, 'record_id': record_id,
            'filename': f.filename, 'storage_path': storage_path, 'note': note,
        }).execute()
        if row.data:
            r = row.data[0]
            r['url'] = sb.storage(_ATT_BUCKET).public_url(storage_path)
            saved.append(r)
    return jsonify({'saved': saved, 'errors': errors})

@app.route('/api/attachments/<int:att_id>', methods=['PUT'])
@login_required
def update_attachment_note(att_id):
    data = request.get_json() or {}
    note = data.get('note', '')
    res = sb.table('repair_attachments').update({'note': note}).eq('id', att_id).execute()
    if not res.data:
        return jsonify({'error': 'not found'}), 404
    return jsonify({'ok': True})

@app.route('/api/attachments/<int:att_id>', methods=['DELETE'])
@login_required
def delete_attachment(att_id):
    res = sb.table('repair_attachments').select('storage_path').eq('id', att_id).execute()
    if not res.data:
        return jsonify({'error': 'not found'}), 404
    storage_path = res.data[0]['storage_path']
    try:
        sb.storage(_ATT_BUCKET).remove([storage_path])
    except Exception:
        pass
    sb.table('repair_attachments').delete().eq('id', att_id).execute()
    return jsonify({'ok': True})

@app.route('/api/staff-work')
@login_required
def staff_work_stats():
    """人員工作記錄統計：各模組每人每月建檔數"""
    try:
        year = int(request.args.get('year', 2026))
    except ValueError:
        year = 2026
    y_start = f"{year}-01-01"
    y_end   = f"{year}-12-31"

    MODULE_CFG = [
        ('維修記錄', 'repair_records',     'form_filler',     'fill_date',   False),
        ('客服換貨', 'exchange_orders',    'cs_staff',        'fill_date',   False),
        ('維修追蹤', 'repair_tracking',    'created_by_name', 'created_at',  True),
        ('退換貨檢', 'inspection_records', 'updated_by_name', 'updated_at',  True),
        ('退費追蹤', 'refund_tracking',    'form_filler',     'order_date',  False),
        ('福利品',   'welfare_products',   'created_by_name', 'created_at',  True),
        ('新增SN',   'sn_records',         'form_filler',     'record_date', False),
    ]

    result     = {}
    all_people = set()

    for mod_name, table, p_col, d_col, is_dt in MODULE_CFG:
        mod_data = {}
        try:
            PAGE   = 1000
            offset = 0
            d_end  = y_end + 'T23:59:59' if is_dt else y_end
            while True:
                rows = (sb.table(table)
                          .select(f'{p_col},{d_col}')
                          .gte(d_col, y_start)
                          .lte(d_col, d_end)
                          .range(offset, offset + PAGE - 1)
                          .execute().data or [])
                for r in rows:
                    person = (r.get(p_col) or '').strip()
                    date_s = str(r.get(d_col) or '')
                    if not person or not date_s:
                        continue
                    try:
                        month = int(date_s[5:7])
                    except Exception:
                        continue
                    if 1 <= month <= 12:
                        if person not in mod_data:
                            mod_data[person] = [0] * 12
                        mod_data[person][month - 1] += 1
                        all_people.add(person)
                if len(rows) < PAGE:
                    break
                offset += PAGE
        except Exception:
            pass
        result[mod_name] = mod_data

    people = sorted(all_people)
    return jsonify({
        'year':    year,
        'people':  people,
        'modules': [c[0] for c in MODULE_CFG],
        'data':    result,
    })


if __name__ == '__main__':
    app.run(debug=True)
