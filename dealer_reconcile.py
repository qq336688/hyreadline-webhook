"""
HyRead 維修記錄系統 — 經銷商對帳模組（獨立 Blueprint）
版本：v1.0  2026-07-21

設計原則：
  - 這支檔案完全獨立於 app.py 的既有邏輯，app.py 只需要在檔案最尾端
    `from dealer_reconcile import dealer_bp` + `app.register_blueprint(dealer_bp)`
    兩行即可掛載，降低對既有大檔案的異動風險。
  - 共用同一個 Flask session 登入機制與 Supabase 專案（sb 由 app.py 建立後於本檔匯入）。
  - 所有 API 路徑皆以 /api/dealer/ 開頭，避免與既有路由衝突。

經銷商格式規則 sheets_config（JSONB）欄位說明（每筆代表一個要解析的分頁）：
  sheet_name       分頁名稱關鍵字（優先完全比對，找不到再用「包含」比對）
  header_row       表頭列在第幾列（1-based）
  row_type_mode    'fixed'：整個分頁都是同一種（銷貨/退貨）
                   'by_column'：依某欄位的值逐列判斷銷貨/退貨（如三創的「異動別」欄）
  row_type_fixed   row_type_mode='fixed' 時使用，值為 'sale' 或 'return'
  col_row_type     row_type_mode='by_column' 時，該欄位的表頭文字
  return_values    逗號分隔，col_row_type 欄位中代表「退貨」的值（如 "退貨"）；其餘視為銷貨
  col_dealer_code  經銷商代碼欄位的表頭文字
  col_dealer_code2 副代碼欄位的表頭文字（可留空）。有些經銷商主代碼不足以唯一識別商品
                   （如富邦momo的「品號」同一品號下常有多個顏色/款式，需搭配「單品編號」
                   才能唯一區分），此時填入第二個欄位，系統會自動組成「主代碼-副代碼」
                   （如 "13524865-002"）再進行比對，避免不同商品被誤判成同一筆
  col_ean          EAN/國際條碼欄位的表頭文字
  col_product_name 商品名稱欄位的表頭文字（必填，用來判斷是否為有效資料列）
  col_product_name2 副商品名稱欄位的表頭文字（可留空）。有些經銷商的「商品名稱」欄不含顏色/款式
                   （如富邦momo同一品號下不同色的商品名稱完全相同，顏色寫在另一欄「單品說明」），
                   此時填入第二個欄位，系統會自動組成「商品名稱 - 副商品名稱」方便人工比對時辨識
  col_qty          數量欄位的表頭文字
  col_unit_price   單價欄位的表頭文字（可留空）
  col_amount       金額欄位的表頭文字（舊版相容用，新規則建議改填下面兩個欄位其中之一）
  col_amount_untaxed 未稅金額欄位的表頭文字（可留空）。此經銷商對帳檔的金額若為未稅，填在這裡
  col_amount_taxed   含稅金額欄位的表頭文字（可留空）。此經銷商對帳檔的金額若為含稅，填在這裡
                   （v2.59起金額欄位拆分為含稅/未稅兩種，同一分頁通常只會填其中一個；
                   若兩個都留空，會退回讀取舊版 col_amount 欄位，但此時 amount_taxed/
                   amount_untaxed 兩個新欄位會是0，財報畫面會明顯看出尚未設定稅別分類）
  col_tax_amount   稅額欄位的表頭文字（可留空）。供財報匯出使用，不影響比對邏輯
"""

import io
import json
import os
import re
import uuid
from datetime import datetime, timedelta
from functools import wraps

import openpyxl
import pdfplumber
from flask import Blueprint, request, jsonify, session, send_file
from werkzeug.utils import secure_filename

from app import sb, now_str, admin_required  # noqa: 延後匯入，app.py 需在 register_blueprint 前已定義好 sb

dealer_bp = Blueprint('dealer_reconcile', __name__, template_folder='templates')

DEALERS = ['三創', '富邦momo', '墊腳石', '親子天下', '熊老闆', '香港經銷商', '展碁', '非銷售開發票', '單位團體(知識)', '書店']


# ============================================================
# 頁面路由（獨立頁面，不掛在 index.html 側邊欄內）
# ============================================================
@dealer_bp.route('/dealer-reconcile')
def dealer_reconcile_page():
    from flask import render_template
    return render_template('dealer_reconcile.html')


# ============================================================
# 權限與共用小工具
# ============================================================
def _login_ok():
    return 'user_id' in session


def _perm_ok(key):
    if session.get('is_admin'):
        return True
    return bool(session.get('permissions', {}).get(key))


def dealer_view_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not (_perm_ok('mod_dealer_view') or _perm_ok('mod_dealer_edit')):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def dealer_edit_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not _perm_ok('mod_dealer_edit'):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


_BOOKSTORE_PLAN_PERM_KEYS = (
    'mod_bookstore_plan_view', 'mod_bookstore_plan_edit', 'mod_bookstore_plan_create',
    'mod_bookstore_plan_update', 'mod_bookstore_plan_delete', 'mod_bookstore_plan_export_template',
    'mod_bookstore_plan_batch_import', 'mod_bookstore_plan_export',
    # 2026-08-12新增：「簡易批次新增」獨立權限，加進來才能讓只被授權這一項的帳號也能打開
    # 「方案清單」頁面（跟其他細分權限一樣，見_any_bookstore_plan_perm()的既有教訓）。
    'mod_bookstore_plan_simple_import',
)


def _any_bookstore_plan_perm():
    """2026-08-08新增：只要方案清單相關的10個細分權限（含檢視/整體編輯）任一個為true，
    就代表這個帳號多少需要用到「方案清單」／「匯入訂單」——修正實際使用時發現的落差：
    「企劃新增方案」這個權限群組只勾了[新增方案][修改]，沒有勾[檢視]，若嚴格只認
    mod_bookstore_plan_view/edit，會讓她連自己有權限新增/修改的方案清單畫面都打不開，
    等於這兩個已勾選的權限形同虛設。"""
    return any(_perm_ok(k) for k in _BOOKSTORE_PLAN_PERM_KEYS)


def bookstore_view_required(f):
    """書店經銷商方案清單「檢視」權限：獨立於 mod_dealer_view/mod_dealer_edit 的權限，可單獨
    開放給特定人員檢視/編輯方案清單，不需要一併給予經銷商對帳模組的其他權限。
    2026-08-08修正：原本只認 mod_bookstore_plan_view／mod_bookstore_plan_edit 兩項，改成只要
    _any_bookstore_plan_perm() 任一細分權限為true即可（理由見該函式註解）。
    2026-08-08再修正：原本這個裝飾器也同時保護「匯入訂單」的分析/匯出API，但使用者實測
    發現「匯入訂單」應該要是完全獨立的權限，不應該因為勾了[新增方案][修改]等方案清單細分
    權限就連帶打開——因此「匯入訂單」改用下面新增的 bookstore_order_view_required，這裡
    只保留給「方案清單」列表本身的檢視使用。"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not _any_bookstore_plan_perm():
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def bookstore_order_view_required(f):
    """2026-08-08新增：「匯入訂單」拆帳分析/匯出的專用檢視權限，獨立於「方案清單」的新增/
    修改/刪除等細分權限之外（見 mod_bookstore_order_view，dealer_reconcile_schema.sql
    v1.19）。背景：先前「匯入訂單」共用 bookstore_view_required（_any_bookstore_plan_perm()），
    導致只被授權「新增方案」「修改」方案清單的帳號，也連帶能使用「匯入訂單」——但這是
    使用者刻意想分開設定的兩個功能，「新增方案/修改」不應該隱含「匯入訂單」的存取權。
    這裡仍然把 mod_bookstore_plan_edit／mod_bookstore_plan_view 算進來（向下相容：既有
    已經設定「完整編輯」或原本「方案清單/匯入訂單(檢視)」的帳號，權限不會因為這次拆分
    而被收回），mod_bookstore_order_view 則是新增的、可以「只開放匯入訂單、不開放方案
    清單其他動作」的獨立選項。"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not (_perm_ok('mod_bookstore_plan_edit') or _perm_ok('mod_bookstore_plan_view')
                or _perm_ok('mod_bookstore_order_view')):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def bookstore_edit_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not _perm_ok('mod_bookstore_plan_edit'):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def bookstore_invoice_required(f):
    """2026-08-12新增：「匯入團體發票」專用權限，跟「方案清單」「匯入訂單」都各自獨立
    （見 mod_bookstore_invoice_view，dealer_reconcile_schema.sql v1.20）。這個功能純粹是
    團體發票的登記/查詢，不跟訂單或方案清單的拆帳資料比對，所以刻意不沿用
    _any_bookstore_plan_perm()／bookstore_order_view_required 的判斷邏輯，只認
    mod_dealer_edit（既有的「完整編輯」master switch）跟這次新增的獨立欄位。"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not (_perm_ok('mod_dealer_edit') or _perm_ok('mod_bookstore_invoice_view')):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def bookstore_action_required(*keys):
    """2026-08-08新增：方案清單細分權限（修改/刪除/新增方案/範本匯出/批次匯入/匯出各自獨立
    勾選），比照使用者需求「各子模組皆能設定權限角色」。設計成 mod_bookstore_plan_edit
    （既有的整體編輯權限）仍然當作「萬用」總開關——只要群組有勾這個，底下所有細分動作都
    照舊可以做，避免這次上線後既有已設定 mod_bookstore_plan_edit=true 的帳號突然失去權限；
    細分權限是給「只想開放單一動作」的情境用的，兩者是 OR 的關係，符合其中一個就放行。
    傳入多個key時（例如「範本/匯出」跟「匯出」兩顆按鈕實際上呼叫同一支API），符合任一個即可。"""
    def deco(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not _login_ok():
                return jsonify({'error': '請先登入'}), 401
            if not (_perm_ok('mod_bookstore_plan_edit') or any(_perm_ok(k) for k in keys)):
                return jsonify({'error': '權限不足'}), 403
            return f(*args, **kwargs)
        return wrapped
    return deco


def sku_view_required(f):
    """2026-08-08新增：內部料號清單「檢視」專用權限，比 dealer_view_required 多開放一個
    來源——只要有方案清單相關任一權限（_any_bookstore_plan_perm()）也算數。背景：「方案
    清單」畫面本身就需要讀取內部料號清單（財報料號品名下拉選單、方案清單欄位、批次匯入/
    匯出範本欄位都直接取自這裡），如果內部料號清單的檢視API仍然只認mod_dealer_view/edit，
    只被授權「方案清單」相關權限、完全沒有「經銷商對帳模組」權限的帳號，會連方案清單自己
    需要的基礎資料都讀不到（下拉選單空白、欄位讀不出來），等於方案清單功能形同壞掉。這裡
    只放寬「檢視」這一支API，內部料號的新增/修改/刪除/排序/批次匯入仍然維持只有
    dealer_edit_required（經銷商對帳模組本身的編輯權限）才能動，不受這次調整影響。"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not (_perm_ok('mod_dealer_view') or _perm_ok('mod_dealer_edit') or _any_bookstore_plan_perm()):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def _audit_new():
    name = session.get('display_name', session.get('username', ''))
    return {
        'created_by': session.get('user_id'),
        'created_by_name': name,
        'updated_by': session.get('user_id'),
        'updated_by_name': name,
    }


def _audit_upd():
    # 2026-08-08修正：補上 updated_at 主動寫入目前時間。既有的呼叫端原本只更新
    # updated_by/updated_by_name，資料庫欄位雖然有 updated_at DEFAULT NOW()，但那只在
    # 新增時套用一次，之後每次更新都不會自動刷新（沒有DB觸發器），導致畫面上如果顯示
    # 「異動時間」，看到的其實永遠是「建立時間」而非真正最後修改時間——這是先前開發
    # bookstore_sku_sync_status（[[v2.92]]）時就已經發現的同一個根因，這裡從源頭修正，
    # 讓所有呼叫 _audit_upd() 的既有CRUD（不只書店經銷商模組）都一併修好。
    name = session.get('display_name', session.get('username', ''))
    return {'updated_by': session.get('user_id'), 'updated_by_name': name, 'updated_at': now_str()}


def _fetch_all_rows(qb_factory, page_size=1000):
    """依 PostgREST 預設 max-rows 上限（本專案實測約1000筆，參見既有 export_sku_list()
    的作法）分頁抓取「全部」資料，避免大型資料表（bookstore_plans／bookstore_plan_items／
    internal_sku_list 皆已成長超過此上限）在單次不帶 .range() 的 select().execute() 時被
    PostgREST 靜默截斷成只有前1000筆左右——且不會丟出任何錯誤，非常容易被誤判成「資料
    真的不存在」。2026-08-08新增：修正「方案清單裡查得到、但『匯入訂單』分析卻顯示查無
    對應方案」的問題，根因就是 _compute_bookstore_order_matrix() 原本用不分頁的
    select('*').execute() 抓 bookstore_plans，資料超過1000筆後，較晚建立/匯入的方案
    （如id較大的1035號）就有可能被截斷掉、比對不到。

    qb_factory：一個「不帶參數、每次呼叫都回傳一個全新 QueryBuilder」的函式（不可傳入已經
    呼叫過 .execute() 的 builder），因為 .range() 每一頁都要重新套用在新的 query 物件上。"""
    all_rows = []
    _off = 0
    while True:
        res = qb_factory().range(_off, _off + page_size - 1).execute()
        batch_rows = res.data or []
        all_rows.extend(batch_rows)
        if len(batch_rows) < page_size:
            break
        _off += page_size
    return all_rows


def _norm(v):
    if v is None:
        return ''
    return str(v).strip()


def _to_num(v):
    try:
        if v is None or v == '':
            return 0
        return float(v)
    except (TypeError, ValueError):
        return 0


def _period_sort_key(batch_rec, dname, prd):
    """對帳查詢矩陣列排序用：「對帳順序」欄位若填了可解析成數字的值，依數字由小到大排最前面
    （使用者手動指定的排序優先權最高，不受下面的拆帳編號排序影響）；未填或無法解析為數字
    時，排在所有已填數字的列之後。
    2026-07-24變更：對帳順序改掛在批次本身，故傳入的是該列對應的批次資料(batch_rec)，
    不再是依(經銷商,期間)查詢的 notes_map。
    2026-08-21變更：「對帳順序」都沒填時，原本退回依(經銷商,期間)字母順序排列，但期間欄位
    格式常常不統一（如「0801-0810」跟「2026/07」混雜），字母排序結果對使用者來說很亂、
    看不出關聯。使用者反饋希望改成依「拆帳編號」(recon_seq_no，格式C+4位數字，同一次
    上傳的批次固定沿用建立時的流水號，不會因為排序變動而改變)由小到大排序——這本來就是
    系統依上傳時間自動產生的序號，用它排序等於「依上傳先後順序」呈現，比字母排序更直覺。
    仍然維持(經銷商,期間)字母順序作為recon_seq_no不存在時（理論上不會發生，只是保險）
    的最後備援排序依據。"""
    ro = (batch_rec or {}).get('reconcile_order') or ''
    seq = (batch_rec or {}).get('recon_seq_no') or ''
    try:
        seq_n = int(seq[1:]) if seq[:1] == 'C' and seq[1:].isdigit() else 999999
    except (TypeError, ValueError):
        seq_n = 999999
    try:
        return (0, float(ro), seq_n, dname, prd)
    except (TypeError, ValueError):
        return (1, 0.0, seq_n, dname, prd)


def _to_num_or_none(v):
    """數字型欄位（產品定價、商品進貨未稅價）用：留空回傳 None，避免覆蓋成 0。"""
    try:
        if v is None or v == '':
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _sku_seq_max_n():
    """回傳 internal_sku_list.sku_seq_no 目前已用過的最大流水號（整數，無資料時回傳0）。
    2026-08-24新增，比照_next_recon_seq_no()「C+4位數流水號」的既有慣例（拆帳編號），
    差別是internal_sku_list資料量已超過1000筆（見list_sku()／export_sku_list()等處皆已
    改用分頁查詢的既有慣例），若像_next_recon_seq_no()一樣用不分頁的單次查詢，會被
    PostgREST靜默截斷、誤判出錯誤（偏小）的下一號，所以這裡改用_fetch_all_rows()分頁
    抓取全部既有序號。供_next_sku_seq_no()（單筆新增）與「批次新增匯入」（一次匯入多筆，
    自行在本地遞增，避免每列都查一次資料庫拿到重複序號）共用。"""
    rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('sku_seq_no').not_.is_('sku_seq_no', 'null'))
    max_n = 0
    for r in rows:
        v = (r.get('sku_seq_no') or '').strip()
        if v[:1] == 'C' and v[1:].isdigit():
            max_n = max(max_n, int(v[1:]))
    return max_n


def _next_sku_seq_no():
    """單筆新增料號(create_sku())專用：產生下一個「料號序號」，格式C+4位數流水號（如
    C0001、C0002…）。批次新增匯入一次匯入多筆時請改呼叫_sku_seq_max_n()自行遞增，不要
    在迴圈裡重複呼叫這支函式——同一批次裡後面的列會重新查到跟前面的列一樣的「目前最大值」，
    導致序號重複。"""
    return f'C{_sku_seq_max_n() + 1:04d}'


# 內部料號批次匯入/新增共用的欄位名稱比對規則——2026-08-24拆分「批次修改匯入」/「批次新增
# 匯入」兩支API時抽出成共用函式，避免兩邊各自維護一份、之後修改欄位規則忘記同步更新。
_SKU_NUMERIC_FIELDS = {'list_price', 'purchase_price_notax'}
_SKU_BOOL_FIELDS = {'is_accessory'}
# 2026-08-24補修正：拆分批次匯入時漏帶了「方案清單顯示」(is_active)/「發票總表顯示」
# (hide_in_invoice_ledger) 這兩個既有欄位（原本就存在於料號編輯畫面/清單頁/篩選條件，
# 只是這兩支批次匯入/範本/匯出是這次拆分時新寫的，漏掉了）。這兩欄在畫面上都是用
# 「顯示」/「隱藏」下拉選單呈現（不是「是」/「否」），所以用獨立的欄位集合、獨立的文字
# 轉換規則（見_get_sku_field()），不能直接併入_SKU_BOOL_FIELDS。
_SKU_VISIBILITY_FIELDS = {'is_active', 'hide_in_invoice_ledger'}


def _map_sku_header(h):
    h = h.strip()
    if '料號序號' in h:  # 2026-08-24新增：系統自動產生的唯一序號(sku_seq_no)，「批次修改匯入」
        return 'sku_seq_no'  # 專用來比對「這一列對應既有資料庫哪一筆」，「批次新增匯入」範本不含此欄
    if '料號順序' in h or '內部料號' in h:  # 料號順序為舊欄名（原本誤用來存內部料號），內部料號為現行欄名
        return 'internal_code'
    if 'EAN' in h.upper():
        return 'ean'
    if '出貨系統商品名稱' in h:  # 2026-08-25新增：供未來「出庫數比對未開發票數」功能比對用
        return 'shipping_system_product_name'
    if 'ERP' in h:  # 舊欄名「ERP後台代碼」、新欄名「ERP順序」皆比對到此
        return 'erp_code'
    if '財報簡稱' in h:
        return 'report_short_name'
    if '方案清單顯示' in h:  # 2026-08-24補修正：見上方_SKU_VISIBILITY_FIELDS說明，須排在「財報」判斷之前
        return 'is_active'
    if '發票總表顯示' in h:
        return 'hide_in_invoice_ledger'
    if '財報' in h:
        return 'report_product_name'
    if '未稅' in h or '進貨' in h:
        return 'purchase_price_notax'
    if '定價' in h:
        return 'list_price'
    if h == '料號' or '型號' in h:  # 型號代碼為舊欄名，料號為新欄名（料號順序/料號序號已於上方優先攔截）
        return 'model_code'
    if '配件' in h:
        return 'is_accessory'
    if '備註' in h:
        return 'note'
    return None


def _build_sku_col_map(raw_headers):
    col_map = {}
    for ci, h in enumerate(raw_headers):
        fk = _map_sku_header(h)
        if fk and fk not in col_map:
            col_map[fk] = ci
    return col_map


def _get_sku_field(row, col_map, fk):
    ci = col_map.get(fk)
    if ci is None or ci >= len(row):
        return None
    v = row[ci]
    if v is None:
        return None
    if fk in _SKU_NUMERIC_FIELDS:
        return _to_num_or_none(v)
    if fk in _SKU_BOOL_FIELDS:
        # 填「是」才算配件，其他非空白值（例如「否」）一律算False；留空維持None，
        # 批次修改模式下留空不覆蓋既有值（跟其他欄位一致）。
        s = str(v).strip()
        if not s:
            return None
        return s in ('是', 'Y', 'y', '1', 'TRUE', 'True', 'true')
    if fk in _SKU_VISIBILITY_FIELDS:
        # 「方案清單顯示」/「發票總表顯示」在畫面上是「顯示」/「隱藏」下拉選單（不是「是」/
        # 「否」），填「顯示」才算顯示，其他非空白值（例如「隱藏」）一律算隱藏；留空維持
        # None，批次修改模式下留空不覆蓋既有值。hide_in_invoice_ledger欄位語意跟畫面顯示
        # 相反（DB存的是「是否隱藏」），這裡要反轉一次，跟前端toggleSkuLedgerVisible()/
        # renderRuleForm()等既有反轉邏輯保持一致。
        s = str(v).strip()
        if not s:
            return None
        shown = s in ('顯示', '是', 'Y', 'y', '1', 'TRUE', 'True', 'true')
        return (not shown) if fk == 'hide_in_invoice_ledger' else shown
    v = str(v).strip()
    return v or None


# ============================================================
# ① 經銷商格式規則 CRUD
# ============================================================
@dealer_bp.route('/api/dealer/rules', methods=['GET'])
@dealer_view_required
def list_dealer_rules():
    res = sb.table('dealer_format_rules').select('*').order('dealer_name').execute()
    return jsonify(res.data)


@dealer_bp.route('/api/dealer/rules', methods=['POST'])
@dealer_edit_required
def create_dealer_rule():
    data = request.json or {}
    rec = {
        'dealer_name': _norm(data.get('dealer_name')),
        'detect_sheet_keywords': data.get('detect_sheet_keywords') or '',
        'sheets_config': data.get('sheets_config') or [],
        'is_active': data.get('is_active', True),
    }
    if not rec['dealer_name']:
        return jsonify({'error': '經銷商名稱必填'}), 400
    rec.update(_audit_new())
    res = sb.table('dealer_format_rules').insert(rec).execute()
    return jsonify(res.data[0] if res.data else {})


@dealer_bp.route('/api/dealer/rules/<int:rid>', methods=['PUT'])
@dealer_edit_required
def update_dealer_rule(rid):
    data = request.json or {}
    rec = {}
    for k in ('dealer_name', 'detect_sheet_keywords', 'sheets_config', 'is_active'):
        if k in data:
            rec[k] = data[k]
    rec.update(_audit_upd())
    sb.table('dealer_format_rules').update(rec).eq('id', rid).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/rules/<int:rid>', methods=['DELETE'])
@dealer_edit_required
def delete_dealer_rule(rid):
    sb.table('dealer_format_rules').delete().eq('id', rid).execute()
    return jsonify({'ok': True})


# ============================================================
# ② 內部料號清單 CRUD（含 sort_order 排序，比照代碼管理慣例）
# ============================================================
@dealer_bp.route('/api/dealer/sku-list', methods=['GET'])
@sku_view_required
def list_sku():
    kw = request.args.get('kw', '').strip()
    # 2026-08-13新增：內部料號清單畫面新增4個篩選條件（配件/方案清單顯示/產品定價/
    # 商品進貨未稅價），皆為選填、可與關鍵字搜尋(kw)同時併用（AND關係）。配件與方案清單
    # 顯示為單選(是/否或顯示/隱藏，留白代表不篩選)，定價/進貨未稅價為「最低～最高」區間。
    f_accessory = request.args.get('is_accessory', '').strip()      # '' / 'true' / 'false'
    f_active    = request.args.get('is_active', '').strip()          # '' / 'true' / 'false'
    # 2026-08-21新增：「發票總表顯示」篩選（hide_in_invoice_ledger），用途同上，供
    # 「料號與規則管理」畫面篩選用，與is_active（方案清單顯示）互相獨立。
    f_ledger_hide = request.args.get('hide_in_invoice_ledger', '').strip()  # '' / 'true' / 'false'
    lp_min = _to_num_or_none(request.args.get('list_price_min', ''))
    lp_max = _to_num_or_none(request.args.get('list_price_max', ''))
    pp_min = _to_num_or_none(request.args.get('purchase_price_min', ''))
    pp_max = _to_num_or_none(request.args.get('purchase_price_max', ''))

    def _build_q():
        q = sb.table('internal_sku_list').select('*')
        if kw:
            q = q.or_(f'internal_code.ilike.%{kw}%,erp_code.ilike.%{kw}%,ean.ilike.%{kw}%,report_product_name.ilike.%{kw}%,report_short_name.ilike.%{kw}%,shipping_system_product_name.ilike.%{kw}%')
        if f_accessory in ('true', 'false'):
            q = q.eq('is_accessory', f_accessory == 'true')
        if f_active in ('true', 'false'):
            q = q.eq('is_active', f_active == 'true')
        if f_ledger_hide in ('true', 'false'):
            q = q.eq('hide_in_invoice_ledger', f_ledger_hide == 'true')
        if lp_min is not None:
            q = q.gte('list_price', lp_min)
        if lp_max is not None:
            q = q.lte('list_price', lp_max)
        if pp_min is not None:
            q = q.gte('purchase_price_notax', pp_min)
        if pp_max is not None:
            q = q.lte('purchase_price_notax', pp_max)
        return q.order('sort_order')

    # ⚠️ 2026-08-08修正：改用 _fetch_all_rows() 分頁抓取——內部料號清單已超過1000筆，
    # 原本不分頁的寫法在無關鍵字搜尋（全部列出）時會被PostgREST靜默截斷。
    rows = _fetch_all_rows(_build_q)
    return jsonify(rows)


@dealer_bp.route('/api/dealer/sku-list', methods=['POST'])
@dealer_edit_required
def create_sku():
    data = request.json or {}
    code = _norm(data.get('internal_code'))
    if not code:
        return jsonify({'error': '內部料號必填'}), 400
    cnt_res = sb.table('internal_sku_list').select('id', count='exact').execute()
    rec = {
        'internal_code': code,
        # 2026-08-24新增：每筆料號皆自動配一組唯一識別號(sku_seq_no，格式C+4位數流水號，
        # 比照拆帳編號_next_recon_seq_no()的既有慣例)，做為「批次修改匯入」比對「這一列
        # Excel資料對應資料庫哪一筆既有紀錄」的新識別鍵，取代原本用internal_code文字比對、
        # 容易因空白/打字誤差而比對不到、進而被誤判成「新增」的舊做法。
        'sku_seq_no': _next_sku_seq_no(),
        'erp_code': data.get('erp_code'),
        'ean': data.get('ean'),
        # 2026-08-25新增：出貨系統匯出檔的「商品名稱」欄位文字，供未來「出庫數比對未開發票數」
        # 功能比對用（定位比照EAN，純輔助比對欄位，非業務主鍵）。
        'shipping_system_product_name': data.get('shipping_system_product_name'),
        'model_code': data.get('model_code'),
        'report_product_name': data.get('report_product_name'),
        'report_short_name': data.get('report_short_name'),
        'list_price': _to_num_or_none(data.get('list_price')),
        'purchase_price_notax': _to_num_or_none(data.get('purchase_price_notax')),
        'sort_order': data.get('sort_order', (cnt_res.count or 0)),
        'note': data.get('note'),
        'is_active': data.get('is_active', True),
        # 2026-08-12新增：「配件」勾選欄位，供「匯入書店訂單」折扣分攤規則判斷用（見
        # dealer_reconcile_schema.sql v1.25 / _compute_bookstore_order_matrix()）。
        'is_accessory': bool(data.get('is_accessory', False)),
        # 2026-08-21新增：「發票總表顯示」開關，供「比對已開/未開發票」寬表檢視功能使用，
        # 與is_accessory/is_active互相獨立（見dealer_reconcile_schema.sql v1.36）。
        'hide_in_invoice_ledger': bool(data.get('hide_in_invoice_ledger', False)),
    }
    rec.update(_audit_new())
    try:
        res = sb.table('internal_sku_list').insert(rec).execute()
    except Exception as e:
        return jsonify({'error': f'新增失敗（可能料號重複）：{e}'}), 400
    _touch_sku_sync()
    return jsonify(res.data[0] if res.data else {})


@dealer_bp.route('/api/dealer/sku-list/<int:sid>', methods=['PUT'])
@dealer_edit_required
def update_sku(sid):
    data = request.json or {}
    rec = {}
    for k in ('internal_code', 'erp_code', 'ean', 'shipping_system_product_name', 'model_code', 'report_product_name', 'report_short_name',
              'sort_order', 'note', 'is_active', 'is_accessory', 'hide_in_invoice_ledger'):
        if k in data:
            rec[k] = data[k]
    for k in ('list_price', 'purchase_price_notax'):
        if k in data:
            rec[k] = _to_num_or_none(data[k])
    rec.update(_audit_upd())
    sb.table('internal_sku_list').update(rec).eq('id', sid).execute()
    _touch_sku_sync()
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/sku-list/<int:sid>', methods=['DELETE'])
@dealer_edit_required
def delete_sku(sid):
    # 2026-08-21新增：刪除按鈕已從列表頁移入編輯畫面內（避免誤刪），這裡額外加上刪除前檢核——
    # 只要這個內部料號已經在「對帳彙總」(dealer_reconcile_summary) 或「比對已開/未開發票」
    # 的分帳明細(sku_invoice_ledger)裡出現過任何一筆月份對帳資料，就一律擋下不能刪除，
    # 避免刪掉已經對過帳的料號、造成歷史資料料號對應斷鏈。比照全系統既有「刪除前先檢查是否
    # 仍被使用中」慣例（如代碼管理FIELD_USAGE_MAP、bookstore_order_import_log等）。
    row_res = sb.table('internal_sku_list').select('internal_code').eq('id', sid).execute()
    rows = row_res.data or []
    if not rows:
        return jsonify({'error': '找不到該筆料號，可能已被刪除'}), 404
    code = rows[0].get('internal_code')
    if code:
        summary_hit = sb.table('dealer_reconcile_summary').select('id').eq('internal_code', code).limit(1).execute()
        if summary_hit.data:
            return jsonify({'error': f'此料號「{code}」在「對帳彙總」已有月份對帳資料，無法刪除，請改用「方案清單顯示」設為隱藏即可'}), 409
        ledger_hit = sb.table('sku_invoice_ledger').select('id').eq('internal_code', code).limit(1).execute()
        if ledger_hit.data:
            return jsonify({'error': f'此料號「{code}」在「比對已開/未開發票」分帳明細已有月份對帳資料，無法刪除，請改用「方案清單顯示」設為隱藏即可'}), 409
    sb.table('internal_sku_list').delete().eq('id', sid).execute()
    _touch_sku_sync()
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/sku-list/reorder', methods=['POST'])
@dealer_edit_required
def reorder_sku():
    """比照代碼管理 moveCode()：本地位置交換後，整批正規化 sort_order。
    支援兩種操作方式：
      - direction='up'/'down'：單筆上下移動一格（原有功能）
      - position=N（1-based）：直接搬移到清單中的第N個位置，避免要移動很多格時得逐次點擊上下鍵
    position 一律以「全部料號依 sort_order 排序」的清單為準，若畫面上有輸入關鍵字篩選，
    輸入的位置對應的是全部清單裡的第N個，不是篩選結果裡的第N個，操作前建議先清除關鍵字篩選。
    """
    data = request.json or {}
    sid = data.get('id')
    direction = data.get('direction')  # 'up' / 'down'
    position = data.get('position')
    # ⚠️ 2026-08-08修正：改用 _fetch_all_rows() 分頁抓取——內部料號清單已超過1000筆，
    # 不分頁會被PostgREST靜默截斷，導致排序清單缺筆、後面的料號完全排不到。
    items = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('id,sort_order').order('sort_order'))
    idx = next((i for i, it in enumerate(items) if it['id'] == sid), None)
    if idx is None:
        return jsonify({'error': '找不到該筆料號'}), 404

    if position is not None:
        try:
            target = int(position) - 1  # 轉為 0-based
        except (TypeError, ValueError):
            return jsonify({'error': '排序位置需為數字'}), 400
        requested = target  # 使用者實際輸入的位置(0-based)，用於跟clamp後的結果比較，才能告知使用者「有沒有被限制」
        target = max(0, min(target, len(items) - 1))
        if target == idx:
            # 不論是「輸入的位置跟目前位置相同」還是「超出範圍被限制在頭/尾、剛好等於目前位置」，
            # 都回傳total/final_position讓前端可以明確告知使用者「總共只有幾筆、目前已經在第幾筆」，
            # 不要讓畫面上看起來像是按了Enter完全沒反應
            return jsonify({'ok': True, 'moved': False, 'total': len(items),
                             'final_position': target + 1, 'requested_position': requested + 1})
        item = items.pop(idx)
        items.insert(target, item)
    else:
        target = idx - 1 if direction == 'up' else idx + 1
        if target < 0 or target >= len(items):
            return jsonify({'ok': True, 'moved': False, 'total': len(items), 'final_position': idx + 1})  # 已在頂端/底端，不動作
        items[idx], items[target] = items[target], items[idx]

    for i, it in enumerate(items):
        if it.get('sort_order') != i:
            sb.table('internal_sku_list').update({'sort_order': i}).eq('id', it['id']).execute()
    _touch_sku_sync()
    return jsonify({'ok': True, 'moved': True, 'total': len(items), 'final_position': target + 1})


# 範本第2列（欄位說明列）的「料號序號」欄文字，供批次修改匯入時識別「這一列是說明列，非真實資料」——
# 2026-08-12修正：改用內容比對（而非列號），避免使用者拿「匯出檔」（真實資料就從第2列開始，沒有說明列）
# 重新批次匯入時，第2列的真實資料被誤判成說明列而永遠被跳過、無法更新。
# 2026-08-24修正：拆分「批次修改匯入」/「批次新增匯入」後，比對鍵改成「料號序號」
# (sku_seq_no)，此說明文字同步改為描述新的比對規則（序號相符才更新、序號空白或找不到
# 一律略過並回報，不再自動新增）。
_SKU_NOTE_ROW_MARKER = '必填，此欄位為系統自動配發的唯一序號，請勿手動輸入或修改；序號與既有資料相符時會更新該筆資料，序號空白或找不到對應資料的列會被略過並列在結果中，不會新增為新資料'


# ── 內部料號 —「批次修改匯入」範本下載（比照兌換碼回填模組慣例：必填欄位黃底、說明列、防呆範例列） ──
# 2026-08-24拆分：原本「批次匯入」（找不到料號就自動新增、找得到就更新）已拆成「批次新增
# 匯入」（見add_template_sku()/add_batch_import_sku()，純新增）跟這支「批次修改匯入」
# （純更新既有資料，找不到對應資料一律略過、不再自動新增）。範本第一欄改成「料號序號」
# （sku_seq_no，系統自動配的C+4位數唯一識別號），做為比對「這一列對應資料庫哪一筆既有
# 紀錄」的新識別鍵，取代舊版用「內部料號」文字比對、容易因空白/打字誤差而比對不到、
# 進而被誤判成「新增」的做法。使用者應直接拿「匯出」下載的檔案修改後送回，不建議手動
# 填寫此欄位。
@dealer_bp.route('/api/dealer/sku-list/template', methods=['GET'])
@dealer_view_required
def sku_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '內部料號批次修改匯入範本'

    # 2026-08-12新增：「配件」欄位（is_accessory），供「匯入書店訂單」折扣分攤規則判斷用，
    # 放在「商品進貨未稅價」跟「備註」中間。
    # 2026-08-24補修正：拆分「批次修改匯入」時漏帶了「方案清單顯示」/「發票總表顯示」這兩個
    # 既有欄位（is_active/hide_in_invoice_ledger），這裡補回來，順序比照匯出檔／畫面清單頁
    # 的欄位順序（見export_sku_list()），方便使用者直接拿匯出檔修改後重新匯入。
    headers = ['料號序號', '內部料號', 'EAN', '出貨系統商品名稱', 'ERP順序', '料號', '財報料號品名', '財報簡稱', '產品定價', '商品進貨未稅價', '配件', '備註', '方案清單顯示', '發票總表顯示']
    required = {'料號序號'}
    notes_row = [
        _SKU_NOTE_ROW_MARKER,
        '僅供對照識別，此欄位異動不會被套用（如需修改內部料號請至畫面上編輯該筆料號）',
        '可留空，廠商條碼（留空不會覆蓋既有值）；上傳對帳檔比對時會用來輔助自動比對內部料號',
        '可留空，出貨系統匯出檔的「商品名稱」欄位文字（留空不會覆蓋既有值）；供未來「出庫數比對未開發票數」功能比對用',
        '可留空（留空不會覆蓋既有值）',
        '可留空（留空不會覆蓋既有值）',
        '可留空，多筆內部料號可共用同一值供財報彙總（留空不會覆蓋既有值）',
        '可留空（留空不會覆蓋既有值）',
        '可留空，請填數字（留空不會覆蓋既有值）',
        '可留空，請填數字（留空不會覆蓋既有值）',
        '可留空，填「是」代表配件、留空或填「否」代表非配件（留空不會覆蓋既有值）',
        '可留空（留空不會覆蓋既有值）',
        '可留空，填「顯示」或「隱藏」（留空不會覆蓋既有值）',
        '可留空，填「顯示」或「隱藏」（留空不會覆蓋既有值）',
    ]

    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    rfill = openpyxl.styles.PatternFill('solid', fgColor='FFF2CC')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    rfont = openpyxl.styles.Font(bold=True, color='7B3F00')
    nfill = openpyxl.styles.PatternFill('solid', fgColor='F2F2F2')
    nfont = openpyxl.styles.Font(italic=True, color='888888')
    widths = [14, 22, 16, 22, 12, 14, 26, 18, 14, 18, 10, 26, 14, 14]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        if h in required:
            c.fill = rfill
            c.font = rfont
        else:
            c.fill = hfill
            c.font = hfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci - 1]

    for ci, n in enumerate(notes_row, 1):
        c = ws.cell(2, ci, n)
        c.fill = nfill
        c.font = nfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.append(['C0001', 'SAMPLE-DEMO-0001（此列為範例，請刪除後填入實際資料）', '4710000000000', '範例出貨系統商品名稱', 'ERP-001', 'MODEL-X', '財報範例品名', '財報簡稱範例', 1990, 1200, '是', '', '顯示', '顯示'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='內部料號批次修改匯入範本.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 內部料號 —「批次新增匯入」範本下載（2026-08-24新增，獨立於「批次修改匯入」之外，
# 純粹新增全新料號，不含料號序號欄位——序號由系統自動配發） ──
@dealer_bp.route('/api/dealer/sku-list/add-template', methods=['GET'])
@dealer_view_required
def add_template_sku():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '內部料號批次新增匯入範本'

    # 2026-08-24補修正：跟批次修改匯入範本一樣漏帶了「方案清單顯示」/「發票總表顯示」，
    # 這裡補回來；留空時新增時分別預設為「顯示」/「顯示」（見add_batch_import_sku()）。
    headers = ['內部料號', 'EAN', '出貨系統商品名稱', 'ERP順序', '料號', '財報料號品名', '財報簡稱', '產品定價', '商品進貨未稅價', '配件', '備註', '方案清單顯示', '發票總表顯示']
    required = {'內部料號'}
    notes_row = [
        '必填，不可與清單中既有內部料號重複；若重複，此列會被略過並列在匯入結果中，不會覆蓋既有資料',
        '可留空，廠商條碼；上傳對帳檔比對時會用來輔助自動比對內部料號',
        '可留空，出貨系統匯出檔的「商品名稱」欄位文字；供未來「出庫數比對未開發票數」功能比對用',
        '可留空', '可留空',
        '可留空，多筆內部料號可共用同一值供財報彙總',
        '可留空',
        '可留空，請填數字', '可留空，請填數字',
        '可留空，填「是」代表配件、留空或填「否」代表非配件',
        '可留空',
        '可留空，填「顯示」或「隱藏」，留空預設為「顯示」',
        '可留空，填「顯示」或「隱藏」，留空預設為「顯示」',
    ]

    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    rfill = openpyxl.styles.PatternFill('solid', fgColor='FFF2CC')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    rfont = openpyxl.styles.Font(bold=True, color='7B3F00')
    nfill = openpyxl.styles.PatternFill('solid', fgColor='F2F2F2')
    nfont = openpyxl.styles.Font(italic=True, color='888888')
    widths = [22, 16, 22, 12, 14, 26, 18, 14, 18, 10, 26, 14, 14]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        if h in required:
            c.fill = rfill
            c.font = rfont
        else:
            c.fill = hfill
            c.font = hfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci - 1]

    for ci, n in enumerate(notes_row, 1):
        c = ws.cell(2, ci, n)
        c.fill = nfill
        c.font = nfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.append(['SAMPLE-DEMO-0001（此列為範例，請刪除後填入實際資料）', '4710000000000', '範例出貨系統商品名稱', 'ERP-001', 'MODEL-X', '財報範例品名', '財報簡稱範例', 1990, 1200, '是', '', '顯示', '顯示'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='內部料號批次新增匯入範本.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 內部料號 — 匯出（欄位順序比照「批次修改匯入」範本，方便直接拿匯出檔修改後重新批次
# 修改匯入；2026-08-24新增「料號序號」欄，這是重新批次修改匯入時的比對依據，請勿手動
# 修改此欄內容） ──
@dealer_bp.route('/api/dealer/sku-list/export', methods=['GET'])
@dealer_view_required
def export_sku_list():
    all_rows = []
    _off = 0
    while True:
        res = sb.table('internal_sku_list').select('*').order('sort_order').range(_off, _off + 999).execute()
        batch_rows = res.data or []
        all_rows.extend(batch_rows)
        if len(batch_rows) < 1000:
            break
        _off += 1000

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '內部料號清單'
    # 2026-08-24補修正：漏帶了「方案清單顯示」/「發票總表顯示」這兩個既有欄位（is_active/
    # hide_in_invoice_ledger），畫面清單頁一直都有顯示這兩欄（見dealer_reconcile.html第594行
    # 表頭），匯出/範本理應一併帶出，否則使用者無法用「匯出→修改→批次修改匯入」的方式
    # 維護這兩個欄位。呈現文字比照畫面下拉選單，統一用「顯示」/「隱藏」（不用「是」/「否」），
    # hide_in_invoice_ledger欄位語意跟顯示文字相反，這裡要反轉一次。
    headers = ['排序', '料號序號', '內部料號', 'EAN', '出貨系統商品名稱', 'ERP順序', '料號', '財報料號品名', '財報簡稱', '產品定價', '商品進貨未稅價', '配件', '備註', '方案清單顯示', '發票總表顯示']
    ws.append(headers)
    for i, s in enumerate(all_rows, 1):
        ws.append([
            i, s.get('sku_seq_no'), s.get('internal_code'), s.get('ean'), s.get('shipping_system_product_name'), s.get('erp_code'), s.get('model_code'),
            s.get('report_product_name'), s.get('report_short_name'),
            s.get('list_price'), s.get('purchase_price_notax'), ('是' if s.get('is_accessory') else '否'), s.get('note'),
            ('隱藏' if s.get('is_active') is False else '顯示'),
            ('隱藏' if s.get('hide_in_invoice_ledger') else '顯示'),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='內部料號清單.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 內部料號 —「批次修改匯入」（2026-08-24拆分自原「批次匯入」，改為純更新既有資料，
# 找不到對應資料一律略過並回報、不再自動新增——避免「比對不到就自行新增」的舊bug類型
# 再度發生，見「批次新增匯入」add_batch_import_sku()處理純新增） ──
# 比對鍵改用「料號序號」(sku_seq_no，系統自動配發的C+4位數唯一識別號)，取代舊版用
# internal_code文字比對、容易因空白/打字誤差而比對不到的做法。
@dealer_bp.route('/api/dealer/sku-list/batch-import', methods=['POST'])
@dealer_edit_required
def batch_modify_import_sku():
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{e}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'error': '檔案無資料列'}), 400

    raw_headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    col_map = _build_sku_col_map(raw_headers)

    def _get(row, fk):
        return _get_sku_field(row, col_map, fk)

    # existing_map：料號序號(sku_seq_no) → id（批次修改匯入的比對鍵，取代舊版的internal_code）
    existing_map = {}
    for r in _fetch_all_rows(lambda: sb.table('internal_sku_list').select('id,sku_seq_no')):
        if r.get('sku_seq_no'):
            existing_map[r['sku_seq_no']] = r['id']

    # 2026-08-24補修正：加入is_active/hide_in_invoice_ledger（方案清單顯示/發票總表顯示），
    # 拆分批次匯入時漏帶了這兩個既有欄位，導致範本/匯出/批次修改匯入三處都無法維護它們。
    _UPDATE_FIELDS = ['erp_code', 'ean', 'shipping_system_product_name', 'model_code', 'report_product_name', 'report_short_name',
                       'list_price', 'purchase_price_notax', 'is_accessory', 'note',
                       'is_active', 'hide_in_invoice_ledger']

    upd_cnt = skip_cnt = err_cnt = unmatched_cnt = 0
    unmatched_detail = []

    for ri, row in enumerate(rows[1:], 2):
        seq_val = _get(row, 'sku_seq_no')
        if not seq_val:
            skip_cnt += 1
            continue
        # 說明列判斷改用內容比對（而非「第2列一律跳過」），因為使用者常拿「匯出檔」重新匯入，
        # 匯出檔第2列就是第一筆真實資料，不能被誤判成說明列而跳過。
        if seq_val in ('料號序號', '（說明）', _SKU_NOTE_ROW_MARKER) or seq_val.upper().startswith('SAMPLE-DEMO'):
            skip_cnt += 1
            continue

        # ── 料號序號比對不到既有資料：略過並回報，不自動新增 ──
        # 這正是使用者要求排除的舊bug行為（「找不到就自行新增」），拆分後這裡改成純更新，
        # 找不到的列一律留給使用者自行確認（序號打錯？還是這本來就是一筆新資料，應改用
        # 「批次新增匯入」處理）。
        if seq_val not in existing_map:
            unmatched_cnt += 1
            unmatched_detail.append({'row': ri, 'sku_seq_no': seq_val,
                                      'internal_code': _get(row, 'internal_code')})
            continue

        # 只更新 Excel 中「有填值」的欄位，空白欄位不覆蓋既有資料
        upd = {}
        for fk in _UPDATE_FIELDS:
            v = _get(row, fk)
            if v is not None and v != '':
                upd[fk] = v
        if not upd:
            skip_cnt += 1
            continue
        upd.update(_audit_upd())
        try:
            sb.table('internal_sku_list').update(upd).eq('id', existing_map[seq_val]).execute()
            upd_cnt += 1
        except Exception:
            err_cnt += 1

    if upd_cnt:
        _touch_sku_sync()
    return jsonify({'ok': 0, 'updated': upd_cnt, 'skip': skip_cnt, 'error': err_cnt,
                    'unmatched': unmatched_cnt, 'unmatched_detail': unmatched_detail})


# ── 內部料號 —「批次新增匯入」（2026-08-24新增，獨立於「批次修改匯入」之外，純粹新增
# 全新料號，不會更動任何既有資料；比照書店經銷商方案清單「簡易批次新增」
# simple_batch_import_bookstore_plans()的既有慣例：內部料號跟系統既有資料/同批次檔案
# 內重複時，該列跳過並回報，其他列照常新增） ──
@dealer_bp.route('/api/dealer/sku-list/add-batch-import', methods=['POST'])
@dealer_edit_required
def add_batch_import_sku():
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{e}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'error': '檔案無資料列'}), 400

    raw_headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    col_map = _build_sku_col_map(raw_headers)

    def _get(row, fk):
        return _get_sku_field(row, col_map, fk)

    # existing_codes：目前資料庫已存在的內部料號（批次新增匯入的重複判斷依據，跟「批次
    # 修改匯入」不同——這裡完全不看sku_seq_no，因為新增的料號本來就還沒有序號）
    existing_codes = {r['internal_code'] for r in _fetch_all_rows(
        lambda: sb.table('internal_sku_list').select('internal_code')) if r.get('internal_code')}

    cnt_res = sb.table('internal_sku_list').select('id', count='exact').execute()
    next_sort = cnt_res.count or 0
    # 批次一次新增多筆時，序號在本地遞增（而非每列都重新查一次資料庫目前最大值），避免
    # 同一批次內後面的列拿到跟前面重複的序號——原因同_next_sku_seq_no()函式說明。
    next_seq_n = _sku_seq_max_n()

    seen_in_file = set()
    ok_cnt = skip_cnt = err_cnt = dup_cnt = 0
    dup_detail = []
    skipped_detail = []
    batch = []

    for ri, row in enumerate(rows[1:], 2):
        code_val = _get(row, 'internal_code')
        if not code_val:
            skip_cnt += 1
            continue
        if code_val in ('料號順序', '內部料號', '（說明）') or code_val.upper().startswith('SAMPLE-DEMO'):
            skip_cnt += 1
            continue
        if code_val in existing_codes or code_val in seen_in_file:
            dup_cnt += 1
            dup_detail.append({'row': ri, 'internal_code': code_val})
            continue

        seen_in_file.add(code_val)
        next_seq_n += 1
        # 2026-08-24補修正：加入is_active/hide_in_invoice_ledger（方案清單顯示/發票總表顯示）—
        # 拆分批次匯入時漏帶了這兩欄，範本上補回來之後這裡也要跟著讀取；Excel留空時分別
        # 預設「顯示」（is_active=True／hide_in_invoice_ledger=False），跟單筆新增create_sku()
        # 的預設值一致。
        is_active_val = _get(row, 'is_active')
        hide_ledger_val = _get(row, 'hide_in_invoice_ledger')
        rec = {
            'internal_code': code_val,
            'sku_seq_no': f'C{next_seq_n:04d}',
            'erp_code': _get(row, 'erp_code'),
            'ean': _get(row, 'ean'),
            'shipping_system_product_name': _get(row, 'shipping_system_product_name'),
            'model_code': _get(row, 'model_code'),
            'report_product_name': _get(row, 'report_product_name'),
            'report_short_name': _get(row, 'report_short_name'),
            'list_price': _get(row, 'list_price'),
            'purchase_price_notax': _get(row, 'purchase_price_notax'),
            'is_accessory': bool(_get(row, 'is_accessory')),
            'note': _get(row, 'note'),
            'sort_order': next_sort,
            'is_active': is_active_val if is_active_val is not None else True,
            'hide_in_invoice_ledger': hide_ledger_val if hide_ledger_val is not None else False,
        }
        rec.update(_audit_new())
        batch.append(rec)
        next_sort += 1

        if len(batch) >= 80:
            try:
                sb.table('internal_sku_list').insert(batch).execute()
                ok_cnt += len(batch)
            except Exception as e:
                err_cnt += len(batch)
                for rec in batch:
                    skipped_detail.append({'row': None, 'reason': f'內部料號「{rec["internal_code"]}」新增失敗：{e}'})
            batch = []

    if batch:
        try:
            sb.table('internal_sku_list').insert(batch).execute()
            ok_cnt += len(batch)
        except Exception as e:
            err_cnt += len(batch)
            for rec in batch:
                skipped_detail.append({'row': None, 'reason': f'內部料號「{rec["internal_code"]}」新增失敗：{e}'})

    if ok_cnt:
        _touch_sku_sync()
    return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt, 'dup': dup_cnt,
                    'dup_detail': dup_detail, 'skipped_detail': skipped_detail})


# ============================================================
# ③ 經銷商品項對照表 CRUD（含歧義代碼自動偵測）
# ============================================================
def _recompute_ambiguous(dealer_name, dealer_code):
    if not dealer_code:
        return
    res = sb.table('dealer_sku_mapping').select('*') \
        .eq('dealer_name', dealer_name).eq('dealer_code', dealer_code).execute()
    rows = res.data or []
    distinct_sku = {r['main_sku'] for r in rows}
    is_ambig = len(distinct_sku) > 1
    for r in rows:
        if r.get('is_ambiguous_code') != is_ambig:
            sb.table('dealer_sku_mapping').update({'is_ambiguous_code': is_ambig}).eq('id', r['id']).execute()


@dealer_bp.route('/api/dealer/mapping', methods=['GET'])
@dealer_view_required
def list_mapping():
    dealer = request.args.get('dealer', '').strip()
    kw = request.args.get('kw', '').strip()
    q = sb.table('dealer_sku_mapping').select('*')
    if dealer:
        q = q.eq('dealer_name', dealer)
    if kw:
        q = q.or_(f'dealer_code.ilike.%{kw}%,ean.ilike.%{kw}%,product_name.ilike.%{kw}%')
    res = q.order('id', desc=True).execute()
    return jsonify(res.data)


def _touch_sku_sync():
    """「內部料號清單」新增/修改/刪除/排序調整/批次匯入後呼叫，更新 bookstore_sku_sync_status
    的時間戳，供「書店經銷商→方案清單」頁面顯示「最後更新」提示（見該表建立時的說明註解）。
    這裡刻意不讓寫入失敗影響原本的CRUD結果（例如表還沒建立時），所以包一層 try/except。"""
    try:
        sb.table('bookstore_sku_sync_status').update({'last_changed_at': now_str()}).eq('id', 1).execute()
    except Exception:
        pass


def _get_sku_last_synced():
    try:
        res = sb.table('bookstore_sku_sync_status').select('last_changed_at').eq('id', 1).execute()
        if res.data:
            return res.data[0].get('last_changed_at')
    except Exception:
        pass
    return None


def _valid_sku_codes():
    """回傳目前「內部料號清單」所有有效的內部料號集合，供經銷商品項對照表存檔前驗證用。
    背景：曾發生「經銷商品項對照表」批次匯入時，Excel填了內部料號清單裡根本不存在的料號
    （無驗證機制），比對命中入庫後，對帳彙總資料出現查無主檔的「孤兒料號」，導致
    「對帳查詢→依財報料號品名彙總」顯示原始料號字串（查無財報品名時的備援顯示），
    看起來跟其他筆的命名風格不一致。"""
    # ⚠️ 2026-08-08修正：改用 _fetch_all_rows() 分頁抓取，避免內部料號清單超過1000筆時
    # 被PostgREST靜默截斷，導致較晚新增的內部料號被誤判成「不存在」。
    rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('internal_code'))
    return {r['internal_code'] for r in rows if r.get('internal_code')}


def _check_sku_refs(main_sku, gift_skus, valid_codes):
    """檢查主料號/贈品料號是否都存在於內部料號清單，不存在則回傳錯誤訊息（None代表通過）"""
    if main_sku and main_sku not in valid_codes:
        return f'主料號「{main_sku}」不存在於「內部料號清單」，請先在該清單新增此料號，或修正為既有料號'
    for gv in gift_skus:
        if gv and gv not in valid_codes:
            return f'贈品料號「{gv}」不存在於「內部料號清單」，請先在該清單新增此料號，或修正為既有料號'
    return None


@dealer_bp.route('/api/dealer/mapping', methods=['POST'])
@dealer_edit_required
def create_mapping():
    data = request.json or {}
    rec = {
        'dealer_name': _norm(data.get('dealer_name')),
        'dealer_code': data.get('dealer_code') or None,
        'ean': data.get('ean') or None,
        'product_name': data.get('product_name') or None,
        'main_sku': _norm(data.get('main_sku')),
        'gift_sku1': data.get('gift_sku1') or None,
        'gift_sku2': data.get('gift_sku2') or None,
        'gift_sku3': data.get('gift_sku3') or None,
        'note': data.get('note'),
    }
    if not rec['dealer_name'] or not rec['main_sku']:
        return jsonify({'error': '來源經銷商與主料號必填'}), 400
    err = _check_sku_refs(rec['main_sku'],
                          [rec['gift_sku1'], rec['gift_sku2'], rec['gift_sku3']],
                          _valid_sku_codes())
    if err:
        return jsonify({'error': err}), 400
    rec.update(_audit_new())
    res = sb.table('dealer_sku_mapping').insert(rec).execute()
    _recompute_ambiguous(rec['dealer_name'], rec.get('dealer_code'))
    return jsonify(res.data[0] if res.data else {})


@dealer_bp.route('/api/dealer/mapping/<int:mid>', methods=['PUT'])
@dealer_edit_required
def update_mapping(mid):
    data = request.json or {}
    # 修改前先取出舊值，若使用者把「經銷商代碼」本身改掉，舊代碼那一群的歧義旗標也要重算，
    # 否則舊代碼群組可能因為少了一筆而不再歧義，卻仍卡在歧義=是
    old_res = sb.table('dealer_sku_mapping').select('dealer_name,dealer_code').eq('id', mid).execute()
    old_row = old_res.data[0] if old_res.data else None

    rec = {}
    for k in ('dealer_name', 'dealer_code', 'ean', 'product_name', 'main_sku',
              'gift_sku1', 'gift_sku2', 'gift_sku3', 'note'):
        if k in data:
            rec[k] = data[k]
    # 只在使用者這次有異動主料號/贈品料號時才檢查（避免只改備註等其他欄位時被無關的舊資料擋下）
    if any(k in rec for k in ('main_sku', 'gift_sku1', 'gift_sku2', 'gift_sku3')):
        cur_res = sb.table('dealer_sku_mapping').select('*').eq('id', mid).execute()
        cur = cur_res.data[0] if cur_res.data else {}
        merged_main = rec.get('main_sku', cur.get('main_sku'))
        merged_gifts = [rec.get(k, cur.get(k)) for k in ('gift_sku1', 'gift_sku2', 'gift_sku3')]
        err = _check_sku_refs(_norm(merged_main), merged_gifts, _valid_sku_codes())
        if err:
            return jsonify({'error': err}), 400
    rec.update(_audit_upd())
    sb.table('dealer_sku_mapping').update(rec).eq('id', mid).execute()

    new_dealer = rec.get('dealer_name') or (old_row or {}).get('dealer_name')
    new_code = rec.get('dealer_code') if 'dealer_code' in rec else (old_row or {}).get('dealer_code')
    if new_dealer:
        _recompute_ambiguous(new_dealer, new_code)
    if old_row and (old_row.get('dealer_name') != new_dealer or old_row.get('dealer_code') != new_code):
        _recompute_ambiguous(old_row.get('dealer_name'), old_row.get('dealer_code'))
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/mapping/<int:mid>', methods=['DELETE'])
@dealer_edit_required
def delete_mapping(mid):
    # 刪除前先取出該筆的經銷商/代碼，刪除後重新計算歧義旗標。
    # 若不重算：刪除掉造成歧義的其中一筆重複/錯誤資料後，剩下那筆的「歧義」標記會永遠卡住不會消失，
    # 使用者會以為刪除沒有生效
    row_res = sb.table('dealer_sku_mapping').select('dealer_name,dealer_code').eq('id', mid).execute()
    row = row_res.data[0] if row_res.data else None
    sb.table('dealer_sku_mapping').delete().eq('id', mid).execute()
    if row:
        _recompute_ambiguous(row.get('dealer_name'), row.get('dealer_code'))
    return jsonify({'ok': True})


# 2026-08-17新增：跟_SKU_NOTE_ROW_MARKER同一目的——批次匯入判斷「這一列是說明列」要用
# 內容比對，不能用列號（見下方batch_import_dealer_mapping()說明）。這裡對應下方
# notes_row第5欄（「主料號」欄的說明文字），跟_get(row, 'main_sku')比對的欄位一致。
_MAPPING_NOTE_ROW_MARKER = '必填，需與「內部料號清單」的內部料號完全一致'


# ── 經銷商品項對照表 — 批次匯入範本下載（比照內部料號清單慣例：必填欄位黃底、說明列、防呆範例列） ──
@dealer_bp.route('/api/dealer/mapping/template', methods=['GET'])
@dealer_view_required
def mapping_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '經銷商品項對照表匯入範本'

    headers = ['來源經銷商', '經銷商代碼', 'EAN', '商品名稱', '主料號',
               '贈品料號1', '贈品料號2', '贈品料號3', '備註']
    required = {'來源經銷商', '主料號'}
    notes_row = [
        '必填，需與「料號與規則管理」的經銷商規則名稱一致（如：三創/富邦momo/墊腳石/親子天下）',
        '可留空（若該經銷商沒有獨立代碼欄，如墊腳石，此欄留空即可，比對會改用EAN）',
        '可留空（建議填寫，比對優先序最高）',
        '可留空，僅供辨識用，不影響比對',
        _MAPPING_NOTE_ROW_MARKER,
        '可留空，組合商品(主機+贈品)的贈品料號1',
        '可留空，組合商品的贈品料號2',
        '可留空，組合商品的贈品料號3（如親子天下等多件式套組需要更多贈品欄位時使用）',
        '可留空',
    ]

    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    rfill = openpyxl.styles.PatternFill('solid', fgColor='FFF2CC')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    rfont = openpyxl.styles.Font(bold=True, color='7B3F00')
    nfill = openpyxl.styles.PatternFill('solid', fgColor='F2F2F2')
    nfont = openpyxl.styles.Font(italic=True, color='888888')
    widths = [14, 16, 16, 30, 18, 16, 16, 16, 20]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        if h in required:
            c.fill = rfill
            c.font = rfont
        else:
            c.fill = hfill
            c.font = hfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci - 1]

    for ci, n in enumerate(notes_row, 1):
        c = ws.cell(2, ci, n)
        c.fill = nfill
        c.font = nfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.append(['墊腳石', '', '4710562710000', '範例商品名稱',
                'SAMPLE-DEMO-0001（此列為範例，請刪除後填入實際資料）', '', '', '', ''])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='經銷商品項對照表匯入範本.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 經銷商品項對照表 — 批次匯入（每列皆為新增一筆對應；與現有資料完全重複者自動跳過） ──
@dealer_bp.route('/api/dealer/mapping/batch-import', methods=['POST'])
@dealer_edit_required
def batch_import_mapping():
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    default_dealer = _norm(request.form.get('dealer_name', ''))
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{e}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'error': '檔案無資料列'}), 400

    raw_headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    def _map_h(h):
        h = h.strip()
        if '來源經銷商' in h or '經銷商名稱' in h:
            return 'dealer_name'
        if '經銷商代碼' in h:
            return 'dealer_code'
        if 'EAN' in h.upper():
            return 'ean'
        if '商品名稱' in h:
            return 'product_name'
        if '贈品料號1' in h or '贈品料號一' in h:
            return 'gift_sku1'
        if '贈品料號2' in h or '贈品料號二' in h:
            return 'gift_sku2'
        if '贈品料號3' in h or '贈品料號三' in h:
            return 'gift_sku3'
        if '主料號' in h:
            return 'main_sku'
        if '備註' in h:
            return 'note'
        return None

    col_map = {}
    for ci, h in enumerate(raw_headers):
        fk = _map_h(h)
        if fk and fk not in col_map:
            col_map[fk] = ci

    def _get(row, fk):
        ci = col_map.get(fk)
        if ci is None or ci >= len(row):
            return None
        v = row[ci]
        if v is None:
            return None
        v = str(v).strip()
        return v or None

    # 既有資料完全重複比對用（避免重複匯入同一份檔案時瘋狂灌入一樣的對應）
    existing_res = sb.table('dealer_sku_mapping').select(
        'dealer_name,dealer_code,ean,main_sku,gift_sku1,gift_sku2,gift_sku3').execute()
    existing_keys = set()
    for r in (existing_res.data or []):
        existing_keys.add((
            _norm(r.get('dealer_name')), _norm(r.get('dealer_code')), _norm(r.get('ean')),
            _norm(r.get('main_sku')), _norm(r.get('gift_sku1')), _norm(r.get('gift_sku2')),
            _norm(r.get('gift_sku3')),
        ))

    # 有效內部料號集合（一次查詢，供逐列驗證用，避免匯入不存在的主料號/贈品料號造成
    # 對帳彙總出現查無主檔的「孤兒料號」——2026-07-23 實際發生過此問題）
    valid_sku_codes = _valid_sku_codes()

    ok_cnt = skip_cnt = err_cnt = invalid_cnt = 0
    batch = []
    touched_pairs = set()  # (dealer_name, dealer_code) 需要重新計算歧義

    for ri, row in enumerate(rows[1:], 2):
        main_sku = _get(row, 'main_sku')
        # 2026-08-17修正：說明列判斷改用內容比對（而非「第2列一律跳過」），比照「內部料號
        # 清單」批次匯入2026-08-12的既有修法（見_SKU_NOTE_ROW_MARKER）——使用者常會把
        # 範本的說明列(第2列)連同範例列一起刪除，讓真實資料往上移到第2列；若還用列號判斷，
        # 這筆真實資料會被誤判成說明列而被跳過，造成「怎麼匯入都是0筆成功」的困惑。
        if main_sku == _MAPPING_NOTE_ROW_MARKER:
            skip_cnt += 1
            continue

        if not main_sku or main_sku.upper().startswith('SAMPLE-DEMO'):
            skip_cnt += 1
            continue

        dealer_name = _get(row, 'dealer_name') or default_dealer
        if not dealer_name:
            skip_cnt += 1
            continue

        dealer_code = _get(row, 'dealer_code')
        ean = _get(row, 'ean')
        product_name = _get(row, 'product_name')
        gift_sku1 = _get(row, 'gift_sku1')
        gift_sku2 = _get(row, 'gift_sku2')
        gift_sku3 = _get(row, 'gift_sku3')
        note = _get(row, 'note')

        # 主料號/贈品料號必須是「內部料號清單」裡實際存在的料號，否則跳過此列
        # （不存在的料號若被存入，日後對帳比對命中入庫時，會在對帳彙總產生查無主檔的孤兒資料）
        if _check_sku_refs(main_sku, [gift_sku1, gift_sku2, gift_sku3], valid_sku_codes):
            invalid_cnt += 1
            continue

        key = (_norm(dealer_name), _norm(dealer_code), _norm(ean),
               _norm(main_sku), _norm(gift_sku1), _norm(gift_sku2), _norm(gift_sku3))
        if key in existing_keys:
            skip_cnt += 1
            continue
        existing_keys.add(key)  # 避免同批次檔案內部重複

        rec = {
            'dealer_name': dealer_name, 'dealer_code': dealer_code, 'ean': ean,
            'product_name': product_name, 'main_sku': main_sku,
            'gift_sku1': gift_sku1, 'gift_sku2': gift_sku2, 'gift_sku3': gift_sku3, 'note': note,
        }
        rec.update(_audit_new())
        batch.append(rec)
        touched_pairs.add((dealer_name, dealer_code))

        if len(batch) >= 80:
            try:
                sb.table('dealer_sku_mapping').insert(batch).execute()
                ok_cnt += len(batch)
            except Exception:
                err_cnt += len(batch)
            batch = []

    if batch:
        try:
            sb.table('dealer_sku_mapping').insert(batch).execute()
            ok_cnt += len(batch)
        except Exception:
            err_cnt += len(batch)

    for dealer_name, dealer_code in touched_pairs:
        if dealer_code:
            _recompute_ambiguous(dealer_name, dealer_code)

    return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt, 'invalid_sku': invalid_cnt})


# ============================================================
# ④ 比對引擎（EAN(經銷商對照表) > EAN(內部料號清單自己的條碼) > 經銷商代碼(排除歧義) > 品名）
# ============================================================
def _load_mapping_index():
    res = sb.table('dealer_sku_mapping').select('*').execute()
    rows = res.data or []
    by_ean, by_code, by_name = {}, {}, {}
    code_variants = {}
    for r in rows:
        dealer = r['dealer_name']
        if r.get('ean'):
            by_ean[(dealer, _norm(r['ean']))] = r
        if r.get('dealer_code'):
            key = (dealer, _norm(r['dealer_code']))
            code_variants.setdefault(key, set()).add(r['main_sku'])
            by_code[key] = r
        if r.get('product_name'):
            by_name[(dealer, _norm(r['product_name']))] = r
    ambiguous = {k for k, v in code_variants.items() if len(v) > 1}

    # 內部料號清單自己的EAN（廠商條碼，同一商品在各經銷商通路下的EAN通常相同，非經銷商專屬）。
    # 用途：當「經銷商品項對照表」(dealer_sku_mapping) 尚無此經銷商代碼/EAN的人工確認紀錄時，
    # 直接拿內部料號清單的EAN當自動比對備援，減少每次對帳都要手動確認內部料號的時間。
    sku_rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('internal_code,ean'))
    sku_by_ean = {}
    ean_ambiguous = set()
    for s in sku_rows:
        if s.get('ean'):
            en = _norm(s['ean'])
            if en in sku_by_ean and sku_by_ean[en] != s['internal_code']:
                ean_ambiguous.add(en)  # 同一EAN對到多個不同內部料號，視為歧義，不參與自動比對
            else:
                sku_by_ean[en] = s['internal_code']
    for en in ean_ambiguous:
        sku_by_ean.pop(en, None)

    return by_ean, by_code, by_name, ambiguous, sku_by_ean


def _match_one(dealer, ean, dealer_code, product_name, idx):
    by_ean, by_code, by_name, ambiguous, sku_by_ean = idx
    ean_n, code_n, name_n = _norm(ean), _norm(dealer_code), _norm(product_name)
    if ean_n and (dealer, ean_n) in by_ean:
        return by_ean[(dealer, ean_n)]
    if ean_n and ean_n in sku_by_ean:
        return {'main_sku': sku_by_ean[ean_n], 'gift_sku1': None, 'gift_sku2': None, 'gift_sku3': None}
    if code_n and (dealer, code_n) not in ambiguous and (dealer, code_n) in by_code:
        return by_code[(dealer, code_n)]
    if name_n and (dealer, name_n) in by_name:
        return by_name[(dealer, name_n)]
    return None


# ============================================================
# ⑤ 經銷商格式偵測與檔案解析
# ============================================================
def _detect_dealer(wb):
    res = sb.table('dealer_format_rules').select('*').eq('is_active', True).execute()
    best, best_score = None, 0
    for rule in res.data or []:
        keywords = [k.strip() for k in (rule.get('detect_sheet_keywords') or '').split(',') if k.strip()]
        if not keywords:
            continue
        score = sum(1 for kw in keywords if any(kw in sn for sn in wb.sheetnames))
        if score > best_score:
            best_score, best = score, rule
    return best


def _find_sheet(wb, name_hint):
    for sn in wb.sheetnames:
        if sn == name_hint:
            return sn
    for sn in wb.sheetnames:
        if name_hint and name_hint in sn:
            return sn
    return None


def _parse_workbook(wb, sheets_config):
    lines = []
    for cfg in sheets_config or []:
        target = _find_sheet(wb, cfg.get('sheet_name', ''))
        if not target:
            continue
        ws = wb[target]
        header_row = int(cfg.get('header_row') or 1)
        headers = {}
        for cell in ws[header_row]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column - 1

        def idx(field):
            h = cfg.get(field)
            return headers.get(str(h).strip()) if h else None

        i_code = idx('col_dealer_code')
        i_code2 = idx('col_dealer_code2')
        i_ean = idx('col_ean')
        i_name = idx('col_product_name')
        i_name2 = idx('col_product_name2')
        i_qty = idx('col_qty')
        i_price = idx('col_unit_price')
        i_amount = idx('col_amount')  # 舊版相容欄位
        i_amount_untaxed = idx('col_amount_untaxed')
        i_amount_taxed = idx('col_amount_taxed')
        i_tax = idx('col_tax_amount')
        i_rt = idx('col_row_type')

        mode = cfg.get('row_type_mode', 'fixed')
        fixed_type = cfg.get('row_type_fixed', 'sale')
        return_values = [v.strip() for v in (cfg.get('return_values') or '').split(',') if v.strip()]

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            def get(i):
                return row[i] if (i is not None and i < len(row)) else None

            if i_name is None or not get(i_name):
                continue

            if mode == 'by_column' and i_rt is not None:
                row_type = 'return' if _norm(get(i_rt)) in return_values else 'sale'
            else:
                row_type = fixed_type

            qty = _to_num(get(i_qty))
            amount = _to_num(get(i_amount))
            amount_untaxed = _to_num(get(i_amount_untaxed)) if i_amount_untaxed is not None else 0
            amount_taxed = _to_num(get(i_amount_taxed)) if i_amount_taxed is not None else 0
            if row_type == 'return':
                qty = -abs(qty)
                amount = -abs(amount)
                amount_untaxed = -abs(amount_untaxed)
                amount_taxed = -abs(amount_taxed)
            # 舊版相容：若新規則兩個欄位都沒設定（尚未依 v2.59 更新規則），退回用舊的 col_amount
            # 值當作「金額」(amount) 顯示，但含稅/未稅兩個新欄位保持0，財報畫面上會明顯看出
            # 這個經銷商尚未設定稅別分類，需要去「料號與規則管理」補設定，不會被靜默誤分類
            if i_amount_untaxed is None and i_amount_taxed is None:
                amount_untaxed = 0
                amount_taxed = 0

            code_main = _norm(get(i_code))
            code_sub = _norm(get(i_code2)) if i_code2 is not None else ''
            if code_main and code_sub:
                dealer_code = code_main + '-' + code_sub
            else:
                dealer_code = code_main or code_sub or None

            name_main = _norm(get(i_name))
            name_sub = _norm(get(i_name2)) if i_name2 is not None else ''
            if name_main and name_sub:
                product_name = name_main + ' - ' + name_sub
            else:
                product_name = name_main or name_sub or None

            tax_amount = _to_num(get(i_tax)) if i_tax is not None else None
            if tax_amount is not None and row_type == 'return':
                tax_amount = -abs(tax_amount)

            lines.append({
                'row_type': row_type,
                'dealer_code': dealer_code,
                'ean': _norm(get(i_ean)) or None,
                'product_name': product_name,
                'qty': qty,
                'unit_price': get(i_price),
                'amount': amount,
                'amount_untaxed': amount_untaxed,
                'amount_taxed': amount_taxed,
                'tax_amount': tax_amount,
                'raw_json': {'sheet': target, 'row': [str(v) if v is not None else None for v in row]},
            })
    return lines


_PDF_DATA_LINE_RE = re.compile(r'^[\d.]+\s+[\d,]+\.?\d*\s+[\d,]+\.?\d*\s+\S+$')


def _parse_pdf_invoice(raw_bytes):
    """解析PDF格式的電子發票（目前用於「香港經銷商」，發票由凌網知識開立，買受人為萬耀科技）。
    回傳格式與 _parse_workbook() 相同的 lines 陣列，讓後續比對/入庫邏輯可以直接共用，不需另外處理。

    版面說明：電子發票證明聯是左右兩欄排版——左欄是證明聯本體(含QR code、發票號碼等)，
    右欄是「交易明細」表格。若直接整頁抽取文字，兩欄內容會依上下位置交錯混在一起無法直接解析，
    因此改用座標篩選：只取 x0>=200（右欄範圍）的文字，再依 y 座標(top)分行重組。

    交易明細每個商品固定佔兩行：第一行是品名，第二行是「數量 單價 金額 稅」。
    已核對過樣本發票：逐行「金額」加總 = 發票的「應稅銷售額」，代表這欄是未稅金額，故存入
    amount_untaxed；發票的稅額是以「應稅銷售額」整張發票的5%計算，不逐行拆分，跟其他經銷商的
    稅額欄位設計邏輯不同，此處刻意不填 tax_amount，避免逐行金額對不齊發票總額造成誤解。

    ⚠️ 2026-07-24修復：原本用「品名行是否以數字開頭」來判斷是否為有效品名列，這個假設在
    香港經銷商的樣本（如「Mini+白 機＋贈亞麻蒼殻」）成立，但展碁的商品名稱常以螢幕吋數開頭
    （如「6吋GAZE MINI+」「7.8 PRO NOTE C」），會被誤判成不是品名而整批跳過，導致解析結果
    是空的。改用更精確的判斷：直接檢查該行文字本身是否完全符合「數量 單價 金額 稅別」這種
    純數字資料列的格式（DATA_LINE_RE），符合才視為資料列跳過；否則一律當品名列處理，不論
    開頭是不是數字。已用展碁與模擬的香港經銷商格式雙重驗證，兩者皆能正確解析。

    「運費」不是實際商品，不能單獨列一筆去比對內部料號，但2026-07-24使用者確認拆帳金額必須
    跟發票金額完全一致（否則對帳金額會少了運費那一截），故改為：運費金額不單獨成一筆，而是
    依這一頁其他商品各自的金額比例，攤加進每個商品品項的 amount/amount_untaxed 裡（見下方
    攤提邏輯），確保商品攤提後金額加總 = 商品原金額 + 運費 = 發票金額，不會有尾差。攤提前後的
    金額都記錄在 raw_json 的 amount_before_freight / freight_allocated，供事後對照追查。

    商品名稱是「機型+贈品」組合文字（如「Mini+白 機＋贈亞麻蒼殻」），沒有料號/EAN可比對，
    因此直接把這段文字當作 dealer_code，交由「經銷商品項對照表」手動維護對應關係——
    跟其他經銷商代碼比對邏輯完全共用同一套機制，差別只是這裡的「代碼」本身是一段中文描述。
    """
    lines_out = []
    with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
        for page in pdf.pages:
            words = [w for w in page.extract_words() if w['x0'] >= 200]
            if not words:
                continue
            rows = {}
            for w in words:
                key = round(w['top'] / 2) * 2
                rows.setdefault(key, []).append(w)
            text_rows = [
                ' '.join(w['text'] for w in sorted(rows[k], key=lambda w: w['x0']))
                for k in sorted(rows.keys())
            ]

            start_i = next((i for i, t in enumerate(text_rows) if '品名' in t and '數量' in t), None)
            end_i = next((i for i, t in enumerate(text_rows) if '應稅銷售額' in t), None)
            if start_i is None or end_i is None:
                continue
            body = text_rows[start_i + 1:end_i]

            page_lines = []
            freight_amt = 0
            i = 0
            while i < len(body):
                first_line = body[i].strip()
                if not first_line or _PDF_DATA_LINE_RE.match(first_line):
                    # 空行，或這行本身就已經是「數量 單價 金額 稅」格式的資料列（代表對齊有問題
                    # 或為孤立資料列），不是品名列，跳過——不再用「是否以數字開頭」判斷，
                    # 因為展碁等經銷商的商品名稱本身就常以數字(螢幕吋數)開頭
                    i += 1
                    continue
                # ⚠️ 2026-08-05修復：商品名稱長度較長時，PDF右欄排版會把品名自動換行成2行甚至
                # 更多行（如「HyRead Gaze Pro Note C 7.8」+「吋螢幕貼」），原本寫死「品名固定佔
                # 1行、緊接著就是資料列」，遇到換行品名時，資料列其實落在第3行，導致這筆商品的
                # 名稱/數量/金額全部解析失敗被跳過（使用者回報「匯入品項比實際發票少很多」，
                # 用DE-51907257這張發票實測，19個品項只解析出10筆，剛好都是換行品名的那9筆消失）。
                # 改為：從品名開始逐行累積，直到遇到符合「數量 單價 金額 稅」格式的資料列為止，
                # 中間累積的所有行合併成完整品名（中文換行處不需要空格，直接接續）
                name_parts = [first_line]
                j = i + 1
                while j < len(body) and not _PDF_DATA_LINE_RE.match(body[j].strip()):
                    part = body[j].strip()
                    if part:
                        name_parts.append(part)
                    j += 1
                if j >= len(body):
                    # 累積到body結尾都找不到對應的資料列（例如末尾殘留的說明文字），放棄這筆
                    i = j
                    continue
                name_line = ''.join(name_parts)
                data_line = body[j].strip()
                m = re.match(r'^([\d.]+)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+(\S+)', data_line)
                i = j + 1
                if not m:
                    continue
                qty = _to_num(m.group(1))
                unit_price = _to_num(m.group(2).replace(',', ''))
                amount = _to_num(m.group(3).replace(',', ''))
                if '運費' in name_line:
                    # 運費不是實際商品，不能單獨列一筆比對，但2026-07-24使用者確認拆帳金額必須
                    # 跟發票金額一致，故不再整筆捨棄，改為記錄下來，等這一頁其他商品都解析完後
                    # 依各商品金額比例攤到每個品項的金額裡（見下方攤提邏輯）
                    freight_amt += amount
                    continue
                page_lines.append({
                    'row_type': 'sale',
                    'dealer_code': name_line,
                    'ean': None,
                    'product_name': name_line,
                    'qty': qty,
                    'unit_price': unit_price,
                    'amount': amount,
                    'amount_untaxed': amount,
                    'amount_taxed': 0,
                    'tax_amount': None,
                    'raw_json': {'source': 'pdf_invoice', 'name_line': name_line, 'data_line': data_line},
                })

            if freight_amt and page_lines:
                # 依各商品「攤提前金額」佔全頁商品金額總數的比例分攤運費，最後一筆用「總運費-已分攤」
                # 反推，確保攤提後全部商品金額加總 = 商品原金額總數 + 運費，跟發票總額對得起來，
                # 不會因為逐筆四捨五入而產生尾差
                total_amt = sum(l['amount'] for l in page_lines)
                allocated_so_far = 0
                for idx, l in enumerate(page_lines):
                    if total_amt:
                        if idx < len(page_lines) - 1:
                            share = round(freight_amt * l['amount'] / total_amt, 4)
                        else:
                            share = round(freight_amt - allocated_so_far, 4)
                    else:
                        share = round(freight_amt / len(page_lines), 4) if idx < len(page_lines) - 1 else round(freight_amt - allocated_so_far, 4)
                    allocated_so_far += share
                    l['raw_json']['amount_before_freight'] = l['amount']
                    l['raw_json']['freight_allocated'] = share
                    l['amount'] = round(l['amount'] + share, 4)
                    l['amount_untaxed'] = l['amount']

            lines_out.extend(page_lines)
    return lines_out


def _parse_shipping_order(raw_bytes):
    """解析香港經銷商的出貨單（.odt格式）。回傳格式與其他parser相同的 lines 陣列。

    背景：發票上的金額不含部分贈品（贈品不會單獨列在發票上），使用者確認「實際出庫數以
    出貨單為主」，因此出貨單負責提供正確的數量（含贈品），發票負責提供正確的金額，
    兩者分開上傳、各自的批次分別入庫，最後在 dealer_reconcile_summary 合併成一筆——
    數量看出貨單、金額看發票，不會互相覆蓋也不會重複加總（見 commit_batch() 的
    source_kind 判斷）。

    版面固定：先有訂購日期/訂單編號/出貨單號/帳號/收件人幾列中繼資料，接著「商品編號｜
    商品名稱｜數量」表頭，之後每一列是一項出貨明細。商品名稱已內嵌型號代碼
    （如「...石墨綠(S001A014W)」）方便使用者人工比對時參考，但比對機制仍統一採用手動
    維護「經銷商品項對照表」（跟發票PDF的比對方式一致），不另外開發以型號代碼自動比對的
    邏輯——因為同一型號常有多個顏色/款式（如S002A014W就有好幾種顏色的側翻殼），光靠型號
    代碼本身無法唯一判斷，容易誤配對到錯的顏色款式。

    使用者確認：實務上一份.odt檔案就是一張出貨單（不會把同期間多張出貨單合併在一個檔案裡），
    但底下解析邏輯仍寫成可以處理多張（若同一檔案出現第二組「商品編號」表頭，會繼續往下解析），
    多一點彈性、不假設檔案一定只有一張。
    """
    from odf.opendocument import load as _odf_load
    from odf.table import Table as _OdfTable, TableRow as _OdfRow, TableCell as _OdfCell
    import odf.teletype as _teletype

    doc = _odf_load(io.BytesIO(raw_bytes))
    lines_out = []
    order_no = ship_no = order_date = None

    for table in doc.getElementsByType(_OdfTable):
        rows_data = []
        for r in table.getElementsByType(_OdfRow):
            cells = r.getElementsByType(_OdfCell)
            rows_data.append([_teletype.extractText(c).strip() for c in cells])

        i = 0
        while i < len(rows_data):
            vals = rows_data[i]
            if len(vals) >= 2 and vals[0] == '訂單編號':
                order_no = vals[1]
                if len(vals) >= 4 and vals[2] == '出貨單號':
                    ship_no = vals[3]
            if len(vals) >= 2 and vals[0] == '訂購日期':
                order_date = vals[1]
            if len(vals) >= 3 and vals[0] == '商品編號' and vals[1] == '商品名稱' and '數量' in vals[2]:
                j = i + 1
                while j < len(rows_data):
                    row = rows_data[j]
                    if len(row) != 3 or not row[0].strip().isdigit() or not row[2].strip().isdigit():
                        break  # 遇到非商品列（如行銷頁尾區塊）就停止，回到外層繼續找下一組表頭
                    code, name, qty = row
                    name = name.strip()
                    lines_out.append({
                        'row_type': 'sale',
                        'dealer_code': name,
                        'ean': None,
                        'product_name': name,
                        'qty': _to_num(qty),
                        'unit_price': None,
                        'amount': 0,
                        'amount_untaxed': 0,
                        'amount_taxed': 0,
                        'tax_amount': None,
                        'raw_json': {'source': 'shipping_order', 'order_no': order_no,
                                     'shipping_no': ship_no, 'order_date': order_date,
                                     'product_code': code},
                    })
                    j += 1
                i = j
                continue
            i += 1
    return lines_out


# ============================================================
# ⑥ 上傳、比對、批次管理
# ============================================================
def _next_recon_seq_no():
    """產生下一個「拆帳編號」，格式C+4位數流水號（如C0001、C0002…），依目前資料庫裡
    已經用過的最大編號往下一號，不會因為批次被刪除而回收/重複利用舊編號。2026-08-18
    新增，供「上傳比對」批次（見_ingest_reconcile_file()）標記每一次上傳的識別碼，
    方便使用者在「對帳查詢」矩陣、「匯入團體發票」列表對照是哪一次上傳/入庫。"""
    res = sb.table('dealer_reconcile_batches').select('recon_seq_no') \
        .not_.is_('recon_seq_no', 'null').execute()
    max_n = 0
    for r in (res.data or []):
        v = (r.get('recon_seq_no') or '').strip()
        if v[:1] == 'C' and v[1:].isdigit():
            max_n = max(max_n, int(v[1:]))
    return f'C{max_n + 1:04d}'


def _attach_recon_seq_info(rows):
    """2026-08-21新增（從list_group_invoices()原本的內嵌邏輯抽出成共用函式）：依每一列的
    'reconciled_upload_group'欄位（可以是None），反查該upload_group底下任一批次的拆帳編號
    （dealer_reconcile_batches.recon_seq_no，同一組批次一定共用同一個編號，見
    _ingest_reconcile_file()），掛回rows裡每一列的'recon_seq_no'欄位；同時判斷這個
    upload_group底下是否還有任何批次存在，沒有的話代表對應批次事後被刪除了（常見於發現是
    重複資料而手動清除），掛回'recon_batch_missing'（True/False），供前端顯示警示圖示。
    呼叫前rows的每個dict都必須已經有'reconciled_upload_group'這個key。
    供list_group_invoices()（發票本身）跟list_shipping_docs()（出貨單，透過
    matched_invoice_id間接取得對應發票的reconciled_upload_group）共用，避免同一段查詢
    邏輯寫兩份、以後改一邊忘了改另一邊。"""
    groups_needed = {r.get('reconciled_upload_group') for r in rows if r.get('reconciled_upload_group')}
    seq_by_group = {}
    existing_groups = set()
    if groups_needed:
        try:
            seq_res = sb.table('dealer_reconcile_batches').select('upload_group,recon_seq_no') \
                .in_('upload_group', list(groups_needed)).execute()
            for br in (seq_res.data or []):
                g = br.get('upload_group')
                if g:
                    existing_groups.add(g)
                if g and br.get('recon_seq_no') and g not in seq_by_group:
                    seq_by_group[g] = br['recon_seq_no']
        except Exception:
            pass
    for r in rows:
        g = r.get('reconciled_upload_group')
        r['recon_seq_no'] = seq_by_group.get(g)
        r['recon_batch_missing'] = bool(g) and g not in existing_groups


def _ingest_reconcile_file(raw_bytes, filename, dealer_override, period, upload_group,
                            user_id, user_name, force_source_kind=None):
    """2026-08-15新增：從 upload_batch() 原本的內容抽出來的共用核心邏輯（解析檔案→比對→
    建立batch+lines），供「上傳比對」手動上傳（upload_batch()本身）以及新增的「從匯入
    團體發票勾選匯入」功能（reconcile_import_from_group_invoices()）共用，避免兩份幾乎
    一樣的解析/比對邏輯各自維護一份、以後改一邊忘了改另一邊。

    回傳 (result_dict, None) 表示成功；(None, (error_body_dict, http_status)) 表示失敗，
    呼叫端自行決定要不要 jsonify() 這個錯誤（upload_batch()直接回傳給前端；勾選匯入那邊
    則是彙整成多筆結果的其中一筆，不會讓其他勾選的發票連帶失敗）。

    force_source_kind：只給新的「從匯入團體發票匯入」路徑使用，強制指定source_kind
    （'invoice_amount_only'／'shipping_qty_only'），不受經銷商名稱是否為「香港經銷商」
    這個既有判斷影響——決策②：發票管金額、出貨單管數量的雙軌計算規則，擴大套用到
    「從匯入團體發票勾選匯入」這條新路徑的所有經銷商，不只香港經銷商。
    upload_batch() 手動上傳這條路徑完全不會傳這個參數（維持None），行為（含既有「只有
    香港經銷商」才會這樣拆分的特判）100%不變，不影響展碁/非銷售開發票等目前正確運作、
    PDF-only且沒有另外的出貨單來源的經銷商。
    """
    filename_lower = (filename or '').lower()
    is_pdf = filename_lower.endswith('.pdf')
    is_odt = filename_lower.endswith('.odt')
    # source_kind 決定 commit_batch() 入庫時，這批資料的數量/金額要不要計入彙總表：
    #   'normal'：數量與金額都計入（現有Excel對帳檔的行為，不受本次異動影響）
    #   'invoice_amount_only'：只計入金額，數量不計入——用於香港經銷商的發票（PDF），因為
    #                          發票的數量不含贈品，正確數量另外由出貨單提供，避免重複加總
    #   'shipping_qty_only'：只計入數量，金額不計入（金額本來就是0，這裡是防呆）——
    #                        用於香港經銷商的出貨單（.odt），提供含贈品的實際出貨數量
    source_kind = 'normal'
    rule = None
    lines = None
    bookstore_auto_adjustment = None  # 只有「書店」且上傳檔案內建報表分頁時才會有值，見下方
    bookstore_qty_mismatches = []  # 同上，逐一財報料號品名的數量落差清單，見下方

    # 2026-08-18新增：「書店」經銷商上傳的是「出貨資料」格式Excel，直接沿用既有書店拆帳
    # 引擎（_compute_bookstore_order_matrix()），走專屬程式碼路徑（比照香港經銷商PDF/ODT），
    # 不透過「經銷商格式規則」(dealer_format_rules) 設定。務必排在is_pdf/is_odt判斷之前——
    # 「書店」上傳的檔案副檔名是.xlsx，若不特判會落入下面else的Excel泛用解析分支，因為
    # dealer_format_rules裡沒有「書店」規則而回報「找不到經銷商規則」的錯誤。
    if dealer_override == '書店':
        try:
            lines = _parse_bookstore_shipment(raw_bytes)
        except Exception as e:
            return None, ({'error': f'無法解析「書店」出貨資料：{e}'}, 400)
        rule = {'dealer_name': '書店'}
        # 2026-08-19新增：使用者反映不想為了核對總額/數量另外多上傳一份參考檔——如果這次
        # 上傳的檔案本身「剛好」就是先前的「書店訂單拆帳」報表（同一活頁簿內建原始出貨
        # 資料＋算好的報表分頁），直接用內建的報表總額/總表分析數量，跟這次拆帳引擎重新
        # 算出來的批次結果比對，有落差就自動加明細列進這個批次，不需要使用者再做任何動作。
        # 注意：qty/金額彙總都要用「原始解析出來的lines」（還沒附加任何調整列之前），
        # 避免調整列本身（product_name不是真正的財報料號品名）污染彙總結果。
        try:
            _wb_for_check = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception:
            _wb_for_check = None
        embedded_ref_total_taxed = _extract_bookstore_report_grand_total_taxed(_wb_for_check) if _wb_for_check else None
        ref_qty_by_name = _extract_bookstore_report_qty_by_name(_wb_for_check) if _wb_for_check else None

        if embedded_ref_total_taxed is not None:
            batch_total_taxed = round(sum(_to_num(ln.get('amount_taxed')) for ln in lines), 2)
            residual_taxed = round(embedded_ref_total_taxed - batch_total_taxed, 2)
            if abs(residual_taxed) > 0.02:
                residual_untaxed = round(residual_taxed / 1.05, 2)
                bookstore_auto_adjustment = {
                    'embedded_ref_total_taxed': embedded_ref_total_taxed,
                    'batch_total_taxed': batch_total_taxed,
                    'residual_taxed': residual_taxed,
                }
                lines.append({
                    'row_type': 'sale', 'dealer_code': '(整體調整)', 'ean': None,
                    'product_name': '整體金額調整（上傳檔案內建「書店訂單拆帳」報表總額 vs 本次拆帳引擎重新'
                                     '計算總額有落差，常見原因是報表產生後「方案清單」設定又被異動過，請人工'
                                     '確認並指定料號）',
                    'qty': 0, 'unit_price': None,
                    'amount': residual_taxed, 'amount_untaxed': residual_untaxed, 'amount_taxed': residual_taxed,
                    'tax_amount': round(residual_taxed - residual_untaxed, 2),
                    'raw_json': {'source': 'bookstore_embedded_report_auto_adjustment',
                                 'note': '上傳檔案內建的「書店訂單拆帳」報表總額(含稅)跟本次重新計算總額不同',
                                 **bookstore_auto_adjustment},
                    '_prematched': True, 'matched_main_sku': None, 'match_status': 'unmatched',
                })

        # 2026-08-19新增：使用者發現部分方案名稱有「xN」數量提示（例如「HyRead Gaze Mini
        # 系列 6吋透明軟膠殼x3」實際上每訂一次要出3個「Mini透明殼」），但「方案清單」品項
        # 設定沒有對應調整，導致拆帳引擎算出來的數量比報表「總表分析」分頁的「料號加總
        # (境內+境外)」列少。這裡逐一財報料號品名比對兩邊數量，有落差就個別加一筆「數量
        # 差異調整」明細列（金額固定0，因為金額落差已經由上面的整體金額調整涵蓋，這裡純粹
        # 是提醒使用者這個料號的數量需要人工確認/補正），使用者可以在批次明細比照其他
        # 「未比對」列直接指定內部料號並確認。
        if ref_qty_by_name:
            batch_qty_by_name = {}
            for ln in lines:
                if ln.get('dealer_code') == '(整體調整)':
                    continue
                name = ln.get('product_name')
                # 2026-08-19修正：'代銷商品'／'電子書方案'這2個集合類品名，
                # _extract_bookstore_report_qty_by_name()本來就特意排除在ref_qty_by_name
                # 之外（這2欄數量欄含意特殊，不适合逐一比對，見該函式註解）——這裡如果沒有
                # 同步排除，這2個品名永遠找不到對應的參考數量(視為0)，會被誤判成「批次比
                # 參考資料多了一整批」的假落差，每次上傳都會誤報，故在這裡也排除，兩邊
                # 排除範圍保持一致。
                if not name or name in _BOOKSTORE_ORDER_AGGREGATE_COL_NAMES:
                    continue
                batch_qty_by_name[name] = batch_qty_by_name.get(name, 0.0) + _to_num(ln.get('qty'))
            all_names = set(ref_qty_by_name.keys()) | set(batch_qty_by_name.keys())
            for name in sorted(all_names):
                ref_qty = round(ref_qty_by_name.get(name, 0.0), 4)
                batch_qty = round(batch_qty_by_name.get(name, 0.0), 4)
                delta_qty = round(ref_qty - batch_qty, 4)
                if abs(delta_qty) <= 0.01:
                    continue
                bookstore_qty_mismatches.append({
                    'product_name': name, 'ref_qty': ref_qty, 'batch_qty': batch_qty, 'delta_qty': delta_qty,
                })
                lines.append({
                    'row_type': 'sale', 'dealer_code': '(數量調整)', 'ean': None,
                    'product_name': f'數量差異調整（{name}）：上傳檔案內建「總表分析」分頁的料號加總數量 vs 本次'
                                     f'拆帳引擎重新計算數量不同，常見原因是方案名稱有「xN」等數量提示但「方案'
                                     f'清單」品項設定未對應調整，請人工確認並指定料號',
                    'qty': delta_qty, 'unit_price': None,
                    'amount': 0, 'amount_untaxed': 0, 'amount_taxed': 0, 'tax_amount': 0,
                    'raw_json': {'source': 'bookstore_embedded_report_qty_check', 'report_product_name': name,
                                 'ref_qty': ref_qty, 'batch_qty': batch_qty, 'delta_qty': delta_qty},
                    '_prematched': True, 'matched_main_sku': None, 'match_status': 'unmatched',
                })
    elif is_pdf:
        # PDF發票（如香港經銷商）目前一定要手動指定經銷商——沒有Excel分頁名稱可用來自動判斷格式，
        # 且PDF發票通常一張只代表一次出貨，同一經銷商同一期間會陸續上傳多張，累加進同一批彙總
        if not dealer_override:
            return None, ({'error': 'PDF發票上傳需要手動指定經銷商，請在上方經銷商欄位選擇'}, 422)
        rule_res = sb.table('dealer_format_rules').select('*') \
            .eq('dealer_name', dealer_override).eq('is_active', True).execute()
        rule = rule_res.data[0] if rule_res.data else None
        if not rule:
            return None, ({'error': f'找不到「{dealer_override}」的經銷商規則，請先至「料號與規則管理」建立'}, 422)
        try:
            lines = _parse_pdf_invoice(raw_bytes)
        except Exception as e:
            return None, ({'error': f'無法解析PDF發票：{e}'}, 400)
        if dealer_override == '香港經銷商':
            source_kind = 'invoice_amount_only'
    elif is_odt:
        # 出貨單（.odt，目前僅香港經銷商使用）——跟PDF發票一樣，沒有版面可自動判斷經銷商，
        # 上傳時務必手動指定
        if not dealer_override:
            return None, ({'error': '出貨單上傳需要手動指定經銷商，請在上方經銷商欄位選擇'}, 422)
        rule_res = sb.table('dealer_format_rules').select('*') \
            .eq('dealer_name', dealer_override).eq('is_active', True).execute()
        rule = rule_res.data[0] if rule_res.data else None
        if not rule:
            return None, ({'error': f'找不到「{dealer_override}」的經銷商規則，請先至「料號與規則管理」建立'}, 422)
        try:
            lines = _parse_shipping_order(raw_bytes)
        except Exception as e:
            return None, ({'error': f'無法解析出貨單：{e}'}, 400)
        if dealer_override == '香港經銷商':
            source_kind = 'shipping_qty_only'
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as e:
            return None, ({'error': f'無法讀取Excel檔案：{e}'}, 400)

        if dealer_override:
            rule_res = sb.table('dealer_format_rules').select('*') \
                .eq('dealer_name', dealer_override).eq('is_active', True).execute()
            rule = rule_res.data[0] if rule_res.data else None
        else:
            rule = _detect_dealer(wb)

        if not rule:
            return None, ({
                'error': '無法自動判斷經銷商格式，請至「料號與規則管理」確認規則，或於上傳時手動指定經銷商',
                'available_dealers': DEALERS,
            }, 422)

        lines = _parse_workbook(wb, rule.get('sheets_config'))

    if force_source_kind is not None:
        source_kind = force_source_kind

    if not lines:
        return None, ({'error': f'依「{rule["dealer_name"]}」規則解析後沒有任何資料列，請檢查規則設定的分頁名稱/表頭列（PDF發票／出貨單請確認版面是否跟預期一致）'}, 422)

    idx = _load_mapping_index()
    # 2026-08-18新增：「拆帳編號」（格式C+4碼流水號，如C0001）——同一次「上傳並比對」選取
    # 的檔案（同一個upload_group，如香港經銷商固定1張出貨單+1張發票一起選取）共用同一個
    # 編號，見_group_batches_by_upload()既有的合併顯示邏輯。同一個upload_group底下的批次
    # 是分開的HTTP請求各自呼叫本函式建立的（doUpload()逐檔案上傳），所以每次都要先查這個
    # upload_group是否已經有其他批次領過編號，有的話沿用同一個，沒有才產生新的一號，
    # 避免同一次上傳的出貨單+發票兩批各自拿到不同編號。
    recon_seq_no = None
    if upload_group:
        existing_grp = sb.table('dealer_reconcile_batches').select('recon_seq_no') \
            .eq('upload_group', upload_group).not_.is_('recon_seq_no', 'null').limit(1).execute()
        if existing_grp.data:
            recon_seq_no = existing_grp.data[0].get('recon_seq_no')
    if not recon_seq_no:
        recon_seq_no = _next_recon_seq_no()
    batch_rec = {
        'dealer_name': rule['dealer_name'],
        'period': period or None,
        'source_filename': filename,
        'status': '待確認',
        'total_lines': len(lines),
        'source_kind': source_kind,
        'upload_group': upload_group,
        'recon_seq_no': recon_seq_no,
    }
    batch_rec.update({'created_by': user_id, 'created_by_name': user_name})
    batch_res = sb.table('dealer_reconcile_batches').insert(batch_rec).execute()
    batch = batch_res.data[0]

    matched, unmatched = 0, 0
    line_recs = []
    for ln in lines:
        # 注意：同一批次 INSERT 的所有列，keys 必須完全一致（否則 PostgREST 會回傳
        # PGRST102 "All object keys must match"）。matched_main_sku 等欄位一律固定存在，
        # 未比對時明確填 None，不可省略欄位。
        rec = {
            'batch_id': batch['id'],
            'row_type': ln['row_type'],
            'dealer_code': ln['dealer_code'],
            'ean': ln['ean'],
            'product_name': ln['product_name'],
            'qty': ln['qty'],
            'unit_price': ln['unit_price'],
            'amount': ln['amount'],
            'amount_untaxed': ln.get('amount_untaxed', 0),
            'amount_taxed': ln.get('amount_taxed', 0),
            'tax_amount': ln.get('tax_amount'),
            'matched_main_sku': None,
            'matched_gift_sku1': None,
            'matched_gift_sku2': None,
            'matched_gift_sku3': None,
            'match_status': 'unmatched',
            'raw_json': ln['raw_json'],
        }
        # 2026-08-18新增：「書店」（_parse_bookstore_shipment()）已經用書店拆帳引擎＋子料號
        # 邏輯自己決定好matched_main_sku/match_status，不透過dealer_sku_mapping這套
        # EAN/經銷商代碼/品名比對機制（那套機制對書店的場景不適用），直接採用它算好的結果。
        if ln.get('_prematched'):
            rec['matched_main_sku'] = ln.get('matched_main_sku')
            rec['match_status'] = ln.get('match_status') or 'unmatched'
            if rec['match_status'] == 'auto_matched':
                matched += 1
            else:
                unmatched += 1
            line_recs.append(rec)
            continue

        m = _match_one(rule['dealer_name'], ln['ean'], ln['dealer_code'], ln['product_name'], idx)
        if m:
            rec['matched_main_sku'] = m['main_sku']
            rec['matched_gift_sku1'] = m.get('gift_sku1')
            rec['matched_gift_sku2'] = m.get('gift_sku2')
            rec['matched_gift_sku3'] = m.get('gift_sku3')
            rec['match_status'] = 'auto_matched'
            matched += 1
        else:
            unmatched += 1
        line_recs.append(rec)

    for i in range(0, len(line_recs), 80):
        sb.table('dealer_reconcile_lines').insert(line_recs[i:i + 80]).execute()

    sb.table('dealer_reconcile_batches').update({
        'matched_lines': matched, 'unmatched_lines': unmatched,
    }).eq('id', batch['id']).execute()

    return {
        'batch_id': batch['id'], 'dealer_name': rule['dealer_name'],
        'total_lines': len(lines), 'matched_lines': matched, 'unmatched_lines': unmatched,
        'bookstore_auto_adjustment': bookstore_auto_adjustment,
        'bookstore_qty_mismatches': bookstore_qty_mismatches,
    }, None


@dealer_bp.route('/api/dealer/upload', methods=['POST'])
@dealer_edit_required
def upload_batch():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '請選擇檔案'}), 400
    dealer_override = request.form.get('dealer_name', '').strip()
    period = request.form.get('period', '').strip()
    upload_group = request.form.get('upload_group', '').strip() or None
    raw_bytes = f.read()
    result, err = _ingest_reconcile_file(
        raw_bytes, f.filename, dealer_override, period or None, upload_group,
        session.get('user_id'), session.get('display_name', session.get('username', '')),
    )
    if err:
        resp, code = err
        return jsonify(resp), code
    return jsonify(result)


# ── 2026-08-15新增：從「匯入團體發票」勾選匯入 ──────────────────────────────
# 背景：「匯入團體發票」（bookstore_group_invoices）透過排程自動收集發票PDF＋出貨單
# 檔案，使用者確認4個決策後要求「上傳比對」新增一個機制，不用手動下載/選檔案上傳，
# 直接勾選已收集齊全的發票登記，系統自動抓Storage裡的檔案跑過同一套解析/比對流程。
# 決策①：每筆發票各自設定「經銷商類型」（dealer_name欄位），不用一張共用的
#        銷售單位→經銷商對照表。
# 決策②：發票管金額、出貨單管數量的雙軌規則，擴大套用到這條新路徑的所有經銷商
#        （用force_source_kind參數達成，見_ingest_reconcile_file()說明；不影響
#        upload_batch()既有的「只有香港經銷商」特判）。
# 決策③：必須發票PDF與出貨單兩者都齊全才能勾選匯入（list_reconcile_candidates()
#        與reconcile_import_from_group_invoices()都各自獨立檢查，不只靠前端）。
# 決策④：已標記「已寫入對帳」(reconciled_at有值)者鎖住，不可重複匯入。
@dealer_bp.route('/api/dealer/group-invoices/reconcile-candidates', methods=['GET'])
@dealer_edit_required
def list_reconcile_candidates():
    # 2026-08-21新增：「已拆帳」欄——使用者反饋候選清單裡有些發票其實已經透過別的方式
    # 處理掉了（如手動上傳），不需要一直留在候選清單提醒，但也不想真的刪除這筆發票登記，
    # 於是新增「封存」勾選+備註（見archive_reconcile_candidate()）。封存後預設從候選
    # 清單隱藏（使用者確認的行為），前端勾選「顯示已封存的候選」時帶include_archived=1
    # 才會連同封存的一起列出（供之後想取消封存時找回來）。
    include_archived = request.args.get('include_archived') in ('1', 'true', 'True')
    invs = _fetch_all_rows(lambda: sb.table('bookstore_group_invoices').select('*').order('id', desc=True))
    invs = [r for r in invs if _norm(r.get('dealer_name')) and not r.get('reconciled_at')]
    if not include_archived:
        invs = [r for r in invs if not r.get('archived')]
    if not invs:
        return jsonify([])
    ids = [r['id'] for r in invs]
    pdf_rows, doc_rows = [], []
    for i in range(0, len(ids), 200):
        chunk = ids[i:i + 200]
        pdf_rows += (sb.table('bookstore_group_invoice_pdfs').select('invoice_id,file_name')
                     .in_('invoice_id', chunk).execute().data or [])
        doc_rows += (sb.table('bookstore_group_invoice_files').select('invoice_id,file_name')
                     .in_('invoice_id', chunk).execute().data or [])
    pdf_map, doc_map = {}, {}
    for r in pdf_rows:
        pdf_map.setdefault(r['invoice_id'], []).append(r['file_name'])
    for r in doc_rows:
        doc_map.setdefault(r['invoice_id'], []).append(r['file_name'])
    out = []
    for r in invs:
        pdfs = pdf_map.get(r['id'], [])
        docs = doc_map.get(r['id'], [])
        if not pdfs or not docs:
            continue  # 決策③：兩者都要有才列入候選，缺一邊就不給勾選
        out.append({
            'id': r['id'], 'seller_name': r['seller_name'], 'dealer_name': r['dealer_name'],
            'invoice_no': r.get('invoice_no'), 'invoice_date': r.get('invoice_date'),
            # 2026-08-21新增：使用者反映勾選候選清單時容易勾錯，希望能多看到「OA訂單編號」
            # 當參考——這個欄位本來就存在於bookstore_group_invoices(見order_no)，這裡一併
            # 回傳給前端顯示，不需要新增欄位/SQL。
            'order_no': r.get('order_no'),
            'amount_incl': r.get('amount_incl'), 'recon_period': r.get('recon_period'),
            'pdf_files': pdfs, 'shipping_files': docs,
            # 2026-08-21新增：見上方list_reconcile_candidates()說明。
            'archived': bool(r.get('archived')), 'archive_note': r.get('archive_note'),
        })
    return jsonify(out)


@dealer_bp.route('/api/dealer/group-invoices/<int:iid>/archive', methods=['PUT'])
@dealer_edit_required
def archive_reconcile_candidate(iid):
    """2026-08-21新增：「從『匯入團體發票』勾選匯入」候選清單的「已拆帳」欄——讓使用者
    針對某筆候選發票標記「封存」+填寫備註。跟「已寫入對帳」(reconciled_at)是完全獨立的
    兩件事：封存只是使用者自己的篩選/備註機制，不會鎖住這筆發票、不會影響它在「匯入團體
    發票」列表本身的顯示或其他欄位，之後隨時可以取消勾選讓它重新出現在候選清單裡。
    只接受archived(bool)/archive_note(text)這2個欄位，不動到發票本身其他資料，避免跟
    update_group_invoice()那支完整編輯API（要求seller_name/amount_incl必填）搞混——
    這裡是候選清單表格裡就地更新，沒有完整表單可以填。"""
    data = request.json or {}
    rec = {}
    if 'archived' in data:
        rec['archived'] = bool(data.get('archived'))
    if 'archive_note' in data:
        rec['archive_note'] = _norm(data.get('archive_note')) or None
    if not rec:
        return jsonify({'ok': True})
    cur = sb.table('bookstore_group_invoices').select('id').eq('id', iid).limit(1).execute()
    if not cur.data:
        return jsonify({'error': '找不到這筆發票'}), 404
    sb.table('bookstore_group_invoices').update(rec).eq('id', iid).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/group-invoices/reconcile-import', methods=['POST'])
@dealer_edit_required
def reconcile_import_from_group_invoices():
    data = request.json or {}
    invoice_ids_in = data.get('invoice_ids') or []
    invoice_ids = [int(x) for x in invoice_ids_in if str(x).strip().lstrip('-').isdigit()]
    if not invoice_ids:
        return jsonify({'error': '請至少勾選一筆發票'}), 400

    user_id = session.get('user_id')
    user_name = session.get('display_name', session.get('username', ''))
    results = []

    for iid in invoice_ids:
        inv_res = sb.table('bookstore_group_invoices').select('*').eq('id', iid).limit(1).execute()
        if not inv_res.data:
            results.append({'invoice_id': iid, 'ok': False, 'error': '找不到這筆發票'})
            continue
        inv = inv_res.data[0]
        if inv.get('reconciled_at'):
            results.append({'invoice_id': iid, 'ok': False, 'error': '此筆發票已匯入對帳過，不可重複匯入'})
            continue
        dealer_name = _norm(inv.get('dealer_name'))
        if not dealer_name:
            results.append({'invoice_id': iid, 'ok': False,
                             'error': '尚未設定「經銷商類型」，請先到「匯入團體發票」編輯這筆發票'})
            continue

        pdf_rows = sb.table('bookstore_group_invoice_pdfs').select('*').eq('invoice_id', iid).execute().data or []
        doc_rows = sb.table('bookstore_group_invoice_files').select('*').eq('invoice_id', iid).execute().data or []
        if not pdf_rows or not doc_rows:
            results.append({'invoice_id': iid, 'ok': False, 'error': '發票PDF或出貨單其中一邊尚未收集齊全，無法匯入'})
            continue

        period = _norm(inv.get('recon_period')) or None
        if not period and inv.get('invoice_date'):
            period = str(inv['invoice_date'])[:7].replace('-', '/')
        upload_group = f'giv{iid}'

        batch_ids = []
        errors = []
        # 2026-08-17新增：累計這張發票（PDF批次＋出貨單批次合計）的自動比對成功/未比對列數，
        # 讓「確認匯入」按下後能立刻在畫面上看到比對品質，不用特地跑去「批次列表」才看得到。
        matched_total, unmatched_total = 0, 0
        pdf_ok = 0
        for pr in pdf_rows:
            try:
                raw = sb.storage(_INVOICE_PDF_BUCKET).download(pr['storage_path'])
            except Exception as e:
                errors.append(f'下載發票PDF「{pr["file_name"]}」失敗：{e}')
                continue
            result, err = _ingest_reconcile_file(
                raw, pr['file_name'], dealer_name, period, upload_group,
                user_id, user_name, force_source_kind='invoice_amount_only',
            )
            if err:
                errors.append(f'發票「{pr["file_name"]}」：{err[0].get("error")}')
            else:
                batch_ids.append(result['batch_id'])
                matched_total += result.get('matched_lines', 0)
                unmatched_total += result.get('unmatched_lines', 0)
                pdf_ok += 1

        doc_ok = 0
        for dr in doc_rows:
            try:
                raw = sb.storage(_SHIPPING_DOC_BUCKET).download(dr['storage_path'])
            except Exception as e:
                errors.append(f'下載出貨單「{dr["file_name"]}」失敗：{e}')
                continue
            result, err = _ingest_reconcile_file(
                raw, dr['file_name'], dealer_name, period, upload_group,
                user_id, user_name, force_source_kind='shipping_qty_only',
            )
            if err:
                errors.append(f'出貨單「{dr["file_name"]}」：{err[0].get("error")}')
            else:
                batch_ids.append(result['batch_id'])
                matched_total += result.get('matched_lines', 0)
                unmatched_total += result.get('unmatched_lines', 0)
                doc_ok += 1

        if pdf_ok > 0 and doc_ok > 0:
            sb.table('bookstore_group_invoices').update({
                'reconciled_at': now_str(), 'reconciled_by': user_id,
                'reconciled_by_name': user_name, 'reconciled_upload_group': upload_group,
            }).eq('id', iid).execute()
            results.append({
                'invoice_id': iid, 'ok': True, 'batch_ids': batch_ids,
                'upload_group': upload_group,
                'matched_lines': matched_total, 'unmatched_lines': unmatched_total,
                'warning': '；'.join(errors) if errors else None,
            })
        else:
            results.append({
                'invoice_id': iid, 'ok': False,
                'error': '；'.join(errors) if errors else '匯入失敗，發票PDF或出貨單皆未成功建立比對批次',
                'batch_ids': batch_ids,
            })

    return jsonify({'results': results})


@dealer_bp.route('/api/dealer/batches', methods=['GET'])
@dealer_view_required
def list_batches():
    res = sb.table('dealer_reconcile_batches').select('*').order('id', desc=True).limit(200).execute()
    return jsonify(res.data)


# ── 刪除批次（僅限尚未入庫的批次；已入庫的批次因為彙總表已引用，直接擋下避免破壞對帳資料）──
@dealer_bp.route('/api/dealer/batches/<int:bid>', methods=['DELETE'])
@dealer_edit_required
def delete_batch(bid):
    batch_res = sb.table('dealer_reconcile_batches').select('*').eq('id', bid).execute()
    if not batch_res.data:
        return jsonify({'error': '找不到批次'}), 404
    batch = batch_res.data[0]
    if batch.get('status') == '已入庫':
        return jsonify({'error': '此批次已入庫，無法直接刪除（會影響已寫入的對帳彙總資料），如需修正請聯絡開發人員手動處理'}), 400
    sb.table('dealer_reconcile_lines').delete().eq('batch_id', bid).execute()
    sb.table('dealer_reconcile_batches').delete().eq('id', bid).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/batches/<int:bid>/lines', methods=['GET'])
@dealer_view_required
def list_batch_lines(bid):
    status = request.args.get('status', '').strip()
    q = sb.table('dealer_reconcile_lines').select('*').eq('batch_id', bid)
    if status:
        q = q.eq('match_status', status)
    res = q.order('id').execute()
    return jsonify(res.data)


@dealer_bp.route('/api/dealer/lines/<int:lid>', methods=['PUT'])
@dealer_edit_required
def confirm_line(lid):
    data = request.json or {}
    main_sku = _norm(data.get('main_sku'))
    if not main_sku:
        return jsonify({'error': '請選擇主料號'}), 400
    gift_in = [data.get('gift_sku1') or None, data.get('gift_sku2') or None, data.get('gift_sku3') or None]
    err = _check_sku_refs(main_sku, gift_in, _valid_sku_codes())
    if err:
        return jsonify({'error': err}), 400

    line_res = sb.table('dealer_reconcile_lines').select('*').eq('id', lid).execute()
    if not line_res.data:
        return jsonify({'error': '找不到該筆明細'}), 404
    ln = line_res.data[0]

    # 已入庫的批次，明細對應改了也不會回頭更新 dealer_reconcile_summary（彙總是入庫當下累加寫入，
    # 不是即時計算），若讓使用者在已入庫批次修改比對結果會造成「畫面看起來改了，但對帳查詢數字沒變」
    # 的落差，故直接擋下，比照 commit_batch() 擋重複入庫、delete_batch() 擋刪除已入庫批次的原則
    batch_res = sb.table('dealer_reconcile_batches').select('*').eq('id', ln['batch_id']).execute()
    batch = batch_res.data[0] if batch_res.data else None
    is_correction = ln.get('match_status') in ('auto_matched', 'manual_confirmed')
    if batch and batch.get('status') == '已入庫' and is_correction:
        return jsonify({'error': '此批次已入庫，修改比對對應不會更新已寫入的對帳彙總資料。如需修正，請聯絡開發人員手動處理，或刪除該經銷商+期間的彙總資料後重新上傳入庫。'}), 400

    rec = {
        'matched_main_sku': main_sku,
        'matched_gift_sku1': data.get('gift_sku1') or None,
        'matched_gift_sku2': data.get('gift_sku2') or None,
        'matched_gift_sku3': data.get('gift_sku3') or None,
        'match_status': 'manual_confirmed',
        'confirmed_by': session.get('user_id'),
        'confirmed_by_name': session.get('display_name', session.get('username', '')),
        'confirmed_at': now_str(),
    }

    # 2026-08-19新增：只有[[v3.66]]/[[v3.67]]自動產生的「整體調整」「數量調整」列
    # （dealer_code是固定的特殊標記，不是真實訂單編號）才允許使用者在確認前修改數量/
    # 金額——這2種列本來就是系統依上傳檔案內建報表估算出來的參考值，讓使用者確認前能
    # 依實際狀況微調合理；一般訂單明細列的數量/金額來自原始上傳檔案，不接受從這個
    # 入口修改，避免跟來源資料兜不起來、比對軌跡失真（前端也只有這2種列會渲染出可
    # 編輯的輸入框，這裡後端再次檢查，不只信任前端）。
    is_adjustment_row = ln.get('dealer_code') in ('(整體調整)', '(數量調整)')
    if is_adjustment_row and ('amount_untaxed' in data or 'amount_taxed' in data or 'qty' in data):
        new_qty = _to_num_or_none(data.get('qty'))
        new_untaxed = _to_num_or_none(data.get('amount_untaxed'))
        new_taxed = _to_num_or_none(data.get('amount_taxed'))
        if new_qty is None or new_untaxed is None or new_taxed is None:
            return jsonify({'error': '調整列的數量/金額必須是數字'}), 400
        rec['qty'] = round(new_qty, 4)
        rec['amount_untaxed'] = round(new_untaxed, 2)
        rec['amount_taxed'] = round(new_taxed, 2)
        rec['amount'] = round(new_taxed, 2)
        rec['tax_amount'] = round(new_taxed - new_untaxed, 2)

    sb.table('dealer_reconcile_lines').update(rec).eq('id', lid).execute()

    dealer_name = batch['dealer_name'] if batch else None

    if data.get('save_mapping') and dealer_name and not is_adjustment_row:
        dealer_code_v = ln.get('dealer_code')
        map_rec = {
            'dealer_name': dealer_name, 'dealer_code': dealer_code_v,
            'ean': ln.get('ean'), 'product_name': ln.get('product_name'),
            'main_sku': main_sku, 'gift_sku1': rec['matched_gift_sku1'],
            'gift_sku2': rec['matched_gift_sku2'], 'gift_sku3': rec['matched_gift_sku3'],
        }
        # 避免重複插入：同一經銷商代碼若每個月對帳都要手動重新確認（例如已被標記歧義、
        # 無法自動比對），且每次都選同一個正確料號，舊版邏輯每次都會新增一筆完全相同的
        # 對應紀錄，長期累積出大量重複列，且因為多筆重複「同料號」與最初那筆錯誤料號同時
        # 存在，歧義狀態會一直卡住解不開。這裡改為：新增前先檢查是否已有完全相同的對應
        # （代碼+EAN+商品名稱+主料號+贈品皆同），若有就不重複新增
        existing_q = sb.table('dealer_sku_mapping').select('*').eq('dealer_name', dealer_name)
        if dealer_code_v:
            existing_q = existing_q.eq('dealer_code', dealer_code_v)
        existing_rows = existing_q.execute().data or []
        is_dup = any(
            _norm(r.get('ean')) == _norm(map_rec['ean']) and
            _norm(r.get('product_name')) == _norm(map_rec['product_name']) and
            _norm(r.get('main_sku')) == _norm(map_rec['main_sku']) and
            _norm(r.get('gift_sku1')) == _norm(map_rec['gift_sku1']) and
            _norm(r.get('gift_sku2')) == _norm(map_rec['gift_sku2']) and
            _norm(r.get('gift_sku3')) == _norm(map_rec['gift_sku3'])
            for r in existing_rows
        )
        if not is_dup:
            map_rec.update(_audit_new())
            sb.table('dealer_sku_mapping').insert(map_rec).execute()
        _recompute_ambiguous(dealer_name, dealer_code_v)

    # 重新計算批次統計
    all_lines_res = sb.table('dealer_reconcile_lines').select('match_status').eq('batch_id', ln['batch_id']).execute()
    all_lines = all_lines_res.data or []
    sb.table('dealer_reconcile_batches').update({
        'matched_lines': sum(1 for l in all_lines if l['match_status'] == 'auto_matched'),
        'manual_lines': sum(1 for l in all_lines if l['match_status'] == 'manual_confirmed'),
        'unmatched_lines': sum(1 for l in all_lines if l['match_status'] == 'unmatched'),
    }).eq('id', ln['batch_id']).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/batches/<int:bid>/commit', methods=['POST'])
@dealer_edit_required
def commit_batch(bid):
    data = request.json or {}
    period = _norm(data.get('period'))
    batch_res = sb.table('dealer_reconcile_batches').select('*').eq('id', bid).execute()
    if not batch_res.data:
        return jsonify({'error': '找不到批次'}), 404
    batch = batch_res.data[0]
    period = period or batch.get('period')
    if not period:
        return jsonify({'error': '請先填寫期間（如 2026/05）才能入庫'}), 400

    # 防止重複入庫：已入庫的批次若再次送出「確認並入庫」，會把同一批明細的數量/金額再疊加一次到
    # dealer_reconcile_summary，造成對帳彙總數字被重複計算。已入庫批次一律擋下，不提供強制覆蓋選項
    # （若確實需要重跑，請先請系統管理員於 Supabase 手動核對/修正 dealer_reconcile_summary 再處理）
    if batch.get('status') == '已入庫':
        return jsonify({
            'error': f'此批次已於 {batch.get("committed_at") or ""} 入庫過，重複入庫會讓對帳彙總的數量與金額被再次疊加，已擋下此次操作。如確實需要重新處理，請聯絡系統管理員先於 Supabase 核對/修正 dealer_reconcile_summary',
            'blocked': True,
        }), 400

    lines_res = sb.table('dealer_reconcile_lines').select('*').eq('batch_id', bid).execute()
    lines = lines_res.data or []
    unmatched = [l for l in lines if l['match_status'] == 'unmatched']
    if unmatched and not data.get('force'):
        return jsonify({'error': f'尚有 {len(unmatched)} 筆未比對成功，請先確認或加上 force 強制入庫（未比對筆數將被略過）'}), 400

    # source_kind（見 upload_batch() 說明）決定這批資料的數量/金額要不要計入彙總表：
    # 香港經銷商的發票只計金額（數量另由出貨單提供，避免贈品重複/漏計），出貨單只計數量
    source_kind = batch.get('source_kind') or 'normal'
    count_qty = source_kind != 'invoice_amount_only'
    count_amount = source_kind != 'shipping_qty_only'

    agg = {}  # internal_code -> {qty, amount, amount_taxed, amount_untaxed, tax_amount}
    for l in lines:
        if l['match_status'] == 'unmatched':
            continue
        qty = (l['qty'] or 0) if count_qty else 0
        amount = (l['amount'] or 0) if count_amount else 0
        amount_untaxed = (l.get('amount_untaxed') or 0) if count_amount else 0
        amount_taxed = (l.get('amount_taxed') or 0) if count_amount else 0
        tax_amount = (l.get('tax_amount') or 0) if count_amount else 0
        main = l['matched_main_sku']
        if main:
            a = agg.setdefault(main, {'qty': 0, 'amount': 0, 'amount_untaxed': 0, 'amount_taxed': 0, 'tax_amount': 0})
            a['qty'] += qty
            a['amount'] += amount
            a['amount_untaxed'] += amount_untaxed
            a['amount_taxed'] += amount_taxed
            a['tax_amount'] += tax_amount
        for gift in (l.get('matched_gift_sku1'), l.get('matched_gift_sku2'), l.get('matched_gift_sku3')):
            if gift:
                a = agg.setdefault(gift, {'qty': 0, 'amount': 0, 'amount_untaxed': 0, 'amount_taxed': 0, 'tax_amount': 0})
                a['qty'] += qty  # 贈品數量比照主料號 1:1 認列，金額與稅額歸零
                a['amount'] += 0
                a['amount_untaxed'] += 0
                a['amount_taxed'] += 0
                a['tax_amount'] += 0

    # ⚠️ 2026-07-24重要變更：不再依「經銷商+期間」跨批次累加彙總。
    # 舊邏輯是先查有沒有同(經銷商,期間,內部料號)的既有列，有就累加、沒有才新增——這在「同一個月
    # 剛好有2次以上出貨/開票」時（如香港經銷商），會把2次不同來源的數量/金額疊在同一列，之後完全
    # 無法拆開查看是哪一次的資料。使用者2026-07-24反饋這是問題，因為同一期間常有多次各經對帳檔。
    # 新邏輯：每個批次的彙總結果一律各自INSERT新的一批列（帶自己的batch_id），不查詢/累加其他
    # 批次的既有列。同一經銷商+期間若有多筆批次，會各自在 dealer_reconcile_summary 產生自己的
    # 列，查詢時（_build_summary_data）加總金額/數量的畫面（依內部料號、依財報品名彙總）依然會
    # 正確地把同期間所有批次的數字加總起來，但「依內部料號名矩陣」這種列表式檢視則改成每個批次
    # 各自一列，方便追溯是哪一次上傳的資料，備註欄位也隨之改成掛在批次本身（見下方
    # dealer_reconcile_batches 的 customer_type 等新欄位），不再掛在「經銷商+期間」上。
    dealer_name = batch['dealer_name']
    summary_batch = []
    for code, v in agg.items():
        summary_batch.append({
            'dealer_name': dealer_name, 'period': period, 'internal_code': code,
            'qty': v['qty'], 'amount': v['amount'],
            'amount_untaxed': v['amount_untaxed'], 'amount_taxed': v['amount_taxed'],
            'tax_amount': v['tax_amount'], 'batch_id': bid,
        })
    if summary_batch:
        sb.table('dealer_reconcile_summary').insert(summary_batch).execute()

    sb.table('dealer_reconcile_batches').update({
        'status': '已入庫', 'period': period, 'committed_at': now_str(),
    }).eq('id', bid).execute()
    return jsonify({'ok': True, 'codes_written': len(agg)})


@dealer_bp.route('/api/dealer/batches/<int:bid>/export', methods=['GET'])
@dealer_view_required
def export_batch(bid):
    batch_res = sb.table('dealer_reconcile_batches').select('*').eq('id', bid).execute()
    if not batch_res.data:
        return jsonify({'error': '找不到批次'}), 404
    batch = batch_res.data[0]
    lines_res = sb.table('dealer_reconcile_lines').select('*').eq('batch_id', bid).order('id').execute()
    lines = lines_res.data or []
    sku_list = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('*').order('sort_order'))

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = '原始明細'
    ws1.append(['類型', '經銷商代碼', 'EAN', '商品名稱', '數量', '單價', '金額', '未稅金額', '含稅金額', '稅額', '比對狀態',
                '主料號', '贈品料號1', '贈品料號2', '贈品料號3'])
    for l in lines:
        ws1.append([l.get('row_type'), l.get('dealer_code'), l.get('ean'), l.get('product_name'),
                    l.get('qty'), l.get('unit_price'), l.get('amount'), l.get('amount_untaxed'), l.get('amount_taxed'),
                    l.get('tax_amount'), l.get('match_status'),
                    l.get('matched_main_sku'), l.get('matched_gift_sku1'), l.get('matched_gift_sku2'),
                    l.get('matched_gift_sku3')])

    agg = {}
    for l in lines:
        for code, is_gift in ((l.get('matched_main_sku'), False),
                               (l.get('matched_gift_sku1'), True),
                               (l.get('matched_gift_sku2'), True),
                               (l.get('matched_gift_sku3'), True)):
            if not code:
                continue
            a = agg.setdefault(code, {'qty': 0, 'amount': 0, 'amount_untaxed': 0, 'amount_taxed': 0, 'tax_amount': 0})
            a['qty'] += l.get('qty') or 0
            if not is_gift:
                a['amount'] += l.get('amount') or 0
                a['amount_untaxed'] += l.get('amount_untaxed') or 0
                a['amount_taxed'] += l.get('amount_taxed') or 0
                a['tax_amount'] += l.get('tax_amount') or 0

    ws2 = wb.create_sheet('依內部料號呈現')
    ws2.append(['內部料號', 'ERP順序', '數量', '金額', '未稅金額', '含稅金額', '稅額'])
    total_qty = total_amt = total_untaxed = total_taxed = total_tax = 0
    yellow = openpyxl.styles.PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    for sku in sku_list:
        code = sku['internal_code']
        v = agg.get(code, {'qty': 0, 'amount': 0, 'amount_untaxed': 0, 'amount_taxed': 0, 'tax_amount': 0})
        row_idx = ws2.max_row + 1
        ws2.append([code, sku.get('erp_code'), v['qty'], v['amount'], v['amount_untaxed'], v['amount_taxed'], v['tax_amount']])
        if v['qty'] or v['amount']:
            for c in range(1, 8):
                ws2.cell(row=row_idx, column=c).fill = yellow
        total_qty += v['qty']
        total_amt += v['amount']
        total_untaxed += v['amount_untaxed']
        total_taxed += v['amount_taxed']
        total_tax += v['tax_amount']
    ws2.append(['合計', '', total_qty, total_amt, total_untaxed, total_taxed, total_tax])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{batch['dealer_name']}_{batch.get('period') or ''}_拆解驗證.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# 2026-08-21新增、同日修正：「總表分析」（依銷售單位）——「上傳比對」批次列表勾選多筆「已入庫」
# 批次後，額外匯出一份格式比照「書店訂單拆帳」報表既有「總表分析」頁籤（見_compute_bookstore_
# order_pivot()/export_bookstore_orders()的ws4）的樞紐分析表。
#
# 分組維度：原本第一版用batch的dealer_name分組，使用者確認要改用「銷售單位」（bookstore_
# group_invoices.seller_name）全稱——批次若是透過「從匯入團體發票勾選匯入」產生，
# dealer_reconcile_batches.upload_group會等於該筆發票的reconciled_upload_group，可以反查
# 回那筆發票拿到seller_name/order_no/invoice_date/invoice_no/amount_incl/shipping_order_no
# （見下方_attach_invoice_info()）；找不到對應發票的批次（例如手動上傳、非團體發票匯入的
# 批次）才退回用dealer_name顯示，避免完全沒有分組標籤可看。
#
# 跟書店版本另外2個刻意不同的地方：
# 1. 沒有「境內未稅/境外含稅」那種依區塊切換稅基的規則——一般經銷商的dealer_reconcile_lines
#    每一列本身就有真實的amount_untaxed（來源發票/出貨單的未稅金額），不需要像書店訂單拆帳
#    那樣「只知道含稅總額、用稅率反推」，這裡統一只呈現未稅單價/未稅金額，避免無意義地套用
#    ×1.05換算。若之後有經銷商需要看含稅金額，可再擴充。
# 2. 沒有書店專屬「代銷商品/電子書方案」集合類品名特例——那是書店特定方案架構才有的概念，
#    一般經銷商的財報料號品名一律當成一般料號、依單價分組呈現。
#
# 2026-08-21再修正：使用者比對書店版「總表分析」後回報2點：
# a) 料號品名欄位比照書店版，固定列出「內部料號清單」裡全部啟用中(is_active!=false)的財報
#    料號品名，不再只列出「這次勾選批次剛好比對到」的品名（否則不同批次匯出時欄位數量/順序
#    會不一致，也看不出這批經銷商完全沒出貨某個品名跟「這品名根本沒設定」的差異）。
# b) 移除最下方「金額合計(未稅)」那幾列——使用者確認不需要，「料號加總(所選銷售單位合計)」
#    這一列留著即可。
# ============================================================
def _attach_invoice_info(batches):
    """依batches(dealer_reconcile_batches列)裡的upload_group，反查bookstore_group_invoices.
    reconciled_upload_group拿到對應那筆團體發票的銷售單位/OA訂單編號/發票日期/發票號碼/
    價格(含稅)/出貨訂單編號，寫回每筆batch dict（key前綴inv_，找不到對應發票則維持None）。
    找不到對應發票很常見（手動上傳/非團體發票匯入的批次本來就沒有upload_group或對應不到），
    不算錯誤。"""
    groups = {b.get('upload_group') for b in batches if b.get('upload_group')}
    invoice_by_group = {}
    if groups:
        inv_res = sb.table('bookstore_group_invoices') \
            .select('seller_name,order_no,invoice_date,invoice_no,amount_incl,shipping_order_no,'
                    'reconciled_upload_group') \
            .in_('reconciled_upload_group', list(groups)).execute()
        for inv in (inv_res.data or []):
            g = inv.get('reconciled_upload_group')
            if g and g not in invoice_by_group:
                invoice_by_group[g] = inv
    for b in batches:
        inv = invoice_by_group.get(b.get('upload_group'))
        b['inv_seller_name'] = inv.get('seller_name') if inv else None
        b['inv_order_no'] = inv.get('order_no') if inv else None
        b['inv_invoice_date'] = inv.get('invoice_date') if inv else None
        b['inv_invoice_no'] = inv.get('invoice_no') if inv else None
        b['inv_amount_incl'] = inv.get('amount_incl') if inv else None
        b['inv_shipping_order_no'] = inv.get('shipping_order_no') if inv else None
    return batches


def _compute_dealer_batch_summary_pivot(batch_ids):
    """依所選批次(batch_ids，須已入庫)，算出「總表分析」（依銷售單位分組）用的樞紐資料。
    回傳 None 代表所選批次裡沒有任何一筆是「已入庫」狀態，無法產生報表。"""
    if not batch_ids:
        return None
    # 2026-08-21新增：防禦性去重——前端理論上用Set()傳id不會重複，這裡多一層保護，避免
    # 萬一query string帶重複id（或Supabase回傳異常重複列），造成「選取批次清單」頁籤同一筆
    # 批次重複列出2次。
    batch_ids = list(dict.fromkeys(batch_ids))
    batches_res = sb.table('dealer_reconcile_batches') \
        .select('id,dealer_name,period,status,source_filename,upload_group').in_('id', batch_ids).execute()
    seen_ids = set()
    batches = []
    for b in (batches_res.data or []):
        if b.get('status') != '已入庫' or b['id'] in seen_ids:
            continue
        seen_ids.add(b['id'])
        batches.append(b)
    if not batches:
        return None
    batches.sort(key=lambda b: b['id'])
    _attach_invoice_info(batches)

    # 分組標籤：有對應到團體發票就用該發票的「銷售單位」全稱，沒有則退回顯示批次的經銷商名稱。
    batch_group = {}
    group_names = []
    for b in batches:
        g = b.get('inv_seller_name') or b.get('dealer_name') or '（未命名）'
        batch_group[b['id']] = g
        if g not in group_names:
            group_names.append(g)
    committed_ids = list(batch_group.keys())

    lines = _fetch_all_rows(lambda: sb.table('dealer_reconcile_lines')
                             .select('batch_id,matched_main_sku,qty,amount_untaxed,match_status')
                             .in_('batch_id', committed_ids))

    # 2026-08-21修正：欄位改成固定使用「內部料號清單」裡全部啟用中(is_active!=false)的財報
    # 料號品名，比照_compute_bookstore_order_matrix()既有規則（見該函式註解）——同一品名底下
    # 只要還有一筆內部料號顯示中就算顯示，全部被設為隱藏才排除；即使某品名目前被隱藏、不出現
    # 在固定欄位清單裡，比對到該品名的金額仍會照算進grouped，只是不會顯示成獨立的一欄。
    sku_rows = _fetch_all_rows(lambda: sb.table('internal_sku_list')
                                .select('internal_code,report_product_name,purchase_price_notax,'
                                        'sort_order,is_active').order('sort_order'))
    name_by_code = {}
    cost_by_name = {}
    active_by_name = {}
    inconsistent_names = set()
    for s in sku_rows:
        code = s.get('internal_code')
        if not code:
            continue
        name = s.get('report_product_name') or code
        name_by_code[code] = name
        cost = _to_num_or_none(s.get('purchase_price_notax'))
        if name not in cost_by_name:
            cost_by_name[name] = cost
        elif cost is not None and cost_by_name[name] is not None and abs(cost - cost_by_name[name]) > 0.01:
            inconsistent_names.add(name)
        if s.get('is_active') is not False:
            active_by_name[name] = True
        elif name not in active_by_name:
            active_by_name[name] = False

    columns = []
    seen_cols = set()
    for s in sku_rows:
        name = s.get('report_product_name') or s.get('internal_code')
        if name and name not in seen_cols and active_by_name.get(name):
            seen_cols.add(name)
            columns.append(name)

    # 2026-08-21修正：使用者實測後回報「同一銷售單位在總表分析裡出現好幾個一模一樣標籤的
    # 區塊，像是重複」——原因是這裡原本比照書店版「依拆帳單價分組，同單價的訂單合併成一列」，
    # 但一般經銷商的明細不像書店訂單有固定的方案定價，同一料號常常只因為四捨五入或每次出貨
    # 議價些微差異，就被拆成好幾個單價幾乎相同的小分組，畫面上看起來像同一個銷售單位重複
    # 出現好幾次。改成不分單價分組——同一(銷售單位, 財報料號品名)底下所有明細直接加總數量與
    # 未稅金額，算出一個「加權平均單價」(=未稅金額加總/數量加總)，每個銷售單位只會出現一次
    # （1個單價列+1個數量列），不會再因為單價略有差異而重複出現同一個銷售單位標籤。
    agg = {}  # (銷售單位, report_product_name) -> {'qty':.., 'amount':..}
    for l in lines:
        if l.get('match_status') == 'unmatched':
            continue
        code = l.get('matched_main_sku')
        if not code:
            continue
        qty = l.get('qty') or 0
        if not qty:
            continue
        group_name = batch_group.get(l.get('batch_id'))
        if not group_name:
            continue
        name = name_by_code.get(code) or code
        a = agg.setdefault((group_name, name), {'qty': 0, 'amount': 0})
        a['qty'] += qty
        a['amount'] += (l.get('amount_untaxed') or 0)

    cell_by_group = {name: {g: None for g in group_names} for name in columns}
    for name in columns:
        cost = cost_by_name.get(name)
        for group_name in group_names:
            a = agg.get((group_name, name))
            if a and a['qty']:
                price = round(a['amount'] / a['qty'], 2)
                cell_by_group[name][group_name] = {
                    'price': price, 'qty': a['qty'],
                    'below_cost': (cost is not None and price < cost),
                }

    col_combined_total = {}
    for name in columns:
        col_combined_total[name] = sum(
            cell_by_group[name][g]['qty'] for g in group_names if cell_by_group[name][g])

    return {
        'group_names': group_names,
        'columns': columns,
        'cost_by_name': cost_by_name,
        'inconsistent_cost_names': sorted(inconsistent_names),
        'cell_by_group': cell_by_group,
        'col_combined_total': col_combined_total,
        'batches': batches,
    }


@dealer_bp.route('/api/dealer/batches/summary-analysis-export', methods=['GET'])
@dealer_view_required
def export_dealer_batch_summary_analysis():
    """2026-08-21新增：見上方_compute_dealer_batch_summary_pivot()說明。前端「批次列表」勾選
    已入庫批次後，用window.open帶?ids=1,2,3觸發下載（比照exportBatch()/exportQuery()既有的
    GET+window.open下載方式，不用另外寫POST+blob）。"""
    ids_raw = request.args.get('ids', '')
    batch_ids = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]
    pivot = _compute_dealer_batch_summary_pivot(batch_ids)
    if pivot is None:
        return jsonify({'error': '所選批次裡沒有任何一筆是「已入庫」狀態，無法產生總表分析（請勾選狀態為已入庫的批次）'}), 400

    group_names = pivot['group_names']
    columns = pivot['columns']
    if not columns:
        return jsonify({'error': '「內部料號清單」目前沒有任何啟用中的財報料號品名，無法產生總表分析'}), 400

    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = '選取批次清單'
    # 2026-08-21新增：額外從對應的「匯入團體發票」帶出銷售單位/OA訂單編號/發票日期/發票號碼/
    # 價格(含稅)/出貨訂單編號（見_attach_invoice_info()）；批次若不是從團體發票匯入產生
    # （例如手動上傳），這幾欄留空。
    ws0.append(['批次id', '經銷商', '期間', '來源檔名', '銷售單位', 'OA訂單編號', '發票日期',
                '發票號碼', '價格(含稅)', '出貨訂單編號'])
    for b in pivot['batches']:
        ws0.append([b['id'], b.get('dealer_name'), b.get('period'), b.get('source_filename'),
                    b.get('inv_seller_name'), b.get('inv_order_no'), b.get('inv_invoice_date'),
                    b.get('inv_invoice_no'), b.get('inv_amount_incl'), b.get('inv_shipping_order_no')])
    for ci, w in enumerate([10, 20, 12, 30, 20, 16, 12, 16, 12, 16], 1):
        ws0.column_dimensions[ws0.cell(1, ci).column_letter].width = w

    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    redfill = openpyxl.styles.PatternFill('solid', fgColor='FFC7CE')
    redfont = openpyxl.styles.Font(color='9C0006')
    totals_fill = openpyxl.styles.PatternFill('solid', fgColor='D9E1F2')
    totals_font = openpyxl.styles.Font(bold=True)
    wrap = openpyxl.styles.Alignment(wrap_text=True, vertical='center', horizontal='center')

    ws4 = wb.create_sheet('總表分析')
    BASE = 2  # A:銷售單位, B:列類型
    data_start = BASE + 1
    total_col2 = data_start + len(columns)

    ws4.cell(1, 1, '總表分析（同一銷售單位+財報料號品名的所有明細合併加總，單價為加權平均'
                    '未稅單價；分組依所選批次對應發票的「銷售單位」，找不到對應發票的批次'
                    '退回顯示經銷商名稱；同一銷售單位的多筆批次會合併呈現成一列，不再依單價'
                    '細分）')

    ws4.cell(2, 2, '未稅進貨價')
    for ci, name in enumerate(columns, data_start):
        cost = pivot['cost_by_name'].get(name)
        if cost is not None:
            _set_int_cell(ws4, 2, ci, cost)

    ws4.cell(3, 1, '銷售單位').fill = hfill
    ws4.cell(3, 1).font = hfont
    ws4.cell(3, 1).alignment = wrap
    ws4.cell(3, 2, '項目').fill = hfill
    ws4.cell(3, 2).font = hfont
    ws4.cell(3, 2).alignment = wrap
    for ci, name in enumerate(columns, data_start):
        c = ws4.cell(3, ci, name)
        c.fill = hfill
        c.font = hfont
        c.alignment = wrap
    c = ws4.cell(3, total_col2, '小計(未稅)')
    c.fill = hfill
    c.font = hfont
    c.alignment = wrap

    # 2026-08-21修正：每個銷售單位固定只佔1個「單價(未稅，加權平均)」+1個「數量」row，不再
    # 依單價分組成多個區塊（理由見_compute_dealer_batch_summary_pivot()裡agg/cell_by_group
    # 的說明），避免同一個銷售單位標籤重複出現好幾次看起來像資料重複。
    col_letter_first = ws4.cell(1, data_start).column_letter
    col_letter_last = ws4.cell(1, total_col2 - 1).column_letter
    r = 4
    for group_name in group_names:
        r_price = r
        ws4.cell(r, 1, group_name)
        ws4.cell(r, 2, '單價(未稅，加權平均)')
        for ci, name in enumerate(columns, data_start):
            cell = pivot['cell_by_group'][name][group_name]
            if cell:
                c = _set_int_cell(ws4, r, ci, cell['price'])
                if cell.get('below_cost'):
                    c.fill = redfill
                    c.font = redfont
        r += 1
        r_qty = r
        ws4.cell(r, 1, group_name)
        ws4.cell(r, 2, '數量')
        for ci, name in enumerate(columns, data_start):
            cell = pivot['cell_by_group'][name][group_name]
            if cell:
                c = _set_int_cell(ws4, r, ci, cell['qty'])
                if cell.get('below_cost'):
                    c.fill = redfill
                    c.font = redfont
        r += 1
        ws4.cell(r_price, total_col2,
                 f'=SUMPRODUCT({col_letter_first}{r_price}:{col_letter_last}{r_price},'
                 f'{col_letter_first}{r_qty}:{col_letter_last}{r_qty})')

    ws4.cell(r, 2, '料號加總(所選銷售單位合計)')
    for ci, name in enumerate(columns, data_start):
        _set_int_cell(ws4, r, ci, pivot['col_combined_total'][name])
    for ci in range(1, total_col2 + 1):
        cc = ws4.cell(r, ci)
        cc.font = totals_font
        cc.fill = totals_fill
    r += 1

    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 18
    for ci in range(data_start, total_col2 + 1):
        ws4.column_dimensions[ws4.cell(3, ci).column_letter].width = 14
    ws4.freeze_panes = ws4.cell(4, data_start).coordinate

    if pivot['inconsistent_cost_names']:
        ws5 = wb.create_sheet('進貨價不一致警示')
        ws5.append(['財報料號品名（底下內部料號進貨未稅價不一致，標紅結果可能不準確，請核對）'])
        for name in pivot['inconsistent_cost_names']:
            ws5.append([name])

    for _ws in wb.worksheets:
        _apply_vertical_top(_ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if len(group_names) == 1:
        fname = f"{group_names[0]}_總表分析.xlsx"
    else:
        fname = f"總表分析_{len(group_names)}個銷售單位.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# ⑦ 對帳查詢
# ============================================================
@dealer_bp.route('/api/dealer/batches/<int:bid>/notes', methods=['PUT'])
@dealer_edit_required
def save_batch_notes(bid):
    """2026-07-24新增：客戶類型/對帳順序/對帳編號/書店對帳編號/對帳區間/對帳備註改為掛在
    「批次」本身（dealer_reconcile_batches 的新增欄位），取代舊版掛在「經銷商+期間」上的做法
    （見下方 get_period_note/save_period_note，現已不再被前端呼叫，保留是為了不動舊資料表，
    不影響現有資料）。改成掛在批次上的原因：同一個經銷商+期間常常有2次以上出貨/開票
    （如香港經銷商），若備註掛在期間上，2次的備註文字會被迫擠在同一欄，混在一起看不清楚；
    掛在批次上，每次上傳都是獨立一筆，各自的備註自然分開，不會互相覆蓋或混淆。
    只更新這次請求「實際有帶」的欄位，避免只改其中一欄時把其他欄位覆蓋成空白。

    2026-07-24補充：對帳查詢矩陣改依 upload_group 合併顯示（見 _group_batches_by_upload），
    畫面上只會讀取「群組代表批次」(batch_id最小者)的備註欄位。若使用者剛好是在「上傳比對」頁
    對同一群組裡「非代表」的那一筆批次（例如發票，若出貨單先建立、id較小）編輯備註，為避免
    存進去卻不會顯示在矩陣畫面上造成困惑，這裡改為：連同該批次的 upload_group 同組其他批次
    一併更新同樣的值，同一組不論編輯哪一筆結果都一致。

    2026-08-07新增：manual_amount_untaxed／manual_amount_taxed（人工填入未稅／含稅金額小計）—
    給像「單位團體(知識)」這種只有出貨單、沒有發票可提供金額的經銷商，出貨單解析出來的金額
    永遠是0（見 _parse_shipping_order()），讓使用者可以直接在「對帳查詢→依內部料號名矩陣」
    畫面補填一個總金額。設計為「覆蓋」而非「加總」：填了數字，畫面小計就直接顯示這個數字
    （取代系統依明細算出來的金額）；清空/留空，則還原顯示系統試算的金額（見
    _build_summary_data() 的覆蓋邏輯）。⚠️ 僅「依內部料號名矩陣」這個檢視會套用，
    query_summary() 的 view=sku/report（依內部料號／依財報品名彙總）是跨批次彙總、沒有單一
    對應的批次可以掛人工金額，不會反映這裡填的值，會固定顯示系統試算金額。
    這兩個欄位是數字，跟其餘文字欄位分開處理：空字串/None 存為 None（代表「未填，維持系統
    試算金額」），其餘值轉成 float，轉換失敗回傳400（避免存入無法計算的垃圾字串）。

    2026-08-21新增：period（對帳月份）——原本只能在「上傳比對」頁批次入庫前(commit_batch())
    填寫一次，入庫後無法修改；使用者反饋「從匯入團體發票勾選匯入」時，若來源發票的
    recon_period 欄位本身填的不是「YYYY/MM」格式（而是像「0801-0810」這種日期區間字串），
    會整段被當成對帳月份帶進批次、入庫時原封不動寫進 dealer_reconcile_summary，之後在
    「對帳查詢」畫面看起來不對但又改不了。這裡開放期後也能修正：
    - 不可留空（跟 commit_batch() 要求一致，避免改成空字串讓資料在查詢時對不到任何期間）
    - 除了更新 dealer_reconcile_batches（連同 upload_group 同組批次），也要同步更新
      dealer_reconcile_summary 裡這幾個 batch_id 已寫入的列的 period 欄位——因為「對帳查詢」
      的「依經銷商矩陣／依內部料號名矩陣」兩種檢視是直接讀批次表的 period 顯示，但
      「依內部料號」「依財報料號品名彙總」以及用期間篩選查詢，讀的是 dealer_reconcile_summary
      自己的 period 欄位（入庫時複製過去、之後不會再跟批次表同步），兩邊沒有一起修正的話，
      同一批資料在不同檢視會看到不同期間，用修正後的期間篩選查詢也會找不到剛剛看到的那筆。"""
    data = request.json or {}
    rec = {}
    if 'period' in data:
        p = _norm(data.get('period'))
        if not p:
            return jsonify({'error': '對帳月份不可留空'}), 400
        rec['period'] = p
    for k in ('reconcile_range', 'note', 'customer_type', 'reconcile_order', 'reconcile_no', 'bookstore_reconcile_no'):
        if k in data:
            rec[k] = data.get(k) or None
    for k in ('manual_amount_untaxed', 'manual_amount_taxed'):
        if k in data:
            v = data.get(k)
            if v is None or (isinstance(v, str) and v.strip() == ''):
                rec[k] = None
            else:
                try:
                    rec[k] = float(v)
                except (TypeError, ValueError):
                    return jsonify({'error': f'{k} 必須是數字'}), 400
    if rec:
        group_bids = [bid]
        sb.table('dealer_reconcile_batches').update(rec).eq('id', bid).execute()
        cur_res = sb.table('dealer_reconcile_batches').select('upload_group').eq('id', bid).execute()
        cur = cur_res.data[0] if cur_res.data else None
        group = cur.get('upload_group') if cur else None
        if group:
            sb.table('dealer_reconcile_batches').update(rec).eq('upload_group', group).neq('id', bid).execute()
            grp_res = sb.table('dealer_reconcile_batches').select('id').eq('upload_group', group).execute()
            group_bids = [r['id'] for r in (grp_res.data or [])] or group_bids
        if 'period' in rec:
            sb.table('dealer_reconcile_summary').update({'period': rec['period']}).in_('batch_id', group_bids).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/dealer/period-notes', methods=['GET'])
@dealer_view_required
def get_period_note():
    """⚠️ 2026-07-24起前端已改用 /api/dealer/batches/<id>/notes（見上方 save_batch_notes），
    這兩個舊端點(GET/PUT)不再被呼叫，保留僅為不動 dealer_reconcile_period_notes 舊資料表，
    避免刪除路由/表格造成不必要的風險。
    依(經銷商,期間)查詢已儲存的對帳區間/對帳備註。供「上傳比對」頁批次明細開啟時預先帶出既有值，
    避免使用者在匯入階段填寫時看到空白，誤以為要重新輸入，結果把之前已存的值蓋掉。"""
    dealer_name = request.args.get('dealer_name', '').strip()
    period = request.args.get('period', '').strip()
    empty = {'reconcile_range': '', 'note': '', 'customer_type': '', 'reconcile_order': '',
             'reconcile_no': '', 'bookstore_reconcile_no': ''}
    if not dealer_name or not period:
        return jsonify(empty)
    res = sb.table('dealer_reconcile_period_notes').select('*') \
        .eq('dealer_name', dealer_name).eq('period', period).execute()
    if res.data:
        r = res.data[0]
        return jsonify({
            'reconcile_range': r.get('reconcile_range') or '', 'note': r.get('note') or '',
            'customer_type': r.get('customer_type') or '', 'reconcile_order': r.get('reconcile_order') or '',
            'reconcile_no': r.get('reconcile_no') or '', 'bookstore_reconcile_no': r.get('bookstore_reconcile_no') or '',
        })
    return jsonify(empty)


@dealer_bp.route('/api/dealer/period-notes', methods=['PUT'])
@dealer_edit_required
def save_period_note():
    """⚠️ 2026-07-24起已不再被前端呼叫（改用 save_batch_notes），保留僅為相容舊資料表。
    依經銷商矩陣檢視用：儲存某個(經銷商,期間)的對帳區間/對帳備註（使用者手動輸入的說明文字）。"""
    data = request.json or {}
    dealer_name = _norm(data.get('dealer_name'))
    period = _norm(data.get('period'))
    if not dealer_name or not period:
        return jsonify({'error': '缺少經銷商或期間'}), 400

    # 只更新這次請求「實際有帶」的欄位（各輸入框各自送出），
    # 避免只改其中一欄時把之前已存的其他欄位覆蓋成空白
    rec = {}
    for k in ('reconcile_range', 'note', 'customer_type', 'reconcile_order', 'reconcile_no', 'bookstore_reconcile_no'):
        if k in data:
            rec[k] = data.get(k) or None
    rec['updated_by'] = session.get('user_id')
    rec['updated_by_name'] = session.get('display_name', session.get('username', ''))
    rec['updated_at'] = now_str()

    existing = sb.table('dealer_reconcile_period_notes').select('id') \
        .eq('dealer_name', dealer_name).eq('period', period).execute()
    if existing.data:
        sb.table('dealer_reconcile_period_notes').update(rec).eq('id', existing.data[0]['id']).execute()
    else:
        rec['dealer_name'] = dealer_name
        rec['period'] = period
        sb.table('dealer_reconcile_period_notes').insert(rec).execute()
    return jsonify({'ok': True})


def _group_batches_by_upload(batches_map):
    """2026-07-24新增：依 upload_group 把同一次「上傳並比對」點擊時選取的多個檔案（香港經銷商
    通常固定是1張出貨單.odt+1張發票.pdf一起選取，見upload_batch()/前端doUpload()）歸成同一組，
    供對帳查詢矩陣合併顯示成一列——這樣使用者才能同時看到一次出貨的「數量(來自出貨單)+
    金額(來自發票)」合併在一起，而不是拆成2個看起來都不完整的列。沒有 upload_group 的批次
    （例如舊資料、或本來就只選了單一檔案上傳）各自視為只有自己一個成員的獨立群組，行為
    等同於前一版「一個批次一列」，不受影響。
    回傳 {group_key: {'batch_ids','rep_id','rep','filenames'}}，rep_id/rep 為該群組代表批次
    （群組內batch_id最小者），代表批次的欄位（dealer_name/period/備註等）用來呈現整個群組的
    資訊；使用者在畫面上編輯備註時，也是寫回這個代表批次（見 saveBatchNoteFor 前端函式）。"""
    groups = {}
    for bid, b in batches_map.items():
        gkey = b.get('upload_group') or f'b{bid}'
        groups.setdefault(gkey, []).append(bid)
    out = {}
    for gkey, bids in groups.items():
        bids_sorted = sorted(bids)
        rep_id = bids_sorted[0]
        names = [batches_map.get(bid, {}).get('source_filename') for bid in bids_sorted]
        names = [n for n in names if n]
        out[gkey] = {
            'batch_ids': bids_sorted,
            'rep_id': rep_id,
            'rep': batches_map.get(rep_id) or {},
            'filenames': ' + '.join(dict.fromkeys(names)),
        }
    return out


def _build_summary_data(dealer, period, code_kw, view, all_sku=False):
    """對帳查詢的共用彙總邏輯，供 JSON 查詢 API 與匯出 Excel 共用，避免邏輯寫兩份。
    回傳值依 view 不同而異：sku/report 回傳 list；matrix 回傳 {'columns','dealers'} 的 dict。
    all_sku=True 時（僅影響 matrix/matrix_report 兩種矩陣檢視）：欄位改為列出「內部料號清單」
    全部料號/財報品名，不論這次查詢範圍內是否有銷售，用途是讓使用者能核對欄位順序是否跟
    「內部料號清單」畫面排序完全一致（平常查詢仍預設只顯示有銷售的欄位，避免表格過寬）。"""
    q = sb.table('dealer_reconcile_summary').select('*')
    if dealer:
        q = q.eq('dealer_name', dealer)
    if period:
        q = q.eq('period', period)
    res = q.execute()
    rows = res.data or []

    sku_rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('*').order('sort_order'))
    sku_map = {s['internal_code']: s for s in sku_rows}

    if code_kw:
        rows = [r for r in rows if code_kw in r['internal_code']]

    if view == 'matrix':
        # 依經銷商矩陣彙總：欄=內部料號（依ERP順序排序），列=「一次批次上傳」（不跨批次加總），
        # 數值=數量加總。用途：比照使用者原本Excel「ALL年度」分頁的檢視習慣
        # （機型為欄、通路/經銷商為列），純看銷售數量，不含金額/財報欄位（財報用途仍用 view=report）
        # ⚠️ 2026-07-24變更：列的粒度從 (dealer_name, period) 改為「上傳群組(upload_group)」——
        # 原本同一個(經銷商,期間)若有2次以上批次會被合併成一列，使用者反饋這樣沒辦法拆開查看是
        # 哪一次的資料（常見於同一個月有多次出貨/開票的經銷商，如香港經銷商）；但改成每個批次
        # 各自一列後，使用者又反饋同一次上傳本來就固定是「1張出貨單+1張發票」一起選取，應該合併
        # 顯示成一列（數量來自出貨單、金額來自發票），拆成2個各自不完整的列反而不好看。因此改為
        # 依「同一次上傳選取的檔案」分組（見 _group_batches_by_upload），群組內的批次合併加總，
        # 不同群組（即使期間相同）仍分開，兩種需求都滿足。
        # 「對帳區間/對帳備註」等欄位改為讀取該群組代表批次的欄位（見 dealer_reconcile_batches）
        batches_res = sb.table('dealer_reconcile_batches').select('*').execute()
        batches_map = {b['id']: b for b in (batches_res.data or [])}
        groups_info = _group_batches_by_upload(batches_map)
        batch_to_group = {bid: gkey for gkey, info in groups_info.items() for bid in info['batch_ids']}

        group_qty = {}    # group_key -> {internal_code: qty}
        codes_seen = set()
        for r in rows:
            bid = r.get('batch_id')
            gkey = batch_to_group.get(bid, f'b{bid}')
            d = group_qty.setdefault(gkey, {})
            code = r['internal_code']
            d[code] = d.get(code, 0) + (r.get('qty') or 0)
            codes_seen.add(code)

        columns = []
        if all_sku:
            # 不限本次查詢範圍內是否有銷售，列出「內部料號清單」全部料號
            for code, sku in sku_map.items():
                columns.append({
                    'internal_code': code,
                    'erp_code': sku.get('erp_code'),
                    'sort_order': sku.get('sort_order') if sku.get('sort_order') is not None else 999999,
                })
        else:
            for code in codes_seen:
                sku = sku_map.get(code, {})
                columns.append({
                    'internal_code': code,
                    'erp_code': sku.get('erp_code'),
                    'sort_order': sku.get('sort_order') if sku.get('sort_order') is not None else 999999,
                })
        columns.sort(key=lambda c: c['sort_order'])

        dealers_out = []
        for gkey, qmap in sorted(group_qty.items(),
                                  key=lambda item: _period_sort_key(
                                      groups_info.get(item[0], {}).get('rep'),
                                      groups_info.get(item[0], {}).get('rep', {}).get('dealer_name', ''),
                                      groups_info.get(item[0], {}).get('rep', {}).get('period', ''))):
            info = groups_info.get(gkey, {})
            rep = info.get('rep', {})
            dealers_out.append({
                'dealer_name': rep.get('dealer_name') or '',
                'period': rep.get('period') or '',
                'batch_id': info.get('rep_id'),  # 代表批次id，編輯備註時寫回這一筆
                'recon_seq_no': rep.get('recon_seq_no') or '',  # 2026-08-18新增：拆帳編號
                'source_filename': info.get('filenames') or '',
                'customer_type': rep.get('customer_type') or '',
                'reconcile_order': rep.get('reconcile_order') or '',
                'reconcile_no': rep.get('reconcile_no') or '',
                'bookstore_reconcile_no': rep.get('bookstore_reconcile_no') or '',
                'reconcile_range': rep.get('reconcile_range') or '',
                'note': rep.get('note') or '',
                'qty_by_code': qmap,
            })
        return {'columns': columns, 'dealers': dealers_out}

    if view == 'matrix_report':
        # 依財報品名矩陣：預設（all_sku=False）欄=財報料號品名，同一財報品名若對應多個內部料號
        # （如同一機型底下有多個福利品分類，共用同一個財報彙總名稱）會合併成一欄，供財報使用。
        # all_sku=True（勾選「顯示全部料號(含無銷售)」）時，改為每個內部料號各自一欄，不再合併
        # ——這是2026-07-24使用者反饋：勾選此checkbox的本意就是要核對「內部料號清單」畫面上的
        # 每一筆料號是否都有對應欄位出現，若多筆料號共用同一財報品名仍被合併成一欄，會讓使用者
        # 誤以為有資料「沒有全部呈現」。故 all_sku=True 時改用 internal_code 當分欄依據（col_key），
        # 財報品名文字（可能同一段文字重複出現在好幾欄）僅作為欄位下方的顯示標籤，不影響是否合併
        # ⚠️ 2026-07-24變更：列=「上傳群組(upload_group)」（不再是「一個經銷商＋一個期間」，也不是
        # 單一批次），理由同上方 view=='matrix' 的說明——同一經銷商+期間常有多次批次，合併成一列
        # 會讓使用者無法拆開查看是哪一次的資料；但同一次上傳選取的1張出貨單+1張發票，使用者希望
        # 合併顯示成一列（數量來自出貨單、金額來自發票），故改依 upload_group 分組（見
        # _group_batches_by_upload）。並在每列右側附上該群組的未稅金額小計/含稅金額小計
        # （跨所有欄位加總，不逐欄拆分金額，避免表格過度複雜）
        batches_res = sb.table('dealer_reconcile_batches').select('*').execute()
        batches_map = {b['id']: b for b in (batches_res.data or [])}
        groups_info = _group_batches_by_upload(batches_map)
        batch_to_group = {bid: gkey for gkey, info in groups_info.items() for bid in info['batch_ids']}

        group_qty = {}       # group_key -> {col_key: qty}
        group_amt = {}       # group_key -> {'amount_untaxed','amount_taxed'}
        col_info = {}        # col_key -> {'report_product_name','erp_code','sort_order'}

        def _col_key(code, sku):
            if all_sku:
                return code  # 每個內部料號各自一欄，不合併
            return sku.get('report_product_name') or code  # 預設：依財報品名合併（財報用途）

        def _upsert_col(ckey, sku, code):
            so = sku.get('sort_order')
            so = so if so is not None else 999999
            if ckey not in col_info or so < col_info[ckey]['sort_order']:
                col_info[ckey] = {
                    'report_product_name': sku.get('report_product_name') or code,
                    'internal_code': code,  # 該欄sort_order最小那筆的內部料號，畫面上改用此欄位顯示（2026-07-24使用者要求，不再顯示財報品名）
                    'erp_code': sku.get('erp_code'),
                    'sort_order': so,
                }

        for r in rows:
            sku = sku_map.get(r['internal_code'], {})
            ckey = _col_key(r['internal_code'], sku)
            bid = r.get('batch_id')
            gkey = batch_to_group.get(bid, f'b{bid}')
            d = group_qty.setdefault(gkey, {})
            d[ckey] = d.get(ckey, 0) + (r.get('qty') or 0)
            a = group_amt.setdefault(gkey, {'amount_untaxed': 0, 'amount_taxed': 0})
            a['amount_untaxed'] += r.get('amount_untaxed') or 0
            a['amount_taxed'] += r.get('amount_taxed') or 0
            _upsert_col(ckey, sku, r['internal_code'])

        if all_sku:
            # 不限本次查詢範圍內是否有銷售，把「內部料號清單」全部料號都列進欄位（一料號一欄）
            for code, sku in sku_map.items():
                _upsert_col(code, sku, code)

        columns = sorted(col_info.keys(), key=lambda k: col_info[k]['sort_order'])

        dealers_out = []
        for gkey, qmap in sorted(group_qty.items(),
                                  key=lambda item: _period_sort_key(
                                      groups_info.get(item[0], {}).get('rep'),
                                      groups_info.get(item[0], {}).get('rep', {}).get('dealer_name', ''),
                                      groups_info.get(item[0], {}).get('rep', {}).get('period', ''))):
            info = groups_info.get(gkey, {})
            rep = info.get('rep', {})
            amt = group_amt.get(gkey, {'amount_untaxed': 0, 'amount_taxed': 0})
            # 2026-08-07新增：人工填入的未稅/含稅金額小計（見 save_batch_notes() 說明）——
            # 有填就覆蓋顯示（取代系統試算金額），沒填(None)則維持系統依明細加總算出來的金額
            manual_u = rep.get('manual_amount_untaxed')
            manual_t = rep.get('manual_amount_taxed')
            final_untaxed = manual_u if manual_u is not None else amt['amount_untaxed']
            final_taxed = manual_t if manual_t is not None else amt['amount_taxed']
            dealers_out.append({
                'dealer_name': rep.get('dealer_name') or '',
                'period': rep.get('period') or '',
                'batch_id': info.get('rep_id'),  # 代表批次id，編輯備註時寫回這一筆
                'recon_seq_no': rep.get('recon_seq_no') or '',  # 2026-08-18新增：拆帳編號
                'source_filename': info.get('filenames') or '',
                'customer_type': rep.get('customer_type') or '',
                'reconcile_order': rep.get('reconcile_order') or '',
                'reconcile_no': rep.get('reconcile_no') or '',
                'bookstore_reconcile_no': rep.get('bookstore_reconcile_no') or '',
                'reconcile_range': rep.get('reconcile_range') or '',
                'note': rep.get('note') or '',
                'qty_by_name': qmap,
                'amount_untaxed': final_untaxed,
                'amount_taxed': final_taxed,
                'system_amount_untaxed': amt['amount_untaxed'],
                'system_amount_taxed': amt['amount_taxed'],
                'manual_amount_untaxed': manual_u,
                'manual_amount_taxed': manual_t,
            })
        return {'columns': [{'col_key': k, 'report_product_name': col_info[k]['report_product_name'],
                              'internal_code': col_info[k]['internal_code'],
                              'erp_code': col_info[k]['erp_code']} for k in columns], 'dealers': dealers_out}

    if view == 'report':
        agg = {}
        for r in rows:
            sku = sku_map.get(r['internal_code'], {})
            key = sku.get('report_product_name') or r['internal_code']
            a = agg.setdefault(key, {'report_product_name': key, 'qty': 0, 'amount': 0,
                                      'amount_untaxed': 0, 'amount_taxed': 0, 'tax_amount': 0})
            a['qty'] += r.get('qty') or 0
            a['amount'] += r.get('amount') or 0
            a['amount_untaxed'] += r.get('amount_untaxed') or 0
            a['amount_taxed'] += r.get('amount_taxed') or 0
            a['tax_amount'] += r.get('tax_amount') or 0
        return sorted(agg.values(), key=lambda x: x['report_product_name'])

    for r in rows:
        sku = sku_map.get(r['internal_code'], {})
        r['erp_code'] = sku.get('erp_code')
        r['sort_order'] = sku.get('sort_order', 999999)
    rows.sort(key=lambda x: x.get('sort_order', 999999))
    return rows


@dealer_bp.route('/api/dealer/summary', methods=['GET'])
@dealer_view_required
def query_summary():
    dealer = request.args.get('dealer', '').strip()
    period = request.args.get('period', '').strip()
    code_kw = request.args.get('code', '').strip()
    view = request.args.get('view', 'sku')
    all_sku = request.args.get('all_sku', '') in ('1', 'true', 'True')
    return jsonify(_build_summary_data(dealer, period, code_kw, view, all_sku))


@dealer_bp.route('/api/dealer/summary/export', methods=['GET'])
@dealer_view_required
def export_summary():
    dealer = request.args.get('dealer', '').strip()
    period = request.args.get('period', '').strip()
    code_kw = request.args.get('code', '').strip()
    view = request.args.get('view', 'sku')
    all_sku = request.args.get('all_sku', '') in ('1', 'true', 'True')
    data = _build_summary_data(dealer, period, code_kw, view, all_sku)

    wb = openpyxl.Workbook()
    ws = wb.active

    if view == 'matrix':
        ws.title = '依經銷商矩陣'
        columns = data.get('columns') or []
        dealers = data.get('dealers') or []
        head_cols = ['經銷商', '來源檔名', '客戶類型', '對帳順序', '對帳月份', '對帳編號', '書店對帳編號', '對帳區間', '對帳備註']
        # 兩列表頭：第一列ERP順序、第二列內部料號（比照畫面呈現）
        # 2026-07-24：新增「來源檔名」欄，因為每一列現在代表一次批次上傳，方便對照是哪個檔案
        ws.append(head_cols + [c.get('erp_code') or '' for c in columns] + ['小計'])
        ws.append([''] * len(head_cols) + [c.get('internal_code') for c in columns] + [''])
        col_totals = [0] * len(columns)
        for d in dealers:
            qmap = d.get('qty_by_code') or {}
            row_vals = []
            row_total = 0
            for i, c in enumerate(columns):
                q = qmap.get(c['internal_code']) or 0
                row_vals.append(q if q else '')
                row_total += q
                col_totals[i] += q
            ws.append([d.get('dealer_name'), d.get('source_filename') or '', d.get('customer_type') or '', d.get('reconcile_order') or '',
                       d.get('period'), d.get('reconcile_no') or '', d.get('bookstore_reconcile_no') or '',
                       d.get('reconcile_range') or '', d.get('note') or '']
                       + row_vals + [row_total])
        ws.append(['總計'] + [''] * (len(head_cols) - 1) + [(v if v else '') for v in col_totals] + [sum(col_totals)])
    elif view == 'matrix_report':
        ws.title = '依內部料號名矩陣'
        columns = data.get('columns') or []
        dealers = data.get('dealers') or []
        head_cols = ['經銷商', '來源檔名', '客戶類型', '對帳順序', '對帳月份', '對帳編號', '書店對帳編號', '對帳區間', '對帳備註']
        # 兩列表頭：第一列ERP順序、第二列內部料號（2026-07-24改為顯示內部料號，不再顯示財報品名，比照畫面呈現）
        # 2026-07-24：新增「來源檔名」欄，因為每一列現在代表一次批次上傳，方便對照是哪個檔案
        ws.append(head_cols + [c.get('erp_code') or '' for c in columns] + ['', ''])
        ws.append([''] * len(head_cols) + [c.get('internal_code') for c in columns] + ['未稅金額小計', '含稅金額小計'])
        col_totals = [0] * len(columns)
        total_untaxed = total_taxed = 0
        for d in dealers:
            qmap = d.get('qty_by_name') or {}
            row_vals = []
            for i, c in enumerate(columns):
                q = qmap.get(c.get('col_key', c.get('report_product_name'))) or 0
                row_vals.append(q if q else '')
                col_totals[i] += q
            ws.append([d.get('dealer_name'), d.get('source_filename') or '', d.get('customer_type') or '', d.get('reconcile_order') or '',
                       d.get('period'), d.get('reconcile_no') or '', d.get('bookstore_reconcile_no') or '',
                       d.get('reconcile_range') or '', d.get('note') or '']
                       + row_vals + [d.get('amount_untaxed') or 0, d.get('amount_taxed') or 0])
            total_untaxed += d.get('amount_untaxed') or 0
            total_taxed += d.get('amount_taxed') or 0
        ws.append(['總計'] + [''] * (len(head_cols) - 1) + [(v if v else '') for v in col_totals] + [total_untaxed, total_taxed])
    elif view == 'report':
        ws.title = '依財報料號品名彙總'
        ws.append(['項次', '財報料號品名', '數量', '未稅金額', '含稅金額'])
        sum_qty = sum_untaxed = sum_taxed = 0
        for i, r in enumerate(data, 1):
            ws.append([i, r.get('report_product_name'), r.get('qty'), r.get('amount_untaxed'), r.get('amount_taxed')])
            sum_qty += r.get('qty') or 0
            sum_untaxed += r.get('amount_untaxed') or 0
            sum_taxed += r.get('amount_taxed') or 0
        ws.append(['小計', '', sum_qty, sum_untaxed, sum_taxed])
    else:
        ws.title = '依內部料號'
        ws.append(['項次', '經銷商', '期間', '內部料號', 'ERP順序', '數量', '未稅金額', '含稅金額'])
        sum_qty = sum_untaxed = sum_taxed = 0
        for i, r in enumerate(data, 1):
            ws.append([i, r.get('dealer_name'), r.get('period'), r.get('internal_code'),
                       r.get('erp_code') or '', r.get('qty'), r.get('amount_untaxed'), r.get('amount_taxed')])
            sum_qty += r.get('qty') or 0
            sum_untaxed += r.get('amount_untaxed') or 0
            sum_taxed += r.get('amount_taxed') or 0
        ws.append(['小計', '', '', '', '', sum_qty, sum_untaxed, sum_taxed])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"對帳查詢_{view}_{period or '全部期間'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# ⑧ 書店經銷商方案清單 CRUD（方案 → 財報料號品名拆帳金額，多對多）
# ============================================================
# 2026-08-07新增，同日改版：籌備「書店」經銷商模組的第一步。方案清單頁面呈現方式比照使用者
# 原本在 Excel「方案」頁籤維護的格式（方案編號/方案中文名稱/售價含稅/備註 + 每個財報品項各
# 一欄的拆帳金額寬表）。
#
# ⚠️ 拆帳明細的鍵值是「財報料號品名」(report_product_name)，不是「內部料號」(internal_code)：
# 「內部料號清單」裡常有多筆內部料號共用同一個財報料號品名（例如同一款機型的白色/淺蔥青/福利
# 等級A/B/S…等變體），這些變體對「書店方案要把多少錢分配到哪個財報品項」這件事沒有差異。
# 一開始的版本曾以 internal_code 作為拆帳明細的鍵值，下拉選單會把每個顏色/福利等級變體都
# 列出來，使用者實際測試後反映「這不是我要的，我希望是可以選擇主料號品名(粗體)的那個，
# 其餘子料號可以隱藏，不顯示」——也就是「財報料號品名」本身才是她要選的東西，不需要逐一
# 挑選底下的變體。因此改為以 report_product_name 為鍵，且因為新表格上線時尚無任何方案資料
# （0筆），直接把 bookstore_plan_items.internal_code 欄位改名為 report_product_name，不需要
# 額外的資料搬遷。料號主檔本身仍完全共用「內部料號清單」(internal_sku_list)，不重複建置，
# 財報料號品名的新增/修改都在「料號與規則管理」那邊維護即可，這裡的寬表欄位會自動跟著變動。
#
# 儲存驗證規則：新增/修改方案時，若有填「售價(含稅)」，則所有財報品項拆帳金額加總必須等於
# 售價(含稅)（容許 ±0.5 元誤差，避免除不盡的四捨五入被誤判）。不相符時後端回傳 409 +
# 差額，前端跳出確認視窗詢問使用者是否仍要儲存；使用者確認後，前端會帶著 force=true 重新
# 送出，後端此時放行（軟性警告，而非強制擋下——因為不排除少數方案本來就有意不需要100%
# 拆帳金額對帳，例如優惠折抵不打算歸屬到任何單一財報品項的情況）。


def _plan_valid_report_names():
    """回傳目前「內部料號清單」裡所有有效的財報料號品名集合（不含空值），
    供書店方案拆帳明細存檔前驗證用。"""
    # ⚠️ 2026-08-08修正：改用 _fetch_all_rows() 分頁抓取，避免內部料號清單超過1000筆時
    # 被PostgREST靜默截斷，導致較晚新增的財報料號品名被誤判成「不存在」而擋下儲存。
    rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('report_product_name'))
    return {r['report_product_name'] for r in rows if r.get('report_product_name')}


# 2026-08-18新增：「子料號」欄位（v1.32，見dealer_reconcile_schema.sql）共用的2個查表函式，
# 分別供①方案清單存檔前驗證子料號是否有效、②「書店」出貨資料比對時判斷「財報料號品名底下
# 只有1個內部料號→自動比對，否則需要子料號」共用，避免各自查一次資料庫、寫兩份邏輯。
def _sku_report_name_by_code():
    """回傳 {內部料號: 財報料號品名}，供驗證「子料號」是否存在、且是否屬於選定的財報料號
    品名底下（避免手滑選到別的品名底下的顏色）。"""
    rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('internal_code,report_product_name'))
    return {r['internal_code']: r.get('report_product_name') for r in rows if r.get('internal_code')}


def _report_name_to_active_codes():
    """回傳 {財報料號品名: [啟用中的內部料號, ...]}（略過is_active=false的料號），供「書店」
    出貨資料比對時判斷：同一財報料號品名底下若只剩1筆啟用中的內部料號，代表沒有色號/等級
    分歧，可以自動比對，不需要方案清單另外指定子料號。"""
    rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('internal_code,report_product_name,is_active'))
    out = {}
    for r in rows:
        if r.get('is_active') is False:
            continue
        name = r.get('report_product_name')
        if name and r.get('internal_code'):
            out.setdefault(name, []).append(r['internal_code'])
    return out


@dealer_bp.route('/api/bookstore/plans', methods=['GET'])
@bookstore_view_required
def list_bookstore_plans():
    kw = request.args.get('kw', '').strip()
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = min(200, max(1, int(request.args.get('per_page', 50) or 50)))

    q = sb.table('bookstore_plans').select('*', count='exact')
    if kw:
        q = q.or_(f'plan_code.ilike.%{kw}%,plan_name.ilike.%{kw}%')
    # 2026-08-08調整：改依「匯入順序」排序（見v1.16欄位），不再依方案編號字串排序——
    # 最近一次批次匯入的整批資料排最上面，批次內部照Excel原始列順序；從未被批次匯入過的
    # 方案（import_seq/last_imported_at皆為空）自然排到最後面，用 id 做穩定排序的最後依據。
    q = q.order('last_imported_at', desc=True, nullslast=True).order('import_seq', nullslast=True).order('id')
    start = (page - 1) * per_page
    res = q.range(start, start + per_page - 1).execute()
    plans = res.data or []
    total = res.count or 0

    plan_ids = [p['id'] for p in plans]
    items_map = {}
    if plan_ids:
        items_rows = _fetch_all_rows(lambda: sb.table('bookstore_plan_items').select('*').in_('plan_id', plan_ids))
        for it in items_rows:
            items_map.setdefault(it['plan_id'], []).append({
                'report_product_name': it['report_product_name'], 'amount': it['amount'],
                'sub_internal_code': it.get('sub_internal_code'),
            })
    for p in plans:
        p['items'] = items_map.get(p['id'], [])
        p['items_total'] = round(sum(i['amount'] for i in p['items']), 2)

    return jsonify({
        'total': total, 'page': page, 'per_page': per_page, 'plans': plans,
        'sku_last_synced_at': _get_sku_last_synced(),
    })


@dealer_bp.route('/api/bookstore/plans/<int:pid>', methods=['GET'])
@bookstore_view_required
def get_bookstore_plan(pid):
    res = sb.table('bookstore_plans').select('*').eq('id', pid).execute()
    if not res.data:
        return jsonify({'error': '找不到此方案'}), 404
    plan = res.data[0]
    items_res = sb.table('bookstore_plan_items').select('*').eq('plan_id', pid).execute()
    plan['items'] = [{
        'report_product_name': i['report_product_name'], 'amount': i['amount'],
        'sub_internal_code': i.get('sub_internal_code'),
    } for i in (items_res.data or [])]
    return jsonify(plan)


def _validate_plan_items(price_incl, items, force):
    """檢查財報料號品名是否都存在於內部料號清單，以及拆帳金額加總是否等於售價(含稅)。
    回傳 (error_response_or_None)。error 為 None 代表通過驗證，可以繼續儲存。
    2026-08-11修正：拿掉「同一財報料號品名在同一方案中不可重複」的檢查——使用者反映
    有些方案是同一料號賣多份，習慣拆成多列各自填金額（例如透明殼×3，分3列各351元，
    而非合併成1列1053元），這是正常用法，不應該擋下儲存。同名多列的金額加總邏輯已在
    _compute_bookstore_order_matrix()／export_bookstore_plans() 一併修正為「加總」而非
    「取最後一筆」，資料庫的 UNIQUE(plan_id, report_product_name) 限制也需要另外執行
    SQL移除，否則這裡驗證通過、實際寫入資料庫時仍會被擋下。"""
    valid_names = _plan_valid_report_names()
    # 2026-08-18新增：子料號驗證（v1.32）——存在性＋是否屬於這一列選定的財報料號品名底下，
    # 避免手滑選到別的品名底下的顏色（例如品名選了「側翻殼」，子料號卻選到「收納套」的顏色）。
    sub_name_by_code = _sku_report_name_by_code() if any(_norm(it.get('sub_internal_code')) for it in items) else {}
    for it in items:
        name = _norm(it.get('report_product_name'))
        if not name:
            return jsonify({'error': '財報料號品名不可留空，請選擇一個財報料號品名或移除該列'}), 400
        if name not in valid_names:
            return jsonify({'error': f'財報料號品名「{name}」不存在於「內部料號清單」，請先在該清單新增此財報料號品名，或修正為既有品名'}), 400
        sub = _norm(it.get('sub_internal_code'))
        if sub:
            sub_name = sub_name_by_code.get(sub)
            if sub_name is None:
                return jsonify({'error': f'子料號「{sub}」不存在於「內部料號清單」'}), 400
            if sub_name != name:
                return jsonify({'error': f'子料號「{sub}」屬於財報料號品名「{sub_name}」，跟這一列選的「{name}」不一致，請重新選擇子料號'}), 400

    if price_incl is not None and items and not force:
        items_total = round(sum(_to_num(it.get('amount')) for it in items), 2)
        diff = round(items_total - float(price_incl), 2)
        if abs(diff) > 0.5:
            return jsonify({
                'error': f'拆帳金額加總（{items_total}）與售價含稅（{price_incl}）不相符，差額 {diff} 元',
                'warning': True,
                'items_total': items_total,
                'diff': diff,
            }), 409
    return None


@dealer_bp.route('/api/bookstore/plans', methods=['POST'])
@bookstore_action_required('mod_bookstore_plan_create')
def create_bookstore_plan():
    data = request.json or {}
    plan_code = _norm(data.get('plan_code'))
    if not plan_code:
        return jsonify({'error': '方案編號必填'}), 400
    # 2026-08-08新增：單筆新增前先明確檢查方案編號是否已存在，回傳清楚的錯誤訊息；
    # 原本只靠資料庫UNIQUE限制擋下重複，錯誤訊息會夾帶原始的資料庫錯誤內容，對使用者不友善。
    dup_res = sb.table('bookstore_plans').select('id').eq('plan_code', plan_code).execute()
    if dup_res.data:
        return jsonify({'error': f'方案編號「{plan_code}」已存在，請使用其他編號，或到清單中編輯該筆既有方案'}), 409
    plan_name = _norm(data.get('plan_name'))
    if not plan_name:
        return jsonify({'error': '方案中文名稱必填'}), 400
    price_incl = _to_num_or_none(data.get('price_incl'))
    if price_incl is None:
        return jsonify({'error': '售價(含稅)必填'}), 400
    items = data.get('items') or []
    force = bool(data.get('force'))

    err = _validate_plan_items(price_incl, items, force)
    if err:
        return err

    rec = {
        'plan_code': plan_code,
        'plan_name': plan_name,
        'price_incl': price_incl,
        'note': data.get('note'),
        # 2026-08-14修正：單筆新增（新增方案按鈕）原本不寫入last_imported_at，導致跟「從
        # 未被批次匯入過」的舊資料混在一起、依id由小到大排在清單最後面，新增的方案要翻到
        # 最後一頁才找得到。這裡仿照批次匯入／簡易批次新增的做法，把單筆新增也蓋上目前時間
        # 當作它自己的last_imported_at，讓list_bookstore_plans()既有的排序（依
        # last_imported_at新到舊）自然把它排到最上面，不需要另外改排序邏輯。
        'last_imported_at': now_str(),
    }
    rec.update(_audit_new())
    try:
        res = sb.table('bookstore_plans').insert(rec).execute()
    except Exception as e:
        return jsonify({'error': f'新增失敗（方案編號可能已存在）：{e}'}), 400
    plan = res.data[0]

    item_rows = [{
        'plan_id': plan['id'], 'report_product_name': _norm(it.get('report_product_name')),
        'amount': _to_num(it.get('amount')), 'sub_internal_code': _norm(it.get('sub_internal_code')) or None,
    } for it in items if _norm(it.get('report_product_name'))]
    if item_rows:
        sb.table('bookstore_plan_items').insert(item_rows).execute()
    plan['items'] = [{
        'report_product_name': r['report_product_name'], 'amount': r['amount'],
        'sub_internal_code': r.get('sub_internal_code'),
    } for r in item_rows]
    return jsonify(plan)


@dealer_bp.route('/api/bookstore/plans/<int:pid>', methods=['PUT'])
@bookstore_action_required('mod_bookstore_plan_update')
def update_bookstore_plan(pid):
    data = request.json or {}
    cur_res = sb.table('bookstore_plans').select('*').eq('id', pid).execute()
    if not cur_res.data:
        return jsonify({'error': '找不到此方案'}), 404
    cur = cur_res.data[0]

    plan_code = _norm(data.get('plan_code', cur['plan_code']))
    if not plan_code:
        return jsonify({'error': '方案編號必填'}), 400
    plan_name = _norm(data['plan_name']) if 'plan_name' in data else _norm(cur.get('plan_name'))
    if not plan_name:
        return jsonify({'error': '方案中文名稱必填'}), 400
    price_incl = _to_num_or_none(data['price_incl']) if 'price_incl' in data else cur.get('price_incl')
    if price_incl is None:
        return jsonify({'error': '售價(含稅)必填'}), 400
    items = data.get('items')
    if items is None:
        items_res = sb.table('bookstore_plan_items').select('*').eq('plan_id', pid).execute()
        items = [{
            'report_product_name': i['report_product_name'], 'amount': i['amount'],
            'sub_internal_code': i.get('sub_internal_code'),
        } for i in (items_res.data or [])]
    force = bool(data.get('force'))

    err = _validate_plan_items(price_incl, items, force)
    if err:
        return err

    rec = {'plan_code': plan_code}
    if 'plan_name' in data:
        rec['plan_name'] = plan_name
    if 'note' in data:
        rec['note'] = data['note']
    if 'price_incl' in data:
        rec['price_incl'] = price_incl
    rec.update(_audit_upd())
    try:
        sb.table('bookstore_plans').update(rec).eq('id', pid).execute()
    except Exception as e:
        return jsonify({'error': f'更新失敗（方案編號可能已存在）：{e}'}), 400

    if 'items' in data:
        sb.table('bookstore_plan_items').delete().eq('plan_id', pid).execute()
        item_rows = [{
            'plan_id': pid, 'report_product_name': _norm(it.get('report_product_name')),
            'amount': _to_num(it.get('amount')), 'sub_internal_code': _norm(it.get('sub_internal_code')) or None,
        } for it in items if _norm(it.get('report_product_name'))]
        if item_rows:
            sb.table('bookstore_plan_items').insert(item_rows).execute()

    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/plans/<int:pid>', methods=['DELETE'])
@bookstore_action_required('mod_bookstore_plan_delete')
def delete_bookstore_plan(pid):
    # 2026-08-08新增：刪除前檢查 bookstore_order_import_log，若此方案已在「匯入訂單」分析
    # 裡被使用過（表示已有真實訂單依賴這個方案的拆帳設定），不可刪除。
    p_res = sb.table('bookstore_plans').select('plan_code').eq('id', pid).execute()
    plan_code = (p_res.data or [{}])[0].get('plan_code') if p_res.data else None
    if plan_code:
        log_res = (sb.table('bookstore_order_import_log')
                   .select('imported_at,file_name')
                   .eq('plan_code', plan_code)
                   .order('imported_at', desc=True).limit(1).execute())
        if log_res.data:
            last = log_res.data[0]
            return jsonify({'error': f'方案「{plan_code}」在「匯入書店訂單」分析中已有使用紀錄'
                                      f'（最近一次：{last.get("imported_at","")}，檔案：{last.get("file_name") or "未命名"}），'
                                      f'不可刪除。如確定要刪除，請先聯絡系統管理員確認並清除相關匯入歷史記錄。'}), 409
    sb.table('bookstore_plans').delete().eq('id', pid).execute()
    return jsonify({'ok': True})


# ── 書店方案清單 — 整批「確認差額」（2026-08-08新增） ──────────────────────────
# 背景：使用者匯入的方案裡有很大一批是延續好幾年的舊方案，拆帳金額加總跟售價本來就對不起來
# （歷史因素），她明確表示不會去修改這些金額，只是希望畫面上不要一直顯示紅色「差額」提醒。
# 這裡刻意不去動 price_incl/拆帳明細本身（那樣等於竄改歷史資料），而是新增一個
# mismatch_ack 標記欄位，整批把「目前金額不符」的方案標記為「已確認」，前端看到這個標記
# 就把紅色改成灰色顯示，資料本身完全不變、之後隨時可以查出原始差額是多少。
@dealer_bp.route('/api/bookstore/plans/ack-all-mismatched', methods=['POST'])
@bookstore_edit_required
def ack_all_mismatched_plans():
    # ⚠️ 2026-08-08修正：改用 _fetch_all_rows() 分頁抓取，避免 bookstore_plans/
    # bookstore_plan_items 超過1000筆時被 PostgREST 靜默截斷，導致部份方案漏算。
    plans = _fetch_all_rows(lambda: sb.table('bookstore_plans').select('id,price_incl,mismatch_ack'))
    plan_ids = [p['id'] for p in plans]
    items_map = {}
    if plan_ids:
        items_rows = _fetch_all_rows(lambda: sb.table('bookstore_plan_items').select('plan_id,amount').in_('plan_id', plan_ids))
        for it in items_rows:
            items_map.setdefault(it['plan_id'], 0)
            items_map[it['plan_id']] += _to_num(it['amount'])

    to_ack = []
    for p in plans:
        if p.get('mismatch_ack'):
            continue
        if p.get('price_incl') is None:
            continue
        items_total = round(items_map.get(p['id'], 0), 2)
        diff = round(items_total - float(p['price_incl']), 2)
        if abs(diff) > 0.5:
            to_ack.append(p['id'])

    CHUNK = 200
    for i in range(0, len(to_ack), CHUNK):
        sb.table('bookstore_plans').update({'mismatch_ack': True}).in_('id', to_ack[i:i + CHUNK]).execute()

    return jsonify({'ok': True, 'acknowledged': len(to_ack)})


# 2026-08-18新增：「子料號」欄位在「範本/匯出」／「批次匯入」裡，每個財報料號品名固定
# 緊接著多一欄「{品名} - 子料號」，讓使用者可以在Excel裡一次批次填寫/修改大量方案的子料號，
# 不需要一個一個方案點進編輯彈窗改（見_resolve_bookstore_sub_sku()子料號比對邏輯說明）。
_BP_SUBCODE_SUFFIX = ' - 子料號'


# ── 書店方案清單 — 範本/匯出（含現有全部資料，可直接編輯後用「批次匯入」整批送回，
#    比照全系統既有慣例：如「內部料號清單」「代碼管理」的範本本身即為現有資料快照） ──
@dealer_bp.route('/api/bookstore/plans/export', methods=['GET'])
@bookstore_action_required('mod_bookstore_plan_export_template', 'mod_bookstore_plan_export')
def export_bookstore_plans():
    # ⚠️ 2026-08-08修正：以下三個查詢都改用 _fetch_all_rows() 分頁抓取，避免
    # bookstore_plans/bookstore_plan_items/internal_sku_list 超過1000筆時被PostgREST
    # 靜默截斷，導致匯出檔漏掉較晚建立的方案或料號。
    sku_rows = _fetch_all_rows(lambda: sb.table('internal_sku_list')
                                .select('report_product_name,erp_code,sort_order,is_active').order('sort_order'))
    # 2026-08-08新增：某財報料號品名底下若「所有」內部料號都被設為隱藏(is_active=false)，
    # 該品名整欄不出現在匯出檔／範本裡，比照方案清單頁面欄位的隱藏規則保持一致（只要還有
    # 一筆是顯示中，就仍視為要匯出）。
    active_by_name = {}
    for s in sku_rows:
        name = s.get('report_product_name')
        if not name:
            continue
        if s.get('is_active') is not False:
            active_by_name[name] = True
        elif name not in active_by_name:
            active_by_name[name] = False

    # 依財報料號品名去重（同一品名可能對應多筆內部料號變體），保留第一次出現（sort_order最小）
    # 那筆的 erp_code 供欄位上方顯示參考
    names = []
    seen_names = set()
    for s in sku_rows:
        name = s.get('report_product_name')
        if name and name not in seen_names and active_by_name.get(name):
            seen_names.add(name)
            names.append({'report_product_name': name, 'erp_code': s.get('erp_code')})

    # 2026-08-08調整：匯出順序改跟方案清單頁面一致（依匯入順序，見v1.16欄位），不再依方案
    # 編號字串排序，避免使用者匯出後排序跟畫面上看到的不一樣造成混淆。
    plans = _fetch_all_rows(lambda: sb.table('bookstore_plans').select('*')
                             .order('last_imported_at', desc=True, nullslast=True)
                             .order('import_seq', nullslast=True).order('id'))
    plan_ids = [p['id'] for p in plans]
    items_map = {}
    sub_map = {}  # plan_id -> {report_product_name: sub_internal_code}，供子料號欄位匯出
    if plan_ids:
        items_rows = _fetch_all_rows(lambda: sb.table('bookstore_plan_items').select('*').in_('plan_id', plan_ids))
        for it in items_rows:
            # 2026-08-11修正：同一方案裡同一財報料號品名現在可以有多列（見_validate_plan_items()
            # 的修正說明），這裡改成「加總」，之前是直接覆蓋，同名多列時只會留下最後一筆讀到的
            # 金額、其他列的金額會悄悄消失在匯出檔裡。
            bucket = items_map.setdefault(it['plan_id'], {})
            name = it['report_product_name']
            bucket[name] = (bucket.get(name) or 0) + _to_num(it['amount'])
            # 2026-08-18新增：子料號同名多列理論上應該一致（同一料號出貨多份），取第一筆
            # 有值的即可，不特別檢查是否所有列都一致（比照全系統既有的寬鬆判斷慣例）。
            sub_bucket = sub_map.setdefault(it['plan_id'], {})
            if it.get('sub_internal_code') and not sub_bucket.get(name):
                sub_bucket[name] = it['sub_internal_code']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '書店方案清單'

    base_headers = ['方案編號', '方案中文名稱', '售價(含稅)', '備註']
    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    rfill = openpyxl.styles.PatternFill('solid', fgColor='FFF2CC')
    rfont = openpyxl.styles.Font(bold=True, color='7B3F00')
    nfill = openpyxl.styles.PatternFill('solid', fgColor='F2F2F2')
    nfont = openpyxl.styles.Font(italic=True, color='888888')
    # 2026-08-18新增：子料號欄位用跟金額欄不同的底色（淺藍），方便肉眼分辨哪些是金額、
    # 哪些是子料號，兩種欄位緊鄰在一起容易看錯。
    subfill = openpyxl.styles.PatternFill('solid', fgColor='D6EAF8')
    subfont = openpyxl.styles.Font(bold=True, color='1A5276')
    center = openpyxl.styles.Alignment(horizontal='center')

    # 每個財報料號品名固定佔用2欄：[品名](金額) + [品名 - 子料號]，欄位從第5欄開始，
    # 每個名稱間隔2欄；col_of(i) 回傳第i個名稱「金額」欄的1-based欄位索引，子料號欄緊接其後。
    def col_of(i):
        return 5 + i * 2

    # Row1：ERP順序（前4欄留空；每個品名的「金額」欄上方顯示ERP順序，方便肉眼核對排序，
    # 子料號欄上方留空；匯入時忽略此列）
    for ci in range(1, 5):
        ws.cell(1, ci, '')
    for i, n in enumerate(names):
        c = ws.cell(1, col_of(i), n.get('erp_code') or '')
        c.font = nfont
        c.alignment = center

    # Row2：欄位標題（方案編號/方案中文名稱/售價(含稅)/備註 + 各財報料號品名+子料號，
    # 批次匯入以此列為標題列）
    for ci, h in enumerate(base_headers, 1):
        c = ws.cell(2, ci, h)
        if h == '方案編號':
            c.fill = rfill
            c.font = rfont
        else:
            c.fill = hfill
            c.font = hfont
        c.alignment = center
    for i, n in enumerate(names):
        ci = col_of(i)
        c = ws.cell(2, ci, n['report_product_name'])
        c.fill = hfill
        c.font = hfont
        c.alignment = center
        c2 = ws.cell(2, ci + 1, n['report_product_name'] + _BP_SUBCODE_SUFFIX)
        c2.fill = subfill
        c2.font = subfont
        c2.alignment = center

    # 2026-08-18新增：子料號欄改用下拉選單（Excel資料驗證），避免手動輸入打錯字、或誤填
    # 到不屬於該財報料號品名的內部料號。選項來源＝該品名底下目前「啟用中」的內部料號，
    # 存放在隱藏工作表「子料號選項」（各品名各佔一欄），下拉選單以跨工作表範圍參照取得。
    ws_opts = wb.create_sheet('子料號選項')
    ws_opts.sheet_state = 'hidden'
    name_to_codes = _report_name_to_active_codes()
    DV_BUFFER_ROWS = 500  # 多預留列數，讓使用者日後在匯出檔手動新增方案列時也有下拉選單可用
    for i, n in enumerate(names):
        codes = name_to_codes.get(n['report_product_name']) or []
        if not codes:
            continue
        opt_col_letter = openpyxl.utils.get_column_letter(i + 1)
        ws_opts.cell(1, i + 1, n['report_product_name'])
        for j, code in enumerate(codes):
            ws_opts.cell(2 + j, i + 1, code)
        dv = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=f"'子料號選項'!${opt_col_letter}$2:${opt_col_letter}${1 + len(codes)}",
            allow_blank=True, showErrorMessage=True,
            errorTitle='子料號不在清單內',
            error='請從下拉選單選擇這個財報料號品名底下的內部料號，或留空（只有1個內部料號時系統會自動比對，不需要另外指定）。'
        )
        ws.add_data_validation(dv)
        sub_col_letter = openpyxl.utils.get_column_letter(col_of(i) + 1)
        dv.add(f"{sub_col_letter}4:{sub_col_letter}{3 + len(plans) + DV_BUFFER_ROWS}")

    # Row3：範例/說明列（固定 SAMPLE-DEMO 前綴防呆，批次匯入時一律跳過，比照全系統慣例）
    c = ws.cell(3, 1, 'SAMPLE-DEMO-0001（此列為範例，請刪除後填入實際資料）')
    c.fill = nfill
    c.font = nfont
    c2 = ws.cell(3, 2, '範例方案，說明：金額欄請填該財報料號品名分配到的金額(含稅)，加總須等於售價(含稅)；子料號欄選填，只有該品名底下有多色/等級變體時才需要指定，否則留空')
    c2.fill = nfill
    c2.font = nfont
    c3 = ws.cell(3, 3, 100)
    c3.fill = nfill
    c3.font = nfont
    if names:
        c5 = ws.cell(3, col_of(0), 100)
        c5.fill = nfill
        c5.font = nfont

    # Row4起：現有方案資料
    r = 4
    for p in plans:
        ws.cell(r, 1, p.get('plan_code'))
        ws.cell(r, 2, p.get('plan_name'))
        ws.cell(r, 3, p.get('price_incl'))
        ws.cell(r, 4, p.get('note'))
        amt_map = items_map.get(p['id'], {})
        sub_bucket = sub_map.get(p['id'], {})
        for i, n in enumerate(names):
            ci = col_of(i)
            name = n['report_product_name']
            v = amt_map.get(name)
            if v is not None:
                ws.cell(r, ci, v)
            sub = sub_bucket.get(name)
            if sub:
                ws.cell(r, ci + 1, sub)
        r += 1

    widths = [22, 22, 12, 20]
    for _ in names:
        widths += [16, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
    ws.freeze_panes = 'E4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='書店方案清單.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 書店方案清單 — 批次匯入（新增方案 + 整筆覆蓋既有方案，依方案編號比對；
#    單列若拆帳金額加總與售價不符則該列跳過並回報，不中斷其他列，比照全系統批次匯入慣例） ──
@dealer_bp.route('/api/bookstore/plans/batch-import', methods=['POST'])
@bookstore_action_required('mod_bookstore_plan_batch_import')
def batch_import_bookstore_plans():
    """
    ⚠️ 效能注意（2026-08-08修正）：這裡的檔案常常是整份「方案」主檔（實測使用者上傳過1600+列、
    46欄的檔案），如果每一列各自呼叫Supabase 2~3次（新增/更新方案主檔 + 刪除舊明細 + 新增新明細），
    1600列會變成3000~5000次逐一往返的HTTP請求，實測會讓整個請求卡在「處理中」數分鐘，最終被
    Render/瀏覽器判定逾時、回傳非JSON內容（前端顯示「伺服器回傳非預期格式」）。修正後改成：
    先在記憶體裡把整份檔案解析完、分類成「要新增」「要更新」兩組，再各自用少數幾次「整批」
    insert/delete/insert（更新方案主檔改用整批upsert，見下方）呼叫處理完，把資料庫來回次數
    從 O(列數) 降到 O(列數/200) 左右。

    ⚠️ 匯入規則（2026-08-08調整）：「拆帳金額加總」與「售價(含稅)」不符的列**不會被跳過**，
    仍照樣新增/更新（含拆帳明細），差額只記錄在回應的 mismatched 供這次匯入結果參考——
    使用者要的是先全部匯進去，之後在方案清單頁面依紅色「差額」徽章自行找出來批次修正，
    不是匯入時就整批擋下。
    """
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{e}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({'error': '檔案無資料'}), 400

    # 找出欄位標題列：該列第1欄文字須包含「方案編號」（避開「範本/匯出」檔案最上方的ERP順序輔助列）
    header_ri = None
    for i, row in enumerate(rows[:5]):
        v = row[0] if row else None
        if v and '方案編號' in str(v):
            header_ri = i
            break
    if header_ri is None:
        return jsonify({'error': '找不到標題列（第1欄應為「方案編號」），請使用「範本/匯出」下載的檔案格式'}), 400

    header = [str(h).strip() if h is not None else '' for h in rows[header_ri]]
    col_name = col_price = col_note = None
    name_cols_raw = {}  # 欄位索引(0-based) -> Excel欄位標題文字（原始，未正規化）
    # 2026-08-18新增：子料號欄位（見export_bookstore_plans()的_BP_SUBCODE_SUFFIX說明），
    # 標題格式固定是「{財報料號品名} - 子料號」，先獨立收集，不跟金額欄混在一起判斷。
    subcode_cols_raw = {}  # 欄位索引(0-based) -> 對應的財報料號品名原始文字（未正規化，已去除後綴）
    for ci, h in enumerate(header):
        if not h or h == '方案編號':
            continue
        elif h == '方案中文名稱':
            col_name = ci
        elif '售價' in h:
            col_price = ci
        elif h == '備註':
            col_note = ci
        elif h.endswith(_BP_SUBCODE_SUFFIX):
            subcode_cols_raw[ci] = h[:-len(_BP_SUBCODE_SUFFIX)]
        else:
            name_cols_raw[ci] = h

    # ⚠️ 空白字元容錯（2026-08-08新增，防禦性保險機制，非本次已確認發生的問題）：
    # 若使用者日後手動複製/編輯欄位標題文字，可能不小心跟「內部料號清單」目前的財報料號
    # 品名之間出現全形/半形空格數量不同的情況，若只用完全相等比對，這種欄位會被誤判成
    # 「不存在的財報料號品名」整欄跳過，造成金額整批漏匯入卻不容易發現。這裡先把兩邊都去除
    # 所有空白字元再比對，比對到後仍一律採用「內部料號清單」裡的正確全名存檔（不是Excel欄位
    # 上的原始文字），確保存進資料庫的名稱永遠跟清單一致。實際檢查使用者提供的1608列真實檔案
    # 後，確認該檔案本身並無空白差異，這次匯入失敗的確認原因是下方的效能問題（N+1查詢逾時）。
    def _name_key(s):
        return re.sub(r'\s+', '', s or '')

    valid_names = _plan_valid_report_names()
    canonical_by_key = {}
    dup_key_names = set()
    for n in valid_names:
        k = _name_key(n)
        if k in canonical_by_key and canonical_by_key[k] != n:
            dup_key_names.add(k)
        canonical_by_key[k] = n

    name_cols = {}       # 欄位索引 -> 正規化後採用的正確財報料號品名
    ignored_cols = []
    for ci, h in name_cols_raw.items():
        k = _name_key(h)
        canonical = canonical_by_key.get(k)
        if canonical and k not in dup_key_names:
            name_cols[ci] = canonical
        else:
            ignored_cols.append(h)
    ignored_cols = sorted(set(ignored_cols))

    # 子料號欄位對照到財報料號品名的邏輯跟金額欄一樣（去空白後比對），比對不到就整欄忽略
    # （併入ignored_cols一起回報，不特別區分，使用者從欄名就看得出是金額欄還是子料號欄）。
    subcode_cols = {}    # 欄位索引 -> 正規化後採用的正確財報料號品名
    for ci, h in subcode_cols_raw.items():
        k = _name_key(h)
        canonical = canonical_by_key.get(k)
        if canonical and k not in dup_key_names:
            subcode_cols[ci] = canonical
        else:
            ignored_cols.append(h + _BP_SUBCODE_SUFFIX)
    ignored_cols = sorted(set(ignored_cols))

    # 子料號還要額外驗證：必須存在於「內部料號清單」，且必須屬於這一欄對應的財報料號品名
    # 底下——比照_validate_plan_items()的驗證邏輯，但批次匯入一律「不阻擋」，查無效/不符
    # 的子料號直接忽略（不寫入該筆），記錄在subcode_ignored供這次匯入結果參考，比照這支
    # 函式一貫「先全部匯入，問題留給使用者事後從畫面上的提示自行修正」的既有慣例。
    # 2026-08-18新增：內部料號常常是人類可讀的描述文字（如「Mini+白機 S001A014W福利S」），
    # 使用者手動輸入/複製貼上到Excel時很容易多打/少打空格，若只用完全相等比對，會讓子
    # 料號誤判成「不存在」而被忽略、卻查不出原因是多了個空格。這裡比照財報料號品名
    # canonical_by_key的做法，去除空白字元後再比對，比對到後一律採用「內部料號清單」
    # 裡的正確完整字串存檔（不是使用者輸入的原始文字），確保存進資料庫的子料號永遠跟
    # 清單一致；同一組去空白後的字串若對應到2筆以上不同的內部料號（理論上少見），視為
    # 無法判斷，一併忽略。
    sub_name_by_code = _sku_report_name_by_code()
    sub_canonical_by_key = {}
    sub_dup_key_codes = set()
    for code in sub_name_by_code:
        k = _name_key(code)
        if k in sub_canonical_by_key and sub_canonical_by_key[k] != code:
            sub_dup_key_codes.add(k)
        sub_canonical_by_key[k] = code
    subcode_ignored = []  # {row, plan_code, report_product_name, sub_internal_code, reason}

    # ⚠️ 2026-08-08修正：改用 _fetch_all_rows() 分頁抓取——bookstore_plans已超過1000筆，
    # 原本不分頁的寫法會被PostgREST靜默截斷，導致部份既有方案被誤判成「新方案」，
    # 匯入時因 plan_code 唯一鍵衝突而失敗。
    existing_rows = _fetch_all_rows(lambda: sb.table('bookstore_plans').select('id,plan_code'))
    existing_map = {p['plan_code']: p['id'] for p in existing_rows}

    # 2026-08-08新增：同一次批次匯入的所有列統一蓋上同一個時間戳（batch_ts），配合
    # import_seq（該列在Excel裡的列號）供 list_bookstore_plans()／export_bookstore_plans()
    # 排序用——讓清單照使用者上傳的Excel原始順序呈現，且最新匯入的整批排在最上面。
    batch_ts = now_str()

    # ── 第一階段：純解析＋驗證（不呼叫資料庫），分類成「要新增」「要更新」「跳過」 ──
    to_create = []
    to_update = []
    seen_in_file = set()
    skipped = 0
    skipped_detail = []   # 2026-08-08新增：{row, reason}，讓「略過」的列不再是無法追查的黑盒子
    mismatched = []
    errors = []

    for ri, row in enumerate(rows[header_ri + 1:], header_ri + 2):
        # ⚠️ 2026-08-08修正：方案編號欄位在Excel裡如果是「數字」格式儲存（而非文字），
        # openpyxl讀出來會是float（例如1035會變成1035.0），原本這裡只單純str()+strip()，
        # 會把方案編號存成「1035.0」而不是「1035」。「匯入訂單」比對方案時用的
        # _norm_plan_code() 有做float轉int的防呆，兩邊字串就會對不起來，導致訂單那邊
        # 明明看得到方案清單存在，比對卻一直「查無」。這裡改用同一套 _norm_plan_code()
        # 邏輯，確保兩邊存的/比對的方案編號格式永遠一致。
        code_val = _norm_plan_code(row[0] if row else None)
        if not code_val:
            skipped += 1
            skipped_detail.append({'row': ri, 'reason': '方案編號為空白'})
            continue
        if code_val.upper().startswith('SAMPLE-DEMO'):
            skipped += 1
            skipped_detail.append({'row': ri, 'reason': '範例列（SAMPLE-DEMO），固定略過'})
            continue

        plan_name = None
        if col_name is not None and col_name < len(row) and row[col_name] is not None:
            plan_name = str(row[col_name]).strip() or None
        price_incl = _to_num_or_none(row[col_price]) if (col_price is not None and col_price < len(row)) else None
        note = None
        if col_note is not None and col_note < len(row) and row[col_note] is not None:
            note = str(row[col_note]).strip() or None

        # 2026-08-18新增：先把這一列各財報料號品名對應的子料號整理成 name -> sub_internal_code
        # 的字典，供下面組items時查詢。子料號欄本身不存在對應的金額欄名稱不會出現在name_cols
        # 裡（理論上不會發生，因為兩欄都是同一份header_map產生），這裡多做防呆用get()。
        sub_by_name = {}
        for ci, name in subcode_cols.items():
            if ci >= len(row):
                continue
            v = row[ci]
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            code_str = _norm(v) if not isinstance(v, str) else v.strip()
            k = _name_key(code_str)
            canonical_code = sub_canonical_by_key.get(k) if k not in sub_dup_key_codes else None
            if canonical_code is None:
                subcode_ignored.append({'row': ri, 'plan_code': code_val, 'report_product_name': name,
                                         'sub_internal_code': code_str, 'reason': '子料號不存在於「內部料號清單」（去除空白字元比對後仍找不到），已略過'})
                continue
            sub_name = sub_name_by_code.get(canonical_code)
            if sub_name != name:
                subcode_ignored.append({'row': ri, 'plan_code': code_val, 'report_product_name': name,
                                         'sub_internal_code': code_str,
                                         'reason': f'子料號屬於「{sub_name}」，跟欄位對應的「{name}」不一致，已略過'})
                continue
            # 一律採用「內部料號清單」裡的正確完整字串存檔，不是使用者輸入的原始文字，
            # 確保存進資料庫的子料號永遠跟清單一致（比照財報料號品名canonical化的做法）。
            sub_by_name[name] = canonical_code

        items = []
        for ci, name in name_cols.items():
            if ci >= len(row):
                continue
            v = row[ci]
            if v is None or v == '':
                continue
            amt = _to_num_or_none(v)
            if amt is None:
                continue
            items.append({'report_product_name': name, 'amount': amt, 'sub_internal_code': sub_by_name.get(name)})

        # ⚠️ 2026-08-08調整：使用者反映「金額加總跟售價不符的列不要跳過不匯入，還是要幫我
        # 匯入，我之後會依照畫面上的差額色塊自己去批次修正」——改成不管加總對不對，都照樣
        # 新增/更新這筆方案（含拆帳明細），只是把差額記錄到 mismatched 清單供這次匯入結果
        # 回報參考。方案清單頁面（loadBookstorePlans()）本來就會依 items_total 跟
        # price_incl 是否一致，自動顯示紅色「差額」徽章或綠色「OK」，不需要額外處理，
        # 使用者可以直接在清單上用顏色找出這些列，逐一或用「批次匯入」再修正一次。
        if price_incl is not None and items:
            items_total = round(sum(it['amount'] for it in items), 2)
            diff = round(items_total - float(price_incl), 2)
            if abs(diff) > 0.5:
                mismatched.append({'row': ri, 'plan_code': code_val, 'items_total': items_total,
                                    'price_incl': price_incl, 'diff': diff})

        if code_val in seen_in_file:
            skipped += 1
            errors.append({'row': ri, 'plan_code': code_val, 'error': '同一檔案內方案編號重複，僅處理第一筆'})
            skipped_detail.append({'row': ri, 'reason': f'方案編號「{code_val}」在檔案內重複，僅處理第一筆出現的那一列'})
            continue
        seen_in_file.add(code_val)

        rec = {'plan_code': code_val, 'plan_name': plan_name, 'price_incl': price_incl, 'note': note,
               'import_seq': ri, 'last_imported_at': batch_ts}
        if code_val in existing_map:
            to_update.append({'pid': existing_map[code_val], 'rec': rec, 'items': items})
        else:
            to_create.append({'rec': rec, 'items': items})

    created = 0
    updated = 0
    CHUNK = 200

    # ── 第二階段：新增（整批insert方案主檔取回id，再整批insert拆帳明細） ──
    for i in range(0, len(to_create), CHUNK):
        chunk = to_create[i:i + CHUNK]
        recs = []
        for c in chunk:
            r = dict(c['rec'])
            r.update(_audit_new())
            recs.append(r)
        try:
            ins = sb.table('bookstore_plans').insert(recs).execute()
        except Exception as e:
            for c in chunk:
                errors.append({'row': None, 'plan_code': c['rec']['plan_code'], 'error': str(e)})
            continue
        code_to_id = {row2['plan_code']: row2['id'] for row2 in (ins.data or [])}
        created += len(ins.data or [])
        item_rows = []
        for c in chunk:
            pid = code_to_id.get(c['rec']['plan_code'])
            if not pid:
                continue
            for it in c['items']:
                item_rows.append({'plan_id': pid, 'report_product_name': it['report_product_name'],
                                   'amount': it['amount'], 'sub_internal_code': it.get('sub_internal_code')})
        for j in range(0, len(item_rows), CHUNK):
            sb.table('bookstore_plan_items').insert(item_rows[j:j + CHUNK]).execute()

    # ── 第三階段：更新 ──
    # ⚠️ 2026-08-08修正：原本這裡對每一筆要更新的方案各自呼叫一次
    # `.update().eq('id', pid).execute()`，1608列裡若大部分是「已存在、要覆蓋」的方案
    # （例如同一份清單重新上傳做修正），就會變成上千次逐筆PATCH，跟原始版本一樣造成逾時
    # ——[[v2.92]]的效能修正只處理到「新增」跟「拆帳明細刪除/新增」兩段，遺漏了「方案主檔
    # 更新」這一段，導致使用者實測時仍卡在匯入中（截圖顯示卡在121筆左右）。
    # 修正做法：改用 `upsert(..., on_conflict='plan_code')` 整批送出（見 supabase_client.py
    # 新增的 on_conflict 參數），依 plan_code 做 `INSERT ... ON CONFLICT (plan_code) DO
    # UPDATE`；因為這裡的方案編號已確認全部存在（existing_map查過），實際上永遠只會走
    # UPDATE分支，等於把「整批更新」用「整批upsert」的方式一次送出，跟新增方案一樣
    # CHUNK=200一批。故意不把 created_by/created_by_name 放進payload，PostgREST的
    # merge-duplicates只會覆蓋payload裡有出現的欄位，不會動到原始建立者資訊。
    update_pids = [u['pid'] for u in to_update]
    for i in range(0, len(update_pids), CHUNK):
        sb.table('bookstore_plan_items').delete().in_('plan_id', update_pids[i:i + CHUNK]).execute()

    for i in range(0, len(to_update), CHUNK):
        chunk = to_update[i:i + CHUNK]
        recs = []
        for u in chunk:
            r = dict(u['rec'])
            r.update(_audit_upd())
            recs.append(r)
        try:
            up = sb.table('bookstore_plans').upsert(recs, on_conflict='plan_code').execute()
            updated += len(up.data or [])
        except Exception as e:
            for u in chunk:
                errors.append({'row': None, 'plan_code': u['rec']['plan_code'], 'error': str(e)})

    all_update_items = []
    for u in to_update:
        for it in u['items']:
            all_update_items.append({'plan_id': u['pid'], 'report_product_name': it['report_product_name'],
                                      'amount': it['amount'], 'sub_internal_code': it.get('sub_internal_code')})
    for i in range(0, len(all_update_items), CHUNK):
        sb.table('bookstore_plan_items').insert(all_update_items[i:i + CHUNK]).execute()

    return jsonify({
        'ok': True, 'created': created, 'updated': updated, 'skipped': skipped,
        'mismatched': mismatched, 'errors': errors, 'ignored_columns': ignored_cols,
        'skipped_detail': skipped_detail,
        # 2026-08-18新增：子料號驗證失敗(不存在/跟品名不符)的清單，該筆金額仍照樣匯入，
        # 只有子料號本身被忽略，比照這支函式一貫「不阻擋，問題留給使用者事後修正」的慣例。
        'subcode_ignored': subcode_ignored,
    })


# ── 書店方案清單 — 簡易批次新增（2026-08-12新增，獨立於上面的「批次匯入」之外）──
# 背景：既有「批次匯入」限管理員使用，欄位很多（每個財報料號品名都要填拆帳金額），對
# 只是想快速登記一批新方案編號/名稱/售價、之後再慢慢補拆帳明細的同仁來說操作門檻較高。
# 這裡另外開一個「簡易批次新增」，只認3欄（方案編號/方案中文名稱/售價(含稅)），不處理
# 拆帳明細，權限也獨立設定（mod_bookstore_plan_simple_import，可以只開放給非管理員同仁
# 使用，不需要連帶開放上面那個完整版批次匯入）。方案編號跟系統既有資料重複時，該列跳過、
# 其他列照常新增（不覆蓋既有資料），比照「匯入團體發票」批次匯入(v1.20)的dup處理慣例——
# 這裡刻意不做「更新」分支，理由是簡易新增的定位是「快速登記新方案」，不是用來覆蓋既有
# 方案資料，要修改既有方案請用「編輯」或既有的完整版批次匯入。
@dealer_bp.route('/api/bookstore/plans/simple-template', methods=['GET'])
@bookstore_action_required('mod_bookstore_plan_simple_import')
def simple_template_bookstore_plans():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '方案簡易批次新增'
    headers = ['方案編號', '方案中文名稱', '售價(含稅)']
    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    nfill = openpyxl.styles.PatternFill('solid', fgColor='F2F2F2')
    nfont = openpyxl.styles.Font(italic=True, color='888888')
    center = openpyxl.styles.Alignment(horizontal='center')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hfill
        c.font = hfont
        c.alignment = center
    c = ws.cell(2, 1, 'SAMPLE-DEMO-0001（此列為範例，請刪除後填入實際資料）')
    c.fill = nfill
    c.font = nfont
    c2 = ws.cell(2, 2, '範例方案，此功能只登記方案編號/名稱/售價，不含拆帳明細')
    c2.fill = nfill
    c2.font = nfont
    c3 = ws.cell(2, 3, 100)
    c3.fill = nfill
    c3.font = nfont
    widths = [22, 30, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
    ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='方案簡易批次新增範本.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@dealer_bp.route('/api/bookstore/plans/simple-batch-import', methods=['POST'])
@bookstore_action_required('mod_bookstore_plan_simple_import')
def simple_batch_import_bookstore_plans():
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{e}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({'error': '檔案無資料'}), 400

    header_ri = None
    for i, row in enumerate(rows[:5]):
        v = row[0] if row else None
        if v and '方案編號' in str(v):
            header_ri = i
            break
    if header_ri is None:
        return jsonify({'error': '找不到標題列（第1欄應為「方案編號」），請使用「下載範本」的檔案格式'}), 400

    header = [str(h).strip() if h is not None else '' for h in rows[header_ri]]
    col_name = col_price = None
    for ci, h in enumerate(header):
        if h == '方案中文名稱':
            col_name = ci
        elif '售價' in h:
            col_price = ci

    # 同理改用 _fetch_all_rows() 分頁抓取全部既有方案編號，避免bookstore_plans超過1000筆時
    # 被PostgREST靜默截斷、誤判既有方案編號為「不存在」。
    existing_rows = _fetch_all_rows(lambda: sb.table('bookstore_plans').select('plan_code'))
    existing_codes = {p['plan_code'] for p in existing_rows}

    batch_ts = now_str()
    to_create = []
    seen_in_file = set()
    skip_cnt = 0
    err_cnt = 0
    dup_cnt = 0
    dup_detail = []
    skipped_detail = []

    for ri, row in enumerate(rows[header_ri + 1:], header_ri + 2):
        code_val = _norm_plan_code(row[0] if row else None)
        if not code_val:
            skip_cnt += 1
            skipped_detail.append({'row': ri, 'reason': '方案編號為空白'})
            continue
        if code_val.upper().startswith('SAMPLE-DEMO'):
            skip_cnt += 1
            skipped_detail.append({'row': ri, 'reason': '範例列（SAMPLE-DEMO），固定略過'})
            continue
        if code_val in existing_codes or code_val in seen_in_file:
            dup_cnt += 1
            dup_detail.append({'row': ri, 'plan_code': code_val})
            continue

        plan_name = None
        if col_name is not None and col_name < len(row) and row[col_name] is not None:
            plan_name = str(row[col_name]).strip() or None
        if not plan_name:
            err_cnt += 1
            skipped_detail.append({'row': ri, 'reason': f'方案編號「{code_val}」缺少方案中文名稱'})
            continue
        price_incl = _to_num_or_none(row[col_price]) if (col_price is not None and col_price < len(row)) else None
        if price_incl is None:
            err_cnt += 1
            skipped_detail.append({'row': ri, 'reason': f'方案編號「{code_val}」缺少售價(含稅)或格式非數字'})
            continue

        seen_in_file.add(code_val)
        to_create.append({
            'plan_code': code_val, 'plan_name': plan_name, 'price_incl': price_incl,
            'import_seq': ri, 'last_imported_at': batch_ts,
        })

    created = 0
    CHUNK = 200
    for i in range(0, len(to_create), CHUNK):
        chunk = to_create[i:i + CHUNK]
        recs = []
        for rec in chunk:
            r = dict(rec)
            r.update(_audit_new())
            recs.append(r)
        try:
            ins = sb.table('bookstore_plans').insert(recs).execute()
            created += len(ins.data or [])
        except Exception as e:
            for rec in chunk:
                skipped_detail.append({'row': None, 'reason': f'方案編號「{rec["plan_code"]}」新增失敗：{e}'})
                err_cnt += 1

    return jsonify({
        'ok': True, 'created': created, 'skip': skip_cnt, 'error': err_cnt, 'dup': dup_cnt,
        'dup_detail': dup_detail, 'skipped_detail': skipped_detail,
    })


# ============================================================
# ⑨ 書店經銷商 — 訂單匯入拆帳分析（境內/境外分帳 + 未稅單價 vs 進貨價 檢核）
# ============================================================
# 2026-08-07新增：書店原始需求裡「依附件[方案]拆帳分析成頁籤[0701-0727A88]」的核心功能——
# 上傳當月訂單Excel後，依訂單裡的「方案」對應到「方案清單」的拆帳明細，把每筆訂單換算成
# 未稅單價（拆帳金額(含稅)×訂單數量÷1.05），依訂單的「運送地區中文」分成境內(台灣)/境外
# (非台灣)兩個區塊，每一筆訂單各佔一列，欄位＝財報料號品名；若某財報品名的未稅單價低於
# 「內部料號清單」裡對應的商品進貨未稅價，該格標記為異常（前端表格與匯出Excel的「書店訂單
# 拆帳」明細頁籤都會標紅；2026-08-11起「總表分析」樞紐彙總頁籤的單價儲存格也會標紅，之前
# 這個頁籤完全沒有做這個比對，只把進貨價印出來當參考，見_compute_bookstore_order_pivot()）。
# 這裡不寫入資料庫，是每次上傳訂單檔即時計算的唯讀分析結果，換算邏輯比照Phase1（先前交付
# 的獨立網頁工具版本）已用真實資料驗證過的公式：未稅單價=拆帳金額(含稅)÷1.05；
# 訂單金額=方案售價×數量。
#
# ⚠️ 進貨未稅價比較規則：使用者確認「同一個財報料號品名底下，內部料號進貨未稅價都一致」，
# 所以直接取該財報品名任一筆內部料號的進貨未稅價即可，不需要取最高/最低/平均值。仍加了一層
# 防呆偵測：若實際資料出現同一財報品名底下進貨價不一致的情況，會回報在 inconsistent_cost_names，
# 提示使用者去「內部料號清單」核對，而不是靜默選錯一個值導致標紅結果不可信。


def _order_col_map(headers):
    """依標題文字比對訂單Excel欄位，回傳 {欄位鍵: 0-based欄位索引}（比對邏輯與批次匯入的
    _map_h() 同樣走「標題文字比對」而非固定欄位順序，避免使用者調整過欄位順序就解析失敗）。
    2026-08-08修正：使用者確認「出貨單號」這欄對分析/匯入沒有意義，不再解析此欄；改為
    新增解析「運費」欄位——「金額」「運送地區中文」「運費」這3個原始欄位會原樣保留到
    分析結果裡（供使用者後續肉眼核對用，不參與任何計算）。
    2026-08-08再修正：使用者要求再保留「折價券」「訂單備註」「購物車折扣的金額」這3個
    原始欄位以利驗證，比照前一批原始保留欄位的做法，同樣原樣保留、不參與任何計算。"""
    m = {}
    for ci, h in enumerate(headers):
        h = (h or '').strip()
        if not h:
            continue
        if h == '訂單編號':
            m['order_no'] = ci
        elif h == '訂購日期':
            m['order_date'] = ci
        elif h == '方案中文':
            m['plan_name_ref'] = ci
        elif h == '方案':
            m['plan_code'] = ci
        elif h == '數量':
            m['qty'] = ci
        elif h == '金額':
            m['amount_ref'] = ci
        elif h == '運費':
            m['freight_ref'] = ci
        elif h == '折價券':
            m['coupon_ref'] = ci
        elif h == '訂單備註':
            m['order_note_ref'] = ci
        elif h == '購物車折扣的金額':
            m['cart_discount_ref'] = ci
        elif '運送地區' in h:
            m['region_raw'] = ci
    return m


# 2026-08-12修正：使用者提供實際訂單備註截圖，格式是「配件折 -NT$72元」——「折」後面
# 還夾了空白＋負號(-)才接到「NT$金額元」，原本的正規表示式沒有考慮負號，抓不到這種格式，
# 整批都落到「折扣待複核」工作表。使用者確認「這種文字描述，統一都是折到配件，對應該筆
# 方案的料號去折價」——也就是折扣目標維持原邏輯不變(該訂單列自己方案底下的料號，不用
# 改動_compute_bookstore_order_matrix()的折扣分攤邏輯)，只需要修正這個正規表示式本身，
# 讓它能認出「折」跟金額中間夾雜負號/貨幣符號的格式。
_ORDER_NOTE_DISCOUNT_RE = re.compile(r'折\s*-?\s*(?:NT\$|NTD|\$)?\s*(\d+(?:\.\d+)?)\s*元')


def _parse_order_discount_incl(order_note_ref, coupon_ref):
    """2026-08-12新增，2026-08-14修正：解析「訂單備註(原始)」「折價券(原始)」這2個原始
    欄位裡的折扣金額（皆為含稅金額，跟訂單「金額」欄位同基礎）。
    - 訂單備註：使用者確認格式是「XXX折xxx元」（例："配件折201元"），用正規表示式抓
      「折」後面緊接的數字＋「元」，例如「配件折201元」→201。只抓得到「折(數字)元」這種
      緊鄰格式，像「79折 -NT$140元」這種「折」和金額中間隔了其他文字的格式抓不到——寧可
      漏抓（不折扣）也不要抓錯（誤折扣），漏抓的訂單會出現在「未折抵訂單備註」清單裡供
      使用者人工複核。
    - 2026-08-14修正：原本用.search()只抓字串中第一個「折xxx元」，若同一筆備註裡有
      2個以上「折xxx元」片語（例如訂單685399備註「白筆加購599元,配件折201元,會員日
      配件折 -NT$131元」，同時有「配件折201元」跟「會員日配件折-131元」2筆折扣），
      只會抓到第一筆201，第二筆131被靜默漏掉，造成該筆訂單少扣131元折扣、對帳出現
      131元落差(這正是使用者回報的境內總額125元落差的主要成因，經逐筆比對「書店訂單
      拆帳」驗證金額欄後定位到此)。修正為用.findall()抓出備註裡所有「折xxx元」片語，
      全部加總扣除，不再只抓第一筆。
    - 折價券：使用者確認這欄本身就是數值，有值就直接當折扣金額用。
    - 若2欄都有值，視為2筆各自獨立的折扣，加總一起扣。
    回傳 (discount_incl, matched) — discount_incl是折扣金額(含稅，尚未除1.05)，
    matched是True代表有從備註文字或折價券欄位抓到折扣（給呼叫端判斷要不要列入複核清單）。"""
    discount = 0.0
    matched = False
    if order_note_ref:
        matches = _ORDER_NOTE_DISCOUNT_RE.findall(str(order_note_ref))
        if matches:
            discount += sum(float(x) for x in matches)
            matched = True
    coupon_amt = _to_num_or_none(coupon_ref)
    if coupon_amt:
        discount += coupon_amt
        matched = True
    return discount, matched


def _norm_plan_code(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _norm_date_display(v):
    if v is None:
        return ''
    try:
        return v.strftime('%Y-%m-%d')
    except AttributeError:
        return str(v)


def _compute_bookstore_order_matrix(rows, col_map):
    # ⚠️ 2026-08-08修正：改用 _fetch_all_rows() 分頁抓取，見該函式註解——bookstore_plans
    # 現已超過1000筆，原本不分頁的 select('*').execute() 會被 PostgREST 靜默截斷，導致
    # 較晚建立/匯入、id較大的方案（例如方案編號1035）完全抓不到，「匯入訂單」比對時就會
    # 誤判成「查無對應方案」——即使該方案在「方案清單」頁面上明明查得到。
    plans_rows = _fetch_all_rows(lambda: sb.table('bookstore_plans').select('*'))
    plans_by_code = {p['plan_code']: p for p in plans_rows}
    plan_ids = [p['id'] for p in plans_by_code.values()]
    items_map = {}
    if plan_ids:
        items_rows = _fetch_all_rows(lambda: sb.table('bookstore_plan_items').select('*').in_('plan_id', plan_ids))
        for it in items_rows:
            items_map.setdefault(it['plan_id'], []).append(it)

    # 2026-08-08修正：欄位改成固定使用「內部料號清單」裡全部啟用中（未隱藏）的財報料號
    # 品名，不再只列出「這次上傳的訂單裡剛好比對到」的品名——使用者要求「依內部料號清單，
    # 沒有隱藏的標題，皆完整呈現」，比照[[v2.96]]方案清單/匯出欄位的隱藏規則保持一致：
    # 同一品名底下只要還有一筆內部料號顯示中(is_active!=false)就算顯示，全部被設為隱藏
    # 才排除。即使某品名目前被隱藏、不出現在固定欄位清單裡，該品名的拆帳金額仍會照算、
    # 計入「合計(未稅)」欄，只是不會顯示成獨立的一欄，避免「隱藏」影響到金額正確性。
    sku_rows = _fetch_all_rows(lambda: sb.table('internal_sku_list')
                                .select('report_product_name,purchase_price_notax,sort_order,is_active,is_accessory')
                                .order('sort_order'))
    cost_by_name = {}
    inconsistent_names = set()
    active_by_name = {}
    # 2026-08-12新增：「配件」品名集合，見下方Pass B的折扣分攤規則說明。跟active_by_name
    # 一樣採「OR」寬鬆判斷——同一財報料號品名底下只要有任一內部料號勾選「配件」，這個品名
    # 就算配件（理論上同一品名底下的配件屬性應該一致，這裡不特別檢查是否所有列都一致）。
    accessory_by_name = {}
    for s in sku_rows:
        name = s.get('report_product_name')
        if not name:
            continue
        cost = _to_num_or_none(s.get('purchase_price_notax'))
        if name not in cost_by_name:
            cost_by_name[name] = cost
        elif cost is not None and cost_by_name[name] is not None and abs(cost - cost_by_name[name]) > 0.01:
            inconsistent_names.add(name)
        if s.get('is_active') is not False:
            active_by_name[name] = True
        elif name not in active_by_name:
            active_by_name[name] = False
        if s.get('is_accessory'):
            accessory_by_name[name] = True
        elif name not in accessory_by_name:
            accessory_by_name[name] = False

    columns_order = []
    seen_cols = set()
    for s in sku_rows:
        name = s.get('report_product_name')
        if name and name not in seen_cols and active_by_name.get(name):
            seen_cols.add(name)
            columns_order.append(name)

    unmatched = []

    # ── Pass A：逐列讀取方案比對/品名彙總/這一列自己的折扣文字解析結果，先不決定折扣真正
    # 要扣在哪一列——同一張訂單編號底下可能有多列（見下方Pass B），要先把同一訂單的所有列
    # 收集齊全才能判斷。row_ctxs 保留原始列順序，供Pass C按原順序組出out_rows。
    row_ctxs = []

    for row in rows:
        def _get(key):
            ci = col_map.get(key)
            if ci is None or ci >= len(row):
                return None
            return row[ci]

        order_no = _get('order_no')
        if order_no is None or str(order_no).strip() == '':
            continue  # 空列略過
        plan_code = _norm_plan_code(_get('plan_code'))
        qty = _to_num_or_none(_get('qty')) or 0
        region_raw = _get('region_raw')
        region = '境內' if (region_raw and str(region_raw).strip() == '台灣') else '境外'

        plan = plans_by_code.get(plan_code)
        if not plan:
            unmatched.append({
                'order_no': str(order_no), 'plan_code': plan_code,
                'plan_name_ref': _get('plan_name_ref'), 'qty': qty, 'region': region,
            })
            continue

        items = items_map.get(plan['id'], [])
        # 2026-08-11新增：同一方案裡同一財報料號品名現在可以有多列——這代表「該方案每賣1份，
        # 這個料號實際出貨的數量」，不是同一顆料號的另一種計價方式。例如方案2372把
        # S002A019W-Mini透明殼拆成3列各351元，意思是這個方案每賣1份，透明殼要出貨3個、
        # 每個未稅單價334.29，不是「出貨1個、未稅單價1002.86」。所以這裡不能把同名列的
        # amount加總成一個數字再算單價（那樣「總表分析」頁籤的單價/數量會失真，跟實際出貨
        # 數量不符）——改成保留每一列自己的單價，彙總到 sub_items，「總表分析」用的樞紐
        # 函式 _compute_bookstore_order_pivot() 會依這裡的sub_items把數量依重複列數乘開。
        # 「書店訂單拆帳」明細表跟前端網頁分析結果那一格顯示的仍是「這筆訂單這個品名的未稅
        # 總金額」（cells[name]['untaxed']），不受這次改法影響，金額打的仍然是所有列加總。
        items_by_name = {}
        for it in items:
            name = it['report_product_name']
            items_by_name.setdefault(name, []).append(_to_num(it['amount']))

        order_note_ref = _get('order_note_ref')
        coupon_ref = _get('coupon_ref')
        row_discount_incl, row_discount_matched = _parse_order_discount_incl(order_note_ref, coupon_ref)
        row_total_amt_incl = sum(sum(amounts) * qty for amounts in items_by_name.values())
        # 2026-08-12新增：這一列的品名是否「全部都是配件」——判斷這一列有沒有資格被Pass B
        # 選為「同一訂單裡單價最高的配件列」，見下方分組邏輯。沒有任何品名(items_by_name
        # 為空)不算配件列。
        is_accessory_line = bool(items_by_name) and all(accessory_by_name.get(n, False) for n in items_by_name)

        row_ctxs.append({
            'order_no': str(order_no), 'order_date': _norm_date_display(_get('order_date')),
            'plan_code': plan_code, 'plan': plan, 'qty': qty, 'region': region,
            'items_by_name': items_by_name, 'row_total_amt_incl': row_total_amt_incl,
            'is_accessory_line': is_accessory_line,
            'row_discount_incl': row_discount_incl, 'row_discount_matched': row_discount_matched,
            'order_note_ref': order_note_ref, 'coupon_ref': coupon_ref,
            'amount_ref': _to_num_or_none(_get('amount_ref')), 'region_raw': region_raw,
            'freight_ref': _to_num_or_none(_get('freight_ref')),
            'cart_discount_ref': _to_num_or_none(_get('cart_discount_ref')),
            # 下面2個欄位由Pass B決定，先給預設值：
            'discount_incl': 0.0, 'discount_review_reasons': [],
            # 2026-08-13新增：下面這2個欄位由Pass B2（運費拆帳）決定，先給預設值：
            'freight_alloc_untaxed': 0.0, 'freight_review_reasons': [],
        })

    # ── Pass B：2026-08-12新增，依「訂單編號」分組，整張訂單只認列一次折扣金額，不再逐列
    # 各自解析/各自套用。
    #
    # 背景（使用者截圖範例，訂單679798）：來源訂單檔裡，同一張訂單編號拆成多列（每個方案
    # 各一列），「訂單備註」欄位內容是整張訂單共用、被複製貼到每一列（例如「會員日配件折
    # -NT$226元」同時出現在同一訂單編號下的3列）——這代表整張訂單只有「一筆」226元折扣，
    # 不是每一列各自一筆226元。舊版本沒有依訂單編號分組，逐列各自解析/套用，會讓同一筆
    # 折扣被多列各自扣抵一次，造成折扣總額被重複扣掉（超過訂單實際折扣金額）。
    #
    # 使用者確認的分攤規則：這筆訂單編號底下的商品如果「全部都是配件」，折扣整筆扣在單價
    # (每1份含稅金額，未乘數量)最高的那一列，其他列不折扣；訂單裡如果有非配件商品（例如
    # 混搭電子書借閱方案），折扣一樣只在「配件」品項裡挑單價最高的一筆扣（也就是候選列一律
    # 限定為「全部品名都是配件」的列，不論訂單裡是否還有其他非配件的列）。「配件」判斷依據
    # 「內部料號清單」的is_accessory勾選欄位（見dealer_reconcile_schema.sql v1.25），不用
    # 品名關鍵字猜測。
    #
    # 如果整張訂單完全沒有配件列（理論上少見，例如整張訂單都是電子書方案），退回舊做法：
    # 依訂單裡各列含稅金額佔比分攤整筆折扣，避免完全不折扣。
    groups = {}
    group_order = []
    for ctx in row_ctxs:
        key = ctx['order_no']
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(ctx)

    for key in group_order:
        group = groups[key]
        # 訂單層級折扣金額：同一訂單編號底下各列各自解析到的折扣金額，正常情況下應該完全
        # 相同（來源文字重複貼在每一列）；若解析出2個以上不同的數字，代表資料有異常，列入
        # 複核、不自動折抵，避免抓錯金額誤折扣（寧可漏折後人工複核，不要抓錯後誤折扣，比照
        # _parse_order_discount_incl()既有原則）。
        discount_values = sorted({round(c['row_discount_incl'], 2) for c in group if c['row_discount_incl']})
        any_matched = any(c['row_discount_matched'] for c in group)
        any_note_has_char = any(c['order_note_ref'] and '折' in str(c['order_note_ref']) for c in group)

        group_review_reasons = []
        discount_incl = 0.0
        if len(discount_values) > 1:
            group_review_reasons.append('同一訂單編號下各列解析到的折扣金額不一致，請人工確認')
        elif len(discount_values) == 1:
            discount_incl = discount_values[0]

        if any_note_has_char and not any_matched:
            group_review_reasons.append('訂單備註含「折」字但抓不到「折(數字)元」格式，請人工確認折扣金額')

        if discount_incl:
            candidates = [c for c in group if c['is_accessory_line']]
            if candidates:
                # 2026-08-12第六輪修正（訂單686146為例：配件2637單價99，折扣108，配件單價
                # 不夠扣）——依單價(每1份含稅金額，未乘數量)高到低排序，逐一嘗試找「第一筆
                # 金額能單獨覆蓋這筆折扣」的配件列；tie時取先出現的列。
                ordered_candidates = sorted(
                    candidates,
                    key=lambda c: sum(sum(a) for a in c['items_by_name'].values()),
                    reverse=True,
                )
                chosen = None
                for c in ordered_candidates:
                    amt = c['row_total_amt_incl'] or 0
                    if discount_incl <= amt + 0.01:
                        chosen = c
                        break
                if chosen is None:
                    # 配件列全部單獨都扣不完這筆折扣，改挑「非配件」品項裡單價最高、且金額
                    # 能單獨覆蓋這筆折扣的那一筆（使用者原話：「該筆訂單無高單價的配件可折，
                    # 要折的配件單價金額小於折扣價，則改折另一個商品內」）。
                    ordered_non_candidates = sorted(
                        (c for c in group if not c['is_accessory_line']),
                        key=lambda c: c['row_total_amt_incl'] or 0,
                        reverse=True,
                    )
                    for c in ordered_non_candidates:
                        amt = c['row_total_amt_incl'] or 0
                        if discount_incl <= amt + 0.01:
                            chosen = c
                            break
                if chosen is not None:
                    chosen['discount_incl'] = discount_incl
                else:
                    # 訂單裡沒有任何一筆(配件或非配件)單獨扣得完這筆折扣，退回依各列含稅
                    # 金額佔比分攤整張訂單的折扣（比照下方「整張訂單沒有配件列」的做法）。
                    group_total = sum(c['row_total_amt_incl'] for c in group)
                    if group_total and discount_incl <= group_total + 0.01:
                        for c in group:
                            c['discount_incl'] = (discount_incl * c['row_total_amt_incl'] / group_total) if group_total else 0
                    else:
                        group_review_reasons.append('折扣金額超過此訂單全部料號含稅總額，無法自動折抵，請人工確認')
            else:
                # 整張訂單沒有配件列可挑，退回依各列含稅金額佔比分攤整筆折扣（沿用舊版單列
                # 邏輯的比例分攤算法，只是改成在整張訂單的範圍內做一次，不是逐列各自做）。
                group_total = sum(c['row_total_amt_incl'] for c in group)
                if group_total and discount_incl <= group_total + 0.01:
                    for c in group:
                        c['discount_incl'] = (discount_incl * c['row_total_amt_incl'] / group_total) if group_total else 0
                else:
                    group_review_reasons.append('折扣金額超過此訂單全部料號含稅總額，無法自動折抵，請人工確認')

        if group_review_reasons:
            for c in group:
                c['discount_review_reasons'] = list(group_review_reasons)

    # ── Pass B2：2026-08-13新增、同日兩度再修正，運費拆帳。使用者最終確認的規則：**只要
    # 這筆訂單編號的運費有值，不論買了1個還是多個商品，運費都要整筆拆進「這筆訂單商品
    # 價格最高那一筆方案」的電子書方案欄位**——不是像折扣那樣依比例分攤到多列/多品名。
    # （2026-08-13第一版曾誤以為只有「買多商品」的訂單才要拆，使用者說明後才確認：
    # 單筆訂單只買1個方案，一樣要把運費拆進這唯一一列的電子書方案，不是「運費已經反映
    # 在這一列」——運費本來就不屬於任何商品金額，原始「金額」欄位不含運費，所以即使只有
    # 1列，也需要額外把運費併進去，不能假設它已經包含在裡面。）
    # 「商品價格最高」比照使用者畫面上實際看得到的「金額(原始，含稅)」欄位（來源訂單檔
    # 案裡原始的「金額」欄，即這裡的amount_ref）逐列比較，取最大值那一列（單筆訂單就是
    # 這唯一一列）；tie時取先出現的列。若某列這欄剛好是空值，退回用這一列所有品名含稅
    # 總額(row_total_amt_incl)比較，避免完全無法比較而整筆略過。
    # 2026-08-13第三度修正（使用者截圖回報境外「電子書方案」欄位算出來的數字對不起來，
    # 例如訂單681615應該是800+1550=2350卻只顯示2312）：原本這裡「境內除以1.05換算未稅、
    # 境外不除1.05直接當未稅用」的做法，混淆了「資料層」跟「顯示層」兩件事——
    # cells[name]['untaxed']這個欄位的資料契約（見Pass C下面商品本身的換算，一律/1.05）
    # 本來就應該是「真未稅」，不分境內境外都一樣；境外要呈現含稅，是在
    # export_bookstore_orders()匯出／顯示的那一層才把整格再乘回1.05（跟「總表分析」樞紐
    # 既有的境外顯示邏輯一致），不應該讓運費在資料層就搶先跳過未稅換算，否則顯示層再統一
    # ×1.05一次時，運費會比商品本身多乘一次1.05、多算了5%。改成不分境內境外，運費一律
    # 先換算成真未稅金額，讓顯示層可以用同一套境內/境外轉換規則，套用在整個儲存格（商品+
    # 運費）上，不用再為運費另開特例。
    for key in group_order:
        group = groups[key]
        freight_values = sorted({round(c['freight_ref'], 2) for c in group if c['freight_ref']})
        if not freight_values:
            continue  # 這筆訂單沒有運費
        if len(freight_values) > 1:
            # 同一訂單編號底下各列解析到的運費金額本應完全相同（來源檔案把運費複製貼到
            # 每一列），若出現不一致代表資料有異常，列入複核、不自動拆運費，避免抓錯金額。
            reason = '同一訂單編號下各列的運費金額不一致，運費未自動拆進電子書方案，請人工確認'
            for c in group:
                c['freight_review_reasons'] = c['freight_review_reasons'] + [reason]
            continue
        freight_ref = freight_values[0]

        def _line_price(c):
            return c['amount_ref'] if c['amount_ref'] is not None else (c['row_total_amt_incl'] or 0)

        target = max(group, key=_line_price)
        # 不分境內境外，運費一律除以1.05換算成真未稅金額再併入cells[name]['untaxed']，
        # 跟商品本身的未稅換算基礎保持一致（境內/境外的顯示差異，交給匯出層統一處理）。
        target['freight_alloc_untaxed'] = round(freight_ref / 1.05, 2)

    # ── Pass C：依原始列順序，用Pass B分配好的discount_incl算出每一列的cells（未稅單價/
    # 低於成本檢核），算法跟舊版完全相同，只是discount_incl的來源改成Pass B的分配結果。
    out_rows = []
    for ctx in row_ctxs:
        qty = ctx['qty']
        items_by_name = ctx['items_by_name']
        discount_incl = ctx['discount_incl']
        row_total_amt_incl = ctx['row_total_amt_incl']
        discount_review_reasons = ctx['discount_review_reasons']

        cells = {}
        untaxed_total = 0
        for name, amounts in items_by_name.items():
            is_aggregate = name in _BOOKSTORE_ORDER_AGGREGATE_COL_NAMES
            cost = cost_by_name.get(name)
            total_amt = sum(amounts)  # 這筆訂單、這個品名，所有列(含稅)加總——決定明細表顯示的總金額
            amt_incl = total_amt * qty
            # 該品名分攤到的折扣(含稅)＝按這一列裡各品名含稅金額佔比分攤（Pass B已決定這一列
            # 整體要扣多少discount_incl，這裡只負責同一列內部、不同品名之間怎麼分）
            name_discount_incl = (discount_incl * amt_incl / row_total_amt_incl) if discount_incl else 0
            # 換算成「每一列(每1份)」要扣的含稅金額，還原到amounts的口徑(amounts是未乘數量的
            # 單份含稅金額)，才能同步調整下面sub_items的單位售價/低於成本判斷。
            name_discount_per_unit_incl = (name_discount_incl / qty) if (name_discount_incl and qty) else 0
            amt_incl_after_discount = amt_incl - name_discount_incl
            untaxed = round(amt_incl_after_discount / 1.05, 2)
            # 2026-08-11修正：below_cost改用「單位未稅單價」（每一列各自的amount/1.05）逐列
            # 跟進貨未稅價比較，不再用加總後的金額比——數量>1或同名多列時，加總金額天生就會
            # 比單位進貨價大很多倍，直接比會失真。代銷商品/電子書方案這2個集合類品名依設計
            # 不做紅字檢核，即使「內部料號清單」剛好幫它們也填了一個購入價，也一律排除。
            sub_items = []
            for a in amounts:
                # 折扣在同品名底下的多列(同料號出貨多份)之間，再依各列自己的含稅金額佔比分攤。
                a_share = (name_discount_per_unit_incl * a / total_amt) if (name_discount_per_unit_incl and total_amt) else 0
                a_after = a - a_share
                unit_price = round(a_after / 1.05, 2)
                sub_items.append({
                    'unit_price': unit_price,
                    'below_cost': (cost is not None and not is_aggregate and unit_price < cost),
                })
            below_cost = any(s['below_cost'] for s in sub_items)
            # 2026-08-19新增：cells[name]['qty']保留「這筆訂單的訂購數量」(qty)，維持不變——
            # _compute_bookstore_order_pivot()（「總表分析」頁籤）依賴這個欄位是「未乘同名
            # 重複列數的原始訂購數量」，自己在逐個sub_item迭代時把它累加len(sub_items)次來
            # 還原「實際出貨數量」，如果這裡改成已經乘開的數字，那邊會被再乘一次造成加倍。
            # 另外新增'total_qty'=qty*len(amounts)，才是「這個財報料號品名這筆訂單實際出貨
            # 的總數量」（例如方案名稱帶「x3」、「方案清單」對應設定3列同品名時，len(amounts)
            # =3）——使用者發現_parse_bookstore_shipment()（「上傳比對→書店」批次明細）
            # 沒有用到這個乘開後的數量，一律照qty（未乘開）計算，導致方案清單已經正確設定
            # 3列同品名的情況下，批次明細數量還是少算成1份的量，這裡補上這個獨立欄位讓
            # _parse_bookstore_shipment()改用，不影響「總表分析」既有的乘開邏輯。
            cells[name] = {'untaxed': untaxed, 'qty': qty, 'total_qty': qty * len(amounts),
                           'below_cost': below_cost, 'cost': cost, 'sub_items': sub_items}
            untaxed_total += untaxed

        # 2026-08-13新增：Pass B2已經決定這一列（如果是該訂單商品價格最高的那一列）要拆進
        # 多少運費(未稅，已依境內/境外做過稅務換算)，這裡實際把金額加進「電子書方案」這個
        # 集合類欄位——若這一列原本沒有電子書方案品項(該方案的料號組合裡沒有這個品名)，
        # 就新建一格；若原本就有，直接在既有未稅金額上累加。qty/below_cost/cost/sub_items
        # 沿用「代銷商品/電子書方案」集合類品名一貫的處理方式（不做低於成本紅字檢核）。
        freight_alloc_untaxed = ctx['freight_alloc_untaxed']
        if freight_alloc_untaxed:
            if '電子書方案' in cells:
                cells['電子書方案']['untaxed'] = round(cells['電子書方案']['untaxed'] + freight_alloc_untaxed, 2)
            else:
                cells['電子書方案'] = {'untaxed': freight_alloc_untaxed, 'qty': 1, 'total_qty': 1,
                                    'below_cost': False, 'cost': None, 'sub_items': []}
            untaxed_total = round(untaxed_total + freight_alloc_untaxed, 2)

        out_rows.append({
            'order_no': ctx['order_no'],
            'order_date': ctx['order_date'],
            'plan_code': ctx['plan_code'], 'plan_name': ctx['plan'].get('plan_name'),
            'qty': qty, 'region': ctx['region'], 'cells': cells,
            'untaxed_total': round(untaxed_total, 2),
            # 2026-08-08新增：保留原始匯入訂單檔的「金額」「運送地區中文」「運費」欄位原始值，
            # 供使用者後續肉眼核對用（不參與任何計算，純顯示）。「出貨單號」使用者確認不需要，
            # 已不再解析/保留此欄。
            # 2026-08-12再修正：使用者回報「金額(原始，含稅)」欄位在Excel裡的境內/境外小計
            # 加總都變成0——根因是原本這裡直接存原始儲存格值，來源檔案裡這欄有時是文字型態
            # 的數字，寫進openpyxl後Excel仍視為文字，SUM()公式會忽略文字儲存格(但H-G這種
            # 算術運算子Excel會自動把文字數字強制轉型，所以個別列的公式沒問題，只有SUM小計
            # 才會出錯，這也是為什麼一開始沒被發現)。改用_to_num_or_none()統一轉成真正的
            # 數值型態(空白仍是None，維持留白，不會顯示成0)，這4個原始金額欄位都要轉。
            'amount_ref': ctx['amount_ref'],
            'region_raw': ctx['region_raw'],
            'freight_ref': ctx['freight_ref'],
            # 2026-08-08再新增：再保留「折價券」「訂單備註」「購物車折扣的金額」3個原始欄位，
            # 同樣只顯示不參與計算。
            'coupon_ref': _to_num_or_none(ctx['coupon_ref']),
            'order_note_ref': ctx['order_note_ref'],
            'cart_discount_ref': ctx['cart_discount_ref'],
            # 2026-08-12新增：這筆訂單實際折抵的含稅金額(=0代表沒有折扣或折扣沒套用成功)，
            # 供匯出時額外標示、方便使用者核對折扣是否有正確套用。
            'discount_applied_incl': round(discount_incl, 2) if discount_incl else 0,
            'discount_needs_review': bool(discount_review_reasons),
            'discount_review_reasons': discount_review_reasons,
            # 2026-08-13新增：這一列實際拆進去的運費未稅金額(=0代表這一列不是該訂單商品
            # 價格最高的那一列，或這筆訂單沒有運費/只買1個商品，不需要拆)，供匯出時額外
            # 標示、方便使用者核對運費是否有正確拆進電子書方案。
            'freight_allocated_untaxed': round(freight_alloc_untaxed, 2) if freight_alloc_untaxed else 0,
            'freight_needs_review': bool(ctx['freight_review_reasons']),
            'freight_review_reasons': ctx['freight_review_reasons'],
        })

    return {
        'columns': columns_order, 'rows': out_rows, 'unmatched': unmatched,
        'inconsistent_cost_names': sorted(inconsistent_names),
        'cost_by_name': cost_by_name,
        # 2026-08-13新增：運費需要人工複核的訂單清單(同一訂單編號各列的運費金額解析出來
        # 不一致)，匯出時額外提示，避免使用者沒發現有訂單漏拆運費。
        'freight_review_rows': [r for r in out_rows if r.get('freight_needs_review')],
        # 2026-08-12新增：折扣需要人工複核的訂單清單(折扣金額比料號總額大扣不完／備註文字
        # 有「折」但抓不到金額)，匯出時額外提示，避免使用者沒發現漏折扣或折扣異常的訂單。
        'discount_review_rows': [r for r in out_rows if r.get('discount_needs_review')],
        # 2026-08-18新增：純附加欄位，不影響既有任何呼叫端（analyze_bookstore_orders()的
        # JSON回應、_compute_bookstore_order_pivot()都只會多看到2個沒在用的key，不受影響）。
        # 供「書店」經銷商上傳比對（_parse_bookstore_shipment()）沿用同一套方案/拆帳明細
        # 查詢結果，不需要另外重複查一次資料庫，也保證跟這次算出來的cells完全一致。
        'plans_by_code': plans_by_code,
        'items_map': items_map,
    }


# ============================================================
# ⑪ 「書店」經銷商——上傳比對（2026-08-18新增，見dealer_reconcile_schema.sql v1.32）
# ============================================================
# 直接沿用既有「書店經銷商」拆帳引擎_compute_bookstore_order_matrix()（不修改引擎本身的
# 拆帳/折扣/運費分攤邏輯），只另外寫一份「出貨資料」格式的欄位對照＋子料號比對邏輯，走
# _ingest_reconcile_file()裡的專屬程式碼路徑（比照香港經銷商PDF/ODT），不透過「經銷商
# 格式規則」(dealer_format_rules) 設定。

_BOOKSTORE_SHIPMENT_HEADER_MAP = {
    '訂單編號': 'order_no',
    '訂購日期': 'order_date',
    '方案名稱': 'plan_name_ref',
    '方案編號': 'plan_code',
    '數量': 'qty',
    '方案訂單金額': 'amount_ref',
    '運送方式': 'region_raw',
    '運送費用': 'freight_ref',
    '折價券': 'coupon_ref',
    '訂單備註': 'order_note_ref',
}


def _bookstore_shipment_col_map(headers):
    """「書店」上傳的「出貨資料」格式，欄位標題跟「匯入書店訂單拆帳」既有訂單Excel
    （_order_col_map()）不完全相同（例如用「方案訂單金額」取代「金額」、「運送費用」取代
    「運費」、「運送方式」取代「運送地區」），故另外寫一份對照表；輸出的欄位鍵名跟
    _order_col_map()完全一致，才能直接沿用既有的_compute_bookstore_order_matrix()拆帳
    引擎，不需要修改引擎本身任何一行邏輯。"""
    m = {}
    for ci, h in enumerate(headers):
        h = (h or '').strip()
        key = _BOOKSTORE_SHIPMENT_HEADER_MAP.get(h)
        if key:
            m[key] = ci
    return m


def _resolve_bookstore_sub_sku(plan_id, name, items_map, name_to_active_codes):
    """決定某筆訂單、某個財報料號品名(name)最終要比對到的具體內部料號(matched_main_sku)：
    1. 該方案項目(bookstore_plan_items)有指定「子料號」→ 直接採用；
    2. 沒指定，但該財報料號品名底下的內部料號清單只有1筆啟用中的料號（沒有色號/等級分歧）
       → 自動採用該筆；
    3. 其餘（查無啟用中料號，或有多筆分歧但未指定子料號）→ 回傳None，match_status標記
       'unmatched'，交由「上傳比對」既有的未比對人工確認流程處理（使用者確認：這種情境要
       擋下，需人工補填子料號或於批次明細手動選擇後才可過帳）。
    回傳 (matched_main_sku, match_status, unmatched_reason)。"""
    items = [it for it in (items_map.get(plan_id) or []) if it.get('report_product_name') == name]
    sub = next((it.get('sub_internal_code') for it in items if it.get('sub_internal_code')), None)
    if sub:
        return sub, 'auto_matched', None
    codes = name_to_active_codes.get(name) or []
    if len(codes) == 1:
        return codes[0], 'auto_matched', None
    if not codes:
        return None, 'unmatched', f'財報料號品名「{name}」在「內部料號清單」查無啟用中的內部料號，請先確認料號設定'
    return None, 'unmatched', (
        f'財報料號品名「{name}」底下有{len(codes)}個內部料號（色號/等級變體），'
        f'請至「方案清單」為此方案的這個品項指定子料號，或於本批次明細手動選擇正確的內部料號'
    )


def _extract_bookstore_report_grand_total_taxed(wb):
    """2026-08-19新增：使用者實務上常常直接把先前用「匯入書店訂單拆帳」產生的「書店訂單
    拆帳」Excel報表檔（同一個活頁簿裡除了她自己要看的'書店訂單拆帳'/'總表分析'分頁，其實
    還內建了當初產生報表用的原始'出貨資料'分頁），原封不動拿來當「上傳比對→書店」建立
    批次的來源檔上傳——這種情況下，同一份檔案裡其實「內建了」一份使用者自己已經看過、
    信任的參考總額（'書店訂單拆帳'分頁的境內小計＋境外小計，「分析後發票」欄＝含稅），
    可以直接跟這次_parse_bookstore_shipment()用同一份檔案的'出貨資料'分頁、透過拆帳引擎
    重新計算出來的批次總額互相比對，抓出兩者是否有落差——常見原因是報表產生後「方案
    清單」的方案設定（單價/贈品/子料號等）又被異動過，同一份原始出貨資料重新計算就會
    跟報表當時的結果不同。

    使用者明確要求：這種落差不需要她另外再上傳第二份檔案給系統比對（那是既有
    bookstore_batch_diff()/「書店批次差異比對」工具的做法，仍保留給只有裸'出貨資料'
    分頁、沒有內建報表分頁的情境使用）——只要她這次上傳的檔案本身「剛好」就內建了
    這份報表分頁，系統就應該自動抓出來比對、自動算出差額，不需要使用者多做任何事。

    回傳：找到'書店訂單拆帳'分頁、且能定位到境內小計／境外小計列時，回傳兩者「分析後
    發票」欄（含稅）的加總（float）；找不到分頁、找不到表頭欄位、或完全沒有小計列時，
    回傳None（代表這份上傳檔案沒有內建報表分頁，呼叫端就不做這個額外比對，不影響既有
    行為）。"""
    sheet_name = _find_sheet(wb, '書店訂單拆帳')
    if not sheet_name:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    if '分析後發票' not in headers:
        return None
    taxed_ci = headers.index('分析後發票')
    total = 0.0
    found_any = False
    for row in rows[1:]:
        for v in row:
            if isinstance(v, str) and v.strip() in ('境內小計', '境外小計'):
                if taxed_ci < len(row):
                    amt = _to_num_or_none(row[taxed_ci])
                    if amt is not None:
                        total += amt
                        found_any = True
                break
    return round(total, 2) if found_any else None


def _extract_bookstore_report_qty_by_name(wb):
    """2026-08-19新增：跟_extract_bookstore_report_grand_total_taxed()同樣的背景（使用者
    把先前的「書店訂單拆帳」報表檔原封不動拿來當「上傳比對→書店」的來源檔上傳），但這裡
    比對的是「數量」而不是金額——使用者發現：拆帳引擎是依「方案清單」設定的每個方案品項
    固定數量去拆，如果方案名稱裡有「xN」這種文字暗示實際應出貨N倍數量（例如方案名稱是
    「HyRead Gaze Mini 系列 6吋透明軟膠殼x3」，代表這個方案每訂一次其實要出3個「Mini透明
    殼」），但「方案清單」的品項設定沒有對應調整成3，拆帳引擎就會照方案清單設定的數量算
    （通常是1），導致特定財報料號品名的總數量比報表上「總表分析」分頁實際加總出來的少。

    「總表分析」分頁的「料號加總(境內+境外)」列，是把每個財報料號品名欄位境內／境外、
    各個拆帳單價分組的「數量」列全部加總後的總數（見_compute_bookstore_order_pivot()），
    可以直接拿來當這份原始出貨資料「真正應該拆出多少數量」的參考基準，不需要使用者
    另外再上傳/手動核對。'代銷商品'／'電子書方案'這2個集合類品名（見既有常數
    _BOOKSTORE_ORDER_AGGREGATE_COL_NAMES）數量欄位含意特殊（不是單純的訂單數量加總），
    不适合拿來做這種逐一料號的數量比對，故排除。

    回傳：{財報料號品名: 數量} 的dict（已排除代銷商品/電子書方案）；找不到「總表分析」
    分頁、找不到表頭列、找不到「料號加總(境內+境外)」列時回傳None。"""
    sheet_name = _find_sheet(wb, '總表分析')
    if not sheet_name:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx, header = None, None
    for i, row in enumerate(rows):
        vals = [str(v).strip() if v is not None else '' for v in row]
        if '項目' in vals and '小計' in vals:
            header, header_idx = vals, i
            break
    if header is None:
        return None
    try:
        item_ci = header.index('項目')
        total_ci = header.index('小計')
    except ValueError:
        return None
    qty_row = None
    for row in rows[header_idx + 1:]:
        if item_ci < len(row) and isinstance(row[item_ci], str) and row[item_ci].strip() == '料號加總(境內+境外)':
            qty_row = row
            break
    if qty_row is None:
        return None
    result = {}
    for ci in range(item_ci + 1, min(total_ci, len(header))):
        name = header[ci]
        if not name or name in _BOOKSTORE_ORDER_AGGREGATE_COL_NAMES:
            continue
        qty = _to_num_or_none(qty_row[ci]) if ci < len(qty_row) else None
        if qty is not None:
            result[name] = qty
    return result if result else None


def _parse_bookstore_shipment(raw_bytes):
    """解析「書店」經銷商上傳的「出貨資料」格式Excel，重用既有的書店拆帳引擎
    _compute_bookstore_order_matrix()：
    1. 讀取「出貨資料」分頁（找不到則退回第一個分頁），依欄位標題對照成引擎需要的col_map。
    2. 「運送方式」欄位是完整描述文字（如「運送至台灣、澎湖、金門、馬祖」），跟引擎既有
       「地區文字是否完全等於『台灣』」的判斷邏輯不相容——這裡在餵給引擎前先正規化：文字
       裡包含「台灣」就換成純文字「台灣」，其餘保持原樣（會被引擎判為境外），藉此不需要
       修改引擎本身的境內/境外判斷邏輯一行程式碼。
    3. 使用者確認「取消/退換貨訂單一律照單全收，不做篩選」——本函式不依「訂單狀態」
       「出貨狀態」「退換貨」等欄位過濾任何一列，全部送進引擎計算。
    4. 引擎算出的每筆訂單、每個財報料號品名一格(cells)，展開成一筆dealer_reconcile_lines，
       matched_main_sku比對邏輯見_resolve_bookstore_sub_sku()。
    5. 訂單裡「方案編號」在「方案清單」查無設定的列，引擎本身會把這種列整列排除在
       result['rows']之外、只列在result['unmatched']（不計入任何金額）——這裡額外把這些列
       還原成一筆dealer_reconcile_lines（標記未比對＋原因，金額取自原始「方案訂單金額」欄），
       避免使用者要求的「全部照單全收，不做篩選」被引擎既有的靜默排除違反。
    回傳值格式與_parse_workbook()等既有解析函式相同的lines陣列（供_ingest_reconcile_file()
    共用批次/入庫邏輯），額外多帶'_prematched'=True、'matched_main_sku'、'match_status'
    3個欄位，讓_ingest_reconcile_file()略過一般經銷商用的_match_one()比對，直接採用這裡
    算好的結果。"""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    sheet_name = _find_sheet(wb, '出貨資料') or wb.sheetnames[0]
    ws = wb[sheet_name]
    raw_rows = list(ws.iter_rows(values_only=True))
    if len(raw_rows) < 2:
        raise ValueError('「出貨資料」分頁無資料列')
    headers = [str(h).strip() if h is not None else '' for h in raw_rows[0]]
    col_map = _bookstore_shipment_col_map(headers)
    missing = [label for label, key in (
        ('訂單編號', 'order_no'), ('方案編號', 'plan_code'), ('數量', 'qty'), ('方案訂單金額', 'amount_ref'),
    ) if key not in col_map]
    if missing:
        raise ValueError(f'「出貨資料」缺少必要欄位：{"、".join(missing)}，請確認欄位標題是否正確')

    ci_region = col_map.get('region_raw')
    data_rows = []
    for row in raw_rows[1:]:
        if ci_region is not None and ci_region < len(row):
            row = list(row)
            v = row[ci_region]
            row[ci_region] = '台灣' if (v and '台灣' in str(v)) else v
            row = tuple(row)
        data_rows.append(row)

    result = _compute_bookstore_order_matrix(data_rows, col_map)
    plans_by_code = result.get('plans_by_code') or {}
    items_map = result.get('items_map') or {}
    name_to_active_codes = _report_name_to_active_codes()

    lines = []
    for r in result['rows']:
        order_no = r['order_no']
        plan = plans_by_code.get(r['plan_code'])
        plan_id = plan['id'] if plan else None
        for name, cell in (r.get('cells') or {}).items():
            sub, status, reason = _resolve_bookstore_sub_sku(plan_id, name, items_map, name_to_active_codes)
            untaxed = cell.get('untaxed') or 0
            amount_taxed = round(untaxed * 1.05, 2)
            # 2026-08-19修正：這裡的數量要用'total_qty'（已經依「方案清單」同一財報料號品名
            # 底下設定的重複列數乘開，例如方案名稱帶「x3」、方案清單對應設定3列同品名時，
            # 這裡就是訂購數量×3），不能只用'qty'（那是未乘開的訂單訂購數量，cells[name]
            # ['qty']保留這個純值是給_compute_bookstore_order_pivot()《總表分析》頁籤自己
            # 逐個sub_item累加用的，兩邊各自用對應的欄位，才不會一邊沒乘、一邊乘兩次）。
            qty_out = cell.get('total_qty', cell.get('qty'))
            raw = {
                'order_no': order_no, 'order_date': r.get('order_date'),
                'plan_code': r.get('plan_code'), 'plan_name': r.get('plan_name'),
                'report_product_name': name, 'region': r.get('region'), 'qty': qty_out,
                'below_cost': cell.get('below_cost'),
                'discount_applied_incl': r.get('discount_applied_incl'),
                'freight_allocated_untaxed': r.get('freight_allocated_untaxed') if name == '電子書方案' else 0,
                'unmatched_reason': reason,
            }
            lines.append({
                'row_type': 'sale', 'dealer_code': order_no, 'ean': None,
                'product_name': name,
                'qty': qty_out, 'unit_price': None,
                'amount': amount_taxed, 'amount_untaxed': untaxed, 'amount_taxed': amount_taxed,
                'tax_amount': round(amount_taxed - untaxed, 2),
                'raw_json': raw,
                '_prematched': True, 'matched_main_sku': sub, 'match_status': status,
            })

    # 「方案編號」在方案清單查無設定的訂單列——照單全收，獨立標記未比對，不靜默排除
    for u in (result.get('unmatched') or []):
        amt = None
        for row in data_rows:
            def _g(key, _row=row):
                ci = col_map.get(key)
                return _row[ci] if (ci is not None and ci < len(_row)) else None
            if str(_g('order_no') or '').strip() == u['order_no'] and _norm_plan_code(_g('plan_code')) == u['plan_code']:
                amt = _to_num_or_none(_g('amount_ref'))
                break
        untaxed = round(amt / 1.05, 2) if amt is not None else None
        lines.append({
            'row_type': 'sale', 'dealer_code': u['order_no'], 'ean': None,
            'product_name': f"[查無方案]{u.get('plan_name_ref') or ''}（方案編號{u['plan_code']}）",
            'qty': u.get('qty'), 'unit_price': None,
            'amount': amt, 'amount_untaxed': untaxed, 'amount_taxed': amt,
            'tax_amount': round(amt - untaxed, 2) if (amt is not None and untaxed is not None) else None,
            'raw_json': {**u, 'unmatched_reason': f"方案編號「{u['plan_code']}」尚未於「方案清單」建立，請先建立方案（或確認方案編號是否輸入錯誤）後，回到此批次手動選擇內部料號，或重新上傳"},
            '_prematched': True, 'matched_main_sku': None, 'match_status': 'unmatched',
        })
    return lines


# ============================================================
# ⑫ 「書店」上傳比對批次 vs. 另一份出貨資料 —— 差異比對＋一次補登（2026-08-18新增）
# ============================================================
# 背景（使用者2026-07書店對帳案例）：「上傳比對→書店」批次與「匯入書店訂單拆帳」Excel
# 報表雖然共用同一套拆帳引擎（_compute_bookstore_order_matrix()／_parse_bookstore_
# shipment()），但如果兩邊分別上傳的「出貨資料」不是同一個時間點的版本（例如其中一份
# 少了幾筆加購/贈品項目），算出來的金額就會對不起來，且引擎本身不會有任何警示（兩份
# 資料各自都比對成功，只是資料範圍不同）。使用者要求：上傳批次時可以順手多上傳一份
# 「當初產生書店訂單拆帳報表用的出貨資料」，系統自動比對兩邊有沒有落差，勾選後一次
# 補登成這個批次自己的明細列，不用一筆一筆手動核對/拆帳。

def _aggregate_bookstore_lines(lines):
    """把「書店」出貨資料算出來的逐筆 lines（可能是 _parse_bookstore_shipment() 回傳的
    list，也可能是 dealer_reconcile_lines 資料表撈出來的list of dict——兩者共用的欄位
    名稱剛好一致：dealer_code/product_name/qty/amount_untaxed/amount_taxed/
    matched_main_sku/match_status），依「訂單編號＋財報料號品名」彙總成一筆。

    需要彙總的原因：同一張訂單如果有多個方案(plan_code)各自都命中到同一個財報料號品名
    （例如訂單裡買了2個不同方案但都含「電子書方案」），_compute_bookstore_order_matrix()
    會拆成好幾列各自的cells——這對「書店訂單拆帳」明細表沒有影響（那邊本來就是逐列呈現），
    但這裡要跟「上傳比對」批次明細（同樣是一張訂單一個品名一列的彙總層級）逐筆核對，
    所以兩邊都先彙總到同一個顆粒度，才能用(訂單編號,品名)當key直接比對，不會因為兩邊
    「同一張訂單同一品名拆成的列數不同」誤判成有落差。

    回傳 dict：key=(訂單編號, 財報料號品名) -> {'qty','amount_untaxed','amount_taxed',
    'matched_main_sku','match_status'}（後兩者取彙總範圍內第一筆非空值，同一訂單同一
    品名理論上都會比對到同一顆內部料號，不會衝突）。"""
    agg = {}
    for ln in lines:
        order_no = str(ln.get('dealer_code') or '').strip()
        product_name = str(ln.get('product_name') or '').strip()
        if not order_no or not product_name:
            continue
        key = (order_no, product_name)
        a = agg.setdefault(key, {'qty': 0.0, 'amount_untaxed': 0.0, 'amount_taxed': 0.0,
                                  'matched_main_sku': None, 'match_status': None})
        a['qty'] += _to_num(ln.get('qty'))
        a['amount_untaxed'] += _to_num(ln.get('amount_untaxed'))
        a['amount_taxed'] += _to_num(ln.get('amount_taxed'))
        if not a['matched_main_sku'] and ln.get('matched_main_sku'):
            a['matched_main_sku'] = ln.get('matched_main_sku')
        if not a['match_status'] and ln.get('match_status'):
            a['match_status'] = ln.get('match_status')
    return agg


@dealer_bp.route('/api/dealer/batches/<int:bid>/bookstore-diff', methods=['POST'])
@dealer_edit_required
def bookstore_batch_diff(bid):
    """比對這個批次目前的明細，跟另外上傳的一份「出貨資料」（通常是產生「書店訂單拆帳」
    Excel報表當時用的來源檔案）之間，逐筆(訂單編號+財報料號品名)有沒有金額/數量落差。
    純讀取比對、不寫入任何資料，可重複呼叫（例如上傳新的比對來源再核對一次）。"""
    batch_res = sb.table('dealer_reconcile_batches').select('*').eq('id', bid).execute()
    if not batch_res.data:
        return jsonify({'error': '找不到批次'}), 404
    batch = batch_res.data[0]
    if batch.get('dealer_name') != '書店':
        return jsonify({'error': '此功能僅供「書店」經銷商的批次使用'}), 400

    f = request.files.get('file')
    if not f:
        return jsonify({'error': '請選擇要比對的出貨資料Excel'}), 400
    try:
        ref_lines = _parse_bookstore_shipment(f.read())
    except Exception as e:
        return jsonify({'error': f'無法解析比對用的出貨資料：{e}'}), 400

    ref_agg = _aggregate_bookstore_lines(ref_lines)
    batch_lines_res = sb.table('dealer_reconcile_lines').select('*').eq('batch_id', bid).execute()
    batch_agg = _aggregate_bookstore_lines(batch_lines_res.data or [])

    TOL = 0.02  # 容忍2位小數的四捨五入雜訊，避免每次都因浮點誤差誤判成有落差
    diff_items = []
    for key in (set(ref_agg.keys()) | set(batch_agg.keys())):
        order_no, product_name = key
        r = ref_agg.get(key) or {'qty': 0.0, 'amount_untaxed': 0.0, 'amount_taxed': 0.0,
                                  'matched_main_sku': None, 'match_status': None}
        b = batch_agg.get(key) or {'qty': 0.0, 'amount_untaxed': 0.0, 'amount_taxed': 0.0,
                                    'matched_main_sku': None, 'match_status': None}
        delta_untaxed = round(r['amount_untaxed'] - b['amount_untaxed'], 2)
        delta_taxed = round(r['amount_taxed'] - b['amount_taxed'], 2)
        delta_qty = round(r['qty'] - b['qty'], 4)
        if abs(delta_untaxed) <= TOL and abs(delta_taxed) <= TOL and abs(delta_qty) <= TOL:
            continue
        # 2026-08-18新增：只有「批次比參考出貨資料少算」(missing_in_batch)才適合自動一次
        # 補登——這種情況直接把差額補成批次自己的明細列，入庫金額才會補齊。「批次比參考
        # 資料還多」(extra_in_batch，delta為負)不透過這裡自動處理，因為可能是這個批次的
        # 來源資料本來就比對方新/完整（比照這次的實際案例），需要人工確認是否要調整，
        # 不適合自動刪減既有明細列，只列出來提醒使用者留意。
        direction = 'missing_in_batch' if delta_untaxed > 0 else 'extra_in_batch'
        diff_items.append({
            'order_no': order_no, 'product_name': product_name,
            'ref_qty': round(r['qty'], 4), 'ref_untaxed': round(r['amount_untaxed'], 2), 'ref_taxed': round(r['amount_taxed'], 2),
            'batch_qty': round(b['qty'], 4), 'batch_untaxed': round(b['amount_untaxed'], 2), 'batch_taxed': round(b['amount_taxed'], 2),
            'delta_qty': delta_qty, 'delta_untaxed': delta_untaxed, 'delta_taxed': delta_taxed,
            'matched_main_sku': r.get('matched_main_sku'), 'match_status': r.get('match_status') or 'unmatched',
            'direction': direction,
        })

    diff_items.sort(key=lambda d: (d['order_no'], d['product_name']))
    missing_count = sum(1 for d in diff_items if d['direction'] == 'missing_in_batch')

    # 2026-08-19新增：使用者反映「書店訂單拆帳」報表的含稅總金額，有時候會跟這個批次
    # 目前明細的總金額對不起來，即使逐筆(訂單編號+財報料號品名)比對後看起來一致或落差
    # 都在容忍範圍內——常見原因是運費等彙總方式差異，不會集中反映在單一筆訂單/品名上，
    # 而是分散在很多筆都各自差一點點（每筆都在TOL容忍範圍內因此不會被上面的逐筆邏輯
    # 抓出來），加總後卻累積成有感的總額落差。使用者要求：這個「逐筆比對後仍無法解釋
    # 的總額落差」也要明確列出來、讓她可以選擇補登一筆調整明細，才能把總金額對平後入庫。
    # 做法：用同一份上傳的參考出貨資料，直接加總「全部」ref_lines/批次全部明細的金額
    # （不透過(訂單編號,品名)分組，避免因分組時濾掉缺料號的列而低估總額），跟逐筆
    # diff_items 加總起來「已經能解釋」的落差相減，剩下的餘額就是要提醒使用者的部分。
    ref_total_untaxed = sum(_to_num(ln.get('amount_untaxed')) for ln in ref_lines)
    ref_total_taxed = sum(_to_num(ln.get('amount_taxed')) for ln in ref_lines)
    batch_all_lines = batch_lines_res.data or []
    batch_total_untaxed = sum(_to_num(ln.get('amount_untaxed')) for ln in batch_all_lines)
    batch_total_taxed = sum(_to_num(ln.get('amount_taxed')) for ln in batch_all_lines)

    grand_delta_untaxed = round(ref_total_untaxed - batch_total_untaxed, 2)
    grand_delta_taxed = round(ref_total_taxed - batch_total_taxed, 2)
    explained_delta_untaxed = round(sum(d['delta_untaxed'] for d in diff_items), 2)
    explained_delta_taxed = round(sum(d['delta_taxed'] for d in diff_items), 2)
    residual_untaxed = round(grand_delta_untaxed - explained_delta_untaxed, 2)
    residual_taxed = round(grand_delta_taxed - explained_delta_taxed, 2)

    total_adjustment = None
    if abs(residual_untaxed) > TOL or abs(residual_taxed) > TOL:
        total_adjustment = {
            'is_total_adjustment': True,
            'order_no': '(整體調整)',
            'product_name': '整體金額調整（運費/總額差異，逐筆比對無法歸屬的部分）',
            'ref_qty': 0, 'ref_untaxed': round(ref_total_untaxed, 2), 'ref_taxed': round(ref_total_taxed, 2),
            'batch_qty': 0, 'batch_untaxed': round(batch_total_untaxed, 2), 'batch_taxed': round(batch_total_taxed, 2),
            'delta_qty': 0, 'delta_untaxed': residual_untaxed, 'delta_taxed': residual_taxed,
            'matched_main_sku': None, 'match_status': 'unmatched',
            'direction': 'missing_in_batch' if residual_untaxed >= 0 else 'extra_in_batch',
        }

    return jsonify({
        'diff_items': diff_items,
        'missing_count': missing_count, 'extra_count': len(diff_items) - missing_count,
        'ref_total_lines': len(ref_lines),
        'total_adjustment': total_adjustment,
        'grand_totals': {
            'ref_untaxed': round(ref_total_untaxed, 2), 'ref_taxed': round(ref_total_taxed, 2),
            'batch_untaxed': round(batch_total_untaxed, 2), 'batch_taxed': round(batch_total_taxed, 2),
            'delta_untaxed': grand_delta_untaxed, 'delta_taxed': grand_delta_taxed,
        },
    })


@dealer_bp.route('/api/dealer/batches/<int:bid>/bookstore-diff/apply', methods=['POST'])
@dealer_edit_required
def bookstore_batch_diff_apply(bid):
    """把使用者在差異比對表勾選的項目，直接補登成這個批次自己的新明細列，讓這次「確認並
    入庫」的總金額直接包含補登的部分，不用再跑第二次入庫。只接受前端送回的
    direction='missing_in_batch'項目（bookstore_batch_diff()回傳的原始資料，前端不應
    自行修改金額欄位）；已入庫的批次比照commit_batch()/confirm_line()既有規則直接擋下，
    避免修改到已經寫入對帳彙總表的資料。"""
    batch_res = sb.table('dealer_reconcile_batches').select('*').eq('id', bid).execute()
    if not batch_res.data:
        return jsonify({'error': '找不到批次'}), 404
    batch = batch_res.data[0]
    if batch.get('status') == '已入庫':
        return jsonify({'error': '此批次已入庫，無法再新增補登明細（會影響已寫入的對帳彙總資料），如需修正請聯絡系統管理員'}), 400

    data = request.json or {}
    # 2026-08-19新增：除了原本的「批次少算」逐筆補登，也接受前端勾選的「整體金額調整」
    # 項目（bookstore_batch_diff()回傳的total_adjustment，is_total_adjustment=True）——
    # 這種項目代表逐筆比對後仍無法歸屬到特定訂單/品名的總額落差（常見原因如運費彙總
    # 方式差異，很多筆訂單各自差一點點、單筆都在容忍範圍內不會被逐筆邏輯抓出來），
    # 使用者確認這種情況也要能一次補登成一筆調整明細，才能把總金額對平後入庫；
    # 因此這類項目不限定方向(missing_in_batch)，正負差額都允許（可能是要加，也可能是要扣）。
    items = [it for it in (data.get('items') or [])
             if it.get('direction') == 'missing_in_batch' or it.get('is_total_adjustment')]
    if not items:
        return jsonify({'error': '請至少勾選一筆「批次少算」或「整體金額調整」的落差項目（批次比參考資料多的一般項目不支援自動補登）'}), 400

    line_recs = []
    for it in items:
        is_adj = bool(it.get('is_total_adjustment'))
        order_no = str(it.get('order_no') or '').strip()
        product_name = str(it.get('product_name') or '').strip()
        if not order_no or not product_name:
            continue
        amount_untaxed = round(_to_num(it.get('delta_untaxed')), 2)
        amount_taxed = round(_to_num(it.get('delta_taxed')), 2)
        line_recs.append({
            'batch_id': bid, 'row_type': 'sale', 'dealer_code': order_no, 'ean': None,
            'product_name': product_name, 'qty': _to_num(it.get('delta_qty')), 'unit_price': None,
            'amount': amount_taxed, 'amount_untaxed': amount_untaxed, 'amount_taxed': amount_taxed,
            'tax_amount': round(amount_taxed - amount_untaxed, 2),
            'matched_main_sku': it.get('matched_main_sku'),
            'matched_gift_sku1': None, 'matched_gift_sku2': None, 'matched_gift_sku3': None,
            'match_status': it.get('match_status') or 'unmatched',
            'raw_json': {'source': 'bookstore_diff_total_adjustment' if is_adj else 'bookstore_diff_supplement',
                         'note': '整體金額調整（逐筆比對無法歸屬的總額差異）' if is_adj else '差異比對後一次補登',
                         'order_no': order_no, 'product_name': product_name,
                         'delta_qty': it.get('delta_qty'), 'delta_untaxed': it.get('delta_untaxed'),
                         'delta_taxed': it.get('delta_taxed')},
        })

    if not line_recs:
        return jsonify({'error': '勾選的項目格式不正確，補登失敗'}), 400

    for i in range(0, len(line_recs), 80):
        sb.table('dealer_reconcile_lines').insert(line_recs[i:i + 80]).execute()

    all_lines_res = sb.table('dealer_reconcile_lines').select('match_status').eq('batch_id', bid).execute()
    all_lines = all_lines_res.data or []
    matched = sum(1 for l in all_lines if l['match_status'] in ('auto_matched', 'manual_confirmed'))
    unmatched = sum(1 for l in all_lines if l['match_status'] == 'unmatched')
    sb.table('dealer_reconcile_batches').update({
        'total_lines': len(all_lines), 'matched_lines': matched, 'unmatched_lines': unmatched,
    }).eq('id', bid).execute()

    return jsonify({'ok': True, 'inserted': len(line_recs)})


# 2026-08-08新增：「總表分析」頁籤裡，這2個報表品名是「集合類」品名(代銷商品/電子書方案，
# 底下對應很多不同商品、每筆訂單分攤到的金額都不太一樣)，使用者要求這2欄不要比照一般料號
# 「拆帳單價相同的擺一列」分組方式，而是「全部加總後，數量填1」——同一區塊(境內/境外)內，
# 不論比對到幾筆訂單、單價各是多少，一律加總成一個數字，只佔一列，數量欄固定顯示1。
_BOOKSTORE_ORDER_AGGREGATE_COL_NAMES = ('代銷商品', '電子書方案')


def _compute_bookstore_order_pivot(result):
    """依「總表分析」頁籤樣式（使用者提供的參考檔[0701-0727A88]頁籤），把 _compute_bookstore_
    order_matrix() 算出來的逐筆訂單資料，重新彙總成「每個報表品名一欄、境內/境外分區、拆帳
    單價相同的訂單合併成一列並加總數量」的樞紐格式：
    - 每個報表品名欄位，境內、境外各自獨立分組：一般料號依「單價(未稅)」分組，同單價的訂單
      數量加總成一列；代銷商品/電子書方案這2個集合類品名固定只有一列，把所有比對到的訂單
      未稅金額全部加總、數量顯示1。
    - 該品名欄位最下面的「數量小計」＝境內＋境外所有列的加總(一般料號是數量加總，代銷商品/
      電子書方案這2欄則是金額加總，比照使用者提供的參考檔邏輯)。
    - 境內、境外各自再各有一列「金額合計」，是該品名欄位在這個區內所有列「單價×數量」的加總
      （集合類品名因為單價本身已經是加總後金額、數量固定1，所以就是那個數字本身）；這兩列
      各自最前面再放一個「該區所有品名合計」的總數，這個數字必須跟「書店訂單拆帳」頁籤裡
      境內/境外小計列最右邊「合計(未稅)」欄的加總一致（因為兩者都是從同一份逐筆訂單資料算
      出來的，只要加總邏輯正確，數字自然會一致，可以互相核對）。

    2026-08-12修正：使用者要求這個頁籤改成「境內顯示未稅、境外顯示含稅」——境外的單價/
    數量小計/金額合計，這裡回傳的price都已經是×1.05換算回含稅後的值，below_cost判斷仍是
    用換算前的未稅價跟未稅進貨價比較。也因此這之後「境外」這欄的金額，已經不會再等於
    「書店訂單拆帳」頁籤境外小計的「合計(未稅)」欄（會差1.05倍），這是使用者這次要求的
    新行為，不是算錯——只有「境內」欄位還維持跟Sheet1境內小計互相核對一致。
    """
    columns = result['columns']
    cost_by_name = result.get('cost_by_name') or {}

    # tiers_by_region[name][region] = [{'price':.., 'qty':..}, ...]（一般料號依單價由小到大排序；
    # 集合類品名固定最多只有1筆）
    tiers_by_region = {name: {'境內': [], '境外': []} for name in columns}

    for name in columns:
        is_aggregate = name in _BOOKSTORE_ORDER_AGGREGATE_COL_NAMES
        cost = cost_by_name.get(name)
        for region in ('境內', '境外'):
            matching = [row for row in result['rows'] if row['region'] == region and name in row['cells']]
            if not matching:
                continue
            # 2026-08-12新增：使用者要求「總表分析」頁籤境內呈現未稅、境外呈現含稅——境外
            # 這邊顯示的單價/金額要換算回含稅(×1.05)。below_cost仍然要用未稅價去跟「內部
            # 料號清單」的未稅進貨價比較（進貨價本身就是未稅基礎，不能拿含稅價去比，否則
            # 標紅門檻會整批被稅金拉高、誤判成沒有低於成本）。
            tax_mult = 1.05 if region == '境外' else 1
            if is_aggregate:
                # 代銷商品/電子書方案為集合類品名，依設計不做紅字檢核（見下方below_cost）
                total_amt_untaxed = round(sum(row['cells'][name]['untaxed'] for row in matching), 2)
                tiers_by_region[name][region] = [
                    {'price': round(total_amt_untaxed * tax_mult, 2), 'qty': 1, 'below_cost': False}]
            else:
                grouped = {}
                for row in matching:
                    cell = row['cells'][name]
                    # 2026-08-11修正：改成逐個sub_item累加數量，而不是整格只算一次「該訂單
                    # 的數量」——同一方案裡這個品名重複幾列(例如透明殼x3列)，這裡就把這筆
                    # 訂單的數量(cell['qty'])累加幾次，才能正確反映「實際出貨數量」，不是
                    # 把重複列合併成1筆單價更高、數量卻沒有跟著變多的錯誤結果。
                    for sub in cell.get('sub_items', []):
                        price = sub['unit_price']  # 未稅單價，分組/below_cost比較基準都用這個
                        grouped[price] = grouped.get(price, 0) + cell['qty']
                # 2026-08-11新增：每個單價分組也帶上below_cost，供「總表分析」頁籤匯出時
                # 比照「書店訂單拆帳」明細頁籤的規則標紅——之前這裡完全沒有這個欄位，
                # 「總表分析」頁籤的單價儲存格因此從來沒有任何顏色標記。
                tiers_by_region[name][region] = [
                    {'price': round(price_untaxed * tax_mult, 2), 'qty': grouped[price_untaxed],
                     'below_cost': (cost is not None and price_untaxed < cost)}
                    for price_untaxed in sorted(grouped.keys())
                ]

    max_tiers = {
        region: max((len(tiers_by_region[name][region]) for name in columns), default=0)
        for region in ('境內', '境外')
    }

    # 每個品名欄位「數量小計」（境內+境外合計；一般料號是數量加總，集合類品名是金額加總）
    col_combined_total = {}
    # 每個品名欄位在單一區內的「金額合計」（單價×數量加總；集合類品名就是那個加總後金額本身）
    col_region_total = {region: {} for region in ('境內', '境外')}
    for name in columns:
        is_aggregate = name in _BOOKSTORE_ORDER_AGGREGATE_COL_NAMES
        combined = 0
        for region in ('境內', '境外'):
            tiers = tiers_by_region[name][region]
            if is_aggregate:
                region_total = sum(t['price'] * t['qty'] for t in tiers)  # qty固定1
                combined += region_total
            else:
                region_total = sum(t['price'] * t['qty'] for t in tiers)
                combined += sum(t['qty'] for t in tiers)
            col_region_total[region][name] = round(region_total, 2)
        col_combined_total[name] = round(combined, 2)

    grand_total = {
        region: round(sum(col_region_total[region].values()), 2) for region in ('境內', '境外')
    }

    return {
        'columns': columns,
        'cost_by_name': cost_by_name,
        'aggregate_names': set(_BOOKSTORE_ORDER_AGGREGATE_COL_NAMES),
        'tiers_by_region': tiers_by_region,
        'max_tiers': max_tiers,
        'col_combined_total': col_combined_total,
        'col_region_total': col_region_total,
        'grand_total': grand_total,
    }


def _parse_order_upload():
    """共用：解析上傳的訂單Excel，回傳 (rows, col_map) 或 (None, error_response)。"""
    if 'file' not in request.files:
        return None, (jsonify({'error': '未上傳檔案'}), 400)
    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return None, (jsonify({'error': f'無法解析 Excel：{e}'}), 400)

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None, (jsonify({'error': '檔案無資料列'}), 400)
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    col_map = _order_col_map(headers)
    if 'plan_code' not in col_map or 'qty' not in col_map:
        return None, (jsonify({'error': '找不到「方案」或「數量」欄位，請確認欄位標題是否正確'}), 400)
    return (rows[1:], col_map), None


def _log_order_import_usage(result, file_name):
    """2026-08-08新增：把這次分析裡「有比對到方案」的訂單列，依方案編號彙總筆數後寫入
    bookstore_order_import_log，供 delete_bookstore_plan() 檢查方案是否已被訂單使用過。
    只記錄彙總筆數，不存訂單明細本身；寫入失敗不應該讓分析結果連帶失敗，故包 try/except。"""
    counts = {}
    for r in (result.get('rows') or []):
        code = r.get('plan_code')
        if code:
            counts[code] = counts.get(code, 0) + 1
    if not counts:
        return
    name = session.get('display_name', session.get('username', ''))
    recs = [{
        'plan_code': code, 'order_count': cnt, 'file_name': file_name,
        'imported_by': session.get('user_id'), 'imported_by_name': name,
    } for code, cnt in counts.items()]
    try:
        sb.table('bookstore_order_import_log').insert(recs).execute()
    except Exception:
        pass


@dealer_bp.route('/api/bookstore/orders/analyze', methods=['POST'])
@bookstore_order_view_required
def analyze_bookstore_orders():
    file_name = request.files.get('file').filename if 'file' in request.files else None
    parsed, err = _parse_order_upload()
    if err:
        return err
    rows, col_map = parsed
    result = _compute_bookstore_order_matrix(rows, col_map)
    _log_order_import_usage(result, file_name)
    return jsonify(result)


def _r0(v):
    """2026-08-13新增：使用者要求「書店訂單拆帳」「總表分析」這兩個匯出Excel頁籤裡的所有
    數值欄位都不要顯示小數位、四捨五入到整數——只影響這個匯出檔案本身的顯示層，不去動
    _compute_bookstore_order_matrix()/_compute_bookstore_order_pivot()內部算好的2位小數
    精度數字（below_cost低於成本比較、/api/bookstore/orders/analyze給網頁即時分析表用的
    JSON，都還是原本的2位小數，只有這支匯出函式最後寫入儲存格前才統一四捨五入到整數）。
    None/空字串安全略過，交還原值，避免把應該留白的儲存格誤寫成0。

    2026-08-14診斷後改回：曾為了排查境內總額差125元的問題，暫時把這裡改成只捨入到小數點
    後2位（保留精度），讓使用者匯出比對「四捨五入到整數、再加總」是否是造成125元落差的
    原因。使用者實測結果：改成2位小數後，境內總額仍是3093000.99，跟原本整數模式的
    3093001幾乎沒有差別，落差依然接近125元——證實**125元不是捨入造成的累積誤差**，
    改回原本「顯示整數」的既有需求即可，不需要再保留2位小數模式。真正成因請改用「書店
    訂單拆帳」頁籤既有的「驗證金額」欄（見export_bookstore_orders()裡的J欄公式與紅字
    標示規則）逐筆核對，找出哪幾張境內訂單的拆帳明細跟「金額(原始，含稅)」對不起來。

    2026-08-24發現：這支函式的「四捨五入後才寫入儲存格」做法，會讓「總表分析」頁籤每個
    單價分組儲存格顯示的都是已經捨去小數的整數單價，使用者自己在Excel用SUMPRODUCT把這些
    「已被捨去的單價」乘以數量重新加總核對時，會跟系統內部「用完全精確的單價相乘、加總完
    才捨入一次」算出來的官方「金額合計(未稅)」對不起來（實測案例：SUMPRODUCT核算比官方
    總額少8元）。這不是加總邏輯寫錯，是「先把每一筆單價捨成整數、再相乘」的四捨五入順序
    問題。使用者確認要改成「先不四捨五入」，故保留這支函式名稱／文件供舊註解對照，但已不
    再被呼叫寫入儲存格——改由下面_set_int_cell()取代（只在Excel顯示格式上隱藏小數位，
    儲存格實際數值維持完全精確，不提前捨入），確保之後不管使用者在哪一格重新加總/核算，
    都會跟系統算出來的總額一致。"""
    if v is None or v == '':
        return v
    try:
        return round(float(v))
    except (TypeError, ValueError):
        return v


def _set_int_cell(ws, row, col, value):
    """2026-08-24新增，取代舊的_r0()寫入方式：儲存格「實際數值」維持完全精確（不做任何
    四捨五入），只套用Excel數字格式'#,##0'，讓畫面上「顯示」為整數、不顯示小數位——滿足
    2026-08-13「不要顯示小數位」的原始需求，但不再提前捨去精確值。這樣使用者不論是用
    SUM、SUMPRODUCT，還是肉眼加總哪一格，重新核算出來的結果都會跟系統內部算出來的官方
    總額完全一致，不會再因為「先把每一筆數字捨成整數才拿去相乘/加總」而產生累積誤差
    （見上方_r0()文件說明的8元落差案例）。
    None/空字串安全略過，直接寫入不設定數字格式，避免把應該留白的儲存格誤寫成0。
    回傳建立好的儲存格物件，呼叫端可以接著設定fill/font等樣式（沿用原本_r0()呼叫慣例）。"""
    c = ws.cell(row, col, value)
    if value is not None and value != '':
        c.number_format = '#,##0'
    return c


def _apply_vertical_top(ws):
    """2026-08-12新增：使用者要求匯出Excel所有工作表的文字對齊方式統一改成「垂直靠上」，
    掃過整張工作表已使用範圍(ws.iter_rows()預設就是走已使用範圍)的每一格，保留原本的水平
    對齊/自動換行/縮排等設定，只把vertical改成'top'。openpyxl的Alignment物件是唯讀的，
    不能單獨改一個屬性，必須整個重新建立一份。"""
    for row in ws.iter_rows():
        for cell in row:
            a = cell.alignment
            cell.alignment = openpyxl.styles.Alignment(
                horizontal=a.horizontal, vertical='top', wrap_text=a.wrap_text,
                text_rotation=a.text_rotation, shrink_to_fit=a.shrink_to_fit, indent=a.indent)


# ── 書店訂單拆帳 — 匯出Excel（境內/境外分區塊，未稅單價低於進貨未稅價的格子標紅，
#    找不到對應方案的訂單另放一個分頁，比照全系統既有的「異常資料另開分頁」慣例） ──
@dealer_bp.route('/api/bookstore/orders/export', methods=['POST'])
@bookstore_order_view_required
def export_bookstore_orders():
    parsed, err = _parse_order_upload()
    if err:
        return err
    rows, col_map = parsed
    result = _compute_bookstore_order_matrix(rows, col_map)
    columns = result['columns']
    pivot = _compute_bookstore_order_pivot(result)  # 2026-08-08新增：「總表分析」頁籤用

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '書店訂單拆帳'

    # 2026-08-08修正：拿掉「出貨單號」（使用者確認此欄無意義，不分析不匯入）；新增保留
    # 原始匯入訂單檔的「金額」「運送地區中文」「運費」3個欄位供後續肉眼核對；欄位清單
    # 最後方新增「合計(未稅)」，是該筆訂單所有拆帳品項未稅金額的加總，方便跟前面保留的
    # 原始「金額」欄位核對換算是否正確。
    # 2026-08-08再修正：再新增保留「折價券」「訂單備註」「購物車折扣的金額」3個原始欄位，
    # 同樣只供肉眼核對、不參與任何計算。
    # 2026-08-12新增：使用者要求在「金額(原始，含稅)」右邊插入「合計(含稅)驗證拆帳用」「驗證
    # 金額」2個新欄位（見下方公式），用來核對這筆訂單的拆帳明細加總換算回含稅後，是否等於
    # 原始收費金額——插入後，原本第8~12欄（運送地區中文起）整批往右移2欄。
    # 2026-08-13改名：原本欄位名稱寫「合計(未稅)檢核用」，但這欄實際上是把未稅合計×1.05
    # 換算回含稅後再核對，數字本身是含稅金額，用「未稅」當標籤名容易誤導，使用者要求正名
    # 為「合計(含稅)驗證拆帳用」，並改用ROUND()四捨五入到整數，不顯示小數點。
    # 2026-08-13新增：使用者截圖回報境外拆帳金額對不起來（例如電子書方案欄位應為800+運費
    # 1500=2300卻算成2312），要求在「金額(原始，含稅)」後面再插入一欄「發票含稅金額」，
    # 提供一個完全不經過拆帳/運費分攤邏輯、只從原始來源欄位直接算出來的「這筆訂單當初開立
    # 發票的含稅總額」，供使用者拿來跟後面「合計(含稅)驗證拆帳用」欄位互相核對、找出拆帳
    # 邏輯哪裡算錯。公式＝SUMIF(同一訂單編號的「金額(原始，含稅)」加總) + 這一列的「運費
    # (原始)」——運費在來源檔案裡本來就是整張訂單複製貼到每一列，同一訂單每一列的值都相同，
    # 所以只取本列自己的運費值相加一次即可，不會因為訂單有多列而重複加總。已用使用者提供的
    # 3個實際案例反推驗證公式正確：681615(單筆)＝12699+1550=14249；683097(2列)＝
    # (6290+295)+1650=8235；685399(境內3列)＝(9890+959+599)+0=11448，皆與使用者回報的
    # 發票金額一致。插入這欄後，原本第8欄（合計含稅驗證拆帳用）起，整批往右移1欄。
    # 2026-08-12再新增：使用者確認訂單備註「XXX折xxx元」／折價券金額，要折抵進對應方案
    # 料號的拆帳單價內（見_compute_bookstore_order_matrix()裡的折扣分攤邏輯），最後方
    # 加一欄「折扣金額(已折抵，含稅)」顯示這筆訂單實際折抵掉的含稅金額，方便使用者核對
    # 折扣是否有正確套用（0代表沒有折扣，或折扣金額比料號總額大扣不完、已改列入下方
    # 「折扣待複核」工作表，未套用折扣）。
    # 2026-08-14再修正：使用者要求把「分析後發票」「分析後運費」這2欄從最右側移到
    # 「金額(原始，含稅)」後面（原本2026-08-14第一次新增時是刻意放在最右側，避免牽動
    # 前面欄位的公式參照——使用者這次直接指定了插入位置，即使會牽動後面所有欄位的欄號
    # /公式參照，也照使用者指定的位置調整）。插入後，原本第8欄起（發票含稅金額）整批
    # 往右移2欄。
    base_headers = ['境內/境外', '訂單編號', '訂購日期', '方案', '方案中文', '數量',
                     '金額(原始，含稅)', '分析後發票', '分析後運費',
                     '發票含稅金額', '合計(含稅)驗證拆帳用', '驗證金額',
                     '運送地區中文(原始)', '運費(原始)',
                     '折價券(原始)', '訂單備註(原始)', '購物車折扣的金額(原始)',
                     '折扣金額(已折抵，含稅)']
    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    redfill = openpyxl.styles.PatternFill('solid', fgColor='FFC7CE')
    redfont = openpyxl.styles.Font(color='9C0006')
    # 2026-08-12新增：「總表分析」頁籤境外區塊底色（使用者要求境外呈現含稅、用橘色底跟境內
    # 未稅區塊做視覺區隔）；below_cost低於成本的儲存格仍優先蓋成紅色（見下方套用順序：
    # 先整列上橘色，below_cost的儲存格再蓋一次紅色，紅色優先權比橘色高）。
    orangefill = openpyxl.styles.PatternFill('solid', fgColor='FCE4D6')
    # 2026-08-12新增：使用者要求表頭列文字自動換行（1.2），改用這個取代原本只有水平置中、
    # 沒有換行的center樣式。
    header_wrap = openpyxl.styles.Alignment(horizontal='center', vertical='top', wrap_text=True)
    totals_fill = openpyxl.styles.PatternFill('solid', fgColor='D9E1F2')
    totals_font = openpyxl.styles.Font(bold=True)
    total_col = len(base_headers) + len(columns) + 1  # 最右側「合計(境內未稅/境外含稅)」欄
    # 2026-08-14：「分析後發票」「分析後運費」現在已經是base_headers裡固定第8/9欄
    # （見上方），不再是插在最右側的獨立欄位，這裡直接寫死欄號常數，取代原本
    # analysis_invoice_col/analysis_freight_col=total_col+1/+2的算法。
    ANALYSIS_INVOICE_COL = 8   # 「分析後發票」
    ANALYSIS_FREIGHT_COL = 9  # 「分析後運費」
    # 2026-08-12新增：欄位字母對照表（1~total_col），供下面組Excel公式字串用，
    # 比逐一呼叫ws.cell(...).column_letter清楚，也避免公式裡欄位字母寫死、跟base_headers
    # 改動時脫節。
    col_letter = {i: openpyxl.utils.get_column_letter(i) for i in range(1, total_col + 1)}

    for ci, h in enumerate(base_headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hfill
        c.font = hfont
        c.alignment = header_wrap
    for ci, name in enumerate(columns, len(base_headers) + 1):
        c = ws.cell(1, ci, name)
        c.fill = hfill
        c.font = hfont
        c.alignment = header_wrap
    # 2026-08-13改名：使用者確認境外商品/電子書方案欄位比照「總表分析」改顯示含稅後，
    # 這欄（連同前面每個財報品名欄）境內是未稅、境外是含稅，原本「合計(未稅)」的名稱
    # 已不準確，改名並在名稱裡標明基礎差異，避免使用者誤會兩個區塊的金額是同一種稅基。
    c = ws.cell(1, total_col, '合計(境內未稅/境外含稅)')
    c.fill = hfill
    c.font = hfont
    c.alignment = header_wrap

    # 2026-08-08新增：每個區塊（境內/境外）最下面加一列「XX小計」——每個財報料號品名欄位
    # 加總該區塊底下所有訂單的金額（境內未稅、境外含稅，見上方欄位改名說明），最右側這欄
    # 則是這些品名小計的總和，方便跟上方逐筆訂單的這欄交叉核對（同一區塊內基礎一致，可以
    # 直接加總；境內、境外兩個小計彼此之間不能直接比較，因為稅基不同）。
    r = 2
    for region in ('境內', '境外'):
        region_rows = [row for row in result['rows'] if row['region'] == region]
        col_totals = {name: 0 for name in columns}
        region_start_row = r
        for row in region_rows:
            ws.cell(r, 1, region)
            ws.cell(r, 2, row['order_no'])
            ws.cell(r, 3, row['order_date'])
            ws.cell(r, 4, row['plan_code'])
            ws.cell(r, 5, row['plan_name'])
            _set_int_cell(ws, r, 6, row['qty'])
            _set_int_cell(ws, r, 7, row['amount_ref'])
            # 2026-08-13新增：「發票含稅金額」＝完全不經過拆帳/運費分攤邏輯，只用SUMIF直接把
            # 同一訂單編號(B欄)所有列的「金額(原始，含稅)」(G欄)加總，再加上本列的「運費
            # (原始)」——運費同一訂單每列數值相同，只取本列這一次，不會重複加總。這是這筆
            # 訂單當初開立發票的原始含稅總額，是「整張訂單」層級的數字（同一訂單編號的每一列
            # 這裡都會顯示同一個總數）。注意這欄是「訂單」顆粒度，下面「合計(含稅)驗證拆帳用」
            # /「驗證金額」是「這一列自己」的顆粒度——訂單只買1個方案時兩者剛好對得起來，
            # 訂單買多個方案（分成多列）時，請把同一訂單編號底下每一列的「合計(含稅)驗證拆帳
            # 用」加總起來，再跟這欄比較，才是正確的核對方式（直接拿單一列去比對這欄，多筆
            # 商品的訂單會出現正常的假差異，不代表算錯）。
            # 2026-08-14再修正：「分析後發票」「分析後運費」現在移到col8/col9（緊接在「金額
            # (原始，含稅)」後面，使用者指定的位置），比它們所依賴的「發票含稅金額」(col10)/
            # 「運費(原始)」(col14)寫入順序早——這在Excel/openpyxl裡沒有問題，公式字串只是
            # 存進儲存格，實際計算是Excel開啟檔案時依公式裡的欄位參照解析，跟Python寫入的
            # 先後順序無關。用COUNTIF($B$2:B{r},B{r})=1判斷這一列是不是這張訂單編號「第一次
            # 出現」，是的話才把「發票含稅金額」/「運費(原始)」的值搬過來，不是第一次出現就
            # 顯示0，這樣同一張訂單不論拆成幾列，SUM這兩欄的結果都只會把訂單的發票/運費各
            # 算一次，可以拿來跟「總表分析」的境內/境外總計互相核對、抓出兩邊金額對不起來是
            # 哪張訂單的問題。
            ws.cell(r, ANALYSIS_INVOICE_COL,
                    f'=IF(COUNTIF($B$2:{col_letter[2]}{r},{col_letter[2]}{r})=1,{col_letter[10]}{r},0)')
            ws.cell(r, ANALYSIS_FREIGHT_COL,
                    f'=IF(COUNTIF($B$2:{col_letter[2]}{r},{col_letter[2]}{r})=1,{col_letter[14]}{r},0)')
            ws.cell(r, 10, f'=SUMIF($B:$B,{col_letter[2]}{r},$G:$G)+{col_letter[14]}{r}')
            # 2026-08-12新增：「合計(含稅)驗證拆帳用」＝這一列最右側「合計(未稅)」欄(這筆訂單
            # 所有拆帳品項未稅金額加總)×1.05換算回含稅，再加上這一列的原始運費——理論上應
            # 該要等於「金額(原始，含稅)」；「驗證金額」＝檢核用欄位－金額(原始，含稅)，>0
            # 代表拆帳分出去的金額比實際收費金額還多（常見成因：訂單有折扣/折價券，但目前
            # 系統還沒有把這筆折扣內容從對應料號單價扣除，見TODO：訂單備註/折價券折扣解析）。
            # 用Excel公式寫入（不是Python算好的死數字），方便使用者直接在Excel裡稽核。
            # 2026-08-13修正：使用者要求不要小數點、四捨五入到整數，改用ROUND(...,0)包住。
            # 2026-08-14再修正：欄號已改為第11欄（「分析後發票」「分析後運費」插到col8/col9後，
            # 原本的第9欄整批往右移2欄）。
            # 2026-08-13再修正（重要）：這裡原本無論境內境外都是「總計欄(未稅)×1.05+運費」，
            # 但2026-08-13新增的Pass B2運費拆帳，已經把運費直接併進「合計(未稅)」欄位本身
            # （見下面SKU欄位迴圈），這裡如果再加一次運費，等於運費被重複計算兩次——這正是
            # 使用者截圖回報681615「合計含稅驗證拆帳用」異常膨脹到15876的根因。改成不再另外
            # +運費：「合計(未稅)」欄本身境內是真未稅、境外(使用者確認比照總表分析)已經在
            # 下面SKU欄位迴圈裡換算成含稅，所以這裡只要依境內/境外決定要不要再乘1.05即可，
            # 換算完就已經包含運費在內，不需要再額外加一次。
            ws.cell(r, 11, f'=ROUND(IF({col_letter[1]}{r}="境內",{col_letter[total_col]}{r}*1.05,'
                           f'{col_letter[total_col]}{r}),0)')
            # 2026-08-13：這裡刻意維持跟「金額(原始，含稅)」(col7，只有商品金額、不含運費)
            # 比對，沒有改成比對「發票含稅金額」(col10，訂單層級加總)——col11是「這一列自己」
            # 的拆帳明細換算，col10卻是「整張訂單」的加總，兩者顆粒度不同，多筆商品的訂單裡
            # 直接拿col11-col10會在兩列都出現不成立的假差異(見下方說明)，比原本的「假差異」更
            # 難懂。運費只會拆進「這張訂單商品價格最高」的那唯一一列，所以：這一列如果剛好
            # 是電子書方案格子有淡藍色底(row.get('freight_allocated_untaxed')有值)的那一列，
            # 這裡算出來的差額約等於這張訂單的運費金額，是預期中的正常現象，不是拆帳算錯；
            # 同一張訂單裡「沒有」淡藍色底的其他列，這裡應該要接近0，若不是0才代表真的有異常。
            # 如果要用「發票含稅金額」核對整張訂單(含多筆商品)有沒有算對，請改看「境內/境外
            # 小計」列，或手動把同一訂單編號的所有列col11加總後跟其中一列的col10比較。
            # 2026-08-14再修正：欄號已改為第12欄（原第10欄，理由同上）。
            ws.cell(r, 12, f'={col_letter[11]}{r}-{col_letter[7]}{r}')
            ws.cell(r, 13, row['region_raw'])
            _set_int_cell(ws, r, 14, row['freight_ref'])
            _set_int_cell(ws, r, 15, row['coupon_ref'])
            ws.cell(r, 16, row['order_note_ref'])
            _set_int_cell(ws, r, 17, row['cart_discount_ref'])
            dc = _set_int_cell(ws, r, 18, row['discount_applied_incl'])
            if row['discount_applied_incl']:
                dc.fill = openpyxl.styles.PatternFill('solid', fgColor='FFF2CC')  # 淡黃：已折抵
            # 2026-08-13新增：使用者確認「書店訂單拆帳」明細表的境外商品/電子書方案欄位，
            # 比照「總表分析」樞紐既有的做法改顯示含稅（境內維持顯示未稅不變）——內部資料
            # (cell['untaxed'])仍然是真未稅、below_cost判斷仍用未稅價比較，只有寫進Excel
            # 儲存格的「顯示值」，境外這裡额外乘回1.05換算成含稅。境外整列額外套上橘色底
            # （below_cost紅色、運費淡藍色的優先權都比橘色高，跟「總表分析」的疊色順序一致），
            # 提醒使用者這幾欄的金額基礎跟境內不同。
            for ci, name in enumerate(columns, len(base_headers) + 1):
                cell = row['cells'].get(name)
                if cell is None:
                    continue
                display_val = cell['untaxed'] * 1.05 if region == '境外' else cell['untaxed']
                c = _set_int_cell(ws, r, ci, display_val)
                if region == '境外':
                    c.fill = orangefill
                if cell['below_cost']:
                    c.fill = redfill
                    c.font = redfont
                # 2026-08-13新增：這一列如果是該訂單商品價格最高、被拆進運費的那一列，
                # 「電子書方案」這格用淡藍色標示，方便使用者肉眼確認運費有沒有正確拆進來
                # （below_cost紅色跟這個不會同時發生——電子書方案是集合類品名，設計上不做
                # 低於成本檢核，見cells建構邏輯）。
                if name == '電子書方案' and row.get('freight_allocated_untaxed'):
                    c.fill = openpyxl.styles.PatternFill('solid', fgColor='DDEBF7')
                col_totals[name] += display_val
            row_display_total = row['untaxed_total'] * 1.05 if region == '境外' else row['untaxed_total']
            tcell = _set_int_cell(ws, r, total_col, row_display_total)
            if region == '境外':
                tcell.fill = orangefill
            r += 1
        if region_rows:
            region_end_row = r - 1
            ws.cell(r, 5, region + '小計')
            # 2026-08-12新增：使用者要求境內/境外小計列，把原本留白的「金額(原始，含稅)」
            # 「合計(含稅)驗證拆帳用」「驗證金額」「運費(原始)」「折價券(原始)」「購物車折扣的
            # 金額(原始)」「折扣金額(已折抵，含稅)」這幾個金額欄也加總填入；文字欄(運送地區
            # 中文/訂單備註)不是金額，加總沒有意義，維持留白。用SUM公式（不是Python算好的
            # 死數字），方便稽核。
            # 2026-08-14再修正：「分析後發票」「分析後運費」(col8/col9)移到這裡跟其他金額欄
            # 一起加總小計——每張訂單只在第一次出現那一列有數字，其他列是0，所以SUM整個
            # 區塊等於這個區塊「每張訂單各算一次」的真總額，可以直接拿來跟「總表分析」的
            # 境內/境外總計核對；其餘欄號因為插入col8/col9，整批往右移2欄（原7,8,9,10,12,
            # 13,15,16→現在7,10,11,12,14,15,17,18）。
            for ci in (7, ANALYSIS_INVOICE_COL, ANALYSIS_FREIGHT_COL, 10, 11, 12, 14, 15, 17, 18):
                L = col_letter[ci]
                ws.cell(r, ci, f'=SUM({L}{region_start_row}:{L}{region_end_row})')
            for ci, name in enumerate(columns, len(base_headers) + 1):
                _set_int_cell(ws, r, ci, col_totals[name])
            _set_int_cell(ws, r, total_col, sum(col_totals.values()))
            for ci in range(1, total_col + 1):
                cc = ws.cell(r, ci)
                cc.font = totals_font
                cc.fill = totals_fill
            r += 1
    last_data_row = r - 1

    # 2026-08-12新增：「驗證金額」欄位>0時整格標紅（見上方欄位說明），用Excel條件式格式設定
    # 而非Python直接上色——因為這欄寫的是公式字串，openpyxl存檔當下不會執行公式、無法預先
    # 知道算出來的值，只能交給Excel開啟時依公式實際結果動態判斷。
    if last_data_row >= 2:
        # 2026-08-14再修正：「分析後發票」「分析後運費」插到col8/col9後，「驗證金額」欄
        # 從第10欄移到第12欄，條件式格式設定的欄位參照要同步改。
        ws.conditional_formatting.add(
            f'{col_letter[12]}2:{col_letter[12]}{last_data_row}',
            openpyxl.formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                 fill=redfill, font=redfont))

    # 2026-08-14再修正：「分析後發票」「分析後運費」寬度(16,16)從最右側移到第8/9欄
    # （緊接在「金額(原始，含稅)」寬度14後面），base_headers其餘寬度依新欄位順序排列，
    # 最右側只留「合計(境內未稅/境外含稅)」1欄的寬度(14)，不再有後面2欄。
    widths = ([10, 14, 12, 10, 30, 8, 14, 16, 16, 16, 16, 16, 16, 10, 12, 20, 18, 18]
              + [14] * len(columns) + [14])
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
    ws.freeze_panes = ws.cell(2, len(base_headers) + 1).coordinate

    if result['unmatched']:
        ws2 = wb.create_sheet('未比對到方案')
        ws2.append(['訂單編號', '方案', '方案中文(訂單上)', '數量', '境內/境外'])
        for u in result['unmatched']:
            ws2.append([u['order_no'], u['plan_code'], u.get('plan_name_ref'), u['qty'], u['region']])

    # 2026-08-12新增：折扣金額比料號含稅總額還大扣不完、或訂單備註有「折」字但抓不到
    # 「折(數字)元」格式的訂單，都不會自動套用折扣，另外列在這個工作表供人工複核，避免
    # 使用者沒發現有訂單漏套用折扣。
    if result.get('discount_review_rows'):
        ws3 = wb.create_sheet('折扣待複核')
        ws3.append(['訂單編號', '方案', '境內/境外', '金額(原始，含稅)', '折價券(原始)',
                    '訂單備註(原始)', '複核原因'])
        for row in result['discount_review_rows']:
            ws3.append([row['order_no'], row['plan_code'], row['region'], row.get('amount_ref'),
                        row.get('coupon_ref'), row.get('order_note_ref'),
                        '；'.join(row.get('discount_review_reasons') or [])])
        for ci, w in enumerate([14, 10, 10, 16, 14, 30, 40], 1):
            ws3.column_dimensions[ws3.cell(1, ci).column_letter].width = w

    # 2026-08-13新增：同一訂單編號底下各列解析到的運費金額不一致的訂單，不會自動拆運費
    # 進電子書方案，另外列在這個工作表供人工複核，避免使用者沒發現有訂單漏拆運費。
    if result.get('freight_review_rows'):
        ws3b = wb.create_sheet('運費待複核')
        ws3b.append(['訂單編號', '方案', '境內/境外', '金額(原始，含稅)', '運費(原始)', '複核原因'])
        for row in result['freight_review_rows']:
            ws3b.append([row['order_no'], row['plan_code'], row['region'], row.get('amount_ref'),
                         row.get('freight_ref'), '；'.join(row.get('freight_review_reasons') or [])])
        for ci, w in enumerate([14, 10, 10, 16, 14, 40], 1):
            ws3b.column_dimensions[ws3b.cell(1, ci).column_letter].width = w

    if result['inconsistent_cost_names']:
        ws3 = wb.create_sheet('進貨價不一致警示')
        ws3.append(['財報料號品名（底下內部料號進貨未稅價不一致，標紅結果可能不準確，請核對）'])
        for name in result['inconsistent_cost_names']:
            ws3.append([name])

    # 2026-08-08新增：「總表分析」頁籤——樣式參考使用者提供的歷史Excel[0701-0727A88]頁籤。
    # 版面：A欄＝境內/境外，B欄＝這一列的類型(單價(未稅)/數量)，C欄起每個報表品名各佔一欄，
    # 最右側再加一欄「合計(未稅)」。
    # 第2列＝該品名的內部料號未稅進貨價；第3列＝品名標題；
    # 第4列起：境內、境外分開兩區塊，同一品名底下「拆帳單價相同」的訂單合併成一列、數量加總；
    # 不同品名各自依自己的單價種類數往下疊，最長的品名決定這個區塊要幾列（等於「不夠可以
    # 往下延伸」）；代銷商品/電子書方案這2個集合類品名，不分單價，全部加總成一列、數量固定1。
    # 再往下：「料號加總(境內+境外)」是每個品名欄位的總計；接著境內、境外各自一列「金額合計」，
    # 最右側「合計(未稅)」欄是這個區塊所有品名合計的總和——這兩個數字必須跟「書店訂單拆帳」
    # 頁籤裡境內/境外小計列最右邊「合計(未稅)」欄一致，可互相核對。
    ws4 = wb.create_sheet('總表分析')
    BASE = 2  # A:境內/境外, B:列類型
    data_start = BASE + 1
    total_col2 = data_start + len(columns)

    # 2026-08-12修正：使用者要求境內顯示未稅、境外顯示含稅(橘色底區塊)後，本頁籤境外的
    # 金額已不再等於「書店訂單拆帳」頁籤境外小計的「合計(未稅)」欄(差1.05倍)，只有境內
    # 還能互相核對，說明文字同步更新，避免使用者誤以為兩邊全部都要一致。
    ws4.cell(1, 1, '總表分析（依拆帳單價分組加總數量；境內為未稅、境外為含稅(橘色底)；'
                    '代銷商品/電子書方案為集合類品名，全部加總後數量以1呈現；'
                    '境內數字應與「書店訂單拆帳」頁籤境內小計加總一致，境外因換算含稅不再相等）')

    ws4.cell(2, 2, '未稅進貨價')
    for ci, name in enumerate(columns, data_start):
        cost = pivot['cost_by_name'].get(name)
        if cost is not None:
            _set_int_cell(ws4, 2, ci, cost)

    # 2026-08-12新增：使用者要求[列3]表頭自動換行（1.3）——境內/境外、項目、合計(未稅)這3個
    # 固定欄位原本沒有套用wrap（只有中間每個財報品名欄有），這次一併補上，統一套用同一個
    # wrap樣式。
    wrap = openpyxl.styles.Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws4.cell(3, 1, '境內/境外').fill = hfill
    ws4.cell(3, 1).font = hfont
    ws4.cell(3, 1).alignment = wrap
    ws4.cell(3, 2, '項目').fill = hfill
    ws4.cell(3, 2).font = hfont
    ws4.cell(3, 2).alignment = wrap
    for ci, name in enumerate(columns, data_start):
        c = ws4.cell(3, ci, name)
        c.fill = hfill
        c.font = hfont
        c.alignment = wrap
    # 2026-08-12修正：這欄現在境內列是未稅小計、境外列是含稅小計(同一欄、依所在列基礎不同)，
    # 標題不再寫死「合計(未稅)」，避免誤導——每一列自己的B欄標籤（單價(未稅)/單價(含稅)）
    # 已經清楚標示這一列的稅別基礎。
    c = ws4.cell(3, total_col2, '小計')
    c.fill = hfill
    c.font = hfont
    c.alignment = wrap

    col_letter_first = ws4.cell(1, data_start).column_letter
    col_letter_last = ws4.cell(1, total_col2 - 1).column_letter
    r = 4
    for region in ('境內', '境外'):
        # 2026-08-12新增：境內顯示未稅、境外顯示含稅——這裡的「單價」列標籤跟著區塊切換；
        # 境外整個區塊(單價列+數量列)先鋪一層橘色底，跟境內做視覺區隔，below_cost的紅色
        # 標記仍優先於橘色（下面per-cell迴圈會在橘色之上再蓋一次紅色）。
        price_label = '單價(含稅)' if region == '境外' else '單價(未稅)'
        row_fill = orangefill if region == '境外' else None
        max_t = pivot['max_tiers'][region]
        for t in range(max_t):
            r_price = r
            ws4.cell(r, 1, region)
            ws4.cell(r, 2, price_label)
            if row_fill:
                for ci in range(1, total_col2 + 1):
                    ws4.cell(r, ci).fill = row_fill
            for ci, name in enumerate(columns, data_start):
                tiers = pivot['tiers_by_region'][name][region]
                if t < len(tiers):
                    tier = tiers[t]
                    c = _set_int_cell(ws4, r, ci, tier['price'])
                    # 2026-08-11新增：單價低於「內部料號清單」進貨未稅價時標紅，比照「書店
                    # 訂單拆帳」明細頁籤既有規則——之前這個頁籤完全沒有這段判斷。
                    if tier.get('below_cost'):
                        c.fill = redfill
                        c.font = redfont
            r += 1
            r_qty = r
            ws4.cell(r, 2, '數量')
            if row_fill:
                for ci in range(1, total_col2 + 1):
                    ws4.cell(r, ci).fill = row_fill
            for ci, name in enumerate(columns, data_start):
                tiers = pivot['tiers_by_region'][name][region]
                if t < len(tiers):
                    tier = tiers[t]
                    c = _set_int_cell(ws4, r, ci, tier['qty'])
                    if tier.get('below_cost'):
                        c.fill = redfill
                        c.font = redfont
            r += 1
            # 2026-08-11新增：紅框那欄「合計(未稅)」原本這幾列都是空的，使用者要求補上
            # 加總——「單價(未稅)」列用SUMPRODUCT(單價列×對應數量列)算出這個單價層級的
            # 金額小計。用Excel公式寫入（不是Python算好的死數字），方便使用者直接在Excel裡
            # 稽核每一格怎麼來的。
            # 2026-08-12修正：「數量」列（3.1）使用者要求這一橫列不用小計——不同財報品名的
            # 「件數」單位不一致（有的是1份方案、有的是1個配件），橫向加總沒有意義，改成
            # 不寫入這一格（維持空白），只保留「單價(未稅)」列的SUMPRODUCT金額小計。
            ws4.cell(r_price, total_col2,
                     f'=SUMPRODUCT({col_letter_first}{r_price}:{col_letter_last}{r_price},'
                     f'{col_letter_first}{r_qty}:{col_letter_last}{r_qty})')

    ws4.cell(r, 2, '料號加總(境內+境外)')
    for ci, name in enumerate(columns, data_start):
        _set_int_cell(ws4, r, ci, pivot['col_combined_total'][name])
    for ci in range(1, total_col2 + 1):
        cc = ws4.cell(r, ci)
        cc.font = totals_font
        cc.fill = totals_fill
    r += 1

    # 2026-08-12新增、2026-08-13移除：原本這裡在「料號加總(境內+境外)」下方額外加「運費
    # 境內小計」「運費境外小計」2列參考數字，是當時對「運費怎麼拆」需求還不夠明確時的暫定
    # 做法（獨立算一次運費總額，不併入任何一列的正式金額，純參考）。2026-08-13使用者完整
    # 確認實際規則後（見_compute_bookstore_order_matrix()裡的Pass B2）：運費應該整筆拆進
    # 「該訂單商品價格最高那一筆方案」的電子書方案欄位——這個拆帳結果已經是
    # _compute_bookstore_order_matrix()回傳的row['cells']資料本身的一部分，上面
    # tiers_by_region/col_combined_total/col_region_total的加總邏輯是直接從這份資料算出來
    # 的，會自動把這筆運費含在正確的「電子書方案」小計/金額合計裡，不需要再另外算一次、
    # 另外列一個參考小計（那樣反而會讓人誤以為運費要另外加總，跟已經正確併入小計的金額
    # 搞混）。運費有沒有正確拆進去，請到「書店訂單拆帳」頁籤看電子書方案欄位裡標淡藍色底
    # 的那一格（該筆訂單商品價格最高的那一列）。

    r += 1  # 空一列

    for region in ('境內', '境外'):
        # 2026-08-12新增：境外這列標籤改成「金額小計(含稅)」（跟境內「金額合計(未稅)」區隔，
        # 也對應這裡的數字現在是換算回含稅後的金額），底色也改用橘色跟上方境外區塊呼應。
        ws4.cell(r, 1, region)
        ws4.cell(r, 2, '金額小計(含稅)' if region == '境外' else '金額合計(未稅)')
        for ci, name in enumerate(columns, data_start):
            _set_int_cell(ws4, r, ci, pivot['col_region_total'][region][name])
        _set_int_cell(ws4, r, total_col2, pivot['grand_total'][region])
        for ci in range(1, total_col2 + 1):
            cc = ws4.cell(r, ci)
            cc.font = totals_font
            cc.fill = orangefill if region == '境外' else totals_fill
        r += 1

    ws4.column_dimensions['A'].width = 10
    ws4.column_dimensions['B'].width = 18
    for ci in range(data_start, total_col2 + 1):
        ws4.column_dimensions[ws4.cell(3, ci).column_letter].width = 14
    ws4.freeze_panes = ws4.cell(4, data_start).coordinate

    # 2026-08-12新增：使用者要求匯出Excel所有表格的文字對齊方式統一改成「垂直靠上」(1.1)。
    # 用一個共用的收尾動作，掃過每張有資料的工作表已使用範圍的每一格，只把vertical改成
    # 'top'、其餘既有的水平置中/自動換行設定原樣保留（openpyxl的Alignment是唯讀物件，
    # 不能只改一個屬性，要整個重新建立一份）。放在所有內容都寫完、最後存檔前執行，才不會
    # 被前面任何一段還沒執行到的alignment設定覆蓋掉。
    for _ws in wb.worksheets:
        _apply_vertical_top(_ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='書店訂單拆帳.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# ⑩ 書店經銷商 — 匯入團體發票（純登記查詢用，不跟訂單/方案清單比對金額）
# ============================================================
# 2026-08-12新增：登記開給團體客戶(學校/機關等)的發票，供之後查詢/核對用。跟「匯入訂單」
# 的拆帳分析是兩件獨立的事，這裡不做任何金額比對，單純CRUD+批次匯入。「未稅價」一律由
# 後端算好（價格(含稅)÷1.05），不接受前端傳入的值；「對帳區間」批次匯入時整批共用同一個
# 使用者輸入值，不是從Excel檔案裡讀取。

def _round2(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _parse_invoice_date(v):
    """發票日期欄位解析：Excel日期格式的儲存格經openpyxl(data_only=True)讀出來已經是
    datetime/date物件，直接格式化；字串格式(如'2026/7/13'、'2026-07-13')也一併容錯，
    比照全系統其他日期欄位的解析慣例。無法解析則回傳None（該列此欄留空，不擋整列匯入）。"""
    if v is None or v == '':
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


# ============================================================
# 2026-08-12新增：「匯入團體發票」出貨單附件（可多筆，PDF/.docx/.odt，單檔≤2MB）
# ============================================================
# 檔案本體存進 Supabase Storage（bucket名稱見下方常數，需使用者先手動在Supabase後台建立，
# 建議設為private——這裡走後端代理下載，不直接把bucket設public對外開放，權限一樣掛
# bookstore_invoice_required），bookstore_group_invoice_files表只存中繼資料(檔名/路徑/
# 大小)，不是這個模組原本既有的做法，是這次全新加上去的，日後如果其他模組也要做檔案
# 上傳，可以直接抄這一段的bucket/表設計方式。
_SHIPPING_DOC_BUCKET = 'bookstore-shipping-docs'
_SHIPPING_DOC_ALLOWED_EXT = {'.pdf', '.docx', '.odt'}
_SHIPPING_DOC_MAX_SIZE = 2 * 1024 * 1024  # 使用者確認的單檔大小上限：2MB


def _shipping_doc_ext_ok(filename):
    if not filename or '.' not in filename:
        return False
    ext = '.' + filename.rsplit('.', 1)[-1].lower()
    return ext in _SHIPPING_DOC_ALLOWED_EXT


@dealer_bp.route('/api/bookstore/group-invoices/<int:iid>/files', methods=['GET'])
@bookstore_invoice_required
def list_group_invoice_files(iid):
    rows = sb.table('bookstore_group_invoice_files').select('*') \
        .eq('invoice_id', iid).order('created_at').execute()
    return jsonify(rows.data or [])


@dealer_bp.route('/api/bookstore/group-invoices/<int:iid>/files', methods=['POST'])
@bookstore_invoice_required
def upload_group_invoice_files(iid):
    """支援一次選多檔（前端用同一個input multiple，或分次呼叫都可以，這裡不限制單次
    呼叫的檔案數）。每個檔案獨立檢查格式/大小，某幾個檔案沒過不會擋下其他檔案，回傳的
    uploaded/errors讓前端可以顯示「3個成功、1個失敗：xxx」這種混合結果。"""
    inv = sb.table('bookstore_group_invoices').select('id').eq('id', iid).limit(1).execute()
    if not inv.data:
        return jsonify({'error': '找不到這筆發票，請先儲存發票後再上傳出貨單'}), 404
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '請選擇要上傳的檔案'}), 400
    uploaded, errors = [], []
    for f in files:
        if not f or not f.filename:
            continue
        if not _shipping_doc_ext_ok(f.filename):
            errors.append(f'{f.filename}：僅接受 PDF、Word(.docx)、OpenDocument(.odt) 格式')
            continue
        data = f.read()
        if len(data) > _SHIPPING_DOC_MAX_SIZE:
            errors.append(f'{f.filename}：檔案超過2MB上限')
            continue
        if not data:
            errors.append(f'{f.filename}：檔案是空的，略過')
            continue
        # 2026-08-12新增：storage路徑用uuid當前綴避免同名檔案互相覆蓋，原始檔名維持
        # 只存在DB的file_name欄位（下載時用這個當顯示檔名），不受storage路徑限制。
        safe_name = secure_filename(f.filename) or 'file'
        storage_path = f'{iid}/{uuid.uuid4().hex}_{safe_name}'
        try:
            sb.storage(_SHIPPING_DOC_BUCKET).upload(
                storage_path, data, f.mimetype or 'application/octet-stream')
        except Exception as e:
            errors.append(f'{f.filename}：上傳失敗（{e}）')
            continue
        rec = {
            'invoice_id': iid, 'file_name': f.filename, 'storage_path': storage_path,
            'file_size': len(data),
            'uploaded_by': session.get('user_id'),
            'uploaded_by_name': session.get('display_name', session.get('username', '')),
        }
        try:
            res = sb.table('bookstore_group_invoice_files').insert(rec).execute()
            uploaded.append(res.data[0])
        except Exception as e:
            # DB紀錄寫入失敗，storage那份孤兒檔要清掉，避免佔空間又查不到、只能手動去
            # Supabase後台清理。
            try:
                sb.storage(_SHIPPING_DOC_BUCKET).remove([storage_path])
            except Exception:
                pass
            errors.append(f'{f.filename}：儲存紀錄失敗（{e}）')
    return jsonify({'uploaded': uploaded, 'errors': errors})


@dealer_bp.route('/api/bookstore/group-invoices/files/<int:fid>/download', methods=['GET'])
@bookstore_invoice_required
def download_group_invoice_file(fid):
    """bucket刻意設計成private，不直接給前端Supabase Storage的公開網址，一律透過這個
    後端路由代理下載——確保沒登入/沒權限的人拿不到出貨單檔案（跟其他API一樣掛
    bookstore_invoice_required）。"""
    row = sb.table('bookstore_group_invoice_files').select('*').eq('id', fid).limit(1).execute()
    if not row.data:
        return jsonify({'error': '找不到這個檔案'}), 404
    rec = row.data[0]
    try:
        data = sb.storage(_SHIPPING_DOC_BUCKET).download(rec['storage_path'])
    except Exception as e:
        return jsonify({'error': f'下載失敗：{e}'}), 500
    return send_file(io.BytesIO(data), as_attachment=True, download_name=rec['file_name'])


@dealer_bp.route('/api/bookstore/group-invoices/files/<int:fid>', methods=['DELETE'])
@bookstore_invoice_required
def delete_group_invoice_file(fid):
    row = sb.table('bookstore_group_invoice_files').select('*').eq('id', fid).limit(1).execute()
    if not row.data:
        return jsonify({'error': '找不到這個檔案'}), 404
    rec = row.data[0]
    try:
        sb.storage(_SHIPPING_DOC_BUCKET).remove([rec['storage_path']])
    except Exception:
        pass  # storage刪除失敗不擋DB紀錄刪除，避免留下使用者點了會404的殭屍紀錄
    sb.table('bookstore_group_invoice_files').delete().eq('id', fid).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/group-invoices', methods=['GET'])
@bookstore_invoice_required
def list_group_invoices():
    # 2026-08-12調整：改依「匯入順序」排序，比照方案清單list_bookstore_plans()的做法（見
    # dealer_reconcile_schema.sql v1.22）——最近一次批次匯入的整批排最上面，批次內部照
    # Excel原始列順序呈現；用「新增發票」單筆新增的資料（import_seq/last_imported_at皆
    # 為空）自然排到最後面，用id做穩定排序的最後依據。
    rows = _fetch_all_rows(lambda: sb.table('bookstore_group_invoices').select('*')
                            .order('last_imported_at', desc=True, nullslast=True)
                            .order('import_seq', nullslast=True)
                            .order('id'))
    # 2026-08-12新增：列表頁要在每一列顯示已上傳的出貨單附件（可多筆），這裡一次把全部
    # 附件metadata抓回來依invoice_id分組掛回對應那一列，避免前端每一列各自呼叫一次
    # /files API造成N+1次請求（發票筆數多的話會很慢）。
    # 2026-08-12修正：這裡原本沒有try/except——如果使用者還沒執行v1.24的建表SQL（
    # bookstore_group_invoice_files尚未存在），這段查詢會丟例外，導致整支API連本來就
    # 存在的發票資料都一起回傳失敗（前端loadGroupInvoices()的catch會把這個當成清單是空的
    # 處理，畫面顯示「共0筆」，讓人誤以為資料不見了——其實資料還在，只是這個新查詢炸掉
    # 拖累了整個API）。改成失敗時容錯：附件功能暫時顯示不出來，但既有發票清單一定要正常
    # 顯示，不能因為一個新功能的資料表還沒建，就讓舊功能整個掛掉。
    try:
        file_rows = _fetch_all_rows(lambda: sb.table('bookstore_group_invoice_files')
                                     .select('id,invoice_id,file_name,file_size,created_at')
                                     .order('created_at'))
    except Exception:
        file_rows = []
    files_by_invoice = {}
    for fr in file_rows:
        files_by_invoice.setdefault(fr['invoice_id'], []).append({
            'id': fr['id'], 'file_name': fr['file_name'],
            'file_size': fr.get('file_size'), 'created_at': fr.get('created_at'),
        })
    for r in rows:
        r['files'] = files_by_invoice.get(r['id'], [])

    # 2026-08-12新增：同理把Gmail自動抓/暫存區手動附加的「發票PDF」也一次分組掛回每一列
    # （比照上面出貨單files的做法，一樣包try/except容錯，理由相同：新表未建不能拖垮舊功能）。
    try:
        pdf_rows = _fetch_all_rows(lambda: sb.table('bookstore_group_invoice_pdfs')
                                    .select('id,invoice_id,file_name,file_size,source,created_at')
                                    .order('created_at'))
    except Exception:
        pdf_rows = []
    pdfs_by_invoice = {}
    for pr in pdf_rows:
        pdfs_by_invoice.setdefault(pr['invoice_id'], []).append({
            'id': pr['id'], 'file_name': pr['file_name'], 'file_size': pr.get('file_size'),
            'source': pr.get('source'), 'created_at': pr.get('created_at'),
        })
    for r in rows:
        r['invoice_pdfs'] = pdfs_by_invoice.get(r['id'], [])

    # 2026-08-18新增：已寫入對帳的發票，回填「拆帳編號」（見dealer_reconcile_batches.
    # recon_seq_no）——匯入時這筆發票對應的批次的upload_group存在reconciled_upload_group
    # 欄位（見reconcile_import_from_group_invoices()），這裡反查同一個upload_group底下
    # 任一批次的recon_seq_no（同一組批次一定共用同一個編號，見_ingest_reconcile_file()），
    # 供列表頁「已寫入對帳」旁顯示，方便對照是哪一次批次。
    # 2026-08-21：抽成共用函式_attach_recon_seq_info()，供list_shipping_docs()共用。
    _attach_recon_seq_info(rows)

    return jsonify(rows)


# 2026-08-17新增：使用者反映「Gmail自動抓發票PDF排程明明有掃到信，發票PDF卻一直沒
# 自動掛上」，追查後找到一種常見成因（不是排程本身壞掉）：Gmail排程掃到信件的當下，
# `_process_one_invoice_email()`(見下方Gmail排程區塊) 只有在這個發票號碼「當時」就已經
# 登記在`bookstore_group_invoices`裡，才會自動掛檔；沒登記的話，PDF只能先存進
# `bookstore_invoice_staging`（狀態'pending'），等使用者自己到「檢視暫存區」手動點
# 「附加」。問題是：watermark水位機制（見project_gmail_invoice_autofetch筆記）會讓
# 已經檢查過的信件UID永遠不會再被重新掃描——如果使用者是「先收到Gmail通知信，過了
# 幾天才把這張發票登記進系統」（例如這次批次匯入/單筆新增都是事後才做），那筆PDF就會
# 永遠卡在暫存區「待處理」，排程再怎麼跑也不會主動追溯過去已經跳過的信件、不會自動補上。
# 這支函式補上這個缺口：在新增/編輯/批次匯入發票、確定這筆發票有填發票號碼的當下，
# 回頭查一次暫存區有沒有現成、狀態還是pending的同號碼PDF，有的話直接幫使用者掛上，
# 不用每次都得自己想到要去暫存區手動找。找不到符合的暫存資料時安靜地什麼都不做（多數
# 發票本來就沒有對應的暫存PDF，這是正常情況，不能讓查詢失敗拖垮發票新增/編輯本身）。
def _auto_attach_pending_staging(invoice_id, invoice_no, user_id=None, user_name=None):
    if not invoice_no:
        return
    try:
        # 2026-08-17再修正：使用者實測回報，重新儲存後「發票PDF」欄位出現同一個檔名
        # 重複2份——根因是同一個invoice_no在暫存區其實有不只1筆status='pending'的紀錄
        # （這張發票的通知信在「還沒登記進系統」前被轉寄/處理了不只一次；bug6原本的重複
        # 檔名比對只有在「當下已登記」時才會生效，未登記時完全沒有去重機制，每一次都各自
        # 存成一筆獨立的pending，見上方函式說明）。改成用`.order('id')`固定拿到「最早」
        # 那筆，只有它會被實際掛檔；其餘同發票號碼的pending紀錄視為重複，只更新狀態、
        # 不會再各自呼叫一次insert()，避免「匯入團體發票」列表同一張發票的PDF重複顯示。
        pending = sb.table('bookstore_invoice_staging').select('*') \
            .eq('invoice_no', invoice_no).eq('status', 'pending').order('id').execute()
    except Exception:
        return  # 暫存區資料表可能還沒建或查詢失敗，不影響發票本身新增/編輯的主流程
    rows = pending.data or []
    if not rows:
        return
    first, rest = rows[0], rows[1:]
    try:
        sb.table('bookstore_group_invoice_pdfs').insert({
            'invoice_id': invoice_id, 'file_name': f"{invoice_no}.pdf",
            'storage_path': first['storage_path'], 'source': 'staging_backfill',
        }).execute()
        sb.table('bookstore_invoice_staging').update({
            'status': 'attached', 'attached_invoice_id': invoice_id,
            'attached_at': now_str(), 'attached_by': user_id,
            'attached_by_name': user_name or '系統自動回填',
        }).eq('id', first['id']).execute()
    except Exception:
        return  # 掛檔失敗就整批維持pending，不去動其餘重複紀錄的狀態，之後仍可重試
    for s in rest:
        try:
            sb.table('bookstore_invoice_staging').update({
                'status': 'duplicate_skipped', 'attached_invoice_id': invoice_id,
                'attached_at': now_str(),
            }).eq('id', s['id']).execute()
        except Exception:
            pass  # 標記失敗不影響主流程，這筆頂多留在暫存區列表繼續顯示pending


@dealer_bp.route('/api/bookstore/group-invoices', methods=['POST'])
@bookstore_invoice_required
def create_group_invoice():
    data = request.json or {}
    seller_name = _norm(data.get('seller_name'))
    if not seller_name:
        return jsonify({'error': '銷售單位為必填'}), 400
    amount_incl = _to_num_or_none(data.get('amount_incl'))
    if amount_incl is None:
        return jsonify({'error': '價格(含稅)為必填，且需為數字'}), 400
    invoice_no = _norm(data.get('invoice_no')) or None
    # 2026-08-12新增：發票號碼不可重複登記（空白不受此限制，允許多筆都留空）。
    if invoice_no:
        dup = sb.table('bookstore_group_invoices').select('id').eq('invoice_no', invoice_no).limit(1).execute()
        if dup.data:
            return jsonify({'error': f'發票號碼「{invoice_no}」已存在，不可重複登記'}), 400
    # 2026-08-15新增：經銷商類型——供「上傳比對」勾選匯入使用，選填，須為DEALERS清單內的值。
    dealer_name = _norm(data.get('dealer_name')) or None
    if dealer_name and dealer_name not in DEALERS:
        return jsonify({'error': f'經銷商類型「{dealer_name}」不在允許清單內'}), 400
    rec = {
        'seller_name': seller_name,
        'dealer_name': dealer_name,
        'order_no': _norm(data.get('order_no')) or None,
        'shipping_order_no': _norm(data.get('shipping_order_no')) or None,
        'invoice_date': _parse_invoice_date(data.get('invoice_date')),
        'invoice_no': invoice_no,
        'amount_incl': amount_incl,
        'amount_notax': _round2(amount_incl / 1.05),
        'note': _norm(data.get('note')) or None,
        'recon_period': _norm(data.get('recon_period')) or None,
        # 2026-08-17新增：應收憑單/請購單號，選填純文字，供會計流程記錄用（見schema v1.31）
        'receivable_doc_no': _norm(data.get('receivable_doc_no')) or None,
        'created_by': session.get('user_id'),
        'created_by_name': session.get('display_name', session.get('username', '')),
    }
    try:
        res = sb.table('bookstore_group_invoices').insert(rec).execute()
        new_id = res.data[0]['id']
        # 2026-08-17新增：新增發票時如果填了發票號碼，順便檢查暫存區有沒有現成PDF可以
        # 直接掛上（見_auto_attach_pending_staging()說明）。
        _auto_attach_pending_staging(new_id, invoice_no, session.get('user_id'),
                                      session.get('display_name', session.get('username', '')))
        return jsonify({'ok': True, 'id': new_id}), 201
    except Exception as e:
        return jsonify({'error': f'新增失敗：{e}'}), 500


# 2026-08-12新增：補上編輯功能（v1.20上線時使用者只要求新增/批次匯入，這次補齊修改）。
# 邏輯跟create_group_invoice()幾乎一樣（同樣的必填驗證、未稅價一律後端重新算好，不接受
# 前端傳入值），差異只在改成update並補上異動人資訊（見dealer_reconcile_schema.sql v1.21
# 新增的updated_at/updated_by/updated_by_name欄位）。
@dealer_bp.route('/api/bookstore/group-invoices/<int:iid>', methods=['PUT'])
@bookstore_invoice_required
def update_group_invoice(iid):
    data = request.json or {}
    seller_name = _norm(data.get('seller_name'))
    if not seller_name:
        return jsonify({'error': '銷售單位為必填'}), 400
    amount_incl = _to_num_or_none(data.get('amount_incl'))
    if amount_incl is None:
        return jsonify({'error': '價格(含稅)為必填，且需為數字'}), 400
    invoice_no = _norm(data.get('invoice_no')) or None
    # 2026-08-12新增：發票號碼不可重複（排除自己這一筆），空白不受此限制。
    if invoice_no:
        dup = sb.table('bookstore_group_invoices').select('id').eq('invoice_no', invoice_no).neq('id', iid).limit(1).execute()
        if dup.data:
            return jsonify({'error': f'發票號碼「{invoice_no}」已存在，不可重複登記'}), 400
    # 2026-08-15新增：經銷商類型——同create_group_invoice()的驗證規則。
    dealer_name = _norm(data.get('dealer_name')) or None
    if dealer_name and dealer_name not in DEALERS:
        return jsonify({'error': f'經銷商類型「{dealer_name}」不在允許清單內'}), 400
    rec = {
        'seller_name': seller_name,
        'dealer_name': dealer_name,
        'order_no': _norm(data.get('order_no')) or None,
        'shipping_order_no': _norm(data.get('shipping_order_no')) or None,
        'invoice_date': _parse_invoice_date(data.get('invoice_date')),
        'invoice_no': invoice_no,
        'amount_incl': amount_incl,
        'amount_notax': _round2(amount_incl / 1.05),
        'note': _norm(data.get('note')) or None,
        'recon_period': _norm(data.get('recon_period')) or None,
        # 2026-08-17新增：應收憑單/請購單號，選填純文字，供會計流程記錄用（見schema v1.31）
        'receivable_doc_no': _norm(data.get('receivable_doc_no')) or None,
        'updated_by': session.get('user_id'),
        'updated_by_name': session.get('display_name', session.get('username', '')),
        'updated_at': now_str(),
    }
    try:
        sb.table('bookstore_group_invoices').update(rec).eq('id', iid).execute()
        # 2026-08-17新增：編輯發票時如果填了發票號碼（例如這次才第一次補上發票號碼），
        # 順便檢查暫存區有沒有現成PDF可以直接掛上（見_auto_attach_pending_staging()說明）。
        _auto_attach_pending_staging(iid, invoice_no, session.get('user_id'),
                                      session.get('display_name', session.get('username', '')))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': f'修改失敗：{e}'}), 500


@dealer_bp.route('/api/bookstore/group-invoices/<int:iid>', methods=['DELETE'])
@admin_required
def delete_group_invoice(iid):
    try:
        # 2026-08-12新增：刪除發票前，先清掉Supabase Storage裡掛在這筆發票底下的出貨單
        # 檔案本體——bookstore_group_invoice_files有ON DELETE CASCADE，發票刪除時DB的
        # 中繼資料列會自動跟著清掉，但Storage裡的實際檔案物件不會自動連動刪除，不手動清
        # 會變成佔空間又永遠查不到的孤兒檔案。
        file_rows = sb.table('bookstore_group_invoice_files').select('storage_path') \
            .eq('invoice_id', iid).execute()
        paths = [r['storage_path'] for r in (file_rows.data or [])]
        if paths:
            try:
                sb.storage(_SHIPPING_DOC_BUCKET).remove(paths)
            except Exception:
                pass  # storage清理失敗不擋刪除本身，避免殘留檔案卡住使用者刪不掉發票
        sb.table('bookstore_group_invoices').delete().eq('id', iid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': f'刪除失敗：{e}'}), 500


# 範本第2列（欄位說明列）的「銷售單位」欄文字，供批次匯入時識別「這一列是說明列，非真實資料」——
# 2026-08-17修正：改用內容比對（而非列號），比照「內部料號清單」2026-08-12的既有修法
# （見_SKU_NOTE_ROW_MARKER）——使用者常會把範本的說明列(第2列)連同範例列(第3列)一起刪除，
# 讓真實資料往上移到第2列；若還用「第2列一律跳過」判斷，這筆真實資料會被誤判成說明列而被
# 跳過，造成「怎麼匯入都是0筆成功」的困惑。
_GROUP_INVOICE_NOTE_ROW_MARKER = '必填'


# ── 匯入團體發票 — 批次匯入範本下載（比照全系統慣例：必填欄位黃底、說明列、防呆範例列） ──
@dealer_bp.route('/api/bookstore/group-invoices/template', methods=['GET'])
@bookstore_invoice_required
def group_invoice_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '團體發票匯入範本'

    # 2026-08-12調整：「訂單編號」改名「OA訂單編號」（比照使用者要求，同步修改列表頁/範本/
    # 匯入/單筆新增），並新增「出貨訂單編號」欄位（緊接在OA訂單編號後面，跟單筆新增/編輯
    # 彈窗的欄位排列順序一致）。
    headers = ['銷售單位', 'OA訂單編號', '出貨訂單編號', '發票日期', '發票號碼', '價格(含稅)', '備註']
    required = {'銷售單位', '價格(含稅)'}
    notes_row = [
        _GROUP_INVOICE_NOTE_ROW_MARKER,
        '可留空',
        '可留空',
        '可留空，格式建議YYYY-MM-DD或YYYY/M/D',
        '可留空',
        '必填，數字，含稅金額——未稅價由系統自動算好(÷1.05)，此範本不需要也不能填未稅價欄',
        '可留空',
    ]
    # 「對帳區間」刻意不放進範本：整批匯入時在畫面上統一填一次，不是逐列從檔案裡讀取。

    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    rfill = openpyxl.styles.PatternFill('solid', fgColor='FFF2CC')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    rfont = openpyxl.styles.Font(bold=True, color='7B3F00')
    nfill = openpyxl.styles.PatternFill('solid', fgColor='F2F2F2')
    nfont = openpyxl.styles.Font(italic=True, color='888888')
    widths = [22, 16, 16, 14, 16, 12, 24]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        if h in required:
            c.fill = rfill
            c.font = rfont
        else:
            c.fill = hfill
            c.font = hfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = widths[ci - 1]

    for ci, n in enumerate(notes_row, 1):
        c = ws.cell(2, ci, n)
        c.fill = nfill
        c.font = nfont
        c.alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.append(['範例銷售單位股份有限公司', 'SAMPLE-DEMO-0001（此列為範例，請刪除後填入實際資料）',
                '', '2026-01-01', 'AB12345678', 10000, ''])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='團體發票匯入範本.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 匯入團體發票 — 批次匯入（每列皆為新增一筆登記；對帳區間整批共用同一個使用者輸入值） ──
@dealer_bp.route('/api/bookstore/group-invoices/batch-import', methods=['POST'])
@bookstore_invoice_required
def batch_import_group_invoices():
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    recon_period = _norm(request.form.get('recon_period', ''))
    if not recon_period:
        return jsonify({'error': '請填寫這批要共用的對帳區間'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{e}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'error': '檔案無資料列'}), 400

    raw_headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    def _map_h(h):
        h = h.strip()
        if '銷售單位' in h:
            return 'seller_name'
        # 2026-08-12新增：「出貨訂單編號」必須排在「訂單編號」判斷之前——因為新標題
        # 「出貨訂單編號」本身也包含「訂單編號」這個子字串，若順序顛倒會被誤判成OA訂單編號欄。
        if '出貨訂單編號' in h:
            return 'shipping_order_no'
        if 'OA訂單編號' in h or '訂單編號' in h:  # 「訂單編號」為舊欄名（改版前），保留相容
            return 'order_no'
        if '發票日期' in h:
            return 'invoice_date'
        if '發票號碼' in h:
            return 'invoice_no'
        if '價格' in h or '含稅' in h:
            return 'amount_incl'
        if '備註' in h:
            return 'note'
        return None

    col_map = {}
    for ci, h in enumerate(raw_headers):
        fk = _map_h(h)
        if fk and fk not in col_map:
            col_map[fk] = ci

    def _get(row, fk):
        ci = col_map.get(fk)
        if ci is None or ci >= len(row):
            return None
        return row[ci]

    ok_cnt = skip_cnt = err_cnt = dup_cnt = 0
    dup_detail = []   # 2026-08-12新增：重複的發票號碼跳過時，記錄是第幾列、哪個號碼，方便使用者對照
    batch = []
    audit = {'created_by': session.get('user_id'),
              'created_by_name': session.get('display_name', session.get('username', ''))}
    # 2026-08-12新增：同一次批次匯入的所有列統一蓋上同一個時間戳（batch_ts），配合
    # import_seq（該列在Excel裡的列號）供list_group_invoices()排序用——比照方案清單
    # 批次匯入的做法（見dealer_reconcile_schema.sql v1.22），讓清單照Excel原始列順序
    # 呈現，且最新匯入的整批排在最上面。
    batch_ts = now_str()

    # 2026-08-12新增：發票號碼不可重複——先用_fetch_all_rows()分頁抓「全部既有」發票號碼
    # （比照bookstore_plans existing_map的做法，避免資料超過1000筆時被PostgREST靜默截斷），
    # 檔案內同一發票號碼也只認第一筆出現的那一列，其餘視為重複、跳過不匯入。空白發票號碼
    # 不受此限制（允許多筆都留空）。
    existing_invoice_rows = _fetch_all_rows(lambda: sb.table('bookstore_group_invoices').select('invoice_no'))
    existing_invoice_nos = {r['invoice_no'] for r in existing_invoice_rows if r.get('invoice_no')}
    seen_invoice_nos = set()

    for ri, row in enumerate(rows[1:], 2):
        seller_name_raw = _get(row, 'seller_name')
        # 2026-08-17修正：說明列判斷改用內容比對（而非「第2列一律跳過」），比照「內部料號
        # 清單」批次匯入2026-08-12的既有修法（見_GROUP_INVOICE_NOTE_ROW_MARKER）——使用者常會
        # 把範本的說明列(第2列)連同範例列(第3列)一起刪除，讓真實資料往上移到第2列；若還用
        # 列號判斷，這筆真實資料會被誤判成說明列而被跳過，造成「怎麼匯入都是0筆成功」的困惑。
        if seller_name_raw == _GROUP_INVOICE_NOTE_ROW_MARKER:
            skip_cnt += 1
            continue

        seller_name = _norm(seller_name_raw)
        if not seller_name or seller_name.upper().startswith('SAMPLE-DEMO') or \
                (_get(row, 'order_no') and str(_get(row, 'order_no')).upper().startswith('SAMPLE-DEMO')):
            skip_cnt += 1
            continue

        amount_incl = _to_num_or_none(_get(row, 'amount_incl'))
        if amount_incl is None:
            err_cnt += 1
            continue

        invoice_no = _norm(_get(row, 'invoice_no')) or None
        if invoice_no and (invoice_no in existing_invoice_nos or invoice_no in seen_invoice_nos):
            dup_cnt += 1
            dup_detail.append({'row': ri, 'invoice_no': invoice_no})
            continue
        if invoice_no:
            seen_invoice_nos.add(invoice_no)

        rec = {
            'seller_name': seller_name,
            'order_no': _norm(_get(row, 'order_no')) or None,
            'shipping_order_no': _norm(_get(row, 'shipping_order_no')) or None,
            'invoice_date': _parse_invoice_date(_get(row, 'invoice_date')),
            'invoice_no': invoice_no,
            'amount_incl': amount_incl,
            'amount_notax': _round2(amount_incl / 1.05),
            'note': _norm(_get(row, 'note')) or None,
            'recon_period': recon_period,
            'import_seq': ri,
            'last_imported_at': batch_ts,
        }
        rec.update(audit)
        batch.append(rec)

        if len(batch) >= 80:
            try:
                res = sb.table('bookstore_group_invoices').insert(batch).execute()
                ok_cnt += len(batch)
                # 2026-08-17新增：批次匯入的每一列如果有發票號碼，順便檢查暫存區有沒有
                # 現成PDF可以直接掛上（見_auto_attach_pending_staging()說明）——這是使用者
                # 反映「發票PDF排程掃到信但沒自動掛上」最常見的成因：Gmail掃到信件時這筆
                # 發票根本還沒批次匯入進系統，只能先存暫存區等著，這裡補上「匯入的當下回頭
                # 找暫存區」這一步，不用每次都得自己想到要去暫存區手動附加。
                for row in (res.data or []):
                    _auto_attach_pending_staging(row['id'], row.get('invoice_no'),
                                                  audit['created_by'], audit['created_by_name'])
            except Exception:
                err_cnt += len(batch)
            batch = []

    if batch:
        try:
            res = sb.table('bookstore_group_invoices').insert(batch).execute()
            ok_cnt += len(batch)
            for row in (res.data or []):
                _auto_attach_pending_staging(row['id'], row.get('invoice_no'),
                                              audit['created_by'], audit['created_by_name'])
        except Exception:
            err_cnt += len(batch)

    # 2026-08-12新增：dup（發票號碼重複而跳過，含與既有資料重複／檔案內部自己重複兩種情形）
    # 獨立回報，不併入skip，方便使用者一眼看出是「範例列/銷售單位空白」還是「發票號碼重複」
    # 被跳過；dup_detail附上是第幾列、哪個發票號碼，供前端顯示明細。
    return jsonify({'ok': ok_cnt, 'skip': skip_cnt, 'error': err_cnt, 'dup': dup_cnt, 'dup_detail': dup_detail})


# ============================================================
# 2026-08-12新增：「匯入團體發票」Gmail自動抓發票PDF排程（限admin使用）
# ============================================================
# 完整背景/架構決定過程見專案筆記project_gmail_invoice_autofetch，這裡只放程式碼。
# 核心設計：
#   - 讀信用IMAP + 應用程式密碼，直接複用Render既有的SMTP_USER/SMTP_PASS環境變數
#     （app.py._send_notification_mail()寄信也是用這組帳密），不需要OAuth2/Google
#     Cloud Console，也不會有token過期問題。
#   - 背景排程用APScheduler，每5分鐘檢查一次「是否到了設定的抓取時間、今天排程還沒
#     跑過」，到了才真正動作——刻意不用APScheduler的cron trigger直接指定時間，是
#     因為使用者可以隨時在畫面上改抓取時間，用interval檢查+DB狀態判斷，不需要每次
#     改設定就重新註冊排程工作，也天然可以撐過程式重啟（下次interval檢查會自己
#     從DB讀到最新設定）。
#   - Render用gunicorn 2 workers，每個worker各自會啟動一份APScheduler，同一天的
#     排程可能被兩個worker都判斷「到時間了」而各自嘗試執行——防重複做法見
#     bookstore_gmail_sync_runs的partial unique index（run_date, trigger_type=
#     'scheduled'），insert失敗的那個worker直接跳過，不會真的跑兩次。
# ============================================================

import imaplib
import email as email_lib
from email.header import decode_header
from email.utils import parsedate_to_datetime
from zoneinfo import ZoneInfo

_TAIPEI_TZ = ZoneInfo('Asia/Taipei')
_INVOICE_PDF_BUCKET = 'bookstore-invoice-files'
_GMAIL_IMAP_HOST = 'imap.gmail.com'
_GMAIL_SUBJECT_KEYWORD = '電子發票開立通知'
# 2026-08-13修正：原本用「寄件人=ebookmail@hyread.com.tw」當IMAP搜尋條件，但使用者
# 實際收到的信是同事轉寄(Fwd:)進來的，寄件人變成轉寄的人（例如她自己或同事），不是
# ebookmail@hyread.com.tw，導致IMAP搜尋不到任何信件（掃到0封）。改用「最近N天內的信」
# 當IMAP搜尋範圍（避免每次都要掃描整個信箱），主旨關鍵字比對交給下面Python判斷
# （不論是原始信件還是Fwd:轉寄，主旨都會保留「電子發票開立通知(發票號碼)」這段文字，
# 轉寄只是在最前面多加"Fwd:"，不影響比對）。已處理過的信件用message_uid唯一索引擋
# 重複，所以就算天數抓寬一點也不會重複處理，只是多花一點點時間查暫存區。
# 2026-08-13再修正：「立即執行一次」是在使用者按下按鈕的同一個HTTP請求裡同步執行
# （不像排程走背景執行緒），如果單次要檢查的信件數量太多，逐封信兩次IMAP來回(先抓
# 主旨再抓整封信)加總的時間可能超過gunicorn預設30秒逾時，導致又是一次WORKER TIMEOUT
# 崩潰、且這次執行紀錄會卡在status='running'永遠不會變成success/error（因為process
# 被強制終止，finally區塊的收尾更新沒機會執行完）。
# 2026-08-13三修（已被四修取代，僅留紀錄）：一度改成「SINCE最近N天」+「只對還沒處理
# 過的信件套用_GMAIL_MAX_UIDS_PER_RUN上限」，但後來發現：主旨不符合關鍵字的信件永遠
# 不會被寫進bookstore_invoice_staging，所以在「還沒處理過」的判斷下永遠算「新的」，
# 每次執行都要重新對這些不相關的信件做一次主旨檢查——天數放寬後這類信件一多，反而
# 又逼近逾時，使用者回報「再次發生執行失敗」正是這個原因。
# 2026-08-13四修（目前版本）：改用「已處理到哪個UID」的水位機制（見_run_gmail_
# invoice_sync()內`last_processed_uid`相關程式碼），每次只搜尋比水位更新的信件——
# 不論主旨符不符合關鍵字，只要嘗試處理過就推進水位，不相關的信件只會被檢查「一次」，
# 之後永遠不會再進入搜尋範圍，徹底解決重複掃描的效能問題。`_GMAIL_SINCE_DAYS`現在
# 只在「水位還是空值」時使用（這支程式第一次執行、或使用者尚未執行v1.28建表SQL），
# 當作第一次執行時的追趕範圍，涵蓋較久的轉寄延遲（例如使用者回報的發票DE51907332）；
# 追趕完成、水位推進之後，後續每次執行都只需要處理「上次水位之後的新信件」，通常
# 數量很小，不會再有逾時風險。`_GMAIL_MAX_UIDS_PER_RUN`是每次執行最多處理的信件數
# （不論是第一次追趕還是之後的日常執行），超過時優先處理最舊的、較新的留到下次執行
# （水位機制要求依序處理，見下方truncation註解）。
_GMAIL_SINCE_DAYS = 30
_GMAIL_MAX_UIDS_PER_RUN = 100
_GMAIL_UID_CHECK_CHUNK = 200
_INVOICE_NO_RE = re.compile(r'\(([^()]+)\)\s*$')


def _decode_mime_words(s):
    """信件主旨/附件檔名可能是MIME encoded-word格式（=?UTF-8?B?...?=），要先解碼成
    一般字串才能用正規表示式解析發票號碼、比對附件檔名。"""
    if not s:
        return ''
    try:
        parts = decode_header(s)
    except Exception:
        return s
    out = []
    for text, enc in parts:
        if isinstance(text, bytes):
            try:
                out.append(text.decode(enc or 'utf-8', errors='replace'))
            except (LookupError, TypeError):
                out.append(text.decode('utf-8', errors='replace'))
        else:
            out.append(text)
    return ''.join(out)


def _run_gmail_invoice_sync(trigger_type='scheduled', user_id=None, user_name=None):
    """執行一次Gmail發票抓取。trigger_type='scheduled'時受bookstore_gmail_sync_runs的
    每日一次限制（見schema partial unique index）；'manual'（畫面「立即執行一次」）
    不受限制，可以重複按。回傳一個dict摘要，所有錯誤都包在回傳值/DB紀錄裡，不對
    呼叫端拋出例外（呼叫端可能是APScheduler背景執行緒，也可能是API endpoint）。"""
    today_str = datetime.now(_TAIPEI_TZ).strftime('%Y-%m-%d')
    try:
        claim = sb.table('bookstore_gmail_sync_runs').insert({
            'run_date': today_str, 'trigger_type': trigger_type, 'status': 'running',
            'started_by': user_id, 'started_by_name': user_name,
        }).execute()
        run_id = claim.data[0]['id']
    except Exception:
        # scheduled情境下insert失敗＝今天已經有其他worker搶到執行權（partial unique
        # index擋下），直接跳過不算錯誤；manual理論上不會撞到，若真的發生也視為略過
        # 即可，不影響使用者再按一次。
        return {'skipped': True, 'reason': '今天排程已執行過，或有其他程序正在執行中'}

    user = os.environ.get('SMTP_USER')
    pwd = os.environ.get('SMTP_PASS')
    scanned = attached = staged = 0
    error_msg = None
    status = 'error'
    summary = None
    imap = None
    # 2026-08-13四修：先給None，確保就算連login都還沒成功就出例外，最後更新設定檔時
    # 這個變數仍然有定義（維持原本的水位不動，不會誤寫入None覆蓋掉既有進度）。
    new_watermark = None
    hit_failure = False
    try:
        if not user or not pwd:
            raise RuntimeError('尚未設定 SMTP_USER / SMTP_PASS 環境變數，無法登入Gmail讀信')

        # 2026-08-13新增：IMAP連線加上20秒逾時——萬一Gmail那端無回應（例如網路異常/
        # 短暫被限流），連線會在20秒後主動拋出socket.timeout例外，被下面的except接住、
        # 正常標記這次執行失敗，而不是無限期卡住等到gunicorn強制關閉整個worker
        # （那樣連這筆執行紀錄的status都來不及更新，會一直卡在running）。
        imap = imaplib.IMAP4_SSL(_GMAIL_IMAP_HOST, timeout=20)
        imap.login(user, pwd)
        imap.select('INBOX')

        # 2026-08-13四修（重要重構）：改用「已處理到哪個UID」的水位(watermark)取代單純
        # 「最近N天」的日期範圍搜尋。背景：三修把SINCE天數從3天放寬到30天、上限改成只套用
        # 在「bookstore_invoice_staging裡還沒有紀錄的信件」上，結果發現一個沒考慮到的副
        # 作用——主旨不符合關鍵字的信件(例如日常業務信、電子報)永遠不會被寫進
        # bookstore_invoice_staging，所以在「還沒處理過」的判斷下永遠都算「新的」，每次
        # 執行都要重新對這些不相關的信件做一次主旨檢查(header fetch)，30天內信件一多，
        # 又會逼近gunicorn 30秒逾時、卡回「執行失敗」——這正是使用者這次回報「再次發生」
        # 的根因。
        # 改法：`bookstore_gmail_sync_config.last_processed_uid`記錄「已經確認處理完畢
        # 的最後一個UID」，每次執行只搜尋比這個UID更新的信件（`UID {last+1}:*`），不論
        # 主旨符不符合關鍵字，只要嘗試處理過就會推進這個水位——不相關的信件只會被檢查
        # 「一次」，之後永遠不會再出現在搜尋範圍內，徹底解決重複掃描的效能問題。
        # 只有這個欄位還是空值（這支程式第一次執行、或使用者尚未執行v1.28建表SQL）時，
        # 才退回舊的SINCE日期範圍搜尋作為「第一次執行的追趕範圍」（`_GMAIL_SINCE_DAYS`
        # 天，涵蓋較久的轉寄延遲，例如使用者回報的發票DE51907332）；追趕完成、水位推進
        # 之後，後續每次執行都只需要處理「上次水位之後的新信件」，通常量很小，不會再有
        # 逾時風險，不需要再靠限縮天數來換取效能。
        cfg_res = sb.table('bookstore_gmail_sync_config').select('*').eq('id', 1).limit(1).execute()
        cfg_row = cfg_res.data[0] if cfg_res.data else {}
        last_processed_uid = cfg_row.get('last_processed_uid')

        if last_processed_uid:
            typ, uid_data = imap.uid('search', None, f'(UID {int(last_processed_uid) + 1}:*)')
        else:
            since_date = (datetime.now(_TAIPEI_TZ) - timedelta(days=_GMAIL_SINCE_DAYS)).strftime('%d-%b-%Y')
            typ, uid_data = imap.uid('search', None, f'(SINCE "{since_date}")')
        if typ != 'OK':
            raise RuntimeError(f'IMAP搜尋失敗：{typ}')
        all_uids = uid_data[0].split() if uid_data and uid_data[0] else []

        # 這裡仍保留「批次查詢是否已處理過」當作第二道防線——正常情況下水位機制應該已經
        # 排除掉所有真正處理過的信件，但如果上一次執行中途失敗、水位來不及推進到最新，
        # 這裡可以避免同一封信被重複下載PDF/寫入暫存區。
        already_processed = set()
        all_uid_strs = [u.decode() for u in all_uids]
        for i in range(0, len(all_uid_strs), _GMAIL_UID_CHECK_CHUNK):
            chunk = all_uid_strs[i:i + _GMAIL_UID_CHECK_CHUNK]
            if not chunk:
                continue
            try:
                rows = sb.table('bookstore_invoice_staging').select('message_uid') \
                    .in_('message_uid', chunk).execute()
                already_processed.update(r['message_uid'] for r in rows.data)
            except Exception:
                pass  # 批次查詢失敗就當作都還沒處理過，交給逐封信處理，最多就是多做工

        candidate_uids = [u for u in all_uids if u.decode() not in already_processed]

        # 2026-08-13四修：IMAP search結果依UID由小到大排列(＝舊到新)——這裡改成優先處理
        # 「最舊」的那幾筆(而不是三修時保留最新的做法)，超過上限的部分(較新的信件)留到
        # 下次執行。這是搭配水位機制的必要改法：水位只能安全地推進到「這次確實處理過的
        # 最後一個UID」，如果像三修那樣跳過中間一段直接處理最新的，中間被跳過的那些較舊
        # 信件的UID會落在新水位之前，之後就永遠不會再被搜尋到、形同永久漏信。改成舊到新
        # 依序處理，水位只會往前推進到「連續處理完畢」的那個UID，較新的信件會留在下次
        # 執行的搜尋範圍內，不會有任何UID被跳過或漏處理。
        new_uids = candidate_uids
        truncated_count = 0
        if len(new_uids) > _GMAIL_MAX_UIDS_PER_RUN:
            truncated_count = len(new_uids) - _GMAIL_MAX_UIDS_PER_RUN
            new_uids = new_uids[:_GMAIL_MAX_UIDS_PER_RUN]

        # 2026-08-13四修：水位只能推進到「目前為止都成功嘗試處理過」的最後一個UID——
        # 一旦某封信處理途中拋例外(hit_failure=True)，即使後面還有信件處理成功，水位也
        # 不會再往前推進，確保下次執行一定會重新嘗試那封失敗的信(以及它之後的信件)，
        # 不會因為推進水位而永久跳過一封只是「這次剛好失敗」的信。
        new_watermark = int(last_processed_uid) if last_processed_uid else None

        # 2026-08-13四修：把原本迴圈內的逐封信處理邏輯抽成一個巢狀函式，用return取代
        # continue——原因是外層迴圈需要在「這封信處理完全沒有拋例外」時才推進水位
        # (new_watermark)，但原本每個業務判斷點(主旨不符合關鍵字/找不到發票號碼/找不到
        # 附件等)都是直接continue跳過整個迴圈本體，沒辦法在這些正常的「略過」路徑之後
        # 還執行「推進水位」這行程式碼。改成return之後，這些正常略過的信件一樣會讓
        # 外層迴圈的try區塊正常結束(視同沒有例外)，水位才能正確推進；只有真正拋出例外
        # 的信件(IMAP逾時/解析錯誤等)才會讓水位卡住不動，確保下次執行一定會重試。
        def _process_one_invoice_email(uid_bytes, uid):
            nonlocal scanned, attached, staged, error_msg
            # 是否已處理過在上面已經用批次查詢排除掉了，這裡不用再逐封信查一次。

            # 2026-08-13新增：先只抓信件主旨（不含附件），主旨符合關鍵字才真正抓整封信
            # （含PDF附件），避免搜尋範圍內每一封不相關的信都白白下載一次附件內容。
            typ, hdr_data = imap.uid('fetch', uid_bytes, '(BODY.PEEK[HEADER.FIELDS (SUBJECT)])')
            if typ != 'OK' or not hdr_data or not hdr_data[0]:
                return
            header_msg = email_lib.message_from_bytes(hdr_data[0][1])
            subject = _decode_mime_words(header_msg.get('Subject', ''))
            if _GMAIL_SUBJECT_KEYWORD not in subject:
                return
            scanned += 1

            typ, msg_data = imap.uid('fetch', uid_bytes, '(RFC822)')
            if typ != 'OK' or not msg_data or not msg_data[0]:
                return
            raw = msg_data[0][1]
            msg = email_lib.message_from_bytes(raw)

            m = _INVOICE_NO_RE.search(subject.strip())
            if not m:
                error_msg = f'信件主旨「{subject}」找不到括號內的發票號碼，略過'
                return
            invoice_no = m.group(1).strip()
            if not invoice_no:
                return

            target_name = f'{invoice_no}.pdf'

            # 2026-08-14新增：使用者回報「發票PDF」欄位出現同一個檔名重覆兩份的情況（例如
            # 同一張發票的通知信被系統重寄/轉寄了第二次，或水位機制追趕期間同一張發票被
            # 掃到2次）。修正原則：這張發票如果已經有一筆同檔名的發票PDF（不論來源是
            # Gmail自動掛檔還是暫存區手動附加），代表已經比對過、掛過檔了，這裡直接跳過，
            # 不再重覆下載/上傳/掛檔一樣的PDF——在真正解析附件內容之前就先判斷，省下不必要
            # 的下載/Storage寫入。
            date_hdr = msg.get('Date')
            try:
                source_date = parsedate_to_datetime(date_hdr).isoformat() if date_hdr else None
            except Exception:
                source_date = None
            dup_invoice = sb.table('bookstore_group_invoices').select('id') \
                .eq('invoice_no', invoice_no).limit(1).execute()
            if dup_invoice.data:
                dup_iid = dup_invoice.data[0]['id']
                dup_pdf = sb.table('bookstore_group_invoice_pdfs').select('id,storage_path') \
                    .eq('invoice_id', dup_iid).eq('file_name', target_name).limit(1).execute()
                if dup_pdf.data:
                    # storage_path欄位是NOT NULL，這裡沿用「已經掛檔那份」的storage_path
                    # 做紀錄（邏輯上是同一份文件，不用再另外存一份重覆的檔案內容）。
                    try:
                        sb.table('bookstore_invoice_staging').insert({
                            'invoice_no': invoice_no, 'message_uid': uid,
                            'storage_path': dup_pdf.data[0]['storage_path'],
                            'source_email_date': source_date,
                            'status': 'duplicate_skipped', 'attached_invoice_id': dup_iid,
                            'attached_at': now_str(),
                        }).execute()
                    except Exception:
                        pass
                    return

            pdf_bytes = None
            for part in msg.walk():
                fname = part.get_filename()
                if not fname:
                    continue
                fname_dec = _decode_mime_words(fname)
                if fname_dec.strip().lower() == target_name.lower():
                    pdf_bytes = part.get_payload(decode=True)
                    break
            if not pdf_bytes:
                error_msg = f'發票{invoice_no}：信件裡找不到附件{target_name}，略過'
                return

            storage_path = f'{invoice_no}/{uuid.uuid4().hex}_{target_name}'
            try:
                sb.storage(_INVOICE_PDF_BUCKET).upload(
                    storage_path, pdf_bytes, 'application/pdf')
            except Exception as e:
                error_msg = f'發票{invoice_no}上傳Storage失敗：{e}'
                return

            # dup_invoice已在上面查過（同一個invoice_no），這裡直接沿用，不用再查一次。
            staging_status = 'pending'
            attached_invoice_id = None
            if dup_invoice.data:
                iid = dup_invoice.data[0]['id']
                try:
                    sb.table('bookstore_group_invoice_pdfs').insert({
                        'invoice_id': iid, 'file_name': target_name,
                        'storage_path': storage_path, 'file_size': len(pdf_bytes),
                        'source': 'gmail_auto',
                    }).execute()
                    staging_status = 'attached'
                    attached_invoice_id = iid
                    attached += 1
                except Exception as e:
                    error_msg = f'發票{invoice_no}掛檔失敗（已存暫存區，可手動再附加一次）：{e}'
                    staged += 1
            else:
                staged += 1

            try:
                sb.table('bookstore_invoice_staging').insert({
                    'invoice_no': invoice_no, 'message_uid': uid,
                    'storage_path': storage_path, 'source_email_date': source_date,
                    'status': staging_status, 'attached_invoice_id': attached_invoice_id,
                    'attached_at': now_str() if staging_status == 'attached' else None,
                }).execute()
            except Exception as e:
                error_msg = f'發票{invoice_no}：暫存區紀錄寫入失敗：{e}'

        for uid_bytes in new_uids:
            uid = uid_bytes.decode()
            try:
                _process_one_invoice_email(uid_bytes, uid)
            except Exception as e:
                # 2026-08-13新增：單一封信處理失敗（IMAP逾時/解析錯誤等）只記錄，不中斷
                # 整批處理，剩下的信件繼續跑；但水位(new_watermark)從此不再往前推進
                # （見上方hit_failure說明），確保這封信下次執行會被重新嘗試。
                error_msg = f'信件(uid={uid})處理失敗，略過：{e}'
                hit_failure = True
                continue
            if not hit_failure:
                new_watermark = int(uid)

        status = 'success' if not error_msg else 'error'
        summary = f'掃到{scanned}封、自動掛檔{attached}張、新增暫存{staged}張'
        if truncated_count:
            # 2026-08-13四修：改成優先處理最舊的信件（見上方水位機制說明），所以留到下次
            # 執行的是「較新」的那些，不是先前版本的「較舊」信件。
            summary += f'（本次信件數量較多，僅處理最舊{_GMAIL_MAX_UIDS_PER_RUN}封，尚有{truncated_count}封較新信件留到下次執行）'
        if error_msg:
            summary += f'（部分錯誤：{error_msg}）'
    except Exception as e:
        status = 'error'
        error_msg = str(e)
        summary = f'執行失敗：{error_msg}'
    finally:
        if imap is not None:
            try:
                imap.logout()
            except Exception:
                pass

    try:
        sb.table('bookstore_gmail_sync_runs').update({
            'finished_at': now_str(), 'status': status, 'scanned_count': scanned,
            'attached_count': attached, 'staged_count': staged, 'error_message': error_msg,
        }).eq('id', run_id).execute()
        cfg_update = {
            'last_run_at': now_str(), 'last_run_status': status, 'last_run_summary': summary,
        }
        # 2026-08-13四修：把這次推進到的水位存回設定檔，下次執行才能只搜尋比這個UID更新
        # 的信件。new_watermark維持None的情況（例如login都還沒成功就出例外、或使用者
        # 尚未執行v1.28建表SQL導致這欄根本不存在）就不寫入這個欄位，避免覆蓋掉既有進度，
        # 也避免PostgREST因為欄位不存在而回錯誤。
        if new_watermark is not None:
            cfg_update['last_processed_uid'] = new_watermark
        try:
            sb.table('bookstore_gmail_sync_config').update(cfg_update).eq('id', 1).execute()
        except Exception:
            # 2026-08-13四修：若使用者還沒執行v1.28建表SQL，last_processed_uid欄位不存在，
            # PostgREST會回錯誤讓這次update整筆失敗（連last_run_at都沒更新到）。這裡退而
            # 求其次，去掉這個新欄位再更新一次，確保至少既有的執行狀態/摘要能正常顯示，
            # 不會讓一個新欄位還沒建的問題，波及到既有功能。
            cfg_update.pop('last_processed_uid', None)
            sb.table('bookstore_gmail_sync_config').update(cfg_update).eq('id', 1).execute()
    except Exception:
        pass

    return {'skipped': False, 'status': status, 'scanned': scanned, 'attached': attached,
            'staged': staged, 'error': error_msg}


def _gmail_sync_tick():
    """APScheduler每5分鐘呼叫一次。檢查排程是否開啟、是否到了設定的抓取時間、今天
    是否已經跑過，到了才真正呼叫_run_gmail_invoice_sync()。這支函式本身不拋出例外
    （APScheduler背景執行緒裡的例外不會有人看到，全部要在這裡攔下來印log）。"""
    try:
        res = sb.table('bookstore_gmail_sync_config').select('*').eq('id', 1).limit(1).execute()
        cfg = res.data[0] if res.data else None
        if not cfg or not cfg.get('enabled'):
            return
        now = datetime.now(_TAIPEI_TZ)
        fetch_hour = cfg.get('fetch_hour', 8)
        fetch_minute = cfg.get('fetch_minute', 0)
        # 到了設定時間之後、當天結束前都算「該執行」，讓即使某次interval檢查漏掉
        # （例如程式剛好在那5分鐘重啟），下一次檢查仍然抓得到；真正防止重複執行的
        # 是bookstore_gmail_sync_runs的每日一次unique constraint，不是這裡的時間窗判斷。
        target_minutes = fetch_hour * 60 + fetch_minute
        now_minutes = now.hour * 60 + now.minute
        if now_minutes < target_minutes:
            return
        print(f'[gmail-sync] 到了排程時間({fetch_hour:02d}:{fetch_minute:02d})，嘗試執行')
        result = _run_gmail_invoice_sync(trigger_type='scheduled')
        print(f'[gmail-sync] 排程執行結果：{result}')
    except Exception as e:
        print(f'[gmail-sync] _gmail_sync_tick 例外：{e}')


def start_gmail_sync_scheduler():
    """由app.py在啟動時呼叫一次。每個gunicorn worker process都會各自呼叫到、各自
    啟動一份APScheduler背景執行緒——這是刻意允許的（見本區塊最上方的架構說明），
    同一天只會有一個worker真正執行到底，其餘會在_run_gmail_invoice_sync()裡因
    insert失敗而提早跳過。"""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except ImportError:
        print('[gmail-sync] 未安裝 APScheduler 套件，背景排程未啟動（請確認 requirements.txt）')
        return
    scheduler = BackgroundScheduler(timezone=str(_TAIPEI_TZ))
    # 刻意延後30秒才第一次執行，避免worker剛啟動時跟主執行緒處理第一個請求（健康檢查）
    # 同時搶著載入Python模組造成import鎖死（曾發生過WORKER TIMEOUT崩潰，詳見專案筆記）
    scheduler.add_job(_gmail_sync_tick, 'interval', minutes=5, id='gmail_invoice_sync_tick',
                       next_run_time=datetime.now(_TAIPEI_TZ) + timedelta(seconds=30))
    scheduler.start()
    print('[gmail-sync] APScheduler背景排程已啟動（每5分鐘檢查一次）')


# ── API：排程設定 ──
@dealer_bp.route('/api/bookstore/gmail-sync/config', methods=['GET'])
@admin_required
def get_gmail_sync_config():
    res = sb.table('bookstore_gmail_sync_config').select('*').eq('id', 1).limit(1).execute()
    data = res.data[0] if res.data else {}
    # 2026-08-17新增：附上「今天排程額度是否已用掉」旗標，供前端提示使用者——排程
    # 一天只能成功觸發一次（見bookstore_gmail_sync_runs的partial unique index：
    # (run_date) WHERE trigger_type='scheduled'），使用者如果在額度用掉後才改抓取
    # 時間，畫面上看起來像是「改了時間也沒反應」，其實是要等明天才會用新時間生效，
    # 不是bug；這裡直接查today有沒有一筆scheduled紀錄，讓前端能顯示對應提示文字。
    try:
        today_str = datetime.now(_TAIPEI_TZ).strftime('%Y-%m-%d')
        used = sb.table('bookstore_gmail_sync_runs').select('id') \
            .eq('run_date', today_str).eq('trigger_type', 'scheduled').limit(1).execute()
        data['today_scheduled_used'] = bool(used.data)
    except Exception:
        data['today_scheduled_used'] = False
    return jsonify(data)


@dealer_bp.route('/api/bookstore/gmail-sync/config', methods=['PUT'])
@admin_required
def update_gmail_sync_config():
    data = request.get_json(force=True) or {}
    try:
        fetch_hour = int(data.get('fetch_hour'))
        fetch_minute = int(data.get('fetch_minute'))
    except (TypeError, ValueError):
        return jsonify({'error': '抓取時間格式錯誤'}), 400
    if not (0 <= fetch_hour <= 23 and 0 <= fetch_minute <= 59):
        return jsonify({'error': '抓取時間超出範圍'}), 400
    rec = {
        'enabled': bool(data.get('enabled')),
        'fetch_hour': fetch_hour, 'fetch_minute': fetch_minute,
        'updated_at': now_str(), 'updated_by': session.get('user_id'),
        'updated_by_name': session.get('display_name', session.get('username', '')),
    }
    sb.table('bookstore_gmail_sync_config').update(rec).eq('id', 1).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/gmail-sync/run-now', methods=['POST'])
@admin_required
def run_gmail_sync_now():
    result = _run_gmail_invoice_sync(
        trigger_type='manual', user_id=session.get('user_id'),
        user_name=session.get('display_name', session.get('username', '')))
    return jsonify(result)


@dealer_bp.route('/api/bookstore/gmail-sync/logs', methods=['GET'])
@admin_required
def list_gmail_sync_logs():
    rows = sb.table('bookstore_gmail_sync_runs').select('*') \
        .order('started_at', desc=True).limit(30).execute()
    return jsonify(rows.data or [])


# ── API：發票PDF暫存區 ──
@dealer_bp.route('/api/bookstore/invoice-staging', methods=['GET'])
@admin_required
def list_invoice_staging():
    """?invoice_no=xxx 篩選特定發票號碼（給「新增/編輯發票」彈窗查詢用）；
    ?status=pending 篩選狀態；都不帶就回傳全部（依建立時間新到舊）。"""
    qb = sb.table('bookstore_invoice_staging').select('*')
    invoice_no = request.args.get('invoice_no')
    status = request.args.get('status')
    if invoice_no:
        qb = qb.eq('invoice_no', invoice_no.strip())
    if status:
        qb = qb.eq('status', status)
    rows = qb.order('created_at', desc=True).limit(200).execute()
    return jsonify(rows.data or [])


@dealer_bp.route('/api/bookstore/invoice-staging/<int:sid>/attach', methods=['POST'])
@admin_required
def attach_invoice_staging(sid):
    data = request.get_json(force=True) or {}
    invoice_id = data.get('invoice_id')
    if not invoice_id:
        return jsonify({'error': '請指定要附加到哪一筆發票'}), 400
    srow = sb.table('bookstore_invoice_staging').select('*').eq('id', sid).limit(1).execute()
    if not srow.data:
        return jsonify({'error': '找不到這筆暫存資料'}), 404
    s = srow.data[0]
    inv = sb.table('bookstore_group_invoices').select('id').eq('id', invoice_id).limit(1).execute()
    if not inv.data:
        return jsonify({'error': '找不到這筆發票，請確認發票是否已儲存'}), 404
    try:
        sb.table('bookstore_group_invoice_pdfs').insert({
            'invoice_id': invoice_id, 'file_name': f"{s['invoice_no']}.pdf",
            'storage_path': s['storage_path'], 'source': 'staging_attach',
        }).execute()
    except Exception as e:
        return jsonify({'error': f'掛檔失敗：{e}'}), 500
    sb.table('bookstore_invoice_staging').update({
        'status': 'attached', 'attached_invoice_id': invoice_id, 'attached_at': now_str(),
        'attached_by': session.get('user_id'),
        'attached_by_name': session.get('display_name', session.get('username', '')),
    }).eq('id', sid).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/invoice-staging/<int:sid>/ignore', methods=['POST'])
@admin_required
def ignore_invoice_staging(sid):
    sb.table('bookstore_invoice_staging').update({'status': 'ignored'}).eq('id', sid).execute()
    return jsonify({'ok': True})


# ── API：發票PDF（已掛在某筆「匯入團體發票」列上的檔案）──
@dealer_bp.route('/api/bookstore/group-invoices/<int:iid>/invoice-pdfs', methods=['GET'])
@bookstore_invoice_required
def list_group_invoice_pdfs(iid):
    rows = sb.table('bookstore_group_invoice_pdfs').select('*') \
        .eq('invoice_id', iid).order('created_at').execute()
    return jsonify(rows.data or [])


@dealer_bp.route('/api/bookstore/group-invoices/invoice-pdfs/<int:fid>/download', methods=['GET'])
@bookstore_invoice_required
def download_group_invoice_pdf(fid):
    row = sb.table('bookstore_group_invoice_pdfs').select('*').eq('id', fid).limit(1).execute()
    if not row.data:
        return jsonify({'error': '找不到這個檔案'}), 404
    rec = row.data[0]
    try:
        data = sb.storage(_INVOICE_PDF_BUCKET).download(rec['storage_path'])
    except Exception as e:
        return jsonify({'error': f'下載失敗：{e}'}), 500
    return send_file(io.BytesIO(data), as_attachment=True, download_name=rec['file_name'])


@dealer_bp.route('/api/bookstore/group-invoices/invoice-pdfs/<int:fid>', methods=['DELETE'])
@bookstore_invoice_required
def delete_group_invoice_pdf(fid):
    row = sb.table('bookstore_group_invoice_pdfs').select('*').eq('id', fid).limit(1).execute()
    if not row.data:
        return jsonify({'error': '找不到這個檔案'}), 404
    rec = row.data[0]
    try:
        sb.storage(_INVOICE_PDF_BUCKET).remove([rec['storage_path']])
    except Exception:
        pass
    sb.table('bookstore_group_invoice_pdfs').delete().eq('id', fid).execute()
    return jsonify({'ok': True})


# ============================================================
# 2026-08-14新增：「上傳團體出貨單」——「匯入團體發票」的姊妹功能。使用者的出貨單附件
# 目前只能在「匯入團體發票」的新增/編輯彈窗裡逐筆手動上傳（見上方bookstore_group_invoice_
# files），但出貨單往往比發票早到、或由不同同仁（倉管/物流）先收到，不方便每次都要先找到
# 對應那筆發票才能上傳。這裡新增一個獨立的「上傳團體出貨單」登記頁面，先以「OA訂單編號」
# 為key把出貨單收進來，再由下面的排程「自動比對出貨單檔至團體發票」，依OA訂單編號比對到
# 「匯入團體發票」的既有列，自動把出貨訂單編號／出貨單附件寫回對應的發票列——跟「匯入
# 團體發票」原有的出貨單上傳功能是兩條互補的路徑：這裡是「先收集、後比對」，原本那個是
# 「已經知道是哪筆發票、直接上傳」，兩者互不影響。
#
# 使用者2026-08-14確認的3個關鍵規則（用AskUserQuestion問過，不是自行推論）：
# 1. 覆蓋規則：同一筆「上傳團體出貨單」再次被排程比對到時，一律用這裡最新的出貨訂單
#    編號／附件覆蓋掉已比對過的發票列（不是「已經比對過就不再處理」）——所以每次執行
#    排程都會對「全部」上傳團體出貨單重新跑一次比對，不是只跑「還沒比對過」的。
# 2. 比對基數：同一個OA訂單編號如果在「匯入團體發票」裡對應到多筆發票列，只附加到
#    「第一筆比對到的發票」（依id由小到大取第一筆），不是全部發票列都附加。
# 3. 列表頁新增「比對狀態」欄位（使用者要求的4個欄位之外，額外加這個方便追蹤）：
#    待比對／已比對／查無對應發票。
#
# 檔案共用既有的 bookstore-shipping-docs bucket（見上方_SHIPPING_DOC_BUCKET），不需要
# 使用者另外在Supabase後台建立新bucket；這裡的storage路徑用「shipdoc-{doc_id}/」前綴，
# 跟「匯入團體發票」既有的「{invoice_id}/」前綴區隔，避免同樣數字的doc_id/invoice_id
# 互相踩到對方的檔案路徑。
# ============================================================
_BOOKSTORE_SHIPPING_PERM_KEYS = (
    'mod_bookstore_shipping_view', 'mod_bookstore_shipping_create',
    'mod_bookstore_shipping_update', 'mod_bookstore_shipping_delete',
)


def _any_bookstore_shipping_perm():
    return any(_perm_ok(k) for k in _BOOKSTORE_SHIPPING_PERM_KEYS)


def bookstore_shipping_view_required(f):
    """「上傳團體出貨單」檢視權限——比照bookstore_action_required的萬用總開關設計，
    mod_dealer_edit（既有整體編輯權限）OR 這個功能任一細分權限，即可打開頁面/列表。"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not (_perm_ok('mod_dealer_edit') or _any_bookstore_shipping_perm()):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def bookstore_shipping_action_required(*keys):
    """「上傳團體出貨單」新增/修改/刪除細分權限，各自獨立勾選（比照方案清單
    bookstore_action_required的設計）。使用者要求「刪除」限admin帳號才可使用——這裡
    刻意不寫死@admin_required，而是沿用跟新增/修改一致的細分權限勾選機制：實務上只要
    「刪除」這顆勾選框只勾給admin群組，效果就等同限admin使用，作法跟「方案清單」
    mod_bookstore_plan_delete、「批次匯入」mod_bookstore_plan_batch_import（使用者說的
    「限管理員使用」）完全一致，不需要另外用硬編碼的admin_required例外處理這一個按鈕。"""
    def deco(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not _login_ok():
                return jsonify({'error': '請先登入'}), 401
            if not (_perm_ok('mod_dealer_edit') or any(_perm_ok(k) for k in keys)):
                return jsonify({'error': '權限不足'}), 403
            return f(*args, **kwargs)
        return wrapped
    return deco


_SHIPPING_MATCH_STATUS_LABELS = {'pending': '待比對', 'matched': '已比對', 'no_match': '查無對應發票'}


@dealer_bp.route('/api/bookstore/shipping-docs', methods=['GET'])
@bookstore_shipping_view_required
def list_shipping_docs():
    rows = _fetch_all_rows(lambda: sb.table('bookstore_shipping_docs').select('*')
                            .order('id', desc=True))
    # 2026-08-21新增：「封存」勾選+備註——跟「從匯入團體發票勾選匯入」候選清單同一套設計
    # （見list_reconcile_candidates()），讓使用者可以把已經處理完、不需要再持續留意的出貨單
    # 登記標記起來，預設從列表隱藏，前端勾選「顯示已封存」時才連同封存的一起列出。
    include_archived = request.args.get('include_archived') in ('1', 'true', 'True')
    if not include_archived:
        rows = [r for r in rows if not r.get('archived')]
    doc_ids = [r['id'] for r in rows]
    files_map = {}
    if doc_ids:
        # 比照list_group_invoices()既有的容錯寫法（try/except，新表萬一還沒建立不能拖垮
        # 這支清單API），一次把全部檔案metadata抓回來依doc_id分組掛回每一列，避免前端
        # 每一列各自呼叫一次。
        try:
            file_rows = _fetch_all_rows(lambda: sb.table('bookstore_shipping_doc_files')
                                         .select('*').in_('doc_id', doc_ids))
            for fr in file_rows:
                files_map.setdefault(fr['doc_id'], []).append(fr)
        except Exception:
            pass
    for r in rows:
        r['files'] = files_map.get(r['id'], [])
        r['match_status_label'] = _SHIPPING_MATCH_STATUS_LABELS.get(r.get('match_status'), r.get('match_status'))

    # 2026-08-21新增：使用者反饋「上傳團體出貨單」列表看不出這筆出貨單對應的發票有沒有
    # 已經寫入對帳，每次都要另外切去「匯入團體發票」頁面查。這裡依match_status='matched'
    # 時記錄下來的matched_invoice_id，反查對應的bookstore_group_invoices是否已
    # reconciled_at，並沿用list_group_invoices()同一套_attach_recon_seq_info()回填拆帳
    # 編號/批次遺失警示，讓這個列表也能直接看到「已寫入對帳」狀態，不用切頁查。
    # 尚未比對到任何發票（match_status不是'matched'，或matched_invoice_id空）的列，
    # reconciled_at等欄位維持None，前端顯示「—」區隔「有比對到發票但還沒入庫」（顯示
    # 「尚未匯入」）跟「根本還沒比對到發票」這兩種不同狀況。
    matched_ids = list({r['matched_invoice_id'] for r in rows if r.get('matched_invoice_id')})
    invoices_map = {}
    if matched_ids:
        try:
            # 2026-08-21新增：多帶invoice_date/invoice_no——使用者要求列表也能看到這筆出貨單
            # 比對到的發票日期/發票號碼，不用切去「匯入團體發票」查。
            inv_res = sb.table('bookstore_group_invoices') \
                .select('id,reconciled_at,reconciled_by_name,reconciled_upload_group,invoice_date,invoice_no') \
                .in_('id', matched_ids).execute()
            invoices_map = {iv['id']: iv for iv in (inv_res.data or [])}
        except Exception:
            invoices_map = {}
    for r in rows:
        inv = invoices_map.get(r.get('matched_invoice_id'))
        r['reconciled_at'] = inv.get('reconciled_at') if inv else None
        r['reconciled_by_name'] = inv.get('reconciled_by_name') if inv else None
        r['reconciled_upload_group'] = inv.get('reconciled_upload_group') if inv else None
        r['invoice_date'] = inv.get('invoice_date') if inv else None
        r['invoice_no'] = inv.get('invoice_no') if inv else None
    _attach_recon_seq_info(rows)

    return jsonify(rows)


@dealer_bp.route('/api/bookstore/shipping-docs/<int:did>/archive', methods=['PUT'])
@bookstore_shipping_action_required('mod_bookstore_shipping_update')
def archive_shipping_doc(did):
    """2026-08-21新增：「上傳團體出貨單」列表的「封存」欄——跟候選發票清單那邊
    （archive_reconcile_candidate()）同一套設計，讓使用者標記某筆出貨單登記「已經處理過、
    不需要再持續留意」，預設從列表隱藏，跟這筆登記本身的比對狀態/已寫入對帳完全無關，
    只是使用者自己的篩選/備註機制，隨時可以取消。只接受archived(bool)/archive_note(text)
    這2個欄位。"""
    data = request.json or {}
    rec = {}
    if 'archived' in data:
        rec['archived'] = bool(data.get('archived'))
    if 'archive_note' in data:
        rec['archive_note'] = _norm(data.get('archive_note')) or None
    if not rec:
        return jsonify({'ok': True})
    cur = sb.table('bookstore_shipping_docs').select('id').eq('id', did).limit(1).execute()
    if not cur.data:
        return jsonify({'error': '找不到這筆出貨單登記'}), 404
    sb.table('bookstore_shipping_docs').update(rec).eq('id', did).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/shipping-docs', methods=['POST'])
@bookstore_shipping_action_required('mod_bookstore_shipping_create')
def create_shipping_doc():
    data = request.json or {}
    oa_order_no = _norm(data.get('oa_order_no'))
    if not oa_order_no:
        return jsonify({'error': 'OA訂單編號必填'}), 400
    rec = {
        'oa_order_no': oa_order_no,
        'shipping_order_no': _norm(data.get('shipping_order_no')) or None,
        'note': data.get('note'),
        'match_status': 'pending',
    }
    rec.update(_audit_new())
    res = sb.table('bookstore_shipping_docs').insert(rec).execute()
    return jsonify(res.data[0])


@dealer_bp.route('/api/bookstore/shipping-docs/<int:did>', methods=['PUT'])
@bookstore_shipping_action_required('mod_bookstore_shipping_update')
def update_shipping_doc(did):
    data = request.json or {}
    cur_res = sb.table('bookstore_shipping_docs').select('id,match_status').eq('id', did).limit(1).execute()
    if not cur_res.data:
        return jsonify({'error': '找不到這筆出貨單登記'}), 404
    oa_order_no = _norm(data.get('oa_order_no'))
    if not oa_order_no:
        return jsonify({'error': 'OA訂單編號必填'}), 400
    rec = {
        'oa_order_no': oa_order_no,
        'shipping_order_no': _norm(data.get('shipping_order_no')) or None,
        'note': data.get('note'),
    }
    # 2026-08-14修正：使用者確認新規則——已經比對成功的登記，排程之後會直接跳過（見
    # _run_shipping_doc_match()），不再像原本一律重跑全部登記。但如果使用者事後又編輯了
    # 這筆登記的OA訂單編號/出貨訂單編號，代表比對條件可能已經變了，必須讓它退回「待比對」
    # 狀態，才能被下一次排程重新抓到、重新比對一次，不會卡在已比對的舊狀態一直沒人處理。
    if cur_res.data[0].get('match_status') == 'matched':
        rec['match_status'] = 'pending'
        rec['matched_invoice_id'] = None
        rec['matched_at'] = None
    rec.update(_audit_upd())
    sb.table('bookstore_shipping_docs').update(rec).eq('id', did).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/shipping-docs/<int:did>', methods=['DELETE'])
@admin_required
def delete_shipping_doc(did):
    # 刪除前先清掉Supabase Storage裡掛在這筆登記底下的出貨單檔案本體——
    # bookstore_shipping_doc_files有ON DELETE CASCADE，這筆登記刪除時DB紀錄會自動一併
    # 清掉，但Storage裡的實體檔案不會自動跟著刪，需要在刪DB紀錄前先手動清乾淨（比照
    # delete_group_invoice()既有的做法）。這裡不動已經比對搬過去、掛在「匯入團體發票」
    # 那邊的複製檔案(bookstore_group_invoice_files)，那些檔案是獨立的Storage物件，
    # 刪除這筆「上傳團體出貨單」登記不影響已經比對成功、附加在發票列上的檔案。
    file_rows = sb.table('bookstore_shipping_doc_files').select('storage_path') \
        .eq('doc_id', did).execute()
    for fr in (file_rows.data or []):
        try:
            sb.storage(_SHIPPING_DOC_BUCKET).remove([fr['storage_path']])
        except Exception:
            pass
    sb.table('bookstore_shipping_docs').delete().eq('id', did).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/shipping-docs/<int:did>/files', methods=['GET'])
@bookstore_shipping_view_required
def list_shipping_doc_files(did):
    rows = sb.table('bookstore_shipping_doc_files').select('*') \
        .eq('doc_id', did).order('created_at').execute()
    return jsonify(rows.data or [])


@dealer_bp.route('/api/bookstore/shipping-docs/<int:did>/files', methods=['POST'])
@bookstore_shipping_action_required('mod_bookstore_shipping_create', 'mod_bookstore_shipping_update')
def upload_shipping_doc_files(did):
    """比照upload_group_invoice_files()既有做法：一次可選多檔，每個檔案獨立檢查格式/
    大小，部分失敗不擋下其他檔案。格式限制沿用系統既有「出貨單」的格式慣例（PDF/.docx/
    .odt，單檔≤2MB，見_SHIPPING_DOC_ALLOWED_EXT/_SHIPPING_DOC_MAX_SIZE），跟「匯入
    團體發票」的出貨單附件共用同一個Storage bucket（見上方區塊說明的路徑前綴區隔）。"""
    doc = sb.table('bookstore_shipping_docs').select('id,match_status').eq('id', did).limit(1).execute()
    if not doc.data:
        return jsonify({'error': '找不到這筆出貨單登記，請先儲存後再上傳檔案'}), 404
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '請選擇要上傳的檔案'}), 400
    uploaded, errors = [], []
    for f in files:
        if not f or not f.filename:
            continue
        if not _shipping_doc_ext_ok(f.filename):
            errors.append(f'{f.filename}：僅接受 PDF、Word(.docx)、OpenDocument(.odt) 格式')
            continue
        data = f.read()
        if len(data) > _SHIPPING_DOC_MAX_SIZE:
            errors.append(f'{f.filename}：檔案超過2MB上限')
            continue
        if not data:
            errors.append(f'{f.filename}：檔案是空的，略過')
            continue
        safe_name = secure_filename(f.filename) or 'file'
        storage_path = f'shipdoc-{did}/{uuid.uuid4().hex}_{safe_name}'
        try:
            sb.storage(_SHIPPING_DOC_BUCKET).upload(
                storage_path, data, f.mimetype or 'application/octet-stream')
        except Exception as e:
            errors.append(f'{f.filename}：上傳失敗（{e}）')
            continue
        rec = {
            'doc_id': did, 'file_name': f.filename, 'storage_path': storage_path,
            'file_size': len(data),
            'uploaded_by': session.get('user_id'),
            'uploaded_by_name': session.get('display_name', session.get('username', '')),
        }
        try:
            res = sb.table('bookstore_shipping_doc_files').insert(rec).execute()
            uploaded.append(res.data[0])
        except Exception as e:
            try:
                sb.storage(_SHIPPING_DOC_BUCKET).remove([storage_path])
            except Exception:
                pass
            errors.append(f'{f.filename}：儲存紀錄失敗（{e}）')
    # 2026-08-14修正：同一原則——已比對成功的登記如果之後又補上傳新的出貨單附件，代表
    # 內容已經變了，退回「待比對」讓下一次排程重新抓到、重新把新檔案搬過去對應的發票，
    # 不會因為已經是matched狀態而被_run_shipping_doc_match()永遠跳過。
    if uploaded and doc.data[0].get('match_status') == 'matched':
        sb.table('bookstore_shipping_docs').update({
            'match_status': 'pending', 'matched_invoice_id': None, 'matched_at': None,
        }).eq('id', did).execute()
    return jsonify({'uploaded': uploaded, 'errors': errors})


@dealer_bp.route('/api/bookstore/shipping-docs/files/<int:fid>/download', methods=['GET'])
@bookstore_shipping_view_required
def download_shipping_doc_file(fid):
    row = sb.table('bookstore_shipping_doc_files').select('*').eq('id', fid).limit(1).execute()
    if not row.data:
        return jsonify({'error': '找不到這個檔案'}), 404
    rec = row.data[0]
    try:
        data = sb.storage(_SHIPPING_DOC_BUCKET).download(rec['storage_path'])
    except Exception as e:
        return jsonify({'error': f'下載失敗：{e}'}), 500
    return send_file(io.BytesIO(data), as_attachment=True, download_name=rec['file_name'])


@dealer_bp.route('/api/bookstore/shipping-docs/files/<int:fid>', methods=['DELETE'])
@bookstore_shipping_action_required('mod_bookstore_shipping_create', 'mod_bookstore_shipping_update')
def delete_shipping_doc_file(fid):
    row = sb.table('bookstore_shipping_doc_files').select('*').eq('id', fid).limit(1).execute()
    if not row.data:
        return jsonify({'error': '找不到這個檔案'}), 404
    rec = row.data[0]
    try:
        sb.storage(_SHIPPING_DOC_BUCKET).remove([rec['storage_path']])
    except Exception:
        pass
    sb.table('bookstore_shipping_doc_files').delete().eq('id', fid).execute()
    # 2026-08-14修正：同一原則，刪除附件也代表這筆登記的內容變了，已比對成功的登記要
    # 退回「待比對」，等下一次排程重新處理（見upload_shipping_doc_files()的對應說明）。
    try:
        doc = sb.table('bookstore_shipping_docs').select('id,match_status') \
            .eq('id', rec['doc_id']).limit(1).execute()
        if doc.data and doc.data[0].get('match_status') == 'matched':
            sb.table('bookstore_shipping_docs').update({
                'match_status': 'pending', 'matched_invoice_id': None, 'matched_at': None,
            }).eq('id', rec['doc_id']).execute()
    except Exception:
        pass
    return jsonify({'ok': True})


# ============================================================
# 2026-08-14新增：「自動比對出貨單檔至團體發票」排程——比照Gmail自動抓發票PDF排程
# （bookstore_gmail_sync_config/_runs）的架構，同樣限admin使用，同樣用APScheduler
# interval檢查+DB狀態判斷（可隨時改設定不需重啟排程）、同樣用partial unique index
# 防止Render多worker同一天重複執行。
# ============================================================
def _run_shipping_doc_match(trigger_type='scheduled', user_id=None, user_name=None):
    """執行一次「出貨單檔比對到團體發票」。trigger_type='scheduled'受bookstore_
    shipping_match_runs的每日一次限制；'manual'（畫面「立即執行一次」）不受限制。

    2026-08-14規則修正（原規則見下方，已由使用者明確改為新規則）：只處理
    match_status IN ('pending','no_match')的登記——已經是'matched'狀態的登記，
    代表上次已經成功比對過、附件也已經搬到對應的發票列上了，這次直接跳過，不用再
    重覆下載/上傳一樣的檔案、也不用再覆蓋一次發票的出貨訂單編號。
    2026-08-12原規則（已改）：「一律覆蓋成上傳團體出貨單裡最新的資料」，每次執行都
    重新處理「全部」登記。**這個舊規則沒有被完全丟棄**：如果使用者後續在update_
    shipping_doc()/upload_shipping_doc_files()/delete_shipping_doc_file()這幾個
    編輯動作裡，動到了一筆已經是'matched'狀態的登記，那幾支函式會主動把它的
    match_status退回'pending'，讓它重新進入這裡的處理範圍——這樣同時滿足「已比對
    的不用每天重複處理」跟「使用者事後編輯了內容，最後還是要能重新比對更新到發票」
    這兩個需求，不衝突。

    ⚠️ 這支函式可能在APScheduler背景執行緒裡被呼叫（trigger_type='scheduled'時），
    背景執行緒沒有Flask的session/request context，因此全程不讀取session，需要的
    使用者資訊一律由呼叫端以user_id/user_name參數傳入（manual時由呼叫端的route
    從session取值後傳進來，見run_shipping_match_now()）。"""
    today_str = datetime.now(_TAIPEI_TZ).strftime('%Y-%m-%d')
    try:
        claim = sb.table('bookstore_shipping_match_runs').insert({
            'run_date': today_str, 'trigger_type': trigger_type, 'status': 'running',
            'started_by': user_id, 'started_by_name': user_name,
        }).execute()
        run_id = claim.data[0]['id']
    except Exception:
        # scheduled情境下insert失敗＝今天已經有其他worker搶到執行權（partial unique
        # index擋下），直接跳過不算錯誤；manual理論上不會撞到，若真的發生也視為略過即可。
        return {'skipped': True, 'reason': '今天排程已執行過，或有其他程序正在執行中'}

    scanned = matched = no_match = skipped_already_matched = 0
    error_msg = None
    status = 'success'
    try:
        all_docs = _fetch_all_rows(lambda: sb.table('bookstore_shipping_docs').select('*'))
        # 2026-08-14修正：已經是'matched'狀態的登記直接跳過不處理（見上方函式說明），
        # 只處理'pending'/'no_match'（含尚未設定match_status的舊資料，一併當作待處理）。
        docs = [d for d in all_docs if d.get('match_status') != 'matched']
        skipped_already_matched = len(all_docs) - len(docs)
        scanned = len(docs)
        for doc in docs:
            oa_order_no = doc.get('oa_order_no')
            if not oa_order_no:
                continue
            inv_res = (sb.table('bookstore_group_invoices').select('id')
                       .eq('order_no', oa_order_no).order('id').limit(1).execute())
            if not inv_res.data:
                sb.table('bookstore_shipping_docs').update({
                    'match_status': 'no_match', 'matched_invoice_id': None,
                }).eq('id', doc['id']).execute()
                no_match += 1
                continue
            invoice_id = inv_res.data[0]['id']

            # 出貨訂單編號：一律覆蓋成這筆「上傳團體出貨單」目前的值（使用者確認規則1）。
            sb.table('bookstore_group_invoices').update({
                'shipping_order_no': doc.get('shipping_order_no'),
                'updated_at': now_str(), 'updated_by_name': '系統自動比對',
            }).eq('id', invoice_id).execute()

            # 附件：先清掉上一次這筆登記自動搬過去的舊複製檔（source_doc_id=doc['id']），
            # 避免每次執行都無限疊加重複檔案；使用者手動在「匯入團體發票」自己上傳的出貨單
            # （source_doc_id為空）不受影響，不會被這裡清掉。
            old_copies = (sb.table('bookstore_group_invoice_files').select('storage_path')
                          .eq('invoice_id', invoice_id).eq('source_doc_id', doc['id']).execute())
            for oc in (old_copies.data or []):
                try:
                    sb.storage(_SHIPPING_DOC_BUCKET).remove([oc['storage_path']])
                except Exception:
                    pass
            sb.table('bookstore_group_invoice_files').delete() \
                .eq('invoice_id', invoice_id).eq('source_doc_id', doc['id']).execute()

            src_files = sb.table('bookstore_shipping_doc_files').select('*') \
                .eq('doc_id', doc['id']).execute()
            for sf in (src_files.data or []):
                try:
                    raw = sb.storage(_SHIPPING_DOC_BUCKET).download(sf['storage_path'])
                except Exception:
                    continue  # 來源檔案下載失敗（例如storage物件已不存在），跳過這一份
                safe_name = secure_filename(sf['file_name']) or 'file'
                new_path = f'{invoice_id}/{uuid.uuid4().hex}_{safe_name}'
                try:
                    sb.storage(_SHIPPING_DOC_BUCKET).upload(new_path, raw, 'application/octet-stream')
                    sb.table('bookstore_group_invoice_files').insert({
                        'invoice_id': invoice_id, 'file_name': sf['file_name'],
                        'storage_path': new_path, 'file_size': sf.get('file_size'),
                        'source_doc_id': doc['id'],
                    }).execute()
                except Exception:
                    try:
                        sb.storage(_SHIPPING_DOC_BUCKET).remove([new_path])
                    except Exception:
                        pass

            sb.table('bookstore_shipping_docs').update({
                'match_status': 'matched', 'matched_invoice_id': invoice_id, 'matched_at': now_str(),
            }).eq('id', doc['id']).execute()
            matched += 1
    except Exception as e:
        status = 'error'
        error_msg = str(e)

    summary = f'共{scanned}筆待處理，成功比對{matched}筆，查無對應發票{no_match}筆'
    if skipped_already_matched:
        summary += f'（另有{skipped_already_matched}筆已比對過，本次略過未重新處理）'
    try:
        sb.table('bookstore_shipping_match_runs').update({
            'finished_at': now_str(), 'status': status,
            'scanned_count': scanned, 'matched_count': matched, 'no_match_count': no_match,
            'error_message': error_msg,
        }).eq('id', run_id).execute()
        sb.table('bookstore_shipping_match_config').update({
            'last_run_at': now_str(), 'last_run_status': status, 'last_run_summary': summary,
        }).eq('id', 1).execute()
    except Exception:
        pass

    return {'skipped': False, 'status': status, 'scanned': scanned, 'matched': matched,
            'no_match': no_match, 'error': error_msg}


def _shipping_match_tick():
    """APScheduler每5分鐘呼叫一次，邏輯完全比照_gmail_sync_tick()。"""
    try:
        res = sb.table('bookstore_shipping_match_config').select('*').eq('id', 1).limit(1).execute()
        cfg = res.data[0] if res.data else None
        if not cfg or not cfg.get('enabled'):
            return
        now = datetime.now(_TAIPEI_TZ)
        run_hour = cfg.get('run_hour', 8)
        run_minute = cfg.get('run_minute', 0)
        target_minutes = run_hour * 60 + run_minute
        now_minutes = now.hour * 60 + now.minute
        if now_minutes < target_minutes:
            return
        print(f'[shipping-match] 到了排程時間({run_hour:02d}:{run_minute:02d})，嘗試執行')
        result = _run_shipping_doc_match(trigger_type='scheduled')
        print(f'[shipping-match] 排程執行結果：{result}')
    except Exception as e:
        print(f'[shipping-match] _shipping_match_tick 例外：{e}')


def start_shipping_match_scheduler():
    """由app.py在啟動時呼叫一次，比照start_gmail_sync_scheduler()。"""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except ImportError:
        print('[shipping-match] 未安裝 APScheduler 套件，背景排程未啟動（請確認 requirements.txt）')
        return
    scheduler = BackgroundScheduler(timezone=str(_TAIPEI_TZ))
    scheduler.add_job(_shipping_match_tick, 'interval', minutes=5, id='shipping_doc_match_tick',
                       next_run_time=datetime.now(_TAIPEI_TZ) + timedelta(seconds=30))
    scheduler.start()
    print('[shipping-match] APScheduler背景排程已啟動（每5分鐘檢查一次）')


@dealer_bp.route('/api/bookstore/shipping-match/config', methods=['GET'])
@admin_required
def get_shipping_match_config():
    res = sb.table('bookstore_shipping_match_config').select('*').eq('id', 1).limit(1).execute()
    data = res.data[0] if res.data else {}
    # 2026-08-17新增：跟get_gmail_sync_config()同一個修法——附上「今天排程額度是否
    # 已用掉」旗標，讓前端可以提示使用者「改了時間但今天不會再跑，要等明天」，避免
    # 誤以為是bug。
    try:
        today_str = datetime.now(_TAIPEI_TZ).strftime('%Y-%m-%d')
        used = sb.table('bookstore_shipping_match_runs').select('id') \
            .eq('run_date', today_str).eq('trigger_type', 'scheduled').limit(1).execute()
        data['today_scheduled_used'] = bool(used.data)
    except Exception:
        data['today_scheduled_used'] = False
    return jsonify(data)


@dealer_bp.route('/api/bookstore/shipping-match/config', methods=['PUT'])
@admin_required
def update_shipping_match_config():
    data = request.get_json(force=True) or {}
    try:
        run_hour = int(data.get('run_hour'))
        run_minute = int(data.get('run_minute'))
    except (TypeError, ValueError):
        return jsonify({'error': '排程時間格式錯誤'}), 400
    if not (0 <= run_hour <= 23 and 0 <= run_minute <= 59):
        return jsonify({'error': '排程時間超出範圍'}), 400
    rec = {
        'enabled': bool(data.get('enabled')),
        'run_hour': run_hour, 'run_minute': run_minute,
        'updated_at': now_str(), 'updated_by': session.get('user_id'),
        'updated_by_name': session.get('display_name', session.get('username', '')),
    }
    sb.table('bookstore_shipping_match_config').update(rec).eq('id', 1).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/bookstore/shipping-match/run-now', methods=['POST'])
@admin_required
def run_shipping_match_now():
    result = _run_shipping_doc_match(
        trigger_type='manual', user_id=session.get('user_id'),
        user_name=session.get('display_name', session.get('username', '')))
    return jsonify(result)


@dealer_bp.route('/api/bookstore/shipping-match/logs', methods=['GET'])
@admin_required
def list_shipping_match_logs():
    rows = sb.table('bookstore_shipping_match_runs').select('*') \
        .order('started_at', desc=True).limit(30).execute()
    return jsonify(rows.data or [])
    return jsonify({'ok': True})


# ============================================================
# 「比對已開/未開發票」新模組 — 歷史總表匯入（sku_invoice_ledger）
# v1.0  2026-08-21
#
# 背景：獨立Excel「歷年每月開發票與未開發票總表」要一次性搬進sku_invoice_ledger表
# （schema見 dealer_reconcile_schema.sql v1.33），原本規劃用本機腳本
# import_invoice_ledger_history.py 執行，但使用者電腦是公司鎖定電腦、不能裝任何軟體，
# 雲端執行環境對Supabase的連線又被防火牆擋掉——兩條路都走不通，因此改成正式功能：
# 直接在網頁上傳Excel、由Render上實際在跑的Flask服務端解析+寫入，不需要使用者電腦
# 具備任何額外工具。
#
# 這裡的wide→long解析邏輯（找「未稅金額小計」欄判斷料號欄範圍結束、掃描A欄找
# 「系統出庫數」關鍵字判斷驗證區起點）跟 import_invoice_ledger_history.py 完全一致，
# 刻意不寫死欄位/列數，因為總表每個月都會增列。
# ============================================================
_SIL_SHEET_NAME_HINT = 'ALL(2019'
_SIL_SKU_END_MARKER = '未稅金額小計'      # 這欄(不含)之後不再是料號欄
# 2026-08-25新增：sku_invoice_ledger.raw_sku_label是NOT NULL欄位（見schema），不能對
# 「無明細品項」佔位紀錄（見_sil_extract_records()）直接寫None，否則Supabase會丟出
# 23502 not-null violation（跟v3.84/v3.85那次upsert()漏帶internal_code是同一類問題）。
# 改用這個固定字串當標記，internal_code仍然維持None（該欄位本來就允許NULL）。
_SIL_NO_ITEM_RAW_LABEL = '（無明細品項—僅登記客戶類型/期間等中繼資料）'
_SIL_VERIFY_BLOCK_MARKER = '系統出庫數'    # 客戶類型欄出現這關鍵字，該列起視為驗證區，全部跳過
_SIL_EXTRA_COL_HEADERS = {
    'reconcile_order': '對帳順序',
    'reconcile_no': '對帳編號',
    'bookstore_reconcile_no': '書店對帳編號',
    # 2026-08-21新增：GA欄「對帳訂單編」（比對原始檔案確認實際文字缺了「號」字，用
    # _sil_norm_header() 正規化去掉結尾「號」字再比對，容忍這種缺字/未來補回的差異）
    'order_no': '對帳訂單編號',
}

# 2026-08-21新增：「科技退/知識退 科技2024/7以前」這組專屬對帳欄位(FX/FY/FZ)的錨點欄
# 表頭文字。這組欄位的FY「對帳順序」／FZ「對帳編號」表頭文字剛好跟上面_SIL_EXTRA_COL_HEADERS
# 的reconcile_order/reconcile_no完全相同，不能放進同一份用「表頭文字比對」的字典裡——若
# 放進去，程式掃到第一個「對帳順序」文字時，reconcile_order跟techkt_reconcile_order兩個key
# 會同時搶到同一格，導致這組專屬欄位跟通用欄位指向錯誤的同一欄。因此改用「先找到FX錨點欄，
# 再用往右1欄/2欄的相對位置」取得FY/FZ，見 _sil_parse_workbook()。
#
# 2026-08-24修正（bug根因）：這個常數原本寫成「科技通/知識通 科技 2024/7以前」，但拿實際
# 總表檔案程式化解析表頭後發現，Excel實際文字是「科技退/知識退\n科技2024/7以前」——用字
# 「退」≠「通」、且欄位中間是換行字元不是空白、也沒有中間那個「/」，導致下面exact match
# 從未成立過，techkt_anchor_idx永遠是None，害這3個專屬欄位(techkt_note/techkt_reconcile_
# order/techkt_reconcile_no)从2026-08-21上線以來、包含2026-08-23那次13,408筆正式歷史
# 匯入，全部都被寫成NULL（用程式驗證：FX/FY/FZ實體欄位本身其實有值，只是沒被歸類進這3個
# techkt_*欄位）。改用_sil_norm_ws()去除所有空白字元(含換行)後再比對，同時把常數改成比對
# 實際觀察到的文字，避免同類「換行/空白差異」問題再犯。
_SIL_TECHKT_ANCHOR_LABEL = '科技退/知識退科技2024/7以前'


def _sil_norm_header(s):
    """去掉字串結尾的「號」字後比對，容忍「對帳訂單編」與「對帳訂單編號」這類差1個字的情況。"""
    return (s or '').rstrip('號')


def _sil_norm_ws(s):
    """移除所有空白字元(含換行/Tab)後比對，容忍表頭因手動編輯造成的空白/換行差異
    （2026-08-24新增，修正_SIL_TECHKT_ANCHOR_LABEL exact match從未成立的bug）。"""
    return re.sub(r'\s+', '', s or '')


def _sil_header_matches(text, label):
    """判斷Excel表頭文字是否對應到_SIL_EXTRA_COL_HEADERS裡的某個欄位。2026-08-24新增：
    原本只用_sil_norm_header()（只處理結尾缺「號」字）比對「對帳訂單編號」等欄位，沒有
    一併做_sil_norm_ws()的空白/換行正規化——但同一天才發現_SIL_TECHKT_ANCHOR_LABEL的exact
    match從未成立，根因正是實際表頭含換行字元、單純字串比對比不出來。為避免「對帳訂單編號」
    等既有欄位也潛藏同一類還沒被發現的空白/換行差異問題，這裡統一套用跟techkt錨點比對相同
    的「先去除所有空白字元，再比較（含容忍結尾缺「號」字）」規則，兩者都通過才判定不吻合。"""
    t = _sil_norm_ws(text)
    l = _sil_norm_ws(label)
    if t == l:
        return True
    return _sil_norm_header(t) == _sil_norm_header(l)


def _sil_find_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith(_SIL_SHEET_NAME_HINT):
            return wb[name]
    return None


def _sil_parse_workbook(ws):
    """回傳 (sku_cols, extra_cols, verify_row_idx) 或在解析失敗時 raise ValueError(說明文字)。"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    sku_cols = []
    extra_cols = {}
    end_idx = None
    techkt_anchor_idx = None
    for idx, val in enumerate(header_row):  # 0-based idx；欄位=idx+1
        text = (str(val).strip() if val is not None else '')
        if text == _SIL_SKU_END_MARKER and end_idx is None:
            end_idx = idx
        if _sil_norm_ws(text) == _SIL_TECHKT_ANCHOR_LABEL and techkt_anchor_idx is None:
            techkt_anchor_idx = idx
        for key, label in _SIL_EXTRA_COL_HEADERS.items():
            if key in extra_cols:
                continue
            if _sil_header_matches(text, label):
                extra_cols[key] = idx

    if end_idx is None:
        raise ValueError(f'找不到「{_SIL_SKU_END_MARKER}」欄，無法判斷料號欄範圍結束位置，請確認總表結構是否變動')

    # 2026-08-21新增：「科技通/知識通 科技 2024/7以前」專屬對帳欄位，用錨點欄往右1/2欄的
    # 相對位置取得（不能用表頭文字比對，理由見_SIL_TECHKT_ANCHOR_LABEL上方註解）。找不到
    # 錨點欄時視為這份總表沒有這個區塊，這3個欄位全部維持None，不視為錯誤（避免舊版本的
    # 總表檔案因為缺這個區塊而完全無法匯入）。
    if techkt_anchor_idx is not None:
        extra_cols['techkt_note'] = techkt_anchor_idx
        extra_cols['techkt_reconcile_order'] = techkt_anchor_idx + 1
        extra_cols['techkt_reconcile_no'] = techkt_anchor_idx + 2

    # H欄=index 7（0-based），到 end_idx（不含）都是料號欄
    for idx in range(7, end_idx):
        text = header_row[idx]
        if text is not None and str(text).strip() != '':
            sku_cols.append((idx, str(text).strip()))

    # 找驗證區起點：掃描A欄(客戶類型)，第一個出現VERIFY_BLOCK_MARKER的列
    verify_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        val = row[0]
        if val is not None and _SIL_VERIFY_BLOCK_MARKER in str(val):
            verify_row_idx = i
            break
    if verify_row_idx is None:
        raise ValueError(f'掃描A欄找不到「{_SIL_VERIFY_BLOCK_MARKER}」關鍵字，無法判斷驗證區起點，請確認總表是否變動格式')

    return sku_cols, extra_cols, verify_row_idx


def _sil_extract_records(ws, sku_cols, extra_cols, verify_row_idx, valid_codes, source_ref):
    """wide -> long 轉換。valid_codes：目前有效的內部料號集合（internal_code），
    比對到的填入internal_code且is_legacy_column=False，比對不到的internal_code=None、
    is_legacy_column=True（代表這欄可能是已歸檔舊料號，或料號主檔還沒補上這個新料號）。"""
    records = []
    skipped_blank_customer_type = 0
    max_col = max(idx for idx, _ in sku_cols) + 1 if sku_cols else 0
    for extra_idx in extra_cols.values():
        max_col = max(max_col, extra_idx + 1)

    for row_order, row in enumerate(ws.iter_rows(min_row=2, max_row=verify_row_idx - 1, max_col=max_col, values_only=True), start=2):
        # row_order：該列在Excel原始檔案中的實際列號（從2開始，因第1列是表頭），
        # 用來讓「寬表檢視」的縱向列順序可以還原成匯入當時Excel的原始排列，而非重新排序。
        customer_type = row[0]
        if customer_type is None or str(customer_type).strip() == '':
            skipped_blank_customer_type += 1
            continue
        customer_type = str(customer_type).strip()
        reconcile_name = row[1] if len(row) > 1 else None  # B欄「對帳名稱」，比客戶類型更細的子標籤
        period = row[2] if len(row) > 2 else None
        period_range = row[3] if len(row) > 3 else None
        note = row[4] if len(row) > 4 else None
        reconcile_order = row[extra_cols['reconcile_order']] if 'reconcile_order' in extra_cols else None
        reconcile_no = row[extra_cols['reconcile_no']] if 'reconcile_no' in extra_cols else None
        bookstore_reconcile_no = row[extra_cols['bookstore_reconcile_no']] if 'bookstore_reconcile_no' in extra_cols else None
        order_no = row[extra_cols['order_no']] if 'order_no' in extra_cols else None  # GA欄「對帳訂單編」
        # 2026-08-21新增：FX/FY/FZ「科技通/知識通 科技 2024/7以前」專屬對帳欄位
        techkt_note = row[extra_cols['techkt_note']] if 'techkt_note' in extra_cols else None
        techkt_reconcile_order = row[extra_cols['techkt_reconcile_order']] if 'techkt_reconcile_order' in extra_cols else None
        techkt_reconcile_no = row[extra_cols['techkt_reconcile_no']] if 'techkt_reconcile_no' in extra_cols else None

        # 2026-08-25新增：先把這一列共用的中繼資料（跟料號欄無關的欄位）組成一份dict，
        # 給下面「有數量的料號欄」跟「整列都沒有數量、補一筆佔位紀錄」兩個分支共用，
        # 避免同一組欄位字串處理邏輯要維護兩份。
        row_meta = {
            'customer_type': customer_type,
            'reconcile_name': str(reconcile_name).strip() if reconcile_name is not None else None,
            'period': str(period).strip() if period is not None else None,
            'period_range': str(period_range).strip() if period_range is not None else None,
            'note': str(note).strip() if note is not None else None,
            'reconcile_order': str(reconcile_order).strip() if reconcile_order is not None else None,
            'reconcile_no': str(reconcile_no).strip() if reconcile_no is not None else None,
            'bookstore_reconcile_no': str(bookstore_reconcile_no).strip() if bookstore_reconcile_no is not None else None,
            'order_no': str(order_no).strip() if order_no is not None else None,
            'techkt_note': str(techkt_note).strip() if techkt_note is not None else None,
            'techkt_reconcile_order': str(techkt_reconcile_order).strip() if techkt_reconcile_order is not None else None,
            'techkt_reconcile_no': str(techkt_reconcile_no).strip() if techkt_reconcile_no is not None else None,
            'source_type': 'excel_import',
            'source_ref': source_ref,
            'source_row_order': row_order,
        }

        row_has_qty = False
        for col_idx, raw_label in sku_cols:
            qty = row[col_idx] if col_idx < len(row) else None
            if qty is None or not isinstance(qty, (int, float)) or qty == 0:
                continue
            row_has_qty = True
            matched = raw_label in valid_codes
            internal_code = raw_label if matched else None
            is_legacy = not matched
            records.append({
                'internal_code': internal_code,
                'raw_sku_label': raw_label,
                'is_legacy_column': is_legacy,
                'qty': qty,
                **row_meta,
            })

        # 2026-08-25新增：使用者要求匯入結果要完整對應原始總表的每一列，供核對用——原本
        # 這一列如果所有料號欄數量都是0/空白，上面的迴圈完全不會產生任何紀錄，這一列就
        # 整列從sku_invoice_ledger「憑空消失」，即使客戶類型/對帳名稱/期間/備註等欄位
        # 明明都有內容（常見於像「MOMO賠償」這種只登記理賠單號、沒有對應實體出貨數量的
        # 備註性質列）。改成沒有任何非零品項時，仍然補一筆「無明細品項」的佔位紀錄
        # （internal_code=None、raw_sku_label=_SIL_NO_ITEM_RAW_LABEL固定字串、qty=0——
        # raw_sku_label不能直接寫None，該欄位在schema是NOT NULL），讓「寬表檢視」分組
        # （用source_ref+source_row_order當分組鍵，見_wide_view_build()）時這一列還是能
        # 單獨成一列呈現，底下的料號欄位全部維持0/空白，不會整列被略過不見。
        # _wide_view_build()的欄位彙總/孤兒欄位偵測那兩段都已同步加上判斷式，跳過
        # raw_sku_label等於這個sentinel常數的紀錄，不會因此多冒出一欄空白的孤兒欄位。
        if not row_has_qty:
            records.append({
                'internal_code': None,
                'raw_sku_label': _SIL_NO_ITEM_RAW_LABEL,
                'is_legacy_column': False,
                'qty': 0,
                **row_meta,
            })
    return records, skipped_blank_customer_type


def _sil_summarize(records, skipped, sku_cols):
    from collections import Counter
    by_type = Counter(r['customer_type'] for r in records)
    legacy_cnt = sum(1 for r in records if r['is_legacy_column'])

    # 2026-08-21新增：把「比對不到現有內部料號」的明細，依raw_sku_label（總表原始欄名）
    # 彙總成清單，讓使用者能肉眼核對這些欄位是不是真的都是已歸檔舊料號（而不是比對邏輯
    # 出錯，把仍在銷售的料號誤判成比對不到）。同時列出「比對得到」的欄位清單方便對照，
    # 兩份清單的欄位數加起來應該等於sku_col_count。
    legacy_labels = Counter()
    legacy_qty = {}
    matched_labels = Counter()
    matched_qty = {}
    # 2026-08-25新增：no_item_row_count統計「無明細品項」佔位紀錄（見_sil_extract_records()
    # 新增的補列邏輯）筆數，供匯入預覽畫面顯示「這次總共會補幾筆沒有數量、只有客戶類型/
    # 對帳名稱/期間等中繼資料的列」，讓使用者能肉眼確認補列數量跟自己數的紅框筆數是否吻合。
    no_item_row_count = 0
    for r in records:
        label = r['raw_sku_label']
        if label == _SIL_NO_ITEM_RAW_LABEL:
            no_item_row_count += 1
            continue  # 佔位紀錄沒有對應任何總表欄名，不計入下面的料號欄位比對統計
        if r['is_legacy_column']:
            legacy_labels[label] += 1
            legacy_qty[label] = legacy_qty.get(label, 0) + r['qty']
        else:
            matched_labels[label] += 1
            matched_qty[label] = matched_qty.get(label, 0) + r['qty']

    # 2026-08-22新增：使用者反饋原本「比對不到內部料號的欄位清單」只列出比對不到的，
    # 沒辦法跟「比對得到」的欄位放在一起肉眼核對總表欄位是否被正確判斷（例如某欄其實
    # 應該比對不到、卻誤配到錯的內部料號，或反過來）。這裡把「全部」總表欄位名稱彙總成
    # 一份清單，每筆額外帶出比對到的內部料號（is_legacy_column=False時），供前端整合
    # 顯示成一張「比對結果總表」，可依「總表欄位名稱」逐一核對，不用在兩份清單間切換比對。
    label_stats = {}
    for r in records:
        label = r['raw_sku_label']
        if label == _SIL_NO_ITEM_RAW_LABEL:
            continue  # 同上，佔位紀錄不進「料號比對結果總表」
        st = label_stats.get(label)
        if st is None:
            st = {'raw_sku_label': label, 'is_legacy_column': r['is_legacy_column'],
                  'internal_codes': set(), 'count': 0, 'qty_sum': 0}
            label_stats[label] = st
        st['count'] += 1
        st['qty_sum'] += r['qty']
        if r.get('internal_code'):
            st['internal_codes'].add(r['internal_code'])
    all_label_breakdown = [{
        'raw_sku_label': st['raw_sku_label'],
        'is_legacy_column': st['is_legacy_column'],
        'internal_code': '、'.join(sorted(st['internal_codes'])) if st['internal_codes'] else None,
        'count': st['count'],
        'qty_sum': st['qty_sum'],
    } for st in label_stats.values()]
    # 2026-08-23修正：使用者要求「料號比對結果總表」也要依Excel總表欄位原始順序排列
    # （原本是依筆數由大到小排），方便逐一比對總表時視覺上跟原始欄位順序一致。sku_cols
    # 本身就是_sil_parse_workbook()依左到右欄位順序建立的清單，這裡用它建一個
    # 「欄位名稱→順序值」對照表（同一欄名若剛好重複出現，只取第一次出現的位置），
    # 排序依據從「筆數由大到小」改成「總表欄位原始順序」；理論上records裡的raw_sku_label
    # 一定都來自sku_cols，找不到的極端情況(不應發生)排到最後面當保險。
    label_col_order = {}
    for idx, (_, col_label) in enumerate(sku_cols):
        label_col_order.setdefault(col_label, idx)
    all_label_breakdown.sort(key=lambda x: label_col_order.get(x['raw_sku_label'], len(sku_cols)))

    return {
        'record_count': len(records),
        'no_item_row_count': no_item_row_count,
        'qty_sum': sum(r['qty'] for r in records),
        'skipped_blank_customer_type': skipped,
        'customer_type_count': len(by_type),
        'customer_type_breakdown': [{'customer_type': ct, 'count': cnt}
                                     for ct, cnt in sorted(by_type.items(), key=lambda x: -x[1])],
        'legacy_count': legacy_cnt,
        'legacy_ratio': round(legacy_cnt / len(records), 4) if records else 0,
        'legacy_label_count': len(legacy_labels),
        'matched_label_count': len(matched_labels),
        'legacy_label_breakdown': [{'raw_sku_label': lbl, 'count': cnt, 'qty_sum': legacy_qty[lbl]}
                                    for lbl, cnt in sorted(legacy_labels.items(), key=lambda x: -x[1])],
        'matched_label_breakdown': [{'raw_sku_label': lbl, 'count': cnt, 'qty_sum': matched_qty[lbl]}
                                     for lbl, cnt in sorted(matched_labels.items(), key=lambda x: -x[1])],
        'all_label_breakdown': all_label_breakdown,
    }


@dealer_bp.route('/api/dealer/invoice-ledger/status', methods=['GET'])
@dealer_view_required
def get_invoice_ledger_status():
    """回傳 sku_invoice_ledger 目前各來源類型的筆數/數量統計，讓上傳頁面能顯示
    「歷史總表是否已經匯入過」，避免使用者不確定而重複匯入。"""
    from collections import Counter
    rows = _fetch_all_rows(lambda: sb.table('sku_invoice_ledger')
                            .select('source_type,qty,source_ref,created_at'))
    by_source = {}
    for r in rows:
        st = r.get('source_type') or '(未分類)'
        agg = by_source.setdefault(st, {'source_type': st, 'count': 0, 'qty_sum': 0,
                                         'source_refs': set(), 'latest_created_at': None})
        agg['count'] += 1
        agg['qty_sum'] += (r.get('qty') or 0)
        if r.get('source_ref'):
            agg['source_refs'].add(r['source_ref'])
        ca = r.get('created_at')
        if ca and (agg['latest_created_at'] is None or ca > agg['latest_created_at']):
            agg['latest_created_at'] = ca
    out = []
    for agg in by_source.values():
        agg['source_refs'] = sorted(agg['source_refs'])
        out.append(agg)
    return jsonify({
        'total_count': len(rows),
        'by_source_type': sorted(out, key=lambda x: -x['count']),
        'has_history_import': any(r.get('source_type') == 'excel_import' for r in rows),
    })


@dealer_bp.route('/api/dealer/invoice-ledger/import-history', methods=['POST'])
@dealer_edit_required
def import_invoice_ledger_history():
    """上傳「歷年每月開發票與未開發票總表.xlsx」，解析後寫入 sku_invoice_ledger。

    表單參數：
      file    必填，總表xlsx檔
      mode    'preview'（預設）：只解析、回傳統計數字，不寫入任何資料，供使用者核對筆數/
              加總是否合理
              'commit'：正式寫入資料庫
      force   '1'：mode=commit時，若偵測到source_type='excel_import'的既有資料，先刪除
              舊資料再重新匯入（避免手動去Supabase後台清空）；不帶這個參數且已有既有資料
              時會擋下，回400錯誤，比照 commit_batch() 既有「擋重複入庫、不自動覆蓋」的慣例。
    """
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    mode = request.form.get('mode', 'preview')
    force = request.form.get('force') == '1'

    try:
        # 2026-08-22修正：Render OOM（used over 512MB）——這份「歷年每月開發票與未開發票
        # 總表.xlsx」橫跨多年、上百個料號欄位，用預設（非read_only）模式載入openpyxl會把
        # 整份活頁簿所有分頁、每一格的樣式/格式資訊都一次性建成Python物件常駐記憶體，
        # 對這種大檔案很吃記憶體，在Render Starter方案(512MB)上會直接把instance打爆。
        # 這個匯入功能全程只用_sil_parse_workbook()/_sil_extract_records()裡的
        # ws.iter_rows(values_only=True)循序讀值，完全不需要存取儲存格樣式或反覆跳著讀，
        # 改用read_only=True（串流逐列讀取，不會把整份活頁簿物件化）大幅降低記憶體用量，
        # 不影響既有解析邏輯。
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'error': f'無法解析 Excel：{e}'}), 400

    ws = _sil_find_sheet(wb)
    if ws is None:
        return jsonify({'error': f'找不到分頁名稱開頭為「{_SIL_SHEET_NAME_HINT}」的工作表，'
                                  f'實際分頁：{wb.sheetnames}'}), 400

    try:
        sku_cols, extra_cols, verify_row_idx = _sil_parse_workbook(ws)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    existing = sb.table('sku_invoice_ledger').select('id').eq('source_type', 'excel_import').limit(1).execute()
    has_existing = bool(existing.data)

    if mode == 'commit' and has_existing and not force:
        return jsonify({
            'error': '已有歷史匯入資料（source_type=excel_import），為避免重複疊加已擋下。'
                     '若確定要用這份檔案重新匯入（會先刪除舊的歷史匯入資料），請勾選「覆蓋重新匯入」後再送出。',
            'has_existing': True,
        }), 409

    sku_list_rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select('id,internal_code,sort_order'))
    valid_codes = {r['internal_code'] for r in sku_list_rows if r.get('internal_code')}
    id_by_code = {r['internal_code']: r['id'] for r in sku_list_rows if r.get('internal_code')}
    max_sort_order = max((r.get('sort_order') or 0) for r in sku_list_rows) if sku_list_rows else 0

    source_ref = f'{secure_filename(f.filename) or "歷年每月開發票與未開發票總表.xlsx"}'
    records, skipped = _sil_extract_records(ws, sku_cols, extra_cols, verify_row_idx, valid_codes, source_ref)
    summary = _sil_summarize(records, skipped, sku_cols)
    summary['sku_col_count'] = len(sku_cols)
    summary['extra_cols_matched'] = list(extra_cols.keys())
    summary['verify_row_idx'] = verify_row_idx
    # 2026-08-21新增：「已歸檔/比對不到」的欄位，之後commit時系統會自動在「內部料號清單」
    # 補一筆佔位資料（is_active=False、hide_in_invoice_ledger=True），讓使用者能在「料號與
    # 規則管理」個別設定這些已歸檔欄位的顯示狀態（見下方commit分支）。這裡在preview階段就
    # 先讓使用者知道會新增幾筆，避免commit後才意外看到料號清單多了一堆資料。
    summary['legacy_labels_to_provision'] = sorted({r['raw_sku_label'] for r in records if r['is_legacy_column']})

    if mode != 'commit':
        summary['mode'] = 'preview'
        summary['has_existing'] = has_existing
        # 預覽模式額外回傳前50筆明細，方便使用者肉眼核對是否合理
        summary['sample_records'] = records[:50]
        return jsonify(summary)

    if not records:
        return jsonify({'error': '解析結果為0筆，未寫入任何資料，請確認檔案內容/格式是否正確'}), 400

    if has_existing and force:
        # 分批刪除，避免單次刪除範圍過大；PostgREST的delete()需要至少一個過濾條件，
        # 這裡用source_type='excel_import'限定範圍，不會動到dealer_sync/shipment_diff的資料
        sb.table('sku_invoice_ledger').delete().eq('source_type', 'excel_import').execute()

    # 2026-08-21新增：自動在「內部料號清單」建立佔位資料——總表欄名比對不到現有內部料號
    # 時（多半是已歸檔的舊料號），過去只會把records標記is_legacy_column=True、
    # internal_code=None，使用者完全無法在「料號與規則管理」找到這個欄位、也就無從設定
    # 「發票總表顯示」開關。這裡改成：commit當下就把這些欄名自動補成一筆
    # internal_sku_list佔位資料（預設is_active=False、hide_in_invoice_ledger=True，
    # 代表「已歸檔、預設不顯示」），並回填到records的internal_code欄位，讓寬表檢視功能
    # 之後能直接用internal_code對應到這筆佔位資料的hide_in_invoice_ledger設定。
    # 刻意保留is_legacy_column=True不變——這是「當初匯入時是否比對得到」的歷史事實，跟
    # 「現在internal_code有沒有對應資料」是兩件事，即使之後補了佔位資料，這批仍應被視為
    # 曾經是已歸檔欄位。
    # 2026-08-23修正：原本用sorted({...})對欄位名稱做文字排序，建出來的佔位資料
    # sort_order跟Excel總表欄位由左到右的實際順序對不起來，使用者反饋希望改成依總表欄位
    # 原始順序排列，方便事後對照。sku_cols本身就是_sil_parse_workbook()依左到右欄位順序
    # 建立的清單（見該函式「H欄=index 7...到end_idx都是料號欄」那段），這裡直接沿用它的
    # 順序來排legacy_labels，不再對文字做sorted()排序；同一欄名若剛好在總表裡重複出現
    # 多次，只取第一次出現的位置（不影響去重結果，只影響排序）。
    legacy_label_set = {r['raw_sku_label'] for r in records if r['is_legacy_column']}
    seen_legacy_labels = set()
    legacy_labels = []
    for _, col_label in sku_cols:
        if col_label in legacy_label_set and col_label not in seen_legacy_labels:
            seen_legacy_labels.add(col_label)
            legacy_labels.append(col_label)
    provisioned = []
    provision_error = None
    if legacy_labels:
        still_missing = [lbl for lbl in legacy_labels if lbl not in valid_codes]
        if still_missing:
            uname = session.get('display_name', session.get('username', ''))
            # 2026-08-24新增：這裡自動建立的佔位資料也是internal_sku_list的一筆真實記錄，
            # 同樣要配到唯一的「料號序號」(sku_seq_no)，否則之後這幾筆佔位資料會完全無法
            # 被「批次修改匯入」比對到（序號比對不到=一律略過，不會自動新增/回填序號）。
            # 批次內在本地遞增（原因同_next_sku_seq_no()函式說明：避免同一批次內後面的列
            # 重新查到跟前面一樣的「目前最大值」而拿到重複序號）。
            _seq_n = _sku_seq_max_n()
            placeholder_batch = []
            for i, lbl in enumerate(still_missing):
                _seq_n += 1
                placeholder_batch.append({
                    'internal_code': lbl,
                    'sku_seq_no': f'C{_seq_n:04d}',
                    'report_product_name': lbl,
                    'is_active': False,
                    'hide_in_invoice_ledger': True,
                    'note': '由「已開/未開發票總表」歷史匯入自動建立的已歸檔料號佔位資料，'
                            '總表原始欄名比對不到現有料號時系統自動產生',
                    'sort_order': max_sort_order + 1 + i,
                    'created_by': session.get('user_id'),
                    'created_by_name': uname,
                })
            try:
                ins_res = sb.table('internal_sku_list').insert(placeholder_batch).execute()
                provisioned = still_missing
                valid_codes.update(still_missing)
                # 2026-08-23修正：把剛建立的佔位資料id也補進id_by_code，供下面
                # invoice_ledger_col_order批次回寫使用（改成批次upsert前，這裡漏補
                # id_by_code的話，剛建立的佔位料號那一欄排序會被誤判成「找不到id」而跳過，
                # 排到寬表檢視最後面，並非本意——這些新佔位料號一樣要有正確的欄位順序）。
                for row in (ins_res.data or []):
                    if row.get('internal_code'):
                        id_by_code[row['internal_code']] = row['id']
            except Exception as e:
                # 佔位資料建立失敗不影響歷史匯入本身繼續進行——records仍會照is_legacy_column
                # 原樣寫入sku_invoice_ledger（internal_code=None），只是使用者這批「已歸檔」
                # 欄位這次沒辦法在料號清單看到，之後可以再重新匯入一次（有existing資料時
                # 用force覆蓋）讓系統重試建立佔位資料。
                provision_error = str(e)
            if provisioned:
                provisioned_set = set(provisioned)
                for r in records:
                    if r['is_legacy_column'] and r['raw_sku_label'] in provisioned_set:
                        r['internal_code'] = r['raw_sku_label']

    # 2026-08-23新增：使用者要求「已開/未開發票總表」寬表檢視的料號欄位順序改成依照
    # 這份Excel總表欄位由左到右的原始順序顯示，方便核對，不要沿用internal_sku_list既有的
    # sort_order（那個排序同時也服務「方案清單」等其他畫面，跟總表欄位順序不一定一致，
    # 不能直接重用，理由比照schema v1.36 hide_in_invoice_ledger不重用is_active）。
    # sku_cols本身就是_sil_parse_workbook()依左到右欄位順序建立的清單，這裡用enumerate
    # 取得每個欄位在總表裡的順序值，寫回internal_sku_list.invoice_ledger_col_order
    # （schema v1.38）。因為比對到/剛建立佔位資料的internal_code都等於raw_sku_label
    # 文字本身（見上面比對邏輯），可以直接拿raw_sku_label當internal_code去更新對應那筆
    # 料號清單資料；valid_codes此時已包含剛才新建立的佔位料號，一併涵蓋在內。
    # 2026-08-23修正：原本這裡是逐欄逐筆update（一個料號欄一次網路來回），總表欄位一多
    # （例如這次168欄）就要168次序列化的DB round trip，實測光是這段就可能拖到超過
    # Render的請求逾時限制，導致「確認匯入」整個請求逾時、瀏覽器收到非JSON回應
    # （伺服器其實還在背景繼續跑，跑到一半才被中斷，造成sku_invoice_ledger只寫入一部分
    # 批次就中止，需要使用者重新整理才看得出實際寫入到哪裡）。改成用id當key、只帶
    # id/invoice_ledger_col_order兩個欄位分批(每批300筆)呼叫upsert()——PostgREST的
    # upsert在Prefer=merge-duplicates（supabase-py預設）下只會覆蓋payload裡有出現的
    # 欄位，不會動到其他既有欄位、也不會因為帶了id而誤觸新增分支，效果等同原本逐筆
    # update，但168欄現在只要1次網路來回（分批的話也只要一兩次），大幅縮短耗時。
    # 2026-08-23再修正：改成批次upsert後實測跳出Supabase 400 (23502 not-null violation)——
    # PostgREST的upsert()底層是產生真正的「INSERT ... ON CONFLICT(id) DO UPDATE」SQL，
    # 即使最終一定會走DO UPDATE分支（id已存在），Postgres在規劃INSERT那一步仍然要先用
    # payload裡帶的欄位組出一列候選資料、檢查NOT NULL等資料表約束，payload只帶
    # id/invoice_ledger_col_order兩欄，internal_code（NOT NULL UNIQUE）沒帶到，检查
    # 階段就直接失敗，跟id是否真的衝突無關。這正是原本逐筆update()的寫法要特別避開
    # upsert的理由（見上面被取代前的舊註解）。修法：payload多帶一個internal_code
    # （值就是raw_label本身，等於這筆資料本來就有的值，不會真的改到內容），滿足
    # INSERT分支的NOT NULL檢查，DO UPDATE時只會覆蓋payload出現的欄位
    # （invoice_ledger_col_order，以及重複寫回同一個internal_code值，無副作用）。
    col_order_error = None
    col_order_updated = 0
    try:
        col_order_payload = []
        for col_order, (_, raw_label) in enumerate(sku_cols):
            row_id = id_by_code.get(raw_label)
            if row_id is not None:
                col_order_payload.append({
                    'id': row_id,
                    'internal_code': raw_label,
                    'invoice_ledger_col_order': col_order,
                })
        CHUNK = 300
        for i in range(0, len(col_order_payload), CHUNK):
            sb.table('internal_sku_list').upsert(col_order_payload[i:i + CHUNK], on_conflict='id').execute()
        col_order_updated = len(col_order_payload)
    except Exception as e:
        col_order_error = str(e)

    BATCH = 500
    inserted = 0
    for i in range(0, len(records), BATCH):
        chunk = records[i:i + BATCH]
        for r in chunk:
            r['created_by'] = session.get('user_id')
            r['created_by_name'] = session.get('display_name', session.get('username', ''))
        sb.table('sku_invoice_ledger').insert(chunk).execute()
        inserted += len(chunk)

    summary['mode'] = 'commit'
    summary['inserted'] = inserted
    summary['replaced_existing'] = bool(has_existing and force)
    summary['sku_placeholders_created'] = provisioned
    summary['sku_placeholder_count'] = len(provisioned)
    if provision_error:
        summary['sku_placeholder_error'] = provision_error
    summary['col_order_updated_count'] = col_order_updated
    if col_order_error:
        summary['col_order_error'] = col_order_error
    return jsonify(summary)


# ============================================================
# 「比對已開/未開發票」— 寬表檢視（2026-08-21新增）
# ============================================================
# 把 sku_invoice_ledger 的 long-format 明細，依篩選條件重新組回「使用者原本習慣的
# 寬表」樣式（一列＝一筆對帳事件，橫向每個料號各佔一欄）。跟原始Excel總表不同的是：
# 欄位順序/是否顯示可由「內部料號清單」的 hide_in_invoice_ledger + sort_order 動態
# 控制，不用像Excel那樣手動調欄位。

_WIDE_VIEW_FROZEN_LABELS = ['客戶類型', '對帳名稱', '期間', '期間區間']


def _wide_view_build(args):
    """共用查詢/彙總邏輯，JSON檢視與Excel匯出都呼叫這個函式。
    args：一個支援 .get(key, default) 與 .getlist(key) 的物件（Flask的request.args）。
    回傳 dict：{'columns':[...], 'rows':[...], 'subtotal':{...}, 'grand_total':n,
                'filter_options':{...}, 'row_count':n}
    """
    customer_types = [x for x in args.getlist('customer_types') if x]
    source_type = (args.get('source_type') or '').strip()
    keyword = (args.get('keyword') or '').strip()
    period_kw = (args.get('period_kw') or '').strip()
    show_hidden = args.get('show_hidden') == '1'
    only_nonzero_cols = args.get('only_nonzero_cols') == '1'
    # 2026-08-24修正：預設值改成不合併——使用者反映匯入結果應該要比照Excel原始每一列
    # 獨立呈現，不應該預設就把同客戶類型/對帳名稱/期間的多列自動合併成一列，畫面上的
    # checkbox預設也已同步改成不勾選（見dealer_reconcile.html iwMergeByName）。這裡改
    # 預設值是為了防呆（萬一有其他呼叫路徑沒帶這個參數），實際上目前前端一律會明確帶入
    # 這個參數，不受此預設值影響。
    merge_by_name = args.get('merge_by_name', '0') == '1'

    # 2026-08-23修正：料號欄位順序改成優先依invoice_ledger_col_order排（總表歷史匯入
    # 「確認匯入」時依Excel欄位由左到右順序寫入，見import_invoice_ledger_history()），
    # 不再用sort_order（那個排序服務「方案清單」等其他畫面，跟總表欄位順序不一定一致）。
    # 從未出現在任何一次總表匯入的料號，invoice_ledger_col_order是NULL，用nullslast排到
    # 最後面；sort_order留著當次要排序值時的次要排序鍵（tie-break），不影響主要排序依據。
    sku_rows = _fetch_all_rows(lambda: sb.table('internal_sku_list').select(
        'internal_code,report_product_name,sort_order,hide_in_invoice_ledger,is_active,'
        'invoice_ledger_col_order'
    ).order('invoice_ledger_col_order', nullslast=True).order('sort_order'))
    sku_by_code = {r['internal_code']: r for r in sku_rows if r.get('internal_code')}

    ledger_rows = _fetch_all_rows(lambda: sb.table('sku_invoice_ledger').select(
        'internal_code,raw_sku_label,is_legacy_column,customer_type,reconcile_name,period,'
        'period_range,note,qty,reconcile_order,reconcile_no,bookstore_reconcile_no,order_no,'
        'techkt_note,techkt_reconcile_order,techkt_reconcile_no,'
        'source_type,source_ref,source_row_order'))

    # ---- 篩選 ----
    all_customer_types = sorted({r['customer_type'] for r in ledger_rows if r.get('customer_type')})
    all_source_types = sorted({r['source_type'] for r in ledger_rows if r.get('source_type')})

    def _kw_hit(r):
        if not keyword:
            return True
        hay = ' '.join(str(r.get(k) or '') for k in
                        ('reconcile_name', 'note', 'order_no', 'reconcile_no', 'bookstore_reconcile_no',
                         'techkt_note', 'techkt_reconcile_order', 'techkt_reconcile_no'))
        return keyword.lower() in hay.lower()

    def _period_hit(r):
        if not period_kw:
            return True
        hay = f"{r.get('period') or ''} {r.get('period_range') or ''}"
        return period_kw.lower() in hay.lower()

    filtered = []
    for r in ledger_rows:
        if customer_types and r.get('customer_type') not in customer_types:
            continue
        if source_type and r.get('source_type') != source_type:
            continue
        if not _kw_hit(r) or not _period_hit(r):
            continue
        filtered.append(r)

    # ---- 欄位（依內部料號清單 sort_order 排序；比對不到現有料號的欄名另外附加在最後） ----
    columns = []
    col_key_set = set()
    for sku in sku_rows:
        code = sku['internal_code']
        if not show_hidden and sku.get('hide_in_invoice_ledger'):
            continue
        columns.append({
            'key': code,
            'label': sku.get('report_product_name') or code,
            'sub_label': code,
            'is_legacy': False,
        })
        col_key_set.add(code)

    orphan_labels = []
    seen_orphan = set()
    for r in filtered:
        code = r.get('internal_code')
        label = r.get('raw_sku_label') or code
        # 2026-08-25新增：跳過_sil_extract_records()新增的「無明細品項」佔位紀錄
        # （raw_sku_label是_SIL_NO_ITEM_RAW_LABEL這個固定字串、internal_code為None，
        # 代表這一列在Excel裡所有料號欄數量都是0/空白，純粹是備註/對帳性質的列，見上方
        # 註解）——這種紀錄沒有對應任何真實或已歸檔的料號，不該被誤判成一個查無對應料號
        # 的孤兒欄位（否則會多冒出一欄顯示這串提示文字的空白欄位）。
        if r.get('raw_sku_label') == _SIL_NO_ITEM_RAW_LABEL:
            continue
        key = code if (code and code in col_key_set) else label
        if code and code in col_key_set:
            continue  # 已經是正常欄位
        if key in seen_orphan:
            continue
        seen_orphan.add(key)
        orphan_labels.append(key)
    for key in orphan_labels:
        columns.append({'key': key, 'label': key, 'sub_label': '已歸檔／查無對應料號', 'is_legacy': True})
        col_key_set.add(key)

    # ---- 分組 ----
    groups = {}
    order_seq = []

    for r in filtered:
        code = r.get('internal_code')
        raw_label = r.get('raw_sku_label') or code
        col_key = code if (code and code in col_key_set and not (code not in sku_by_code)) else raw_label
        # 保險：若 internal_code 存在但目前料號清單已查不到這筆（如被刪除），退回用 raw_sku_label 當欄位鍵
        if code and code not in sku_by_code:
            col_key = raw_label

        if merge_by_name:
            gkey = (r.get('customer_type'), r.get('reconcile_name'), r.get('period'), r.get('period_range'))
        elif r.get('source_ref') and r.get('source_row_order') is not None:
            # 2026-08-24修正：使用者取消勾選「合併成一列」後，期望的是完全比照Excel匯入時
            # 的原始列順序、每一列都獨立呈現，不要有任何合併。舊寫法用一組business欄位
            # （note/reconcile_order/order_no/techkt_*等）當分組鍵，這仍然是一種「內容相同
            # 就合併」的邏輯——只要恰好這些欄位都相同（例如同一對帳名稱/期間內兩列都沒填
            # 備註/對帳訂單編號等選填欄位），還是會被誤合併成一列，跟使用者想要的「每列都
            # 獨立」不符。改直接用(source_ref, source_row_order)——即「這一列在哪個匯入
            # 檔案的第幾列」，這是每一筆明細記錄天生就有、對應到單一Excel原始列的鍵值，
            # 不會有任兩個不同列被誤判成相同的疑慮。source_row_order是v1.39才新增的欄位，
            # 舊資料或非Excel匯入來源(如手動新增)可能沒有這個值，這種情況才退回舊的
            # business欄位分組法（見下方else分支）。
            gkey = ('_row', r.get('source_ref'), r.get('source_row_order'))
        else:
            gkey = (r.get('customer_type'), r.get('reconcile_name'), r.get('period'), r.get('period_range'),
                    r.get('note'), r.get('reconcile_order'), r.get('reconcile_no'),
                    r.get('bookstore_reconcile_no'), r.get('order_no'), r.get('source_type'), r.get('source_ref'),
                    r.get('techkt_note'), r.get('techkt_reconcile_order'), r.get('techkt_reconcile_no'))

        g = groups.get(gkey)
        if g is None:
            g = {
                'customer_type': r.get('customer_type'),
                'reconcile_name': r.get('reconcile_name'),
                'period': r.get('period'),
                'period_range': r.get('period_range'),
                'notes': [], 'order_nos': [], 'source_types': [],
                # 2026-08-22新增：使用者反饋寬表檢視畫面上看不到FX/FY/FZ「科技通/知識通」
                # 專屬欄位，確認後改為要在畫面上新增獨立可見欄位（跟通用reconcile_order/
                # reconcile_no/bookstore_reconcile_no維持「只用於搜尋/去重、不顯示」不同，
                # 這三個是使用者主動要求要顯示的），比照notes/order_nos的收集+合併寫法。
                'techkt_notes': [], 'techkt_reconcile_orders': [], 'techkt_reconcile_nos': [],
                'values': {},
                # 2026-08-24新增：使用者要求「已開/未開發票總表」寬表檢視的縱向列順序，
                # 要能還原成匯入當時Excel的原始上下排列，而非依客戶類型/期間等文字重新排序
                # （欄位順序v3.81已改用invoice_ledger_col_order解決，這裡是另外的「列」順序）。
                # 做法：每筆明細記錄本身在寫入sku_invoice_ledger時就帶有source_row_order
                # （schema v1.39，即_sil_extract_records()記錄的Excel實際列號），同一個
                # 寬表列（同一組customer_type/reconcile_name/period/period_range）通常對應
                # 到Excel同一列的多個料號欄，理論上source_row_order會相同；用min()取最小值
                # 是為了保險（例如merge_by_name=False時，不同料號欄仍可能落在不同細節但同一
                # 群組的情況），並非代表允許同一群組橫跨多個Excel列。
                'min_row_order': None,
            }
            groups[gkey] = g
            order_seq.append(gkey)
        if r.get('note') and r.get('note') not in g['notes']:
            g['notes'].append(r.get('note'))
        if r.get('order_no') and r.get('order_no') not in g['order_nos']:
            g['order_nos'].append(r.get('order_no'))
        if r.get('source_type') and r.get('source_type') not in g['source_types']:
            g['source_types'].append(r.get('source_type'))
        if r.get('techkt_note') and r.get('techkt_note') not in g['techkt_notes']:
            g['techkt_notes'].append(r.get('techkt_note'))
        if r.get('techkt_reconcile_order') and r.get('techkt_reconcile_order') not in g['techkt_reconcile_orders']:
            g['techkt_reconcile_orders'].append(r.get('techkt_reconcile_order'))
        if r.get('techkt_reconcile_no') and r.get('techkt_reconcile_no') not in g['techkt_reconcile_nos']:
            g['techkt_reconcile_nos'].append(r.get('techkt_reconcile_no'))
        if col_key is not None and r.get('raw_sku_label') != _SIL_NO_ITEM_RAW_LABEL:
            g['values'][col_key] = g['values'].get(col_key, 0) + (r.get('qty') or 0)
        ro = r.get('source_row_order')
        if ro is not None and (g['min_row_order'] is None or ro < g['min_row_order']):
            g['min_row_order'] = ro

    rows = []
    for gkey in order_seq:
        g = groups[gkey]
        row_total = sum(g['values'].values())
        rows.append({
            'customer_type': g['customer_type'],
            'reconcile_name': g['reconcile_name'],
            'period': g['period'],
            'period_range': g['period_range'],
            'note': '、'.join(g['notes']),
            'order_no': '、'.join(g['order_nos']),
            'source_type': '、'.join(g['source_types']),
            # 2026-08-22新增：對應上方techkt_notes/techkt_reconcile_orders/techkt_reconcile_nos
            'techkt_note': '、'.join(g['techkt_notes']),
            'techkt_reconcile_order': '、'.join(g['techkt_reconcile_orders']),
            'techkt_reconcile_no': '、'.join(g['techkt_reconcile_nos']),
            'values': g['values'],
            'row_total': row_total,
            '_row_order': g.get('min_row_order'),
        })

    # 2026-08-24修正：原本依(customer_type, period, reconcile_name)文字排序，跟使用者
    # Excel原始上下排列不一致，改成依Excel實際列號(_row_order)排序，還原匯入當時的
    # 縱向順序，方便使用者對照原始總表核對。舊資料（v1.39上線前匯入、或source_row_order
    # 尚未回填的資料）沒有_row_order，用float('inf')排到最後面，而非讓整批排序失效。
    rows.sort(key=lambda x: x['_row_order'] if x['_row_order'] is not None else float('inf'))
    for row in rows:
        row.pop('_row_order', None)

    # ---- 只顯示有數量的欄位 ----
    if only_nonzero_cols:
        nonzero_keys = set()
        for row in rows:
            for k, v in row['values'].items():
                if v:
                    nonzero_keys.add(k)
        columns = [c for c in columns if c['key'] in nonzero_keys]

    # ---- 小計 ----
    subtotal = {c['key']: 0 for c in columns}
    grand_total = 0
    for row in rows:
        for c in columns:
            v = row['values'].get(c['key'], 0)
            subtotal[c['key']] += v
        grand_total += row['row_total']

    return {
        'columns': columns,
        'rows': rows,
        'subtotal': subtotal,
        'grand_total': grand_total,
        'row_count': len(rows),
        'filter_options': {
            'customer_types': all_customer_types,
            'source_types': all_source_types,
        },
    }


def _wide_view_build_safe(args):
    """2026-08-21新增：包一層try/except呼叫_wide_view_build()。這個功能依賴的
    sku_invoice_ledger表／internal_sku_list.hide_in_invoice_ledger欄位，是分別在
    schema v1.35／v1.36才新增的，若Supabase尚未執行對應SQL migration，_wide_view_build()
    裡的查詢會拋出PostgREST例外——沒有這層防護的話，就會變成未捕捉例外→Flask預設500 HTML
    錯誤頁→前端fetch().json()解析失敗，畫面顯示難以理解的「Unexpected token '<'...
    is not valid JSON」。改成回傳明確的中文錯誤訊息，比照[[v3.23]]「任何依賴新表/新欄位
    的查詢都要包容錯，降級成清楚訊息」的既有教訓。回傳 (data, error_message)，
    error_message為None代表成功。"""
    try:
        return _wide_view_build(args), None
    except Exception as e:
        return None, ('讀取「已開/未開發票總表」寬表資料失敗，最常見原因是資料庫尚未執行最新的'
                       'SQL migration（請確認已在Supabase SQL Editor執行過'
                       ' dealer_reconcile_schema.sql 的 v1.35／v1.36 區塊：'
                       '建立 sku_invoice_ledger 表、internal_sku_list 新增'
                       ' hide_in_invoice_ledger 欄位）。詳細錯誤：' + str(e))


@dealer_bp.route('/api/dealer/invoice-ledger/wide-view', methods=['GET'])
@dealer_view_required
def invoice_ledger_wide_view():
    data, err = _wide_view_build_safe(request.args)
    if err:
        return jsonify({'error': err}), 500
    return jsonify(data)


@dealer_bp.route('/api/dealer/invoice-ledger/wide-view/export', methods=['GET'])
@dealer_view_required
def invoice_ledger_wide_view_export():
    data, err = _wide_view_build_safe(request.args)
    if err:
        return jsonify({'error': err}), 500
    columns = data['columns']
    rows = data['rows']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '已開未開發票寬表'

    hfill = openpyxl.styles.PatternFill('solid', fgColor='1A5276')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF')
    legacyfill = openpyxl.styles.PatternFill('solid', fgColor='F2F2F2')
    legacyfont = openpyxl.styles.Font(italic=True, color='888888')
    subfill = openpyxl.styles.PatternFill('solid', fgColor='FFF2CC')
    subfont = openpyxl.styles.Font(bold=True, color='7B3F00')
    center = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 2026-08-22新增：科技通/知識通專屬FX/FY/FZ三欄，比照上面_wide_view_build()的說明，
    # 使用者要求要顯示成獨立欄位，故一併加進Excel匯出。
    base_headers = ['客戶類型', '對帳名稱', '期間', '期間區間', '備註', '對帳訂單編號',
                     '科技通備註', '科技通對帳順序', '科技通對帳編號']
    all_headers = base_headers + [c['label'] for c in columns] + ['小計', '來源']
    for ci, h in enumerate(all_headers, 1):
        c = ws.cell(1, ci, h)
        c.font = hfont
        c.fill = hfill
        c.alignment = center

    # 第2列：料號的sub_label（内部料號/已歸檔標示），左側固定欄與小計/來源欄留空
    for ci in range(1, len(base_headers) + 1):
        ws.cell(2, ci, '')
    for i, col in enumerate(columns):
        c = ws.cell(2, len(base_headers) + 1 + i, col['sub_label'])
        c.alignment = center
        if col['is_legacy']:
            c.fill = legacyfill
            c.font = legacyfont
        else:
            c.font = openpyxl.styles.Font(size=9, color='888888')
    ws.cell(2, len(base_headers) + len(columns) + 1, '')
    ws.cell(2, len(base_headers) + len(columns) + 2, '')

    # 2026-08-25修正（bug根因）：原本用ws.cell(3, ...).coordinate只是想算出凍結窗格的
    # 座標字串（如'J3'），但openpyxl的ws.cell()一經呼叫就會實際在工作表建立/觸碰該儲存格，
    # 副作用是把ws.max_row從2（此時只寫了表頭列1、料號sub_label列2）直接墊高到3——導致
    # 緊接在後面用ws.append(vals)寫入的第一筆真實資料，被迫從第4列開始寫，永遠空出第3列
    # 一整列（每一欄都是空的）。改用openpyxl.utils.get_column_letter()純字串運算組座標，
    # 不去碰觸工作表本身，避免這個副作用。
    ws.freeze_panes = f'{openpyxl.utils.get_column_letter(len(base_headers) + 1)}3'

    for row in rows:
        vals = [row['customer_type'], row['reconcile_name'], row['period'], row['period_range'],
                row['note'], row['order_no'],
                row['techkt_note'], row['techkt_reconcile_order'], row['techkt_reconcile_no']]
        vals += [row['values'].get(c['key']) or None for c in columns]
        vals += [row['row_total'], row['source_type']]
        ws.append(vals)

    sub_row = ['小計（共%d列）' % len(rows), '', '', '', '', '', '', '', '']
    sub_row += [data['subtotal'].get(c['key']) or None for c in columns]
    sub_row += [data['grand_total'], '']
    r_idx = ws.max_row + 1
    ws.append(sub_row)
    for ci in range(1, len(all_headers) + 1):
        cc = ws.cell(r_idx, ci)
        cc.fill = subfill
        cc.font = subfont

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    for i in range(len(columns)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(len(base_headers) + 1 + i)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(base_headers) + len(columns) + 1)].width = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(base_headers) + len(columns) + 2)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='已開未開發票寬表.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# 「物料清單」新模組（material_master_list） — 2026-08-24新增
#
# 背景：使用者要求在「經銷商對帳模組→資料維護」新增一個獨立模組，維護實體物料
# （機器/殼套/螢幕貼/週邊/觸控筆）的料號主檔，可單筆新增/刪除/修改，列表顯示
# 序號/料號/物料品名/新增時間/建檔人員，並可透過權限管理設定此模組權限（見
# dealer_reconcile_schema.sql v1.40）。刻意不重用「料號與規則管理」的
# internal_sku_list——那張表是拆帳比對專用主檔，欄位/用途跟這裡單純的實體物料
# 主檔管理完全不同，混用會互相污染（理由同v1.40 SQL註解）。
#
# 「料號」使用者要求「系統自動產生」：使用者只選「類型」＋填「公司別」(可留空)，
# 系統依下方_MATERIAL_CATEGORIES規則找出該類型目前最大流水號＋1，組出新料號，
# 不開放使用者自行輸入/事後修改料號，避免打錯格式或跟既有料號重複。
# ============================================================
_MATERIAL_CATEGORIES = [
    ('S001A', '機器'),
    ('S002A', '殼套'),
    ('S003A', '螢幕貼'),
    ('S004A', '週邊'),
    ('S005A', '觸控筆'),
]
_MATERIAL_CATEGORY_BY_LABEL = {label: prefix for prefix, label in _MATERIAL_CATEGORIES}

_MATERIAL_PERM_KEYS = (
    'mod_material_view', 'mod_material_create', 'mod_material_update', 'mod_material_delete',
)


def _any_material_perm():
    return any(_perm_ok(k) for k in _MATERIAL_PERM_KEYS)


def material_view_required(f):
    """「物料清單」檢視權限：mod_dealer_edit（既有整體編輯權限）OR 這個功能任一細分
    權限，即可打開頁面/列表，比照bookstore_shipping_view_required的設計。"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not (_perm_ok('mod_dealer_edit') or _any_material_perm()):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def material_action_required(*keys):
    """「物料清單」新增/修改/刪除細分權限，各自獨立勾選，比照
    bookstore_shipping_action_required()的設計。"""
    def deco(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not _login_ok():
                return jsonify({'error': '請先登入'}), 401
            if not (_perm_ok('mod_dealer_edit') or any(_perm_ok(k) for k in keys)):
                return jsonify({'error': '權限不足'}), 403
            return f(*args, **kwargs)
        return wrapped
    return deco


def _material_next_seq(prefix):
    """掃描目前資料庫裡所有以這個類型前綴開頭的料號，取出緊接在前綴後面的3碼數字
    流水號部分，回傳「目前最大值+1」。直接掃描實際料號字串（而非只信任seq_no欄位），
    避免seq_no欄位萬一跟料號本身不同步時算錯下一號。"""
    rows = sb.table('material_master_list').select('material_code') \
        .like('material_code', prefix + '%').execute()
    max_seq = 0
    for r in (rows.data or []):
        code = r.get('material_code') or ''
        rest = code[len(prefix):]
        m = re.match(r'^(\d{3})', rest)
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return max_seq + 1


@dealer_bp.route('/api/material-list', methods=['GET'])
@material_view_required
def list_material_list():
    rows = _fetch_all_rows(lambda: sb.table('material_master_list').select('*').order('id'))
    return jsonify(rows)


@dealer_bp.route('/api/material-list/categories', methods=['GET'])
@material_view_required
def list_material_categories():
    """提供前端「新增」彈窗的類型下拉選單選項，跟後端_MATERIAL_CATEGORIES共用同一份
    定義，避免前後端各自維護一份分類規則、日後改規則時漏改其中一邊。"""
    return jsonify([{'prefix': prefix, 'label': label} for prefix, label in _MATERIAL_CATEGORIES])


@dealer_bp.route('/api/material-list', methods=['POST'])
@material_action_required('mod_material_create')
def create_material():
    data = request.json or {}
    category = _norm(data.get('category'))
    company_code = _norm(data.get('company_code')).upper()
    material_name = _norm(data.get('material_name'))
    if not material_name:
        return jsonify({'error': '物料品名必填'}), 400
    prefix = _MATERIAL_CATEGORY_BY_LABEL.get(category)
    if not prefix:
        return jsonify({'error': '請選擇有效的物料類型'}), 400
    last_err = None
    # 正常情況下不會真的衝突（只有同一瞬間兩個人同時新增同一類型才可能撞號），這裡
    # 保留重試只是防呆，避免極端併發狀況下直接500給使用者看不懂的錯誤。
    for _ in range(5):
        seq = _material_next_seq(prefix)
        code = f'{prefix}{seq:03d}{company_code}'
        rec = {
            'material_code': code,
            'category': category,
            'company_code': company_code or None,
            'seq_no': seq,
            'material_name': material_name,
        }
        rec.update(_audit_new())
        try:
            res = sb.table('material_master_list').insert(rec).execute()
            return jsonify(res.data[0]), 201
        except Exception as e:
            last_err = e
            if '23505' not in str(e) and 'duplicate' not in str(e).lower():
                return jsonify({'error': f'新增失敗：{e}'}), 500
            continue
    return jsonify({'error': f'新增失敗，料號持續衝突，請重新整理後再試一次：{last_err}'}), 409


@dealer_bp.route('/api/material-list/<int:mid>', methods=['PUT'])
@material_action_required('mod_material_update')
def update_material(mid):
    """料號一經產生即不可修改（避免跟其他模組引用的料號脫鉤），編輯僅允許修改物料品名。"""
    data = request.json or {}
    material_name = _norm(data.get('material_name'))
    if not material_name:
        return jsonify({'error': '物料品名必填'}), 400
    cur = sb.table('material_master_list').select('id').eq('id', mid).limit(1).execute()
    if not cur.data:
        return jsonify({'error': '找不到這筆物料資料'}), 404
    rec = {'material_name': material_name}
    rec.update(_audit_upd())
    sb.table('material_master_list').update(rec).eq('id', mid).execute()
    return jsonify({'ok': True})


@dealer_bp.route('/api/material-list/<int:mid>', methods=['DELETE'])
@material_action_required('mod_material_delete')
def delete_material(mid):
    cur = sb.table('material_master_list').select('id').eq('id', mid).limit(1).execute()
    if not cur.data:
        return jsonify({'error': '找不到這筆物料資料'}), 404
    sb.table('material_master_list').delete().eq('id', mid).execute()
    return jsonify({'ok': True})


# ============================================================
# 「銷退清單」新模組（sales_return_list / sales_return_items） — 2026-08-24新增
#
# 背景：使用者要求在「經銷商對帳模組→資料維護」新增一個獨立模組，登記銷退（折讓）
# 應收憑單資料，可單筆新增/刪除/修改。每筆主資料底下可有多列「財報料號品名＋子料號
# (選填)＋拆帳金額(含稅)」明細，架構比照「書店經銷商-方案清單」(bookstore_plans/
# bookstore_plan_items) 的主檔+明細拆帳設計，品名/子料號下拉選單直接沿用既有的
# _plan_valid_report_names()/_sku_report_name_by_code()（見上方③區段），不重複寫一份。
#
# 跟bookstore_plans驗證邏輯的關鍵差異（使用者2026-08-24明確要求）：財報料號品名/子料號
# 本身無效（不存在、或子料號跟品名對不上）時仍然擋下儲存（資料完整性問題）；但「明細
# 金額加總」跟「折讓金額(含稅)」不相符時，只回傳警示訊息，不擋下儲存（跟bookstore_plans
# 需要force=true才能強制儲存的做法不同，這裡從一開始就不擋）。
# ============================================================
_SALES_RETURN_PERM_KEYS = (
    'mod_sales_return_view', 'mod_sales_return_create',
    'mod_sales_return_update', 'mod_sales_return_delete',
)


def _any_sales_return_perm():
    return any(_perm_ok(k) for k in _SALES_RETURN_PERM_KEYS)


def sales_return_view_required(f):
    """「銷退清單」檢視權限：mod_dealer_edit（既有整體編輯權限）OR 這個功能任一細分權限，
    即可打開頁面/列表，比照material_view_required()的設計。"""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not _login_ok():
            return jsonify({'error': '請先登入'}), 401
        if not (_perm_ok('mod_dealer_edit') or _any_sales_return_perm()):
            return jsonify({'error': '權限不足'}), 403
        return f(*args, **kwargs)
    return wrapped


def sales_return_action_required(*keys):
    """「銷退清單」新增/修改/刪除細分權限，各自獨立勾選，比照material_action_required()
    的設計。"""
    def deco(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not _login_ok():
                return jsonify({'error': '請先登入'}), 401
            if not (_perm_ok('mod_dealer_edit') or any(_perm_ok(k) for k in keys)):
                return jsonify({'error': '權限不足'}), 403
            return f(*args, **kwargs)
        return wrapped
    return deco


def _validate_sales_return_items(items):
    """檢查每一列的財報料號品名是否存在於「內部料號清單」、子料號(選填)是否存在且屬於
    該品名底下，跟_validate_plan_items()前半段的品名/子料號驗證規則完全一致（直接沿用
    同一份_plan_valid_report_names()/_sku_report_name_by_code()，避免兩份規則各自維護
    日後失去同步）。回傳 (error_response_or_None)；這裡刻意不含金額加總比對——那部分是
    軟性警示、不擋存檔，交給呼叫端另外處理，見下方_sales_return_amount_diff()。"""
    valid_names = _plan_valid_report_names()
    sub_name_by_code = _sku_report_name_by_code() if any(_norm(it.get('sub_internal_code')) for it in items) else {}
    for it in items:
        name = _norm(it.get('report_product_name'))
        if not name:
            return jsonify({'error': '財報料號品名不可留空，請選擇一個財報料號品名或移除該列'}), 400
        if name not in valid_names:
            return jsonify({'error': f'財報料號品名「{name}」不存在於「內部料號清單」，請先在該清單新增此財報料號品名，或修正為既有品名'}), 400
        sub = _norm(it.get('sub_internal_code'))
        if sub:
            sub_name = sub_name_by_code.get(sub)
            if sub_name is None:
                return jsonify({'error': f'子料號「{sub}」不存在於「內部料號清單」'}), 400
            if sub_name != name:
                return jsonify({'error': f'子料號「{sub}」屬於財報料號品名「{sub_name}」，跟這一列選的「{name}」不一致，請重新選擇子料號'}), 400
    return None


def _sales_return_amount_diff(amount_incl, items):
    """回傳(items_total, diff)：items_total為明細金額加總，diff=items_total-折讓金額(含稅)，
    皆四捨五入到小數點後2位。純計算，不擋存檔——使用者明確要求金額不符仍可儲存，只顯示
    警示提醒，是bookstore_plans「不符會擋下」的較嚴格做法故意不採用的地方。"""
    items_total = round(sum(_to_num(it.get('amount')) for it in items), 2)
    diff = round(items_total - float(amount_incl), 2)
    return items_total, diff


def _sales_return_calc_tax(amount_incl):
    """折讓金額(未稅)=含稅/1.05，折讓金額(稅額)=含稅-未稅，皆四捨五入到小數點後2位。
    這2個欄位一律由後端重新計算、不採信前端送來的值，確保永遠跟「折讓金額(含稅)」維持
    公式一致（2026-08-24使用者需求明確指定這個公式）。"""
    amount_notax = round(amount_incl / 1.05, 2)
    tax_amount = round(amount_incl - amount_notax, 2)
    return amount_notax, tax_amount


def _sales_return_row_with_items(row, items):
    row = dict(row)
    row['items'] = [{
        'report_product_name': i['report_product_name'], 'amount': i['amount'],
        'sub_internal_code': i.get('sub_internal_code'),
    } for i in items]
    row['items_total'] = round(sum(i['amount'] for i in items), 2)
    return row


@dealer_bp.route('/api/sales-return', methods=['GET'])
@sales_return_view_required
def list_sales_return():
    kw = request.args.get('kw', '').strip()
    q = sb.table('sales_return_list').select('*')
    if kw:
        q = q.or_(f'ar_doc_no.ilike.%{kw}%,invoice_no.ilike.%{kw}%,order_no.ilike.%{kw}%,'
                  f'credit_note_no.ilike.%{kw}%,recon_summary_no.ilike.%{kw}%')
    rows = _fetch_all_rows(lambda: q.order('credit_date', desc=True).order('id', desc=True))

    return_ids = [r['id'] for r in rows]
    items_map = {}
    if return_ids:
        items_rows = _fetch_all_rows(lambda: sb.table('sales_return_items').select('*').in_('return_id', return_ids))
        for it in items_rows:
            items_map.setdefault(it['return_id'], []).append(it)
    out = [_sales_return_row_with_items(r, items_map.get(r['id'], [])) for r in rows]
    return jsonify(out)


# 2026-08-25新增：「銷退清單」匯出Excel，欄位/順序、金額差異計算方式皆比照畫面列表頁一致
# （見dealer_reconcile.html renderSalesReturnTable()）；kw篩選規則直接複用list_sales_return()
# 同一套（依應收憑單/發票號碼/訂單編號/折讓單號/對帳總表編號比對），搜尋框有輸入關鍵字時
# 按「匯出」只會匯出篩選後的結果，避免使用者匯出後才發現筆數跟畫面對不上。
@dealer_bp.route('/api/sales-return/export', methods=['GET'])
@sales_return_view_required
def export_sales_return():
    kw = request.args.get('kw', '').strip()
    q = sb.table('sales_return_list').select('*')
    if kw:
        q = q.or_(f'ar_doc_no.ilike.%{kw}%,invoice_no.ilike.%{kw}%,order_no.ilike.%{kw}%,'
                  f'credit_note_no.ilike.%{kw}%,recon_summary_no.ilike.%{kw}%')
    rows = _fetch_all_rows(lambda: q.order('credit_date', desc=True).order('id', desc=True))

    return_ids = [r['id'] for r in rows]
    items_map = {}
    if return_ids:
        items_rows = _fetch_all_rows(lambda: sb.table('sales_return_items').select('*').in_('return_id', return_ids))
        for it in items_rows:
            items_map.setdefault(it['return_id'], []).append(it)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '銷退清單'
    redfont = openpyxl.styles.Font(color='9C0006')
    headers = ['序號', '應收憑單', '發票號碼', '訂單編號', '折讓單號', '折讓日期',
               '折讓金額(含稅)', '折讓金額(未稅)', '折讓金額(稅額)', '備註', '金額差異',
               '對帳總表編號', '對帳總表區間', '財報料號品名', '處理人員', '新增日期']
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        items = items_map.get(r['id'], [])
        items_total = round(sum(_to_num(it.get('amount')) for it in items), 2)
        diff = round(items_total - _to_num(r.get('amount_incl')), 2)
        names = '、'.join(it.get('report_product_name') or '' for it in items)
        ws.append([
            i, r.get('ar_doc_no'), r.get('invoice_no'), r.get('order_no'), r.get('credit_note_no'),
            r.get('credit_date'), r.get('amount_incl'), r.get('amount_notax'), r.get('tax_amount'),
            r.get('note'), (diff if abs(diff) > 0.5 else None),
            r.get('recon_summary_no'), r.get('recon_summary_range'), names, r.get('handler_name'),
            (r.get('created_at') or '')[:16].replace('T', ' '),
        ])
        if abs(diff) > 0.5:
            ws.cell(ws.max_row, 11).font = redfont

    widths = [6, 16, 14, 14, 16, 12, 13, 13, 13, 18, 10, 14, 16, 26, 10, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='銷退清單.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@dealer_bp.route('/api/sales-return/<int:rid>', methods=['GET'])
@sales_return_view_required
def get_sales_return(rid):
    res = sb.table('sales_return_list').select('*').eq('id', rid).execute()
    if not res.data:
        return jsonify({'error': '找不到此筆銷退資料'}), 404
    items_res = sb.table('sales_return_items').select('*').eq('return_id', rid).execute()
    return jsonify(_sales_return_row_with_items(res.data[0], items_res.data or []))


_SALES_RETURN_REQUIRED_FIELDS = {
    'ar_doc_no': '應收憑單', 'invoice_no': '發票號碼', 'order_no': '訂單編號',
    'credit_note_no': '折讓單號', 'credit_date': '折讓日期',
}


@dealer_bp.route('/api/sales-return', methods=['POST'])
@sales_return_action_required('mod_sales_return_create')
def create_sales_return():
    data = request.json or {}
    for fk, label in _SALES_RETURN_REQUIRED_FIELDS.items():
        if not _norm(data.get(fk)):
            return jsonify({'error': f'{label}為必填'}), 400
    amount_incl = _to_num_or_none(data.get('amount_incl'))
    if amount_incl is None:
        return jsonify({'error': '折讓金額(含稅)為必填'}), 400
    items = [it for it in (data.get('items') or []) if _norm(it.get('report_product_name'))]
    if not items:
        return jsonify({'error': '財報料號品名為必填，請至少新增一列明細'}), 400

    err = _validate_sales_return_items(items)
    if err:
        return err

    amount_notax, tax_amount = _sales_return_calc_tax(amount_incl)
    items_total, diff = _sales_return_amount_diff(amount_incl, items)

    handler_name = _norm(data.get('handler_name')) or session.get('display_name', session.get('username', ''))
    rec = {
        'ar_doc_no': _norm(data.get('ar_doc_no')),
        'invoice_no': _norm(data.get('invoice_no')),
        'order_no': _norm(data.get('order_no')),
        'credit_note_no': _norm(data.get('credit_note_no')),
        'credit_date': data.get('credit_date'),
        'amount_incl': amount_incl,
        'amount_notax': amount_notax,
        'tax_amount': tax_amount,
        'note': data.get('note'),
        'recon_summary_no': data.get('recon_summary_no'),
        'recon_summary_range': data.get('recon_summary_range'),
        'handler_name': handler_name,
    }
    rec.update(_audit_new())
    try:
        res = sb.table('sales_return_list').insert(rec).execute()
    except Exception as e:
        return jsonify({'error': f'新增失敗：{e}'}), 500
    row = res.data[0]

    item_rows = [{
        'return_id': row['id'], 'report_product_name': _norm(it.get('report_product_name')),
        'amount': _to_num(it.get('amount')), 'sub_internal_code': _norm(it.get('sub_internal_code')) or None,
    } for it in items]
    sb.table('sales_return_items').insert(item_rows).execute()

    out = _sales_return_row_with_items(row, item_rows)
    if abs(diff) > 0.5:
        out['warning'] = f'明細金額加總（{items_total}）與折讓金額(含稅)（{amount_incl}）不相符，差額 {diff} 元，已照樣儲存，請自行確認是否需要修正'
    return jsonify(out), 201


@dealer_bp.route('/api/sales-return/<int:rid>', methods=['PUT'])
@sales_return_action_required('mod_sales_return_update')
def update_sales_return(rid):
    cur_res = sb.table('sales_return_list').select('*').eq('id', rid).execute()
    if not cur_res.data:
        return jsonify({'error': '找不到此筆銷退資料'}), 404
    cur = cur_res.data[0]
    data = request.json or {}

    merged = dict(cur)
    for fk in list(_SALES_RETURN_REQUIRED_FIELDS.keys()) + ['note', 'recon_summary_no', 'recon_summary_range', 'handler_name']:
        if fk in data:
            merged[fk] = data[fk]
    for fk, label in _SALES_RETURN_REQUIRED_FIELDS.items():
        if not _norm(merged.get(fk)):
            return jsonify({'error': f'{label}為必填'}), 400
    amount_incl = _to_num_or_none(data['amount_incl']) if 'amount_incl' in data else cur.get('amount_incl')
    if amount_incl is None:
        return jsonify({'error': '折讓金額(含稅)為必填'}), 400

    items = data.get('items')
    if items is None:
        items_res = sb.table('sales_return_items').select('*').eq('return_id', rid).execute()
        items = items_res.data or []
    items = [it for it in items if _norm(it.get('report_product_name'))]
    if not items:
        return jsonify({'error': '財報料號品名為必填，請至少保留一列明細'}), 400

    err = _validate_sales_return_items(items)
    if err:
        return err

    amount_notax, tax_amount = _sales_return_calc_tax(amount_incl)
    items_total, diff = _sales_return_amount_diff(amount_incl, items)

    rec = {
        'ar_doc_no': _norm(merged.get('ar_doc_no')),
        'invoice_no': _norm(merged.get('invoice_no')),
        'order_no': _norm(merged.get('order_no')),
        'credit_note_no': _norm(merged.get('credit_note_no')),
        'credit_date': merged.get('credit_date'),
        'amount_incl': amount_incl,
        'amount_notax': amount_notax,
        'tax_amount': tax_amount,
        'note': merged.get('note'),
        'recon_summary_no': merged.get('recon_summary_no'),
        'recon_summary_range': merged.get('recon_summary_range'),
        'handler_name': _norm(merged.get('handler_name')) or cur.get('handler_name'),
    }
    rec.update(_audit_upd())
    sb.table('sales_return_list').update(rec).eq('id', rid).execute()

    if 'items' in data:
        sb.table('sales_return_items').delete().eq('return_id', rid).execute()
        item_rows = [{
            'return_id': rid, 'report_product_name': _norm(it.get('report_product_name')),
            'amount': _to_num(it.get('amount')), 'sub_internal_code': _norm(it.get('sub_internal_code')) or None,
        } for it in items]
        if item_rows:
            sb.table('sales_return_items').insert(item_rows).execute()

    out = {'ok': True}
    if abs(diff) > 0.5:
        out['warning'] = f'明細金額加總（{items_total}）與折讓金額(含稅)（{amount_incl}）不相符，差額 {diff} 元，已照樣儲存，請自行確認是否需要修正'
    return jsonify(out)


@dealer_bp.route('/api/sales-return/<int:rid>', methods=['DELETE'])
@sales_return_action_required('mod_sales_return_delete')
def delete_sales_return(rid):
    cur = sb.table('sales_return_list').select('id').eq('id', rid).limit(1).execute()
    if not cur.data:
        return jsonify({'error': '找不到此筆銷退資料'}), 404
    # sales_return_items 有 ON DELETE CASCADE，刪除主資料時明細會一併清除，不需要另外處理。
    sb.table('sales_return_list').delete().eq('id', rid).execute()
    return jsonify({'ok': True})
